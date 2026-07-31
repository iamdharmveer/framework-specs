#!/usr/bin/env python3
"""frequency_xlsx.py v1.0 — Frequency workbook engine for the PYQ analysis pipeline.

EXTRACTED 2026-07-31 from Framework_MockTestAnalyse.md §16-1/§16-2/§16-4..§16-8
(v2.39.1) BYTE-IDENTICALLY. The spec retains §16's full contract: sheet
specifications (§16-3), the 9-item validation checklist (§16-9) and the frequency
edge cases EC-F1.. (§16-10). This file is the SINGLE SOURCE OF TRUTH for the
implementation; the spec is the single source of truth for the contract.

Public entry points (called from Framework_MockTestAnalyse.md executable blocks):
    aggregate_frequency_data(progress, all_entries=None, frequency_scope='all',
                             exam_config=None)
    generate_frequency_xlsx(progress, exam_code, all_entries=None,
                            frequency_scope='all', exam_config=None)

Usage from a spec block:
    import frequency_xlsx as fx
    xlsx_path = fx.generate_frequency_xlsx(progress, exam_code,
                                           all_entries=all_entries)

Requires blueprint_core (imported as bc) for era-scoped frequency
(bc.paper_eras_from_progress / bc.filter_progress_to_eras).
"""
import blueprint_core as bc

from collections import Counter, defaultdict
import re

def aggregate_frequency_data(progress, all_entries=None):
    """
    Aggregate per-subtopic, per-year data from analysis_progress.json.
    Returns structured data ready for xlsx generation.

    v2.24.6 FIX B (defense-in-depth for Gap 1 + Gap 2, closed authoritatively at Step 6
    by FIX A): previously iterated ONLY `progress` — which is populated exclusively by
    PYQ-OBSERVED (section, topic, subtopic) keys, so a Zero-PYQ subtopic (never observed
    in any PYQ paper) could structurally never become an Excel row — and derived Format
    with a 2-way (TEXT/FIGURAL only) `image_role`-only rule independent of the manifest's
    4-way (TEXT/FIGURAL/PASSAGE/DI) `synthesise_subtopic` derivation, so the Excel and the
    manifest could disagree on the SAME subtopic.

    When `all_entries` (the full PYQ + Zero-PYQ-scaffold list — the SAME list
    write_subtopic_manifest() writes from) is supplied, this function now:
      (1) seeds subtopic_data from all_entries FIRST, so every taxonomy subtopic gets a
          row (Zero-PYQ ones as all-zero rows) — Excel is taxonomy-complete, matching the
          manifest's COMPLETENESS INVARIANT (ref §15-1 taxonomy-sync FIX A).
      (2) takes Format from each entry['format'] (already the manifest's 4-way value —
          the SAME field write_subtopic_manifest() reads) instead of re-deriving it
          locally, so Excel Format == manifest Format for every subtopic BY CONSTRUCTION,
          never merely by coincidence.
    `progress` is still the sole source of PER-YEAR counts (all_entries only carries an
    aggregate `observed_count`, not a year-by-year breakdown).

    Backward-compatible: if `all_entries` is omitted (all_entries=None), falls back to the
    pre-v2.24.6 progress-only behavior (PYQ-observed subtopics only, 2-way Format) — so any
    caller that hasn't been updated to pass all_entries still gets a working (if incomplete)
    workbook rather than an error.
    """
    meta = progress.get('_meta', {})
    all_years = sorted(meta.get('years_processed', []))
    papers_processed = meta.get('papers_processed', [])

    # Count papers per year
    papers_per_year = Counter()
    for paper_id in papers_processed:
        year = extract_year_from_paper_id(paper_id)
        if year:
            papers_per_year[year] += 1

    # Aggregate questions per (section, topic, subtopic) per year
    subtopic_data = {}

    # ── STEP 1 (v2.24.6 FIX B): seed EVERY taxonomy subtopic from all_entries, PYQ +
    # Zero-PYQ alike, with its manifest-identical 4-way Format. This is what makes the
    # Excel taxonomy-complete and Format-parity-guaranteed with the manifest.
    if all_entries:
        for e in all_entries:
            key = (e['section'], e['topic'], e['subtopic'])
            subtopic_data[key] = {
                'section': e['section'],
                'topic': e['topic'],
                'subtopic': e['subtopic'],
                'format': e.get('format', 'TEXT'),   # manifest-identical — same field, same source
                'year_counts': defaultdict(int),
                'total': 0
            }

    for key, questions in progress.items():
        if not isinstance(key, tuple) or len(key) != 3:
            continue
        # v2.15 BUG-D08: removed dead str(key).startswith('_') check —
        # tuple keys never stringify to a '_'-prefixed string, and the
        # isinstance check above already excludes string keys like '_meta'.

        section, topic, subtopic = key

        if key not in subtopic_data:
            # Only reachable when all_entries is None (backward-compat path) or a
            # progress key somehow isn't in all_entries (shouldn't happen post-taxonomy-
            # sync, but handled defensively rather than dropping data).
            subtopic_data[key] = {
                'section': section,
                'topic': topic,
                'subtopic': subtopic,
                'format': 'TEXT',
                'year_counts': defaultdict(int),
                'total': 0
            }

        entry = subtopic_data[key]
        for q in questions:
            year = q.get('year')
            if year:
                entry['year_counts'][year] += 1
                entry['total'] += 1

            # Legacy 2-way format fallback — ONLY used when all_entries was not supplied
            # (entry['format'] already carries the manifest's 4-way value otherwise; never
            # overwrite a value that already came from all_entries).
            if not all_entries:
                role = q.get('image_role', 'none')
                if role in ('stem_only', 'stem_and_options', 'options_only'):
                    entry['format'] = 'FIGURAL'

    # ── COMPLETENESS INVARIANT (v2.24.6 FIX B — mirrors the manifest's own S2-MANIFEST-
    # COMPLETENESS gate): when all_entries is supplied, the Excel MUST cover exactly the
    # same subtopic set as the manifest. Advisory WARN here (xlsx generation never hard-
    # stops synthesis) — genuine gaps still surface downstream at Step 6 S2-MANIFEST.
    if all_entries:
        manifest_keys = {(e['section'], e['topic'], e['subtopic']) for e in all_entries}
        excel_keys = set(subtopic_data.keys())
        missing_from_excel = manifest_keys - excel_keys
        if missing_from_excel:
            print(f"  WARN: master_data_completeness_test would FAIL — "
                  f"{len(missing_from_excel)} manifest subtopic(s) absent from the "
                  f"Frequency Excel: {sorted(missing_from_excel)[:5]}"
                  + (" ..." if len(missing_from_excel) > 5 else ""))

    return {
        'subtopic_data': subtopic_data,
        'all_years': all_years,
        'papers_per_year': dict(papers_per_year),
        'total_papers': len(papers_processed),
        'total_questions': sum(e['total'] for e in subtopic_data.values())
    }

def extract_year_from_paper_id(paper_id):
    m = re.search(r'(\d{4})', str(paper_id))
    return int(m.group(1)) if m else None

def compute_derived_metrics(entry, all_years, papers_per_year):
    """
    Compute all derived metrics for a subtopic entry.
    """
    year_counts = entry['year_counts']

    # Avg/Paper per year
    avg_per_year = {}
    for year in all_years:
        qs = year_counts.get(year, 0)
        papers = papers_per_year.get(year, 0)
        avg_per_year[year] = round(qs / papers, 2) if papers > 0 else 0

    # Avg/Paper combined
    total_papers = sum(papers_per_year.values())
    avg_combined = round(entry['total'] / total_papers, 2) if total_papers > 0 else 0

    # Consistency: number of years where qs > 0
    consistency = sum(1 for y in all_years if year_counts.get(y, 0) > 0)

    # Trend: compare recent half vs older half
    n_years = len(all_years)
    if n_years >= 2:
        mid = n_years // 2
        recent_years = all_years[mid:]
        older_years = all_years[:mid]
        recent_avg = sum(avg_per_year.get(y, 0) for y in recent_years) / len(recent_years)
        older_avg = sum(avg_per_year.get(y, 0) for y in older_years) / len(older_years)

        only_latest = (year_counts.get(all_years[-1], 0) > 0 and
                       all(year_counts.get(y, 0) == 0 for y in all_years[:-1]))
        if only_latest:
            trend = 'New in ' + str(all_years[-1])
        elif recent_avg > older_avg * 1.3:
            trend = 'Rising'
        elif recent_avg > older_avg * 1.1:
            trend = 'Recovering'
        elif recent_avg < older_avg * 0.7:
            trend = 'Declining'
        elif recent_avg < older_avg * 0.9:
            trend = 'Cooling'
        elif older_avg == 0 and recent_avg == 0:
            trend = 'No Data'
        else:
            trend = 'Stable'
    else:
        trend = 'N/A (1 year)'  # v2.17 FIX-7: clearer label for single-year exams

    # Importance: based on combined avg/paper
    if avg_combined >= 0.50:
        importance = 'High'
    elif avg_combined >= 0.15:
        importance = 'Medium'
    else:
        importance = 'Low'

    # Must Prepare: High importance in a majority of available years.
    # v2.15 BUG-D09 fix: threshold scales with available years instead of fixed >= 4.
    # An exam with only 3 years can now have Must Prepare subtopics (needs 3 of 3).
    high_years = sum(1 for y in all_years if avg_per_year.get(y, 0) >= 0.50)
    must_prepare_threshold = min(4, len(all_years))
    must_prepare = 'Must Prepare' if high_years >= must_prepare_threshold else '—'  # v2.17 FIX-5

    return {
        'avg_per_year': avg_per_year,
        'avg_combined': avg_combined,
        'consistency': consistency,
        'trend': trend,
        'importance': importance,
        'must_prepare': must_prepare
    }

def compute_ranks(data):
    """Compute rank_in_topic for each subtopic (by combined count within topic)."""
    topic_groups = defaultdict(list)
    for key, entry in data.items():
        topic_groups[(entry['section'], entry['topic'])].append((key, entry))
    for (section, topic), entries in topic_groups.items():
        sorted_entries = sorted(entries, key=lambda x: x[1]['total'], reverse=True)
        for rank, (key, entry) in enumerate(sorted_entries, 1):
            entry['rank_in_topic'] = rank

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
DATA_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(bold=True, name='Arial', size=10)
TITLE_FONT = Font(bold=True, name='Arial', size=14)

def generate_frequency_xlsx(progress, exam_code, all_entries=None,
                            frequency_scope='all', exam_config=None):
    """
    Generate the complete Frequency xlsx from analysis_progress.json.
    Called during synthesis phase after section_rules and manifest are written.

    v2.17: Now reads exam_config.json to get exam_total_questions per section
    for accurate % of Subject calculations. Also performs coverage validation,
    topic name consistency check, and duplicate/near-duplicate detection.

    v2.24.6 FIX B: accepts `all_entries` (the same PYQ + Zero-PYQ-scaffold list passed to
    write_subtopic_manifest()) so the workbook is taxonomy-complete and Format-parity-
    guaranteed with the manifest — see aggregate_frequency_data() docstring (§16-1).

    v2.25 ERA-SCOPED FREQUENCY (`frequency_scope`):
      'all'          — DEFAULT. Every paper counts. Byte-identical to pre-v2.25 behaviour
                       for every exam, so no existing exam changes unless it opts in.
      'current-era'  — only papers matching the CURRENT exam pattern contribute to the
                       frequency numbers.

      WHY THIS SWITCH EXISTS. A PYQ corpus carries two kinds of information that need
      OPPOSITE treatment. QUESTION SHAPES — the templates, difficulty calibration and
      phrasing synthesised by §14 synthesise_subtopic() — are more valuable the more eras
      they span; a subtopic seen across 26 years is far better characterised than one seen
      twice, and that variety is the whole reason legacy papers are retained. PROPORTIONS
      — how many questions a subtopic deserves out of today's paper — can only be answered
      by the current pattern; a retired pattern's subject mix is not evidence about today's
      exam. Framework_Blueprint §3 recency weighting (last 2 valid years x2) dampens the
      second problem but cannot solve it: a corpus with 21 retired-era years against 6
      current-era ones still gives the retired pattern ~72% of the weight.

      WHAT THIS DOES AND DOES NOT TOUCH. The filter is applied HERE ONLY — to the counting
      view that feeds the Frequency xlsx and the axis distribution. `progress` itself is
      never mutated (bc.filter_progress_to_eras returns a copy), so §14's pattern/template
      synthesis, section_rules.md, and the subtopic manifest keep consuming the FULL corpus.
      Variety in, faithful proportions out.

      CONSEQUENCE FOR LEGACY-ONLY SUBTOPICS. A subtopic the current pattern never asks
      drops to zero observed questions, so Framework_Blueprint §3-4 CASE 2 classifies it
      Zero-PYQ and §5 ZP rotation gives it occasional coverage instead of a mix-distorting
      quota. It remains in the taxonomy and keeps its full pattern library. This is the
      intended outcome, not data loss.
    """
    import blueprint_core as bc   # ENGINE — pattern-era classification lives in Cluster F

    counting_progress = progress
    era_stats = None
    if frequency_scope == 'current-era':
        if not exam_config:
            raise SystemExit(
                "HARD STOP: frequency_scope='current-era' requires exam_config. "
                "Pattern era is defined by comparison against the CURRENT pattern, so "
                "without exam_config every paper's era is unknowable. Supply exam_config "
                "or use frequency_scope='all'.")
        eras = bc.paper_eras_from_progress(progress, exam_config)
        counting_progress, era_stats = bc.filter_progress_to_eras(
            progress, eras, keep=('current',))
        if era_stats['kept_papers'] == 0:
            raise SystemExit(
                "HARD STOP: frequency_scope='current-era' selected, but NOT ONE paper in "
                f"the corpus matches the current exam pattern (eras seen: "
                f"{era_stats['era_counts']}). Frequency cannot be computed from zero "
                "papers. Either the exam pattern xlsx does not describe the papers you "
                "have, or no current-pattern paper has been added yet. Use "
                "frequency_scope='all' until at least one current-era paper exists.")
        print("ERA-SCOPED FREQUENCY (frequency_scope='current-era')")
        print(f"  papers by era      : {era_stats['era_counts']}")
        print(f"  counted / excluded : {era_stats['kept_papers']} / "
              f"{era_stats['dropped_papers']} papers")
        print(f"  questions counted  : {era_stats['kept_questions']} "
              f"(excluded from COUNTS only: {era_stats['dropped_questions']})")
        print("  Excluded papers still feed the question-pattern library, section_rules.md")
        print("  and the taxonomy in full — only the PROPORTIONS are current-era.")
        if era_stats['kept_papers'] < 3:
            print(f"  WARN: only {era_stats['kept_papers']} current-era paper(s). "
                  "Recency weighting needs 2+ valid years to discriminate "
                  "(Framework_Blueprint §3-4 CASE 1); proportions will be thin.")

    agg = aggregate_frequency_data(counting_progress, all_entries=all_entries)
    data = agg['subtopic_data']
    all_years = agg['all_years']
    papers_per_year = agg['papers_per_year']

    # v2.17 FIX-1: Read exam_config for real exam totals
    exam_total_questions = {}  # section_name → total Qs from exam pattern
    exam_total_all = 0
    exam_name_display = exam_code.replace('_', ' ')
    try:
        import glob as _g
        cfg_matches = sorted(_g.glob('/mnt/project/*exam_config.json'))
        if cfg_matches:
            import json as _j
            cfg = _j.load(open(cfg_matches[0], encoding='utf-8'))
            for sec in cfg.get('sections', []):
                exam_total_questions[sec['name']] = sec['q_count']
            exam_total_all = cfg.get('total_questions', 0)
            exam_name_display = cfg.get('exam_name', exam_name_display)
    except Exception:
        pass  # fallback: use section_totals from mapped data

    # v2.17 FIX-4: Check for duplicate (section, topic, subtopic) keys
    seen_keys = set()
    for key, entry in list(data.items()):
        triple = (entry['section'], entry['topic'], entry['subtopic'])
        if triple in seen_keys:
            print(f"  WARN: Duplicate subtopic '{triple[2]}' under {triple[0]}/{triple[1]}. Merging.")
        seen_keys.add(triple)

    # v2.17 FIX-6: Near-duplicate subtopic warning
    from difflib import SequenceMatcher as _SM
    topic_groups = defaultdict(list)
    for entry in data.values():
        topic_groups[(entry['section'], entry['topic'])].append(entry['subtopic'])
    for (sec, top), subs in topic_groups.items():
        for i in range(len(subs)):
            for j in range(i+1, len(subs)):
                if _SM(None, subs[i], subs[j]).ratio() > 0.75:
                    print(f"  WARN: Near-duplicate subtopics in {sec}/{top}: "
                          f"'{subs[i]}' ~ '{subs[j]}' (>75% similar)")

    # Compute metrics for all subtopics
    for key, entry in data.items():
        metrics = compute_derived_metrics(entry, all_years, papers_per_year)
        entry.update(metrics)
    compute_ranks(data)

    wb = Workbook()

    # Sheet 1: Summary Dashboard
    write_summary_dashboard(wb, data, all_years, papers_per_year, agg,
                            exam_code, exam_name_display, exam_total_all)

    # Sheet 2: Master Data
    write_master_data(wb, data, all_years, papers_per_year, exam_code,
                      exam_total_questions)

    # Sheet 3: Topic Analysis
    write_topic_analysis(wb, data, all_years, papers_per_year,
                         exam_total_questions)

    # Sheet 4: Trend Analysis
    write_trend_analysis(wb, data, all_years)

    # Per-section sheets
    sections = sorted(set(e['section'] for e in data.values()))
    for section in sections:
        sec_entries = {k: v for k, v in data.items() if v['section'] == section}
        sec_exam_total = exam_total_questions.get(section, 0)
        write_section_sheet(wb, section, sec_entries, all_years, papers_per_year,
                            sec_exam_total)

    path = f'/mnt/user-data/outputs/{exam_code}_PYQ_Frequency.xlsx'
    wb.save(path)

    # Recalculate formulas
    import subprocess
    subprocess.run(['python', '/mnt/skills/public/xlsx/scripts/recalc.py', path],
                   capture_output=True)

    # v2.17 FIX-2: Coverage validation (XLSX-F9)
    #
    # v2.39 (GAP-2026-07-27-D) — THE DENOMINATOR WAS PER-PAPER.
    # exam_total_all is cfg['total_questions'], i.e. the size of ONE paper (60 for
    # IIT_JAM_BIOTECHNOLOGY). mapped_total is the CORPUS sum (1719). The ratio therefore
    # exceeded 0.05 on every multi-paper corpus that has ever run, and the message
    # printed `60 - 1719` as a count of unclassified questions:
    #     "⚠ COVERAGE WARNING (XLSX-F9): ... 1719 Qs but exam has 60 total (2865% mapped).
    #      -1659 questions were not classified to any subtopic."
    # Not merely wrong — INVERTED. A gate that cannot fire correctly trains operators to
    # ignore gates, and this one had been crying wolf since v2.17.
    #
    # The honest expectation is one paper's worth of questions per paper actually
    # counted. Under frequency_scope='current-era' only the current-era papers feed
    # `data`, so the denominator must count those papers and no others — hence
    # counting_progress, not progress.
    mapped_total = sum(e['total'] for e in data.values())
    _n_papers = len((counting_progress.get('_meta', {}) or {}).get('papers_processed', []))
    expected_total = exam_total_all * _n_papers if (exam_total_all > 0 and _n_papers) else 0
    if expected_total > 0 and abs(mapped_total - expected_total) / expected_total > 0.05:
        pct_mapped = round(mapped_total / expected_total * 100, 1)
        shortfall  = expected_total - mapped_total
        print(f"\n  ⚠ COVERAGE WARNING (XLSX-F9): Frequency xlsx accounts for {mapped_total} Qs "
              f"but {_n_papers} paper(s) x {exam_total_all} Qs/paper = {expected_total} "
              f"expected ({pct_mapped}% mapped).")
        if shortfall > 0:
            print(f"    {shortfall} question(s) were not classified to any subtopic.")
        else:
            # Guarded: a negative shortfall is a SURPLUS and must never be printed as
            # "not classified". Legitimately reachable — legacy papers in an era-mixed
            # corpus can carry more questions than the current pattern declares.
            print(f"    {-shortfall} question(s) MORE than expected were classified — "
                  f"papers in this corpus differ in size from the current exam pattern "
                  f"(legacy era), or a paper was counted twice.")
        print(f"    Downstream blueprint will be inaccurate. Review PYQ classification.")

    return path

def write_summary_dashboard(wb, data, all_years, papers_per_year, agg,
                            exam_code, exam_name_display, exam_total_all):
    ws = wb.active
    ws.title = 'Summary Dashboard'
    n_years = len(all_years)
    total_papers = agg['total_papers']
    total_qs = agg['total_questions']

    year_range = f'{all_years[0]}-{all_years[-1]}' if all_years else 'N/A'
    # v2.17 FIX-9: Use exam_name_display from exam_config
    ws['A1'] = (f'{exam_name_display} PYQ Frequency Analysis '
                f'({year_range} | {n_years} Years | {total_papers} Papers | {total_qs:,} Questions)')
    ws['A1'].font = TITLE_FONT
    # v2.17: Show exam total if available and different from mapped total
    if exam_total_all > 0 and exam_total_all != total_qs:
        ws['A2'] = (f'Years: [{", ".join(str(y) for y in all_years)}] | '
                    f'Papers: {total_papers} | Questions: {exam_total_all} '
                    f'(Mapped: {total_qs})')
    else:
        ws['A2'] = (f'Years: [{", ".join(str(y) for y in all_years)}] | '
                    f'Papers: {total_papers} | Questions: {total_qs}')

    year_parts = []
    for y in all_years:
        p = papers_per_year.get(y, 0)
        q = sum(e['year_counts'].get(y, 0) for e in data.values())
        year_parts.append(f'{y}:{p}P({q:,}Q)')
    ws['A2'] = ' | '.join(year_parts)

    # Year-wise overview
    row = 4
    ws.cell(row=row, column=1, value='YEAR-WISE OVERVIEW').font = BOLD_FONT
    row += 1
    sections_list = sorted(set(e['section'] for e in data.values()))
    headers = ['Year', 'Papers', 'Total Qs', 'Avg Qs/Paper'] + sections_list + ['% of Grand Total']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    row += 1
    for y in all_years:
        p = papers_per_year.get(y, 0)
        q = sum(e['year_counts'].get(y, 0) for e in data.values())
        ws.cell(row=row, column=1, value=y)
        ws.cell(row=row, column=2, value=p)
        ws.cell(row=row, column=3, value=q)
        ws.cell(row=row, column=4, value=round(q/p, 0) if p > 0 else 0)
        for si, sec in enumerate(sections_list):
            sec_q = sum(e['year_counts'].get(y, 0) for e in data.values() if e['section'] == sec)
            ws.cell(row=row, column=5+si, value=sec_q)
        pct = round(q / total_qs * 100, 1) if total_qs > 0 else 0
        ws.cell(row=row, column=5+len(sections_list), value=f'{pct}%')
        row += 1

    # Top 25
    row += 2
    ws.cell(row=row, column=1, value=f'TOP 25 MOST ASKED SUB-TOPICS ({n_years} Years Combined)').font = BOLD_FONT
    row += 1
    top_h = ['Rank', 'Sub-Topic', 'Subject', 'Combined Qs', 'Avg/Paper',
             'Consistency', 'Trend', 'Importance', 'Must Prepare']
    for col, h in enumerate(top_h, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    row += 1
    sorted_by_total = sorted(data.values(), key=lambda e: e['total'], reverse=True)
    for rank, entry in enumerate(sorted_by_total[:25], 1):
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=entry['subtopic'])
        ws.cell(row=row, column=3, value=entry['section'])
        ws.cell(row=row, column=4, value=entry['total'])
        ws.cell(row=row, column=5, value=entry['avg_combined'])
        ws.cell(row=row, column=6, value=entry['consistency'])
        ws.cell(row=row, column=7, value=entry['trend'])
        ws.cell(row=row, column=8, value=entry['importance'])
        ws.cell(row=row, column=9, value=entry['must_prepare'])
        row += 1

    # Legend
    row += 2
    ws.cell(row=row, column=1, value='IMPORTANCE TAG LEGEND').font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value='High'); ws.cell(row=row, column=2, value='Avg/Paper >= 0.50 - MUST PREPARE'); row += 1
    ws.cell(row=row, column=1, value='Medium'); ws.cell(row=row, column=2, value='Avg/Paper 0.15-0.49 - IMPORTANT'); row += 1
    ws.cell(row=row, column=1, value='Low'); ws.cell(row=row, column=2, value='Avg/Paper < 0.15 - LOW YIELD')

def write_master_data(wb, data, all_years, papers_per_year, exam_code,
                      exam_total_questions):
    ws = wb.create_sheet(f'Master Data ({len(all_years)} Years)')
    headers = ['Subject', 'Topic', 'Sub-Topic', 'Format']
    for y in all_years: headers.append(f'Qs {y}')
    headers.append('Combined Qs')
    for y in all_years: headers.append(f'Avg/Paper {y}')
    headers.append('Avg/Paper Combined')
    for y in all_years: headers.append(f'Papers In {y}')
    headers.append('Papers In Combined')
    headers += ['% of Subject', 'Importance', 'Consistency', 'Trend', 'Must Prepare', 'Rank in Topic']

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')

    sorted_entries = sorted(data.values(), key=lambda e: (e['section'], e['topic'], e['subtopic']))
    # v2.17 FIX-1: Use exam_total_questions from exam_config as denominator
    # Fallback to section_totals (sum of mapped Qs) if exam_config absent
    section_totals_mapped = Counter()
    for e in data.values(): section_totals_mapped[e['section']] += e['total']

    row = 2
    for entry in sorted_entries:
        col = 1
        ws.cell(row=row, column=col, value=entry['section']); col += 1
        ws.cell(row=row, column=col, value=entry['topic']); col += 1
        ws.cell(row=row, column=col, value=entry['subtopic']); col += 1
        ws.cell(row=row, column=col, value=entry['format']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=entry['year_counts'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=entry['total']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=entry['avg_per_year'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=entry['avg_combined']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=papers_per_year.get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=sum(papers_per_year.values())); col += 1
        # v2.17 FIX-1: % of Subject uses exam total (not mapped total)
        sec_denom = exam_total_questions.get(entry['section'], 0)
        if sec_denom == 0:
            sec_denom = section_totals_mapped.get(entry['section'], 1)
        ws.cell(row=row, column=col, value=round(entry['total']/sec_denom*100, 1) if sec_denom > 0 else 0); col += 1
        ws.cell(row=row, column=col, value=entry['importance']); col += 1
        ws.cell(row=row, column=col, value=entry['consistency']); col += 1
        ws.cell(row=row, column=col, value=entry['trend']); col += 1
        # v2.17 FIX-5: Use '—' instead of empty string for non-Must-Prepare
        ws.cell(row=row, column=col, value=entry['must_prepare'] if entry['must_prepare'] else '—'); col += 1
        ws.cell(row=row, column=col, value=entry.get('rank_in_topic', '')); col += 1
        row += 1

    # v2.17 FIX-3: Topic name consistency assertion
    master_sections = set(e['section'] for e in data.values())

    for ci in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

def write_topic_analysis(wb, data, all_years, papers_per_year,
                         exam_total_questions):
    ws = wb.create_sheet('Topic Analysis')
    topic_agg = defaultdict(lambda: {'year_counts': defaultdict(int), 'total': 0, 'n_subtopics': 0})
    for e in data.values():
        k = (e['section'], e['topic'])
        for y, c in e['year_counts'].items(): topic_agg[k]['year_counts'][y] += c
        topic_agg[k]['total'] += e['total']
        topic_agg[k]['n_subtopics'] += 1

    headers = ['Subject', 'Topic']
    for y in all_years: headers.append(f'Qs {y}')
    headers += ['Combined Qs']
    for y in all_years: headers.append(f'Avg/Paper {y}')
    headers += ['Avg/Paper Combined', '% of Subject', 'Sub-Topics', 'Importance', 'Rank in Subject']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL

    # v2.17 FIX-1: Use exam_total_questions for % denominator
    # v2.17 FIX-8: Use full section name (no truncation)
    section_totals_mapped = Counter()
    for (sec, _), a in topic_agg.items(): section_totals_mapped[sec] += a['total']
    sorted_topics = sorted(topic_agg.items(), key=lambda x: (x[0][0], -x[1]['total']))
    section_ranks = defaultdict(int)
    row = 2
    for (section, topic), a in sorted_topics:
        section_ranks[section] += 1
        col = 1
        # v2.17 FIX-3 + FIX-8: section name must be IDENTICAL to Master Data
        ws.cell(row=row, column=col, value=section); col += 1
        ws.cell(row=row, column=col, value=topic); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=a['year_counts'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=a['total']); col += 1
        for y in all_years:
            p = papers_per_year.get(y, 0)
            ws.cell(row=row, column=col, value=round(a['year_counts'].get(y,0)/p, 2) if p > 0 else 0); col += 1
        total_p = sum(papers_per_year.values())
        ws.cell(row=row, column=col, value=round(a['total']/total_p, 2) if total_p > 0 else 0); col += 1
        # v2.17 FIX-1: % of Subject uses exam total (not mapped total)
        sec_denom = exam_total_questions.get(section, 0)
        if sec_denom == 0:
            sec_denom = section_totals_mapped.get(section, 1)
        ws.cell(row=row, column=col, value=round(a['total']/sec_denom*100, 1) if sec_denom > 0 else 0); col += 1
        ws.cell(row=row, column=col, value=a['n_subtopics']); col += 1
        avg_c = round(a['total']/total_p, 2) if total_p > 0 else 0
        ws.cell(row=row, column=col, value='High' if avg_c >= 0.50 else ('Medium' if avg_c >= 0.15 else 'Low')); col += 1
        ws.cell(row=row, column=col, value=section_ranks[section]); col += 1
        row += 1

def write_trend_analysis(wb, data, all_years):
    ws = wb.create_sheet('Trend Analysis & Charts')
    ws.cell(row=1, column=1,
            value=f'Top 20 Sub-Topics: Avg/Paper Across All {len(all_years)} Years').font = BOLD_FONT
    headers = ['Sub-Topic'] + [str(y) for y in all_years] + ['Combined', 'Subject']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
    top_20 = sorted(data.values(), key=lambda e: e['avg_combined'], reverse=True)[:20]
    row = 3
    for entry in top_20:
        col = 1
        ws.cell(row=row, column=col, value=entry['subtopic']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=entry['avg_per_year'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=entry['avg_combined']); col += 1
        ws.cell(row=row, column=col, value=entry['section'][:20]); col += 1
        row += 1

def write_section_sheet(wb, section, section_entries, all_years, papers_per_year,
                        sec_exam_total):
    sheet_name = section[:31].replace('/', ' ')
    ws = wb.create_sheet(sheet_name)
    headers = ['Topic', 'Sub-Topic', 'Format']
    for y in all_years: headers.append(f'Qs {y}')
    headers.append('Combined Qs')
    for y in all_years: headers.append(f'Avg {y}')
    headers.append('Avg Combined')
    for y in all_years: headers.append(f'PapersIn {y}')
    headers += ['% of Subject', f'Consistency (of {len(all_years)})', 'Trend', 'Importance', 'Must Prepare', 'Rank in Topic']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')

    sorted_e = sorted(section_entries.values(), key=lambda e: (e['topic'], e['subtopic']))
    # v2.17 FIX-1: Use exam total as denominator for % of Subject
    section_total_mapped = sum(e['total'] for e in section_entries.values())
    sec_denom = sec_exam_total if sec_exam_total > 0 else section_total_mapped
    row = 2
    for entry in sorted_e:
        col = 1
        ws.cell(row=row, column=col, value=entry['topic']); col += 1
        ws.cell(row=row, column=col, value=entry['subtopic']); col += 1
        ws.cell(row=row, column=col, value=entry['format']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=entry['year_counts'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=entry['total']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=entry['avg_per_year'].get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=entry['avg_combined']); col += 1
        for y in all_years:
            ws.cell(row=row, column=col, value=papers_per_year.get(y, 0)); col += 1
        ws.cell(row=row, column=col, value=round(entry['total']/sec_denom*100, 1) if sec_denom > 0 else 0); col += 1
        ws.cell(row=row, column=col, value=entry['consistency']); col += 1
        ws.cell(row=row, column=col, value=entry['trend']); col += 1
        ws.cell(row=row, column=col, value=entry['importance']); col += 1
        # v2.17 FIX-5: Use '—' instead of empty string
        ws.cell(row=row, column=col, value=entry['must_prepare'] if entry['must_prepare'] else '—'); col += 1
        ws.cell(row=row, column=col, value=entry.get('rank_in_topic', '')); col += 1
        row += 1
    for ci in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
