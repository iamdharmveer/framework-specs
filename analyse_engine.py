#!/usr/bin/env python3
"""analyse_engine.py v1.0 — Step 5 (PYQExtract) shared extraction engine.

v2.57 stamps (2026-09-02, SYLLABUS-TRANSITION Release B): the three emitted
provenance stamps track Framework_MockTestAnalyse v2.57 per its CROSS-FILE
SYNC RULE (MINOR bump => engine bump; MS-3 STAMP-PARITY enforces).

WHAT THIS IS. The §2 universal extraction primitives (E-1 .. E-11) and the §4 vision
aggregation helpers of Framework_MockTestAnalyse.md, moved here VERBATIM. The spec
retains every contract, edge case, DoD item and word of narrative; this file holds the
code. Same treatment §16 received at v2.39.2, when its implementation moved to
frequency_xlsx.py and its sections kept the contract.

WHY. Framework_MockTestAnalyse.md carried 6,667 lines of executable python that the
model READ INTO CONTEXT and then RE-EMITTED as code. Every other mature pipeline in
this repo — blueprint_core, corpus_io, notes_core, frequency_xlsx — puts that in an
engine that is imported and never read. Step 5 never got the treatment, and the cost was
not merely tokens:

  A model that re-emits code repairs it on the way past. GAP-2026-08-16-STEP5-SYNTHESIS-
  UNRUNNABLE found `collections` used in §5 and imported nowhere in the file, live for
  ten days; GAP-2026-08-16-BASELINE-SUPPRESSED-NAMEERRORS found process_pyq_paper()
  raising NameError on every paper of every exam. Both survived because each session
  silently supplied what the spec lacked, succeeded, and wrote nothing down. An engine
  that is imported cannot be repaired mid-flight. It fails once, loudly, in CI, and is
  fixed at source for all 200 exams at once.

HOW THE BOUNDARY WAS CHOSEN — not by section number, by the call graph. §2+§4 is the
subset that closes: it needs NOTHING from any other extractable section. §2+§3+§4 does
not close (§3's process_pyq_paper calls §5's tag_axes), and §3+§5+§6 needs 16 names from
§2. So §2+§4 ships first and §3+§5+§6 follows, each a set that can be asserted whole.

WHAT MOVED AND WHAT DID NOT. Every top-level node of the §2 and §4 python fences is a
DEFINITION — measured: 33 definitions, ZERO top-level session-flow statements. That is
why these fences could move intact rather than being picked apart. §1 (session start,
40 defs / 31 flow) and §8 (batch orchestration, 20 defs / 1 flow) carry the CLASS: T
tool calls and the session flow, and stay in the spec permanently.

CONTRACT. This module is IMPORTED AND EXECUTED, NEVER READ as part of a session's spec
budget (bootstrap.py verifies its checksum; SKILL Rule 2 does not route .py files).
Callers bind the 20 public names listed in __all__ via the import stubs that replaced
the fences. The 11 remaining names are engine-internal.

Stdlib + repo engines only. No I/O at import time.
"""
import json
import os
import re
import sys
from collections import Counter
from difflib import SequenceMatcher
import hashlib
import math
import unicodedata

import blueprint_core as bc      # Cluster H — pure acquisition/image decisions
import corpus_io                 # I/O shell — Drive fetch, image integrity, governor

# WHY THESE ARE DECLARED HERE AND WERE NOT FREE NAMES IN THE SPEC.
# In Framework_MockTestAnalyse.md the §1 fence imports bc / corpus_io / SequenceMatcher
# at module scope, and every later fence inherits them from the shared session
# namespace. A free-name analysis of §2+§4 ALONE therefore reports them as bound and
# says nothing is missing — which is exactly the shape of GAP-2026-08-16-STEP5-
# SYNTHESIS-UNRUNNABLE D2, where §5 used `collections` that no fence in the file ever
# imported and the defect hid for ten days behind the same inheritance.
# An engine has its own globals and inherits nothing. Every dependency is named here,
# at the top, where a reader and an auditor can both see it. That is the point of
# moving the code, not a side effect of it.

__all__ = [
    # §2 — universal extraction primitives (E-1 .. E-11)
    'detect_question_start', 'is_option', 'clean_option_text',
    'extract_and_map_images', 'enrich_paragraph_with_omml',
    'extract_note_block', 'classify_note_frequency', 'canonical_note_text',
    'detect_linked_groups', 'classify_option_format', 'subtopic_option_format',
    'determine_strip_mode', 'generate_templates',
    'classify_wrong_option_structure',
    # §4 — vision aggregation
    'get_vision_candidates', 'apply_vision_observations', 'aggregate_figural',
    'VISION_PER_SHEET', 'VISION_WORKDIR',
]


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1624-1667 (v2.53.1) — VERBATIM ═══
Q_PATTERNS = [
    r'^Q\.\s*(\d+)\s+',            # Q.1  Q.25  Q. 1
    r'^Q(\d+)\.\s+',               # Q1.  Q25.
    r'^Q\.\s*(\d+)\s*$',           # Q.4   bare label — OMML / figure / empty stem
    r'^Q(\d+)\.\s*$',              # Q4.   bare label, alt form
]

# DELEGATED to the engine (blueprint_core Cluster G). Four specs parse Q-numbers from the
# same documents and must agree exactly; a local copy in any one of them is drift waiting to
# happen. This table mirrors the engine's canonical table EXACTLY and is verified by
# audit_deep.py TABLE-PARITY.
#
# WHY THESE FOUR AND NO MORE (2026-07-25, extended 2026-08-15).
#
# Entries 1-2 are the WITH-CONTENT forms. Three further forms exist in RAW exam sources —
# "Question 1:", bare "1." and "(1)" — and Step 1 detects them via its own
# SOURCE_Q_PATTERNS. They are deliberately ABSENT here and must never be restored. After
# Step 1 every document is NORMALISED: questions read "Q.N" and OPTIONS read "N. text".
# The bare-number pattern therefore matches every option line. Verified by execution on a
# canonical two-question fixture: the two-pattern table finds 2 question starts; the
# five-pattern table finds 10. A 100-question paper would parse as 500.
# Until 2026-07-25 these tables carried all five entries while the engine implemented two,
# and audit_deep TABLE-PARITY could not see it: its extraction regex stopped at the first
# ']', which occurs inside r'^Question\s+(\d+)\s*[:.]', so it compared a silently truncated
# two-entry slice against the engine's two and always passed.
#
# Entries 3-4 are the BARE-LABEL forms (GAP-2026-08-15-BAREQ). python-docx's p.text is
# <w:t>-only, so a stem paragraph whose entire payload is <m:oMath>, a drawing, or nothing
# at all (PYQPrepare S1-4 "empty/corrupt") reads as just "Q.N" — and entries 1-2 require
# whitespace AFTER the digits, applied to already-stripped text, so they can never match.
# Such a question DID NOT EXIST for this parser: its stem, its options and its date label
# were absorbed into the preceding question's body, and every check below passed because
# input and output are counted with the SAME blind detector. Measured on
# IIT_JAM_MATHEMATICS 12-Feb-2017: 4 of 60 questions (Q.4, Q.6, Q.25, Q.27) lost.
# This is the QUESTION half of GAP-2026-08-07-OMML, whose OPTION half shipped as
# corpus_io.BARE_OPT_PATTERNS + is_option(para=); OPT_PATTERNS had a bare-label companion
# and Q_PATTERNS did not.
# The $ anchor is LOAD-BEARING: it admits ONLY a paragraph that is nothing but the label,
# so it can never match an option line (options never begin with Q), an in-passage
# cross-reference "Q.11-15", a date label, a heading, or "Q1 Analysis". Verified by
# execution over a 34-case adversarial fixture in blueprint_core.self_test().
detect_question_start = bc.detect_question_start


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1671-1688 (v2.53.1) — VERBATIM ═══
# ── OPTION PREDICATE — DELEGATED (v2.36, audit_deep [XSPEC-DRIFT]) ────────────
# is_option() was defined in THREE specs (here, PYQSort, PYQAnalyse), each with a
# docstring claiming alignment with the others. v2.34/v2.35 added the image-option
# path HERE only, so those claims became false and PYQSort — which uses the predicate
# to COUNT options — kept undercounting image options. Measured on
# IIT_JAM_BIOTECHNOLOGY 2022: 156 options counted, 160 actual.
# corpus_io now owns the single definition. Do NOT re-localise any of these names.
OPT_PATTERNS      = corpus_io.OPT_PATTERNS
BARE_OPT_PATTERNS = corpus_io.BARE_OPT_PATTERNS
para_has_image    = corpus_io.para_has_image
is_option         = corpus_io.is_option
clean_option_text = corpus_io.clean_option_text

# After collecting options per Q: options = options[:options_count]
# If still < options_count: try E-5 OMML and E-4 image check.
# If still missing: q_incomplete=True, exclude from template extraction.


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1692-1767 (v2.53.1) — VERBATIM ═══
from docx import Document
import os

def extract_and_map_images(doc, paper_id, docx_path, expected_size=None):
    """
    BUG-A03 fix: accepts already-opened Document object, not doc_path.
    Caller (S3-1) opens Document once and passes it here.

    v2.34 — SINGLE GATED PATH. The legacy DOM-only branch is DELETED and docx_path is
    now a REQUIRED POSITIONAL. GAP-2026-07-26-002 DEFECT-1/2/3/5:

      * DEFECT-1. docx_path was optional and the ONLY call site in the entire framework
        (S3-1) never passed it, so `if docx_path:` was always False and every Step-5 run
        on every exam silently took the branch this file itself labelled UNGATED. The
        docstring documented a requirement nothing satisfied. A fallback that no caller
        ever avoids is not a safety net — it IS the default path.
      * DEFECT-2. The two branches returned THREE values each in DIFFERENT orders with
        DIFFERENT meanings — (img_map, imap, q_roles) vs (imap, q_roles, report). Equal
        arity means the interpreter raises nothing; the unpack simply mis-binds and
        q_roles becomes the report dict, so every .get(num,{}).get('role') returns the
        'none' default. There is now ONE return shape: (image_map, q_roles, report).
      * DEFECT-3. The gated branch built {'stem','opts'} but never derived 'role', the
        documented contract read by S3-1 (image_role) and S4-1 (get_vision_candidates).
        Role derivation is DELEGATED to bc.derive_image_roles() so one rule exists in
        one place — the same contract audit_deep.py enforces for detect_question_start.
      * DEFECT-5. expected_size was captured by enumeration and threaded nowhere, so
        IMG-1 evaluated against None and SKIPped on every paper, permanently. It is the
        only check that catches a payload truncated at a ZIP member boundary.

    Measured on the 22-paper IIT_JAM_BIOTECHNOLOGY corpus: this path maps the identical
    329 images the legacy walk did (that corpus carries no table/VML/header images —
    Step 1 normalisation removes them), but assigns the CORRECT role to the image-option
    questions the legacy walk mis-typed as stem_only. All 22 papers PASS IMG-1..IMG-5;
    zero hard stops.

    docx_path is required: the gates read the PACKAGE, not the DOM.
    """
    img_dir = f'/home/claude/pyq_images/{paper_id}'
    extracted = corpus_io.extract_images(docx_path, img_dir)
    mapping   = corpus_io.map_images_to_questions(docx_path)
    verdicts, stats = corpus_io.verify_images(
        docx_path, extracted=extracted, mapping=mapping,
        expected_size=expected_size, workdir=img_dir)

    if not bc.gates_passed(verdicts):
        failed = {k: v for k, v in verdicts.items() if not str(v).startswith(('PASS', 'SKIP'))}
        raise SystemExit(
            f"HARD STOP — image integrity gate failed for {paper_id}:\n  "
            + "\n  ".join(f"{k}: {v}" for k, v in failed.items())
            + "\n\nAn image that is present in the document but missing from the "
              "mapping becomes a question classified TEXT instead of FIGURAL, which "
              "silently corrupts the format distribution used by Step 7. Resolve the "
              "document before continuing; do not process it partially.")

    if stats['preamble']:
        print(f"    note: {stats['preamble']} image(s) before Q.1 (not question figures)")
    if stats['header_footer']:
        print(f"    note: {stats['header_footer']} header/footer image(s) (not question figures)")
    if stats['vector']:
        print(f"    note: {stats['vector']} vector part(s) — rasterised before view()")

    image_map = []
    for q_num, parts in mapping.items():
        if q_num == corpus_io.PREAMBLE:
            continue
        for i, base in enumerate(parts):
            rec = extracted.get(base, {})
            image_map.append({'img_idx': i, 'q_num': q_num,
                              'position': 'stem' if i == 0 else f'opt{i}',
                              'path': rec.get('path'), 'kind': rec.get('kind')})

    # DEFECT-3: ONE role rule, owned by the engine. Never re-implement inline.
    q_roles = bc.derive_image_roles(image_map)
    return image_map, q_roles, {'verdicts': verdicts, 'stats': stats}


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1777-1824 (v2.53.1) — VERBATIM ═══
MATH_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/math'

def omml_to_linear(omath_elem):
    def recurse(el):
        tag = el.tag.split('}')[-1] if '}' in el.tag else el.tag
        if tag == 'r':
            t = el.find(f'{{{MATH_NS}}}t')
            return (t.text or '') if t is not None else ''
        elif tag == 'f':
            n = el.find(f'{{{MATH_NS}}}num'); d = el.find(f'{{{MATH_NS}}}den')
            return f'({recurse(n)})/({recurse(d)})' if n is not None and d is not None else '?/?'
        elif tag == 'sSup':
            b = el.find(f'{{{MATH_NS}}}e'); s = el.find(f'{{{MATH_NS}}}sup')
            return f'{recurse(b)}^{recurse(s)}' if b is not None and s is not None else '?'
        elif tag == 'sSub':
            b = el.find(f'{{{MATH_NS}}}e'); s = el.find(f'{{{MATH_NS}}}sub')
            return f'{recurse(b)}_{recurse(s)}' if b is not None and s is not None else '?'
        elif tag == 'rad':
            d = el.find(f'{{{MATH_NS}}}deg'); b = el.find(f'{{{MATH_NS}}}e')
            btext = recurse(b) if b is not None else '?'
            return f'root({d.text},{btext})' if d is not None and d.text else f'sqrt({btext})'
        else:
            return ''.join(recurse(c) for c in el)
    try:    return recurse(omath_elem).strip()
    except: return '[OMML_FAILED]'

def enrich_paragraph_with_omml(stem_text, paragraph):
    """
    Find all OMML nodes in paragraph, convert to linear text.
    BUG-A05 fix: all_ok starts True and is AND-reduced (not last-write-wins).
    Returns: (enriched_text, all_ok, has_omml)
    Failed OMML nodes are skipped (not appended) to avoid polluting templates.
    """
    additions = []
    all_ok    = True   # True = every OMML node in this paragraph converted OK
    for elem in paragraph._p.iter():
        if elem.tag == f'{{{MATH_NS}}}oMath':
            linear = omml_to_linear(elem)
            if '[OMML_FAILED]' in linear:
                all_ok = False          # propagate failure via AND
                # do NOT append failed output to stem
            else:
                additions.append(linear)
    has_omml = bool(additions) or not all_ok
    enriched = stem_text + (' ' + ' '.join(additions) if additions else '')
    return enriched.strip(), all_ok, has_omml


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1828-1877 (v2.53.1) — VERBATIM ═══
# BUG-A14-related (Unicode): NOTE_PAT is a non-raw string so \u escapes ARE processed.
# v2.16 RIGID-6: expanded from English+Hindi to ALL major Indic scripts.
# Regional exams (Tamil Nadu PSC, AP/TS PSC, WBPSC, KPSC, etc.) use NOTE blocks
# in their respective scripts. Missing these would lose instruction metadata.
NOTE_PAT = re.compile(
    '\\((?:NOTE|Note|INSTRUCTION|Important|Caution'
    '|\u0928\u094b\u091f'           # Hindi/Marathi (Devanagari): नोट
    '|\u092e\u0939\u0924\u094d\u0935\u092a\u0942\u0930\u094d\u0923'  # Hindi: महत्वपूर्ण
    '|\u0b95\u0bc1\u0bb1\u0bbf\u0baa\u0bcd\u0baa\u0bc1'              # Tamil: குறிப்பு
    '|\u0c17\u0c2e\u0c28\u0c3f\u0c15'                                # Telugu: గమనిక
    '|\u09a6\u09cd\u09b0\u09b7\u09cd\u099f\u09ac\u09cd\u09af'       # Bengali: দ্রষ্টব্য
    '|\u0c9f\u0cbf\u0caa\u0ccd\u0caa\u0ca3\u0cbf'                   # Kannada: ಟಿಪ್ಪಣಿ
    '|\u0d15\u0d41\u0d31\u0d3f\u0d2a\u0d4d\u0d2a\u0d4d'             # Malayalam: കുറിപ്പ്
    '|\u0a28\u0a4b\u0a1f'                                            # Punjabi (Gurmukhi): ਨੋਟ
    '|\u0aa8\u0acb\u0a82\u0aa7'                                      # Gujarati: નોંધ
    '|\u0b1f\u0b3f\u0b2a\u0b4d\u0b2a\u0b23\u0b40'                   # Odia: ଟିପ୍ପଣୀ
    ')[^)]{10,}\\)',
    re.DOTALL
)

def extract_note_block(stem_text):
    """Return (clean_stem, note_text, found)."""
    m = NOTE_PAT.search(stem_text)
    if m:
        note    = m.group(0)
        cleaned = (stem_text[:m.start()] + ' ' + stem_text[m.end():]).strip()
        return cleaned, note, True
    return stem_text, '', False

def classify_note_frequency(note_count, total):
    """mandatory >=80% | conditional 20-79% | rare 1-19% | never 0%"""
    if total == 0: return 'never'
    pct = note_count / total * 100
    if pct >= 80: return 'mandatory'
    if pct >= 20: return 'conditional'
    if pct >= 1:  return 'rare'
    return 'never'

def canonical_note_text(notes_by_year):
    """Most common NOTE text from most recent year. Skips None keys."""
    if not notes_by_year: return ''
    int_years = {k: v for k, v in notes_by_year.items() if isinstance(k, int)}
    if int_years:
        recent = max(int_years.keys())
        texts  = int_years[recent]
    else:
        texts  = list(notes_by_year.values())[0] if notes_by_year else []
    return Counter(texts).most_common(1)[0][0] if texts else ''


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1881-1929 (v2.53.1) — VERBATIM ═══
def detect_linked_groups(questions):
    """
    Method 1: reprinted stimulus — verbatim match >=90% (Cloze, DI, Reading Comprehension etc.).
    Method 2: proximity stimulus (non-reprinted passage) applied during extraction
              (long para before run of short-stem Qs — used in CAT, UPSC, CLAT etc.).
    """
    groups, visited = [], set()

    for i, q in enumerate(questions):
        if q['num'] in visited: continue
        visited.add(q['num'])   # mark lead question immediately to prevent re-processing
        stim_q = extract_stimulus(q['stem'])
        if not stim_q: continue
        members = [q['num']]

        for j in range(i+1, min(i+8, len(questions))):
            nq = questions[j]
            if nq['num'] in visited: break
            stim_nq = extract_stimulus(nq['stem'])
            if not stim_nq: break
            if SequenceMatcher(None, stim_q, stim_nq).ratio() >= 0.90:
                members.append(nq['num']); visited.add(nq['num'])
            else:
                break

        if len(members) > 1:
            groups.append({'group_id'      : f'G{len(groups)+1}',
                           'q_numbers'     : members,
                           'stimulus_text' : stim_q,
                           'stimulus_type' : classify_stimulus(stim_q),
                           'word_count'    : len(stim_q.split())})
            visited.update(members)

    return groups

def extract_stimulus(stem):
    """Extract shared stimulus portion (content before question ask)."""
    m = re.search(r'\b(Select|Find|Which|Who|What|How|Identify|Choose)\b', stem)
    if m and m.start() > 100: return stem[:m.start()].strip()
    return stem if len(stem.split()) > 50 else ''

def classify_stimulus(text):
    t = text.lower()
    if '|' in text or 'table' in t: return 'table'
    if any(kw in t for kw in ['clue','sitting','puzzle','arrangement']): return 'puzzle'
    if 'code' in t and 'means' in t: return 'code'
    return 'passage'


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L1933-2017 (v2.53.1) — VERBATIM ═══
def classify_option_format(opts):
    """
    BUG-A23 fix: en-dash U+2013 matched via actual Unicode char (non-raw string).
    BUG-A12 / BUG-B19 fix: empty opts handled before any index access.
    12 types: single_value, value_pair, coordinate_pair, letter_cluster, code_pair,
              sentence, label_only, segment_label, roman_label, image_only,
              value_pair_quad, word_form_number
    """
    cleaned = [o.strip() for o in opts if isinstance(o, str)]
    if not cleaned or all(o == '' for o in cleaned): return 'image_only'

    wnums = {'one','two','three','four','five','six','seven','eight','nine','ten',
             'eleven','twelve','thirteen','fourteen','fifteen','sixteen','seventeen',
             'eighteen','nineteen','twenty'}
    if all(o.lower() in wnums for o in cleaned if o): return 'word_form_number'
    if all(o.lower() in {'i','ii','iii','iv','v','vi'} for o in cleaned if o):
        return 'roman_label'
    if all(len(o)==1 and o.upper() in 'ABCD' for o in cleaned if o):
        return 'segment_label'

    sent_kw = {'only','both','follow','follows','correct','neither','none'}
    if any(any(kw in o.lower() for kw in sent_kw) for o in cleaned if o):
        return 'sentence' if any(len(o.split()) > 3 for o in cleaned if o) else 'label_only'

    if all(re.match(r'^[A-D]-\d+(?:,\s*[A-D]-\d+)*$', o) for o in cleaned if o):
        return 'value_pair_quad'
    if all(re.match(r'^\(\d+,\s*\d+\)$', o) for o in cleaned if o):
        return 'coordinate_pair'

    # BUG-A23 fix: actual en-dash U+2013 embedded as literal character in raw string.
    # Cannot use \u2013 in raw string (r-prefix disables escapes).
    # Instead: compile pattern with en-dash as literal char (avoids SyntaxWarning).
    _VALUE_PAIR_PAT = re.compile(r'\d+\s*[-–]\s*\d+$')
    if all(_VALUE_PAIR_PAT.match(o) for o in cleaned if o):
        return 'value_pair'

    if all(re.match(r'^[A-Z]{2,6}$', o) for o in cleaned if o):
        return 'letter_cluster'
    if all('\u2192' in o or (' - ' in o and bool(re.search(r'[A-Z]{3,}', o)))
           for o in cleaned if o):
        return 'code_pair'
    return 'single_value'

def subtopic_option_format(qs):
    """
    BUG-A12 fix: guard for empty qs list (years[-1] IndexError).
    """
    if not qs:
        return {'primary':'single_value','recent_format':'single_value',
                'changed_recently':False,'all_observed':[]}
    fmts = [classify_option_format(q.get('options', [])) for q in qs]
    by_year = {}
    for q, fmt in zip(qs, fmts):
        # DEFECT D7 (found 2026-08-17 by the B7 mutation batch; see the note below).
        # WAS: by_year.setdefault(q.get('year', '?'), []).append(fmt)
        #
        # The '?' sentinel put a str into a dict whose other keys are ints, and the very
        # next line sorts those keys. So ONE question missing a year — one failed year
        # extraction anywhere in a subtopic — raised
        #     TypeError: '<' not supported between instances of 'str' and 'int'
        # inside synthesise_subtopic, which runs for EVERY subtopic. That is a hard stop
        # of the whole synthesis, in the same family as D1/D2/D6 and invisible on the
        # reference corpus, where every question happens to carry a year.
        #
        # The sentinel also made the `if not years:` guard below UNREACHABLE: by_year
        # gained a key for every question, so `years` was never empty and the branch the
        # author wrote for a year-less corpus could never run. Dropping year-less
        # questions from the RECENCY vote (they cannot inform which format is most
        # recent) both fixes the crash and restores that guard to its intended job.
        #
        # NO ARTEFACT MOVES. With every question carrying a year — the reference corpus
        # and the golden set — by_year is unchanged. With NO question carrying a year the
        # old path returned recent_format = Counter(all fmts).most_common(1) == primary
        # and changed_recently == False, which is exactly what the restored guard returns.
        # The only behaviour that changes is the mixed case, and that case used to crash.
        _y7 = q.get('year')
        if _y7 is None:
            continue
        by_year.setdefault(_y7, []).append(fmt)
    primary = Counter(fmts).most_common(1)[0][0]
    years   = sorted(by_year.keys())
    # GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE (D4) — ARTEFACT NONDETERMINISM.
    # `all_observed` was list(set(fmts)). Set iteration order for str elements
    # depends on PYTHONHASHSEED, which python randomises PER PROCESS, so this list
    # came out in a different order on every run. It is emitted verbatim by
    # write_section_rules as `option_format_all_observed:` (§14), which means TWO
    # runs of identical code over an identical corpus produced two different
    # section_rules.md files — measured: same 433,260 bytes, different sha256,
    # differing at 34 lines, one per subtopic.
    #
    # CONSEQUENCE: no Step 5 artefact was ever reproducible, so no diff between two
    # runs could distinguish a real regression from hash-seed noise, and any
    # byte-identity gate over this pipeline was impossible to build. That is why
    # this is fixed HERE, in the release that establishes the Wave 2 Part C
    # regression floor, rather than deferred: without it the floor cannot exist.
    #
    # WHY IT HID: a single process reuses one seed, so back-to-back runs inside one
    # session agree with each other and look deterministic. Only a fresh process
    # disagrees — and every real run is a fresh process.
    #
    # sorted() not list(): a stable, explicit, reviewable order.
    if not years:
        return {'primary':primary,'recent_format':primary,
                'changed_recently':False,'all_observed':sorted(set(fmts))}
    recent = Counter(by_year.get(years[-1], [])).most_common(1)
    rfmt   = recent[0][0] if recent else primary
    return {'primary':primary,'recent_format':rfmt,
            'changed_recently':rfmt != primary,'all_observed':sorted(set(fmts))}


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L2027-2039 (v2.53.1) — VERBATIM ═══

# score_difficulty adapter RETIRED — GAP-2026-08-27-DIFFICULTY-PROFILE (see blueprint_core Cluster E header)

def determine_strip_mode(section, topic, subtopic):
    """v2.30 — DELEGATED to blueprint_core Cluster E (GAP-2026-07-25-002).
    Second copy removed (GAP-2026-07-25-002). The engine's version carries the
    v2.16 RIGID-5 Devanagari terms, so delegating also means a Hindi-medium exam is
    classified by the same table everywhere rather than by whichever copy ran."""
    import blueprint_core as bc      # local: this block does not share the module alias
    return bc.determine_strip_mode(section, topic, subtopic)

def strip_variables(stem, mode):
    """
    Strip variable parts from stem to produce structural skeleton.
    BUG-A14 fix: currency pattern uses non-raw string with actual \u20b9 (₹),
                 so the Unicode escape IS processed correctly.
    BUG-B10 fix: logical mode preserves trailing punctuation.
    """
    t = stem
    if mode == 'quantitative':
        t = re.sub('\u20b9\\s*[\\d,]+(?:\\.\\d+)?', '\u20b9_P_', t)  # ₹ currency
        t = re.sub(r'\d+(?:\.\d+)?%', '_R_%', t)
        t = re.sub(r'\b\d+\s*(?:years?|months?|days?|hours?|weeks?)\b',
                   '_T_', t, flags=re.IGNORECASE)
        t = re.sub(r'\b\d+(?:,\d+)*(?:\.\d+)?\b', '_NUM_', t)
        t = re.sub(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\b', '_NAME_', t)
    elif mode == 'reasoning':
        t = re.sub(r'\b[A-Z]{3,}\b', '_WORD_', t)
        t = re.sub(r'\b[A-Z]{2,4}(?:\s*:\s*[A-Z]{2,4})+\b', '_LCLUS_:_LCLUS_', t)
        t = re.sub(r'\b\d+(?:,\s*\d+){2,}\b', '_SERIES_', t)
        t = re.sub(r'\b\d+\b', '_NUM_', t)
    elif mode == 'factual':
        t = re.sub(r'\b(19|20)\d{2}\b', '_YEAR_', t)
        t = re.sub(r'\b\d+(?:st|nd|rd|th)?\b', '_NUM_', t)
    elif mode == 'english':
        t = re.sub(r'_{3,}', '_BLANK_', t)
    elif mode == 'logical':
        # BUG-B10 fix: strip trailing punctuation separately, restore after
        quantifiers = {'All','all','Some','some','No','no','None','none',
                       'are','is','can','never','be','being'}
        words  = t.split()
        result = []
        for w in words:
            w_clean  = w.rstrip('.,;:!?')
            trailing = w[len(w_clean):]
            if w_clean in quantifiers or not (w_clean and w_clean[0].isupper()):
                result.append(w)
            else:
                result.append('_NOUN_' + trailing)
        t = ' '.join(result)
    t = re.sub(r'^Q\.?\d+\.?\s*', '', t)
    t = re.sub(r'\[\d{1,2}-\w+-\d{4}[^\]]*\]', '', t)
    return t.strip()

def generate_templates(questions, strip_mode, sig=None):
    """
    BUG-A04 fix: SequenceMatcher imported at module top (§1 S1-1).
    BUG-A13 fix: empty skeletons returns [] — caller handles this case.
    BUG-A20 fix: deprecated flag added to each pattern dict.
    """
    if not questions: return []

    # v2.54.1 (GAP-2026-08-23-STEP5-YEARLESS-TEMPLATES). q['year'] is stamped
    # UNCONDITIONALLY at extraction (process_pyq_paper), so a paper whose filename
    # carries no year stores year=None WITH THE KEY PRESENT — and .get's default
    # never fires for a present key. sorted() over {int, None} raised TypeError
    # inside synthesise_subtopic (which runs for EVERY subtopic): a hard stop of
    # the whole synthesis for any corpus mixing dated and undated paper filenames.
    # SAME FAMILY as D7, one function from D7's own fix, hidden by a dead default
    # that only a key-ABSENT fixture shape could satisfy. Year-less questions drop
    # out of the RECENCY basis only — exactly the rule subtopic_option_format
    # already applies to its recency vote.
    years = sorted({q.get('year') for q in questions if q.get('year') is not None})
    last2 = set(years[-2:]) if len(years) >= 2 else set(years)

    skeletons = []
    for q in questions:
        if not q.get('stem'): continue
        skel = (strip_variables_v2(q['stem'], sig, strip_mode) if sig is not None
                else strip_variables(q['stem'], strip_mode))
        if not skel.strip(): continue
        weight = 2 if q.get('year') in last2 else 1
        # v2.54.1: carry the true value (None allowed); the cluster 'years' set
        # below filters it. The old 2020 default was dead for a present key and
        # would have invented a fake year for an absent one.
        skeletons.append({'skel':skel, 'weight':weight, 'year':q.get('year')})

    if not skeletons: return []

    clusters = []
    for item in skeletons:
        placed = False
        for c in clusters:
            if SequenceMatcher(None, c[0]['skel'], item['skel']).ratio() >= 0.90:
                c.append(item); placed = True; break
        if not placed: clusters.append([item])

    total_w  = sum(s['weight'] for s in skeletons) or 1
    pats_raw = sorted([{
        'skel'     : c[0]['skel'],
        'w_count'  : sum(i['weight'] for i in c),
        'raw_count': len(c),
        'years'    : sorted({i['year'] for i in c if i['year'] is not None}),
    } for c in clusters], key=lambda p: -p['w_count'])

    for p in pats_raw:
        p['raw_pct']   = p['w_count'] / total_w * 100
        p['frequency'] = int(p['raw_pct'])
    deficit = 100 - sum(p['frequency'] for p in pats_raw)
    if deficit > 0:
        for p in sorted(pats_raw, key=lambda p: -(p['raw_pct']-p['frequency']))[:deficit]:
            p['frequency'] += 1

    result = []
    for i, p in enumerate(pats_raw, 1):
        conf = 'observed' if p['raw_count'] >= 3 else 'inferred'
        if p['years'] and all(y in last2 for y in p['years']) and len(p['years']) <= 2:
            conf = 'observed_recent'
        # BUG-A20 fix: deprecated flag for patterns absent from last 2 years
        deprecated = bool(p['years'] and not any(y in last2 for y in p['years']))
        result.append({
            'id'        : f'P{i}',
            'template'  : p['skel'],
            'frequency' : p['frequency'],
            'raw_count' : p['raw_count'],
            'confidence': conf,
            'deprecated': deprecated,
            'years'     : p['years'],
        })
    # Filter out patterns with frequency=0 (can occur for very rare templates
    # when raw_pct < 0.5% and largest-remainder does not assign them +1).
    # Such templates would appear in section_rules.md but never be selected by Step 7.
    result = [p for p in result if p['frequency'] >= 1]
    # Re-normalize if any were removed (ensure sum still = 100)
    if result:
        total_freq = sum(p['frequency'] for p in result)
        if total_freq != 100:
            # Add deficit to highest-frequency pattern (always the first after DESC sort)
            result[0]['frequency'] += (100 - total_freq)
    return result


# ═══ FROM Framework_MockTestAnalyse.md §2, fence L2182-2260 (v2.53.1) — VERBATIM ═══
def classify_wrong_option_structure(subtopic_qs):
    """
    BUG-A08 fix: image-only check before fixed_set detection.
    BUG-A24 fix: empty strings filtered from shared_pool counter.
    BUG-B12 fix: image_only added to DESC dict (11th type).
    """
    all_opt_sets = [q['options'] for q in subtopic_qs if q.get('options')]
    if not all_opt_sets:
        return {'type':'varied','description':'No options to classify'}

    # BUG-A08 fix: detect image-only before entering fixed_set logic
    if all(o.strip() == '' for opts in all_opt_sets for o in opts):
        return {'type':'image_only',
                'description':'All options are images — option text is blank.'}

    frozen  = [frozenset(o.strip().lower() for o in s) for s in all_opt_sets]
    top_s, ct = Counter(frozen).most_common(1)[0]
    if ct / len(frozen) >= 0.80:
        canon = next(s for s in all_opt_sets
                     if frozenset(o.strip().lower() for o in s) == top_s)
        return {'type':'fixed_set',
                'description':'Identical options every question of this type.',
                'fixed_option_texts':[o.strip() for o in canon]}

    # BUG-A24 fix: filter empty strings from shared_pool counter
    all_words = Counter(o.strip().lower() for s in all_opt_sets for o in s if o.strip())
    top4      = [w for w, _ in all_words.most_common(4)]
    coverage  = sum(1 for s in all_opt_sets if any(o.strip().lower() in top4 for o in s))
    if top4 and coverage / len(all_opt_sets) >= 0.50 and len(all_words) < 8:
        return {'type':'shared_pool',
                'description':'Same word pool rotated as options across linked questions.',
                'shared_pool_words':top4}

    votes = Counter(_classify_one_set([o.strip() for o in s if o.strip()])
                    for s in all_opt_sets)
    dom   = votes.most_common(1)[0][0]
    # BUG-B12 fix: image_only added as 11th type in DESC
    DESC  = {
        'adjacent_values': 'Numeric options within +-20% — near-miss calculations.',
        'anagram'        : 'Options are character rearrangements of each other.',
        'alliterative'   : '3+ options share first letter — deliberate distractor.',
        'same_category'  : 'All options are real entities of the same class.',
        'sentence_label' : '"Only I follows" / "Both II and III follow" style.',
        'segment_label'  : 'Options are A/B/C/D referring to labelled stem segments.',
        'roman_label'    : 'Options are i/ii/iii/iv referring to numbered sub-items.',
        'value_pair_quad': 'Matching combinations: A-1, B-2, C-3, D-4.',
        'word_form'      : 'Options are number word forms: One/Two/Three/Four.',
        'image_only'     : 'All options are images — option text is blank.',
        'varied'         : 'No consistent structural pattern detected.',
    }
    return {'type':dom, 'description':DESC.get(dom, 'Pattern detected.')}

def _classify_one_set(opts):
    if not opts: return 'varied'
    if set(o.upper() for o in opts) == {'A','B','C','D'} and all(len(o)==1 for o in opts):
        return 'segment_label'
    if all(o.lower() in {'i','ii','iii','iv','v','vi'} for o in opts): return 'roman_label'
    wn = {'one','two','three','four','five','six','seven','eight','nine','ten'}
    if all(o.lower() in wn for o in opts): return 'word_form'
    kw = {'only','both','follow','correct','neither'}
    if any(any(k in o.lower() for k in kw) for o in opts): return 'sentence_label'
    csets = [Counter(o.upper().replace(' ','')) for o in opts]
    if len(set(frozenset(c.items()) for c in csets)) == 1 and len(opts[0]) <= 6:
        return 'anagram'
    fl = [o[0].upper() for o in opts if o]
    # Guard for empty fl (all opts were empty after filtering)
    most_fl = Counter(fl).most_common(1)
    if most_fl and most_fl[0][1] >= 3: return 'alliterative'
    nums = []
    for o in opts:
        try: nums.append(float(re.sub('[,\u20b9%\\s]', '', o)))
        except: pass
    if len(nums) == len(opts) and nums and max(nums) / max(min(nums), 0.001) <= 1.4:
        return 'adjacent_values'
    cap = sum(1 for o in opts if o and o[0].isupper() and not any(c.isdigit() for c in o))
    if cap >= 3: return 'same_category'
    return 'varied'


# ═══ FROM Framework_MockTestAnalyse.md §4, fence L2875-2919 (v2.53.1) — VERBATIM ═══
def get_vision_candidates(questions, q_roles, imap):
    """
    Returns list of (q, [image_path, ...]) pairs requiring vision analysis.
    Only questions where the image is in the stem (or stem+options).
    options_only: E-11 classifies those from text patterns, no vision needed.
    Text-extractable subtopics (dice, cube, counting shapes etc.): values
    already present in stem text — vision adds nothing for these.
    imap: image mapping list from extract_and_map_images() — passed explicitly.

    EC-V6 (GAP-2026-07-26-003). ALL of a question's stem images are returned, not
    just the first. A 4-panel series is ONE question whose rule can only be read
    across the panels; handing the queue panel 1 alone and asking for the
    transformation guaranteed either a wrong answer or an unreadable verdict. The
    panels are composed into a single cell so they are judged together (S4-2a).
    """
    candidates = []
    # SKIP_TEXT_EXTRACTABLE: subtopics where the image contains countable/readable
    # values already present as text in the stem — vision adds no information.
    # Detected by checking if stem contains numeric/face values AND subtopic name
    # suggests a counting/identification task (dice, cube, counting shapes, etc.).
    # This list is NOT hardcoded — it is inferred from subtopic name keywords.
    SKIP_KEYWORDS = {'dice', 'cube', 'count', 'counting', 'dots', 'faces', 'nets'}

    for q in questions:
        role = q_roles.get(q['num'], {}).get('role', 'none')

        if role not in ('stem_only', 'stem_and_options'):
            continue   # options_only or none: no stem image, skip

        sub_lower = q.get('subtopic', '').lower()
        if any(kw in sub_lower for kw in SKIP_KEYWORDS):
            continue   # stem already has the values; vision not needed

        # Get the path(s) of stem images for this Q (imap passed from E-4), in
        # document order — the series order IS the question's meaning.
        stem_imgs = [entry for entry in imap
                     if entry['q_num'] == q['num']
                     and entry['position'] == 'stem']
        paths = [e['path'] for e in stem_imgs if e.get('path')]
        if paths:
            candidates.append((q, paths))

    return candidates


# ═══ FROM Framework_MockTestAnalyse.md §4, fence L2957-2964 (v2.53.1) — VERBATIM ═══
# Where the queue and its sheets live. One directory per run, NOT per paper: a
# batch's sheets are viewed together in Phase B, and EC-V15 keys every item by
# (paper_id, q_num) so one shared queue cannot confuse two papers.
VISION_WORKDIR   = '/home/claude/pyq_vision'
VISION_PER_SHEET = 6      # 153 figures -> 26 view() calls instead of 153.
                          # A parameter, not a constant of nature: re-tile freely.


# ═══ FROM Framework_MockTestAnalyse.md §4, fence L3024-3064 (v2.53.1) — VERBATIM ═══
def apply_vision_observations(progress, workdir=VISION_WORKDIR):
    """Fold Phase-B observations onto the questions. THE ONLY WRITER of image_clarity.

    Returns the merge stats dict. NEVER raises, NEVER halts: a missing or malformed
    observations file is the ordinary 'Phase B has not run' state and must flow
    through as zero observations so the run completes and QV-14 reports the gap.

    Idempotent (EC-V12): running it twice on the same inputs yields the same fields.
    """
    queue, why = corpus_io.load_vision_queue(workdir)
    if queue is None:
        print(f"    PHASE C: {why} — no figural vision data this run")
        return {'queued': 0, 'observed': 0, 'missing': 0, 'unreadable': 0,
                'unknown': [], 'duplicate': [], 'vision_status': 'not_applicable'}

    observations, why = corpus_io.load_vision_observations(workdir)
    if why:
        print(f"    PHASE C: {why}")

    by_key, stats = bc.merge_vision_observations(queue['items'], observations)

    # Stamp the fields onto the live question dicts, keyed (paper_id, q_num) — EC-V15.
    for key, entries in progress.items():
        if not isinstance(key, tuple):
            continue
        for q in entries:
            rec = by_key.get((str(q.get('paper_id')), q.get('q_num')))
            if rec:
                q.update({'object_type'   : rec['object_type'],
                          'transformation': rec['transformation_type'],
                          'arrangement'   : rec['arrangement'],
                          'complexity'    : rec['complexity'],
                          'image_clarity' : rec['image_clarity']})

    print(f"    PHASE C: {stats['observed']}/{stats['queued']} figure(s) observed "
          f"({stats['vision_status']})"
          + (f", {stats['unreadable']} illegible" if stats['unreadable'] else '')
          + (f", {len(stats['unknown'])} unknown tag(s) ignored" if stats['unknown'] else ''))
    return stats


# ═══ FROM Framework_MockTestAnalyse.md §4, fence L3086-3132 (v2.53.1) — VERBATIM ═══
def aggregate_figural(questions_for_subtopic, q_roles=None):
    """
    Called during synthesis (§5) for FIGURAL subtopics.

    v2.37 (GAP-2026-07-26-003). The aggregation itself is DELEGATED to
    bc.vision_profile(). It used to be re-implemented here as a plain top-3 over
    object_type, which had two defects:

      * it named a dominant type from ANY number of observations. Two figures cannot
        support a claim about what a subtopic "typically" looks like, and Step 7
        generates against that claim (EC-V20).
      * it named a dominant type even when the distribution was FLAT. Six figures of
        six DIFFERENT types produced a "dominant" list of the alphabetically-first
        three — handing the generator a fixation the evidence did not support. Caught
        by end-to-end test, not by inspection (EC-V26).

    One aggregation rule, in the engine, unit-tested. Never re-implement it here.

    image_role stays local: it is derived from q['image_role'] (stored per question in
    S3-1) rather than from q_roles, which is NOT persisted in progress.json across
    sessions. q_roles is retained for backward compatibility and is unused.
    """
    from collections import Counter

    # Dominant image_role from q['image_role'] stored in each question dict during S3-1.
    roles = [q.get('image_role', 'none') for q in questions_for_subtopic
             if q.get('image_role', 'none') != 'none']
    dominant_role = Counter(roles).most_common(1)[0][0] if roles else 'stem_only'

    # Only questions that actually carry a figure enter the vision denominator.
    # EC-V13/EC-V14: a zero-PYQ inferred FIGURAL subtopic and an INHERENTLY-VISUAL
    # keyword override are legitimately FIGURAL with no embedded figure, so they
    # contribute no records and are excluded from QV-14 by construction.
    figural = [q for q in questions_for_subtopic
               if q.get('image_role', 'none') != 'none']

    prof = bc.vision_profile([
        {'image_clarity'      : q.get('image_clarity', 'vision_unavailable'),
         'object_type'        : q.get('object_type'),
         'transformation_type': q.get('transformation'),
         'arrangement'        : q.get('arrangement'),
         'complexity'         : q.get('complexity')}
        for q in figural])
    prof['image_role'] = dominant_role
    return prof



# ═══ FROM Framework_MockTestAnalyse.md §3, fence L1884-2046 (v2.53.2) — VERBATIM ═══

def extract_shift_from_filename(path, session_re=None):
    """v2.16 RIGID-1: session/shift detection from the configurable keyword.

    WAVE 2 PART C B3 — THE ONE SESSION GLOBAL IN THE ENTIRE EXTRACTION SCOPE.
    This read module-level `SESSION_RE`, which §1 builds from exam_config.json's
    `session_keyword` at session start. An engine has its own globals and inherits
    nothing, so the regex is now INJECTED rather than reached for.

    It is a parameter, not module state set by a configure() call, deliberately:
    module state makes the engine non-reentrant and lets a caller that forgot to
    configure it silently use a STALE regex from a previous exam — a wrong answer
    rather than an error, which is this corpus's most expensive failure shape.
    A caller that forgets instead fails immediately and says what to pass, matching
    the v2.39 `vision_pending` precedent in process_pyq_paper below.
    """
    if session_re is None:
        raise RuntimeError(
            "extract_shift_from_filename: session_re not supplied. Pass "
            "session_re=SESSION_RE, the compiled regex §1 builds from "
            "exam_config.json's session_keyword via build_session_re(). It is exam "
            "configuration and cannot be defaulted here — guessing the keyword would "
            "silently mislabel every paper's shift (v2.16 RIGID-1).")
    m = session_re.search(os.path.basename(path))
    return f'S{m.group(1)}' if m else 'S1'

def process_pyq_paper(docx_path, paper_id, exam_code,
                      time_per_q, marks_per_q, options_count, multi_select,
                      progress, expected_size=None, vision_pending=None,
                      session_re=None):
    """
    All parameters auto-detected in S1-3, passed directly — no cfg dict.
    taxonomy not needed: presorted papers carry their own taxonomy headings.
    extract_presorted() is the only extraction path.

    v2.34 (GAP-2026-07-26-002):
      expected_size — DEFECT-5. The byte size Drive enumeration already captured, from
        paper_ref['fileSize']. Threaded to IMG-1, which SKIPped on every paper before
        this because nothing connected the two. None is still accepted (the upload lane
        may not know it) and IMG-1 then SKIPs for a legitimate reason.
    v2.37 (GAP-2026-07-26-003):
      probe_passed is GONE. Its only consumer was the in-loop vision block, which
        could not execute (see S3-1 / S4-2). This function now runs PHASE A only —
        it queues figural work and writes NO vision field. image_clarity has exactly
        one writer, apply_vision_observations() (S4-2c).
    """
    from docx import Document
    doc   = Document(docx_path)   # opened ONCE; passed to all sub-functions
    year  = extract_year_from_filename(docx_path)
    shift = extract_shift_from_filename(docx_path, session_re=session_re)

    # E-4: extract images. docx_path and expected_size are REQUIRED — without them the
    # v2.29 gates cannot run. Unpack order matches the single return shape
    # (image_map, q_roles, report); the old three-way unpack mis-bound silently because
    # both branches returned arity 3 in different orders (GAP-2026-07-26-002 DEFECT-2).
    image_map, q_roles, img_report = extract_and_map_images(
        doc, paper_id, docx_path, expected_size=expected_size)

    # Always presorted — single extraction path
    # v2.39 (GAP-2026-07-27-E): marking_scheme comes from exam_config via _meta and maps
    # an original exam position to a declared question_type. Exam-agnostic — the bands
    # are read at runtime and nothing about any exam is hardcoded. Absent scheme => the
    # positional branch is inert and detection is exactly v2.38.
    _mscheme = (progress.get('_meta', {}) or {}).get('marking_scheme') or []
    questions = extract_presorted(doc, year, shift, paper_id, q_roles,
                                  options_count, multi_select,
                                  marking_scheme=_mscheme)

    linked_groups = detect_linked_groups(questions)
    link_map = {qn: g['group_id'] for g in linked_groups for qn in g['q_numbers']}

    for q in questions:
        # (dead `sm = determine_strip_mode(...)` removed 2026-08-31 — assigned,
        #  never consumed; a leftover of the retired per-question strip stamp.
        #  MS-16 forbids table callers outside the aptitude shim.)
        # Determine marks per question: try 'MCQ' key first (most exams),
        # then try all values and use the max (handles GATE 1-mark/2-mark structure
        # where marks_per_q = {'1-mark':1,'2-mark':2} — use max available mark as default).
        q_marks = marks_per_q.get('MCQ') or marks_per_q.get('mcq') or max(marks_per_q.values(), default=1)
        q['option_format']   = classify_option_format(q.get('options', []))
        q['paper_id']        = paper_id   # needed for max_per_paper/typical_per_paper computation
        # q['difficulty'] (E-9 keyword scorer) RETIRED — GAP-2026-08-27-DIFFICULTY-PROFILE
        q['image_role']      = q_roles.get(q['num'], {}).get('role', 'none')
        q['linked_group_id'] = link_map.get(q['num'])
        q['year']            = year
        q['shift']           = shift
        q['paper_id']        = paper_id
        # v2.23 THREE-AXIS TAGGING: tag Axis-1/2/3 now that linked_group_id + image_role
        # + options are all set (the LINKED gate and FIGURAL/NAT signals need them).
        # Step 8 re-tags GENERATED questions with the SAME AXIS CLASSIFIER v1.0 functions.
        tag_axes(q)

    # IMG-5b: cross-check image presence against Axis-1 now that tag_axes() has run.
    img5b = run_img5b({e['q_num']: [e['path']] for e in image_map}, questions,
                      overrides=tuple(progress.get('_meta', {}).get('inherently_visual', ())))

    # §4: PHASE A. Claude CANNOT view an image from inside this python process — a
    # view() is a TOOL CALL and a tool call can only happen BETWEEN model turns. This
    # loop therefore QUEUES the work and writes contact sheets to disk; Phase B (S4-2b)
    # performs the views at the batch boundary the run already stops at; Phase C
    # (S4-2c) folds the observations back in. See EXECUTION-BOUNDARY LAW.
    #
    # GAP-2026-07-26-003. What stood here was `vision_result = analyse_image_claude(...)`
    # followed by `vision_result.get(...)`. analyse_image_claude() was a `pass` stub, so
    # the call returned None and the next line raised AttributeError — meaning the block
    # could not run AS WRITTEN at all. Every production run therefore executed some
    # SUBSTITUTED body, and the substitution that shipped wrote image_clarity only.
    # Measured on IIT_JAM_BIOTECHNOLOGY: object_type/transformation/arrangement/
    # complexity present on 0 of 1719 questions; 153/153 figural = vision_unavailable;
    # 45/45 FIGURAL subtopics shipped an empty profile; QV-9 returned PASS.
    #
    # NOTHING HERE HALTS. image_clarity is NOT written in this loop — Phase C is its
    # only writer (one writer, no drift). Until Phase C runs the field is simply absent.
    # v2.39 (GAP-2026-07-27-B) — THIS FUNCTION NO LONGER BUILDS THE QUEUE.
    #
    # build_vision_queue() WRITES vision_queue.json and vision_sheet_NNN.png into
    # VISION_WORKDIR, which S4-2a defines as one directory per RUN, not per paper. It
    # never read an existing queue, and both filenames are fixed constants
    # (VISION_QUEUE_NAME, VISION_SHEET_FMT) — so calling it once per paper meant paper N
    # OVERWROTE the queue and sheets of papers 1..N-1. Only the last paper of a batch
    # ever survived into Phase B.
    #
    # MEASURED on IIT_JAM_BIOTECHNOLOGY: 153 figural questions across 22 papers, but
    # _meta.vision recorded queued=8 — the final paper's count alone. Batch 7 queued
    # 11 + 10 + 3 = 24 and 3 survived. Five separate sessions hit this and each invented
    # a DIFFERENT workaround (run-level vision_records.json, vision_items.json rebuilt
    # over a union, carry-forward from disk, two independent hoists); no two agreed, and
    # divergent tag assignment between them would orphan observations silently.
    #
    # The stated Phase A invariant — "every queued item appears on exactly one sheet,
    # and that sheet exists" — was violated in the OPPOSITE direction: sheets existed
    # that no queue referenced.
    #
    # Candidates are accumulated here and the queue is built ONCE, at the batch
    # boundary, in run_batch_loop() immediately before Phase B.
    vision_candidates = get_vision_candidates(questions, q_roles, image_map)
    if vision_pending is None:
        # A caller that forgets the accumulator would silently drop every figure in the
        # paper — the exact failure mode this fix exists to remove. Fail loudly instead.
        raise RuntimeError(
            "process_pyq_paper: vision_pending accumulator not supplied. Phase A "
            "candidates must be collected at the RUN level and passed to "
            "corpus_io.build_vision_queue() ONCE, at the batch boundary (v2.39 "
            "GAP-2026-07-27-B). Pass vision_pending=<list> from run_batch_loop().")
    vision_pending.extend(
        {'paper_id': paper_id, 'q_num': q['num'],
         'srcs': [p for p in srcs if p],
         'subtopic': q.get('subtopic'), 'image_role': q.get('image_role')}
        for q, srcs in vision_candidates)
    if vision_candidates:
        print(f"    PHASE A: {len(vision_candidates)} figural question(s) accumulated "
              f"for this paper ({len(vision_pending)} pending in this batch)")

    for q in questions:
        key = (q.get('section','?'), q.get('topic','?'), q.get('subtopic','?'))
        progress.setdefault(key, []).append(q)

    progress.setdefault('_linked_groups', {})
    for g in linked_groups:
        progress['_linked_groups'][g['group_id']] = g

    meta = progress.setdefault('_meta', {'papers_processed':[],'total_questions':0,
                                          'years_processed':[]})
    meta['papers_processed'].append(paper_id)
    # v2.56 (GAP-2026-08-29-STYLE-FIDELITY §6.2/§6.3): the corpus_hash both style
    # artefacts stamp is sha256 over the SORTED set of these per-paper file
    # hashes. Stamped at the one place the docx bytes are in hand. Additive: a
    # pre-v2.56 progress file simply lacks the key and the synthesis fence falls
    # back to hashing paper_ids (documented DEGRADED source, still deterministic).
    try:
        with open(docx_path, 'rb') as _fh:
            meta.setdefault('paper_file_hashes', {})[paper_id] = \
                hashlib.sha256(_fh.read()).hexdigest()
    except OSError:
        meta.setdefault('paper_file_hashes', {})[paper_id] = \
            hashlib.sha256(str(paper_id).encode()).hexdigest()
    meta['total_questions'] += len(questions)
    if year and year not in meta['years_processed']:
        meta['years_processed'].append(year)

    # GAP-2026-08-16-BASELINE-SUPPRESSED-NAMEERRORS (D5). This line read `n_vision`,
    # which is a LOCAL of run_batch_loop() in §8 — bound nowhere in this function, this
    # section, or at module scope. process_pyq_paper() therefore raised
    # `NameError: name 'n_vision' is not defined` on its second-to-last statement, on
    # EVERY paper of EVERY exam, unguarded, after all the work was done and before
    # `return questions, linked_groups` could run.
    # It survived because spec_name_audit DETECTED it and `n_vision` sat in
    # spec_name_audit_baseline.json as an accepted unbound name — the same untyped
    # baseline that hid D2/D3 in GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE.
    # The per-paper figural count is len(vision_candidates); §8's n_vision is the
    # BATCH-level queue size and was never the right number here. "queued" not
    # "analysed": Phase B does the analysis, this is Phase A accumulation.
    print(f"  {os.path.basename(docx_path)}: {len(questions)} Qs, "
          f"{len(linked_groups)} linked groups, "
          f"{len(vision_candidates)} figural image(s) queued for vision")
    return questions, linked_groups

# ═══ FROM Framework_MockTestAnalyse.md §3, fence L2083-2093 (v2.53.2) — VERBATIM ═══
def vision_liveness(stats):
    """Session vision verdict, derived from Phase-C merge stats. Never raises.

    'not_applicable' — nothing was queued (a text-only exam is not a failure)
    'unavailable'    — figures were queued and NOTHING came back
    'partial'        — some cells observed, some omitted (procedural, re-runnable)
    'observed'       — every queued figure observed
    """
    return stats.get('vision_status', 'not_applicable')

# ═══ FROM Framework_MockTestAnalyse.md §3, fence L2144-2170 (v2.53.2) — VERBATIM ═══
def run_img5b(mapping, questions, overrides=()):
    """Gate IMG-5b — cross-check image presence against Axis-1 classification.

    v2.35 (audit_callgraph C4). corpus_io.figural_consistency() has existed since
    v2.29 and was called from NOWHERE — it appeared only in the prose bullet above,
    so the gate written to catch exactly the DEFECT-3 class of fault never ran. Two
    defects were masking each other: IMG-5b would have detected the missing
    q_roles['role'], but IMG-5b itself was unreachable.

    WARN, not HARD STOP: a mismatch can be legitimate (an INHERENTLY-VISUAL override
    is visual without carrying an embedded figure), and IMG-4 already hard-stops on
    actual image loss. This gate reports classification disagreement.
    """
    q_formats = {q['num']: q.get('format') or q.get('axis1') for q in questions}
    fc = corpus_io.figural_consistency(mapping, q_formats, overrides=overrides)
    if fc['image_not_figural']:
        print(f"    ! IMG-5b WARN — question(s) carry an image but are not FIGURAL: "
              f"{fc['image_not_figural'][:10]}")
        print(f"      An image present in the document but absent from the format "
              f"distribution is the silent corruption IMG-4 cannot see.")
    if fc['figural_no_image']:
        print(f"    ! IMG-5b WARN — question(s) classified FIGURAL with no image: "
              f"{fc['figural_no_image'][:10]}")
        print(f"      Expected only for logged INHERENTLY-VISUAL overrides.")
    return fc

# ═══ FROM Framework_MockTestAnalyse.md §3, fence L2174-2478 (v2.53.2) — VERBATIM ═══
# v2.27 — DELEGATED TO THE ENGINE (blueprint_core Cluster G). These three functions existed
# in BOTH this spec and Framework_PYQAnalyse Phase B, which walk the SAME sorted .docx and
# must agree. Each pair had drifted despite an explicit "keep IDENTICAL" instruction and
# EC-P14 naming the failure mode. One definition now, so they cannot drift again.
# GAP-2026-07-26-001 — next_text IS MANDATORY for sorted-PYQ walks.
# PYQSort EC-S8 emits multi-paragraph stems whose continuation lines are bold, not
# dates, not options and not question starts — character for character the heading
# predicate. In extract_presorted() below this is the DANGEROUS half of the defect:
# the inner loop breaks on a heading, so a continuation line terminated its own
# question mid-body, truncating the stem and orphaning every option after it, with
# no error raised anywhere. Measured on IIT_JAM_BIOTECHNOLOGY (22 papers): 16
# questions truncated, 28 option lines silently discarded, 10 papers affected.
# Step 4's phantom gate stops the run; nothing here stopped anything.
is_taxonomy_heading      = lambda para, next_text=None, colour_available=False: bc.is_taxonomy_heading(para, is_option, next_text, colour_available)
parse_taxonomy_level     = bc.parse_taxonomy_level
extract_year_from_filename = bc.extract_year_from_filename

def detect_blank_position(stem):
    m = re.search(r'_{3,}', stem)
    if not m: return 'none'
    pos = m.start() / max(len(stem), 1)
    return 'start' if pos < 0.2 else ('end' if pos > 0.8 else 'middle')

# ── v2.5 MSQ detection (EC-A root-cause fix) ──────────────────────────────────
# Forgery-resistant: keys on OPTION SHAPE, not stem wording. A statement-
# combination MCQ (EC-9) whose options are predominantly combination-labels
# ("Only 1", "Both 1 and 2", "Neither … nor …", "1 and 3", "None of …",
# "All of the above") is SINGLE-answer and must never be tagged MSQ, even when
# its stem reads "Which is/are correct". Conversely a genuine multi-select stem
# with ordinary content options IS an MSQ. Used only when multi_select=True.
_MSQ_INSTR_RE = re.compile(
    r'select all that apply|select all|one or more (?:options?|are|may)|'
    r'select\s+two|select\s+three|more than one (?:option|correct)|'
    r'which .*\bare correct\b', re.IGNORECASE)
_COMBO_OPT_RE = re.compile(
    r'^\s*(only\b|both\b|neither\b|all of\b|none of\b|'
    r'\d+\s+and\s+\d+\b|\d+\s*,\s*\d+)', re.IGNORECASE)

def _is_statement_combination(options):
    """True when the option SET is predominantly combination-labels (EC-9 MCQ)."""
    opts = [o for o in (options or []) if isinstance(o, str) and o.strip()]
    if not opts:
        return False
    combo = sum(1 for o in opts if _COMBO_OPT_RE.match(o.strip()))
    return combo >= max(2, len(opts) - 1)   # most/all options are combo-labels

def detect_is_msq(full_stem, options, positional_type=None):
    """v2.5 contract detector. Caller already gated on multi_select=True.

    v2.39 (GAP-2026-07-27-E) — POSITIONAL BRANCH ADDED.

    The instruction-phrase test alone measured 24 MSQ across 1,719 questions on an exam
    whose marking scheme reserves Q31-40 for MSQ (~10/paper, so ~120 in the current era
    alone). The cause is cross-step information destruction, not a weak regex: PYQSort
    renumbers questions into taxonomy order, so by the time Step 5 reads the paper the
    exam positions that identify the MSQ band no longer exist. Step 5 could not recover
    what Step 3 had discarded.

    PYQSort v1.18 now stamps the original position into the date label, so the band is
    knowable again. positional_type is the question_type whose marking_scheme q_range
    covers that position — exam-agnostic, read from exam_config at runtime, no hardcoded
    ranges anywhere.

    PRECEDENCE. The EC-A guard still wins: a statement-combination MCQ is NEVER MSQ,
    however it is banded, because its option SHAPE is forgery-resistant evidence about
    the answer mechanism and a band is only evidence about where the question sat. A
    paper that deviates from its own declared scheme is thereby not mis-typed.

    positional_type is None for a pre-v1.18 sorted file, an absent marking_scheme, or a
    position no band covers. All three mean "no positional evidence" and fall back to
    v2.5 behaviour exactly — so no exam changes until its papers are re-sorted.
    """
    if _is_statement_combination(options):   # EC-A guard — highest precedence
        return False
    if _MSQ_INSTR_RE.search(full_stem or ''):
        return True
    return positional_type == 'MSQ'


def _detect_option_label_style(raw_option_lines):
    """
    v2.15 BUG-D07: Detect option LABEL style from raw option lines.
    Returns: '1/2/3/4' | 'A/B/C/D' | '(1)/(2)/(3)/(4)' | '(A)/(B)/(C)/(D)' | 'A)/B)/C)/D)' | 'unknown'
    Distinct from option FORMAT type (single_value etc.) which describes CONTENT shape.
    """
    if not raw_option_lines:
        return 'unknown'
    for line in raw_option_lines:
        t = line.strip()
        if re.match(r'^[1-5]\.\s+', t):       return '1/2/3/4'
        if re.match(r'^[A-E]\.\s+', t):        return 'A/B/C/D'
        if re.match(r'^[a-e]\.\s+', t):        return 'a/b/c/d'
        if re.match(r'^\([1-5]\)\s+', t):      return '(1)/(2)/(3)/(4)'
        if re.match(r'^\([A-E]\)\s+', t):      return '(A)/(B)/(C)/(D)'
        if re.match(r'^\([a-e]\)\s+', t):      return '(a)/(b)/(c)/(d)'
        if re.match(r'^[A-E]\)\s+', t):        return 'A)/B)/C)/D)'
        if re.match(r'^[a-e]\)\s+', t):        return 'a)/b)/c)/d)'
    return 'unknown'


def extract_presorted(doc, year, shift, paper_id, q_roles, options_count, multi_select,
                      marking_scheme=None):
    questions = []
    cur_sec = cur_top = cur_sub = ''
    cur_date_label = None          # v2.39 — carries the stamped original position
    # GAP-2026-07-26-001 + GAP-2026-08-05-001. BLOCK-level: tables are not paragraphs,
    # and an image/equation/object-only paragraph has no TEXT — the old lookahead
    # skipped both and returned the next question's date label, so a bold stem
    # continuation passed the level-3 heading test and TRUNCATED its own question here.
    paras, nxt = bc.sorted_body_lookahead(doc)
    colour_ok  = bc.heading_colour_available(paras)   # D6 — probed ONCE per FILE
    terminated_by_heading = 0     # QV-15 counter (GAP-2026-08-05-001)
    # Per-question flag, RESET at the top of every question body below. Initialised
    # here as well so it is defined even for a question whose body loop never runs.
    _ended_by_heading = False
    i = 0

    while i < len(paras):
        para = paras[i]; text = para.text.strip()
        if not text: i += 1; continue

        if is_taxonomy_heading(para, nxt[i], colour_ok):   # GAP-2026-07-26-001 + -08-05-001
            lv, content = parse_taxonomy_level(text)
            if lv == 1: cur_sec = content
            elif lv == 2: cur_top = content
            else: cur_sub = content
            i += 1; continue

        # BUG-A16 fix: case-insensitive shift tag check via pattern
        # v2.39 (GAP-2026-07-27-E): the label is RETAINED, not merely skipped. PYQSort
        # v1.18 stamps the original exam position into it, which is the only surviving
        # evidence of the MSQ band after Step 3's renumbering.
        # bc.is_position_label is THE predicate (bc.DATE_TAG_RE). The inline literal
        # that stood here is the one bc.DATE_TAG_RE's own comment claims to have
        # replaced at GAP-2026-07-26-001; it had not been. Writer (PYQSort) and reader
        # (here) now share one definition, so they cannot drift apart again.
        if bc.is_position_label(text):
            cur_date_label = text
            i += 1; continue

        q_num = detect_question_start(text)
        if q_num is None: i += 1; continue

        options      = []
        options_raw  = []   # v2.15 BUG-D07: raw option lines for label detection
        omml_present = False
        omml_ok      = True   # BUG-A05 fix: start True, AND-reduce below

        # ── GAP-2026-08-15-BAREQ, defect F-2 — THE STEM PARAGRAPH IS A PARAGRAPH TOO ──
        # Until 2026-08-15 stem_parts[0] was built from `text`, i.e. para.text, which is
        # <w:t>-only. Every CONTINUATION line went through enrich_paragraph_with_omml();
        # the stem paragraph — the one paragraph guaranteed to exist for every question —
        # did not. So the equation in a stem paragraph, which is the NORMAL shape for a
        # mathematics, physics, statistics or quantitative-aptitude paper, was silently
        # dropped from full_stem, and therefore from 'stem', 'stem_raw', detect_is_msq(),
        # detect_blank_position(), the negative-question test, the dedupe/taxonomy keys,
        # PYQ_STEM_PATTERNS and every mock question derived from the item.
        #
        # Measured on IIT_JAM_MATHEMATICS 12-Feb-2017: 46 of 60 stems (77%) lost part or
        # all of their mathematics, and a stem that was ENTIRELY OMML extracted as ''.
        #
        # omml_present was computed from the body loop alone, so a question whose maths
        # lives only in the stem paragraph was recorded omml_present=False — which made
        # QV-8 (OMML recovery >= 80%) skip exactly those questions (om == [] -> PASS).
        # Producer and consumer sharing a blind spot is what made this silent, the same
        # shape as the Q-detector defect above. Seeding both flags from the stem closes it.
        # v2.56 (GAP-2026-08-29-STYLE-FIDELITY §6.1.6 / E-13): the stem TEXT comes from
        # corpus_io.text_of(para, omml_renderer=omml_to_linear) so <w:t> and <m:t> interleave in DOCUMENT ORDER —
        # enrich_paragraph_with_omml is retained ONLY for the omml_present/omml_ok
        # flags (its appended-at-end text is discarded).
        _, _stem_ok, _stem_has_omml = enrich_paragraph_with_omml('', para)
        _stem_txt = re.sub(r'^Q\.?\d+\.?\s*', '', corpus_io.text_of(para, omml_renderer=omml_to_linear)).strip()
        if _stem_has_omml:
            omml_present = True
            omml_ok      = omml_ok and _stem_ok
        stem_parts   = [_stem_txt]
        i += 1

        while i < len(paras):
            nt = paras[i].text.strip()
            if not nt: i += 1; continue
            # ── GAP-2026-08-16-PYQEXTRACT-DATE-LABEL-POSITION ────────────────────
            # A PYQSort position label ALWAYS precedes the next question and can never
            # be part of THIS question's body. Terminating here — rather than absorbing
            # it as a stem continuation — is what lets the OUTER loop refresh
            # cur_date_label PER QUESTION. Without it the outer loop's label branch is
            # UNREACHABLE after the first question of a taxonomy block, so every
            # question in that block inherits the block's FIRST exam position and the
            # label text is concatenated into the preceding stem.
            # MEASURED (IIT_JAM_MATHEMATICS, 3 papers, declared A=90/B=30/C=60): bands
            # read 150/17/13; is_msq 18 against 30; 179 of 180 stems carried embedded
            # label text. After this line: 90/30/60 exact, is_msq 30, 0 pollution,
            # holding to 884/884 across the full 22-paper corpus.
            # WHY THE EXISTING TERMINATORS CANNOT CATCH IT: a label is not a question
            # start, and is_taxonomy_heading() fires on 0 of 60 labels (measured). The
            # loop was structurally incapable of stopping here; this is not tuning.
            # THIS TEST MUST SIT ABOVE THE HEADING TEST. A label is definitionally not
            # a heading, and evaluating it there would let a mis-inference increment
            # QV-15's terminated_by_heading on a paragraph that can never be one.
            # BREAK WITHOUT ADVANCING i. Control returns to the outer loop with i still
            # on the label; the outer loop re-reads it, matches its own branch, assigns
            # and advances. This mirrors the question-start break exactly and keeps
            # cur_date_label at ONE writer. Termination is safe because both tests are
            # the SAME engine predicate, so if the inner fires the outer necessarily
            # fires — divergence is the only spin risk and one definition makes it
            # unreachable. Cost: one redundant predicate call per question.
            if bc.is_position_label(nt):
                break
            # GAP-2026-07-26-001: nxt[i] is what stops a bold STEM CONTINUATION from
            # terminating its own question here. Without it the stem is truncated and
            # every option after this point is silently discarded.
            # _ended_by_heading is initialised to False before this loop (see above),
            # so a question whose body loop never runs can never leave it unbound.
            _ended_by_heading = (detect_question_start(nt) is None and
                                 is_taxonomy_heading(paras[i], nxt[i], colour_ok))
            if detect_question_start(nt) is not None or _ended_by_heading:
                if _ended_by_heading:
                    terminated_by_heading += 1   # QV-15 (GAP-2026-08-05-001)
                break
            if is_option(nt, paras[i]):
                options.append(clean_option_text(nt))
                options_raw.append(nt)   # v2.15: preserve raw line for label detection
            else:
                _, ok, has_omml = enrich_paragraph_with_omml('', paras[i])
                if has_omml:
                    omml_present = True
                    omml_ok      = omml_ok and ok  # BUG-A05 fix: AND not replace
                    # v2.56 §6.1.6 / E-13: in-order <w:t>+<m:t>, not appended-at-end
                    stem_parts.append(corpus_io.text_of(paras[i], omml_renderer=omml_to_linear).strip())
                else:
                    stem_parts.append(nt)
            i += 1

        full_stem  = ' '.join(stem_parts)
        clean_stem, note, note_found = extract_note_block(full_stem)
        blank_pos  = detect_blank_position(clean_stem)
        is_neg     = bool(re.search(r'\b(NOT|INCORRECT|EXCEPT|FALSE|WRONG)\b', full_stem))
        # multi_select is auto-detected in S1-3 from exam pattern + PYQ stems.
        # v2.5 EC-A ROOT-CAUSE FIX: the old detector (r'select all|which.*are correct')
        # false-matched statement-combination MCQs (EC-9: "Which is/are correct?" with
        # combo-label options "1. Only A", "3. Both A and B"), which are SINGLE-answer.
        # The forgery-resistant signal is OPTION SHAPE, not stem wording:
        #   - a stem with a genuine multi-select instruction phrase, AND
        #   - options that are NOT predominantly combination-labels (only/both/neither/
        #     "X and Y"/none-of/all-of) ⇒ MSQ.
        # A statement-combination MCQ (most options are combo-labels) is NEVER MSQ even
        # when its stem says "are correct". Validated empirically (both directions).
        # v2.39 (GAP-2026-07-27-E): resolve the exam position stamped by PYQSort v1.18
        # into a declared question_type. Both helpers live in corpus_io Cluster Q — the
        # same engine PYQSort delegates its stamp to — so writer and reader cannot drift.
        _orig_q = corpus_io.parse_original_q_num(cur_date_label)
        _pos_ty = corpus_io.question_type_for_position(_orig_q, marking_scheme)
        is_msq = bool(multi_select) and detect_is_msq(full_stem, options,
                                                      positional_type=_pos_ty)

        # BUG-D07 fix (v2.15): detect option label style per question
        # from the RAW option lines (before clean_option_text strips the label).
        # This is the LABEL style ('1/2/3/4' vs 'A/B/C/D' etc.), distinct from
        # option FORMAT type ('single_value' etc.) which describes content shape.
        _label = _detect_option_label_style(options_raw)

        # GAP-2026-08-05-001 (QV-15). Record whether an INFERRED heading ended this
        # question's body. When the inference is wrong the stem is truncated and every
        # remaining option is discarded — silently, because nothing counted it. Stamped
        # here so QV-15 can test it per question rather than only in aggregate.
        questions.append({
            'num'         : q_num,
            'terminated_by_heading': _ended_by_heading,
            # v2.37 (GAP-2026-07-26-003 PART 9). BOTH keys are stamped at the ONE
            # emission point. This dict was emitted with 'num' only, while
            # blueprint_core.paper_eras_from_progress() reads q.get('q_num') and the
            # S-SECMAP block reads _q['q_num'] behind a "'q_num' not in _q: continue"
            # guard. The era classifier therefore collected None for every question,
            # filtered to an empty list and labelled all 22 reference papers on a
            # question count of ZERO; S-SECMAP's Stage-1 observation was empty, so
            # every subject fell through to the Stage-3 fallback and every section
            # received every subject. Both failures were SILENT.
            # EC-V15 also depends on this: the vision queue is keyed (paper_id, q_num),
            # and a bare q_num collides across papers — measured on the reference
            # corpus, 153 figural questions carry only 62 distinct q_num values, so
            # keying on q_num alone would have mis-attributed 91 of them.
            'q_num'       : q_num,
            # v2.39 (GAP-2026-07-27-E + GAP-X residual). The ORIGINAL exam position and
            # the type declared for its band are persisted, not just consumed. Two
            # reasons: classify_paper_era()'s type-corroboration branch was inert
            # because question_type never reached the persisted dict (measured:
            # type_checked False on all 22 reference papers, so era rested on question
            # COUNT alone); and a re-synthesis from progress.json alone can now redo
            # positional typing without re-reading any docx. None means "unknown" — a
            # pre-v1.18 sorted file, or a position no band covers — never "position 0".
            'original_q_num': _orig_q,
            'question_type' : _pos_ty,
            'stem'        : clean_stem,
            'stem_raw'    : full_stem,
            'options'     : options,
            'section'     : cur_sec,
            'topic'       : cur_top,
            'subtopic'    : cur_sub,
            'has_note'    : note_found,
            'note_text'   : note,
            'blank_pos'   : blank_pos,
            'is_negative' : is_neg,
            'is_msq'      : is_msq,
            'omml_present': omml_present,
            'omml_failed' : omml_present and not omml_ok,
            'option_label': _label,   # v2.15 BUG-D07: '1/2/3/4' | 'A/B/C/D' | etc.
        })
    return questions

# ═══ FROM Framework_MockTestAnalyse.md §5, fence L2667-2676 (v2.53.2) — VERBATIM ═══
def pre_synthesis_check(progress, taxonomy, target='ALL'):
    for sec, entries in taxonomy.items():
        if target != 'ALL' and sec != target: continue
        for e in entries:
            key   = (sec, e['topic'], e['subtopic'])
            count = len(progress.get(key, []))
            if count == 0:  print(f"ABSENT: {e['subtopic']} -- no PYQ data")
            elif count < 3: print(f"SPARSE: {e['subtopic']} -- only {count} Qs (inferred)")

# ═══ FROM Framework_MockTestAnalyse.md §5, fence L2680-3614 (v2.53.2) — VERBATIM ═══
# ════════════════════════════════════════════════════════════════════════════
# AXIS CLASSIFIER v1.0  (v2.23 — SHARED SINGLE SOURCE OF TRUTH)
# ────────────────────────────────────────────────────────────────────────────
# The canonical, exam-agnostic classifier for the three orthogonal format axes.
# audit_canonical.py MUST re-tag GENERATED questions with THESE SAME
# functions (import/copy verbatim, never re-implement) — the PYQ distribution and
# the generated distribution are only comparable if classified identically.
#
#   Axis 1  STIMULUS/MEDIA  : TEXT | FIGURAL | PASSAGE | DI
#   Axis 2  STEM STRUCTURE  : the exclusive 8-class ladder (below)
#   Axis 3  ANSWER MECHANISM: MCQ | MSQ | NAT
#   negative polarity       : an ORTHOGONAL boolean flag (is_negative), NOT a class
#
# EXCLUSIVITY: every question maps to exactly ONE Axis-2 class (first-match-wins).
# LINKED is a GATE decided by shared-stimulus membership (linked_group_id), never by
# phrasing — an assertion-reason question printed inside a passage is LINKED, and its
# inner shape becomes a secondary detail only. SEQUENCE sits ABOVE STATEMENT because
# "arrange the following statements in order" is fundamentally an ordering task.
# ════════════════════════════════════════════════════════════════════════════

# v2.56 (GAP-2026-08-29-STYLE-FIDELITY §6.1.4 / P-7): AXIS2_CLASSES is defined ONCE,
# in blueprint_core, and imported here. 8 -> 11: IDENTIFY, SELECT_PLOT, RANK inserted
# in ladder order after SEQUENCE and before STATEMENT.
AXIS2_CLASSES = bc.AXIS2_CLASSES                          # ladder order == precedence

# Canonical stem_format_variant → Axis-2 class map. Step 7's STEM_FORMAT_MENU tokens
# MUST map through this table (File 4 wires Step 7 to it) so capability stays consistent.
STEM_FORMAT_TO_AXIS2 = {
    'direct_question': 'DIRECT', 'isolated_word': 'DIRECT', 'phrase_to_word': 'DIRECT',
    'reverse_word_to_phrase': 'DIRECT', 'definition_to_word': 'DIRECT',
    'meaning_of_idiom': 'DIRECT', 'idiom_for_situation': 'DIRECT',
    'sentence_substitution': 'DIRECT', 'sentence_embedded_underlined': 'DIRECT',
    'fill_blank': 'FILL_BLANK', 'fill_in_context_blank': 'FILL_BLANK',
    'assertion_reason': 'ASSERTION_REASON', 'match_the_following': 'MATCH',
    'statement_correctness': 'STATEMENT', 'sequence_ordering': 'SEQUENCE',
    'odd_one_out': 'ODD_ONE_OUT', 'odd_one_out_pair': 'ODD_ONE_OUT',
}

# family → set of Axis-2 classes its Step-7 menu can faithfully render (derived from
# Step 7 STEM_FORMAT_MENU via STEM_FORMAT_TO_AXIS2; kept explicit for round-trip clarity).
FAMILY_AXIS2_MENU = {
    'vocab_single_word'    : {'DIRECT', 'FILL_BLANK', 'ODD_ONE_OUT'},
    'one_word_substitution': {'DIRECT'},
    'idiom_phrase'         : {'DIRECT', 'ODD_ONE_OUT'},
    'fact_recall'          : {'DIRECT', 'FILL_BLANK', 'ASSERTION_REASON', 'MATCH',
                              'ODD_ONE_OUT', 'STATEMENT'},
}

_TABLE_WORD_RE = re.compile(r'(?i)\b(table|tabulated|following data|dataset)\b')
def _looks_like_table_stimulus(stem):
    """v2.24.6 FIX C — SHARED, single-source-of-truth structural table detector.
    Was, independently in TWO places (classify_axis1 here and synthesise_subtopic's
    has_tbl), a naive substring match `'|' in stem or 'table' in stem.lower()` that
    false-positived on any word merely CONTAINING "table" — "vegetable", "acceptable",
    "notable" — with no real tabular data present. Requires either (a) >=2 pipe-delimited
    rows (a real rendered table), or (b) a word-boundary table-keyword match co-occurring
    with >=1 pipe-delimited row. A bare stray '|' or an unrelated "table"-containing word
    alone no longer qualifies. MUST PROPAGATE (byte-identical) to audit_canonical.py's
    verbatim classifier copy — same requirement as classify_axis2's MATCH rule.
    """
    stem = stem or ''
    pipe_rows = sum(1 for ln in stem.splitlines() if ln.count('|') >= 2)
    return pipe_rows >= 2 or (bool(_TABLE_WORD_RE.search(stem)) and pipe_rows >= 1)

def classify_axis1(q):
    """STIMULUS/MEDIA. Priority FIGURAL > PASSAGE > DI > TEXT — identical ordering to
    the per-subtopic `fmt` line in synthesise_subtopic (a linked DI passage resolves
    PASSAGE, matching that function)."""
    if q.get('image_role', 'none') not in ('none', None):
        return 'FIGURAL'
    if q.get('linked_group_id'):
        return 'PASSAGE'
    stem = (q.get('stem') or q.get('stem_raw') or '')
    if _looks_like_table_stimulus(stem):
        return 'DI'
    return 'TEXT'

def classify_axis3(q):
    """ANSWER MECHANISM. NAT = no selectable options (mirrors the answer_type=='numerical'
    detection: zero text options AND no option-images). MSQ = is_msq. Else MCQ."""
    opts = q.get('options', []) or []
    if len(opts) == 0 and q.get('image_role', 'none') not in ('options_only', 'stem_and_options'):
        return 'NAT'
    return 'MSQ' if q.get('is_msq') else 'MCQ'

def _opts_are_combination_labels(opts):
    """EC-A signal: options predominantly combination-labels (Only N / Both N and M /
    Neither…nor / None of / All of the above / "N and M"). Distinguishes STATEMENT and
    MATCH combo-answer stems from genuine free-form options."""
    if not opts:
        return False
    combo = 0
    for o in opts:
        t = (o or '').strip().lower()
        if re.search(r'\b(only|both|neither|none of|all of)\b', t) or \
           re.match(r'^[a-d1-4](\s*(and|,|&|-)\s*[a-d1-4])+$', t):
            combo += 1
    return combo >= max(2, (len(opts) + 1) // 2)

# ── MATCH option-shape backstop (v2.24.2) ──────────────────────────────────────
# A language-agnostic MATCH signal: the OPTIONS are a set of CROSS-DOMAIN label pairs
# (e.g. "A-I, B-III, C-IV, D-II" / "1-C 2-A 3-D 4-B" / "(A)-(i), (B)-(iv) ..."). It fires
# when the stem keywords (match / list-I / column) are ABSENT — the two cases that matter:
#   (a) NON-ENGLISH match papers (Hindi/regional), whose stems carry no English cue;
#   (b) matches whose List-I/List-II body has been rendered into a Word table, so the list
#       labels no longer appear in stem_raw (only the "Match ..." instruction does).
# CROSS-DOMAIN (left label family != right label family) is REQUIRED, so digit:digit ratios
# ("2:3, 4:5"), coordinate pairs and word-word hyphenations never trip it. The family of a
# COLUMN (not a single token) is used so the roman-vs-letter "I" ambiguity resolves from
# context (a column carrying II/III/IV is roman even where a bare I appears).
_MATCH_PAIR_RE = re.compile(
    r'\(?\s*([A-Za-z]{1,4}|\d{1,2})\s*\)?\s*[-\u2010-\u2015:\u2192>]+\s*'
    r'\(?\s*([A-Za-z]{1,4}|\d{1,2})\s*\)?')
_MATCH_PAIR_SUB = (r'\(?\s*(?:[A-Za-z]{1,4}|\d{1,2})\s*\)?\s*[-\u2010-\u2015:\u2192>]+\s*'
                   r'\(?\s*(?:[A-Za-z]{1,4}|\d{1,2})\s*\)?')
_MATCH_OPT_RE = re.compile(r'^\s*' + _MATCH_PAIR_SUB + r'(?:[,;\s]+' + _MATCH_PAIR_SUB + r'){1,}\s*$')

def _label_family(tokens):
    """Family of a same-side label COLUMN: 'digit' | 'roman' | 'alpha' | 'other'.
    Column-level (not per-token) so a bare 'I' resolves to roman when its column also
    carries II/III/IV, and to alpha when its column is A/B/C/D."""
    low = [t.lower() for t in tokens if t]
    if not low:
        return 'other'
    if all(re.fullmatch(r'\d{1,2}', t) for t in low):
        return 'digit'
    romanish = all(re.fullmatch(r'[ivxlcdm]+', t) for t in low)
    if romanish and any(len(t) > 1 for t in low):
        return 'roman'
    if all(re.fullmatch(r'[a-z]', t) for t in low):
        return 'roman' if set(low) <= {'i', 'v', 'x'} else 'alpha'
    if romanish:
        return 'roman'
    if all(re.fullmatch(r'[a-z]{1,4}', t) for t in low):
        return 'alpha'
    return 'other'

def _opts_are_match_pairs(opts):
    """True when a MAJORITY of options are each a set of >=2 CROSS-DOMAIN label pairs that
    consume the whole option text. Threshold mirrors _opts_are_combination_labels. Used by
    classify_axis2 AFTER the keyword rules, so it can only convert a would-be non-MATCH
    class to MATCH, never the reverse (additive + monotone)."""
    if not opts:
        return False
    hits = 0
    for o in opts:
        st = (o or '').strip()
        if not st or not _MATCH_OPT_RE.match(st):
            continue
        pairs = _MATCH_PAIR_RE.findall(st)
        if len(pairs) < 2:
            continue
        lf = _label_family([p[0] for p in pairs])
        rf = _label_family([p[1] for p in pairs])
        if lf == rf or 'other' in (lf, rf):
            continue
        hits += 1
    return hits >= max(2, (len(opts) + 1) // 2)

def _axis2_option_shape(q):
    """v2.56 — option shape as seen by the Axis-2 ladder: the stamped value when
    present, else computed. Kept tiny so classify_axis2 stays order-pure."""
    return q.get('option_shape') or detect_option_shape(q)


def classify_axis2(q):
    """STEM STRUCTURE — the exclusive 11-class ladder (first-match-wins; v2.56). Discrimination
    is by task-verb + option-shape, not ladder position alone, so collisions are rare and
    deterministic. Grounded in EC-8/9/11/12/13; SEQUENCE + ODD_ONE_OUT added in v2.23."""
    # GATE 0 — LINKED: structural, decided by shared-stimulus membership, not phrasing.
    if q.get('linked_group_id'):
        return 'LINKED'
    stem = (q.get('stem_raw') or q.get('stem') or '')
    s    = stem.lower()
    opts = q.get('options', []) or []
    # 1 — ASSERTION_REASON (EC-8): both an Assertion and a Reason clause present.
    if re.search(r'\bassertion\b', s) and re.search(r'\breason\b', s):
        return 'ASSERTION_REASON'
    # 2 — MATCH (EC-13): match/list-I/column stems, OR (v2.24.2) a CROSS-DOMAIN label-pair
    #     option shape. The option-shape backstop is language-agnostic and table-safe (see
    #     _opts_are_match_pairs): it catches non-English matches and matches whose List-I/
    #     List-II body has moved into a Word table. Placed AFTER the keyword rules it is
    #     additive/monotone — it only converts a would-be non-MATCH class to MATCH.
    if re.search(r'\bmatch\b', s) and re.search(r'\b(following|list|column|set)\b', s):
        return 'MATCH'
    if re.search(r'list[\s\-]*i\b|column[\s\-]*(i|a)\b', s):
        return 'MATCH'
    if _opts_are_match_pairs(opts):
        return 'MATCH'
    # 3 — SEQUENCE / ORDERING (v2.23): the OPERATION is arranging (kept above STATEMENT).
    if re.search(r'\b(arrange|rearrange|correct sequence|proper sequence|correct order|'
                 r'logical order|chronological order|sequence of the following|'
                 r'order of the following)\b', s):
        return 'SEQUENCE'
    # 4a — IDENTIFY (v2.56, §6.1.4): an identify-ask whose option shape carries the
    #      meaning — structures, figures, expressions. Detected by cue + shape, so a
    #      plain "which of the following is correct" prose question never lands here.
    if (re.search(r'\bidentify\b|which of the following (is|are) the'
                  r'( correct| major| final)?\s*(product|structure|intermediate|'
                  r'compound|species|expression|form)\b', s)
            and _axis2_option_shape(q) in ('structure_image', 'figure', 'expression')):
        return 'IDENTIFY'
    # 4b — SELECT_PLOT (v2.56): choose-the-graph questions; plot options or an
    #      explicit graph/plot/curve selection ask.
    if (_axis2_option_shape(q) == 'plot_image'
            or re.search(r'which (of the following )?(graph|plot|curve|variation)', s)):
        return 'SELECT_PLOT'
    # 4c — RANK (v2.56): ordering by a property WITHOUT an arrange-operation stem
    #      (those stay SEQUENCE above, byte-preserving the v2.55 ladder).
    if re.search(r'\b(increasing|decreasing|ascending|descending)\s+order\b', s) \
            and max(len(opts), len(_named_entities(stem))) >= 3:
        return 'RANK'
    # 4 — STATEMENT-BASED (EC-9, WIDENED v2.56 §6.1.4): fires when (a) >= 2 options
    #     are full sentences (and the stem carries no cloze blank — a sentence-
    #     completion cloze stays FILL_BLANK), OR (b) the stem cue statement(s)
    #     appears anywhere — labelled "Statement (A):" blocks included — OR
    #     (c) options are combination labels. The v2.55 phrasing-locked detector is
    #     subsumed by (b).
    if (re.search(r'\bstatements?\b', s) and opts) \
       or _opts_are_combination_labels(opts) \
       or (sum(1 for o in opts if _is_full_sentence_option(o)) >= 2
           and q.get('blank_pos', 'none') in ('none', None)
           and not re.search(r'_{3,}', s)):
        return 'STATEMENT'
    # 5 — FILL_BLANK / CLOZE (EC-11): a blank to complete.
    #
    # v2.39 (GAP-2026-07-27-F). Gated on the question HAVING OPTIONS. Every NAT question
    # carries an answer-entry blank as a artefact of the current-era answer line, and
    # detect_blank_position() cannot tell that blank from a cloze gap in the stem.
    #
    # MEASURED (IIT_JAM_BIOTECHNOLOGY, 1719 Qs): FILL_BLANK fired on 218 of 261 NAT
    # questions (83.5%) against 37 of 1434 MCQ (2.6%) — a 32x enrichment that tracks the
    # answer mechanism, not the question form. The decisive control is batch 7: across
    # 300 legacy-era questions, which carry almost no NAT, FILL_BLANK moved by exactly 1.
    #
    # The axes are orthogonal so nothing downstream is corrupted TODAY, but the NAT
    # section's Axis-2 profile reads as ~100% cloze, and Step 7 rendering that literally
    # emits NAT stems as fill-in-the-blank.
    #
    # A NAT question with a genuine cloze stem is real but rare, and is not recoverable
    # from the artefact — it is deliberately classified by its residual form instead of
    # being asserted on evidence that cannot distinguish the two.
    if (opts and (q.get('blank_pos', 'none') not in ('none', None)
                  or re.search(r'_{3,}|\bfill in the blank', s))):
        return 'FILL_BLANK'
    # 6 — ODD_ONE_OUT: genuine "which does not belong" classification (narrowed — mere
    #     negative phrasing is is_negative, handled orthogonally, not this class).
    if re.search(r'\bodd one out\b|does not belong|which one is different|find the odd', s):
        return 'ODD_ONE_OUT'
    # 7 — DIRECT: residual floor.
    return 'DIRECT'

def tag_axes(q):
    """Attach the three exclusive axis labels to a question dict in place. is_negative is
    already set during extraction (EC-12). Idempotent."""
    q['axis1'] = classify_axis1(q)
    q['axis2'] = classify_axis2(q)
    q['axis3'] = classify_axis3(q)
    return q

# family keyword → family key (Step-5 approximation of Step 7 resolve_presentation_family;
# Step 7 refines with CONCEPT_GROUP at its S3-8 join. Exam-agnostic keyword sets.)
_FAMILY_KEYWORDS = {
    'vocab_single_word'    : ('antonym', 'synonym', 'spelling', 'homonym'),
    'one_word_substitution': ('one word substitution', 'one-word'),
    'idiom_phrase'         : ('idiom', 'phrase'),
    'fact_recall'          : ('gk', 'general awareness', 'general knowledge', 'static',
                              'current affairs', 'fact'),
}

def resolve_presentation_family_s5(subtopic, fmt):
    """Lightweight family resolution from the subtopic name + format. Returns a family key
    or None. Mirrors Step 7 PRESENTATION_FAMILIES; used only to seed axis2_capability —
    Step 7 remains authoritative once CONCEPT_GROUP is joined."""
    name = (subtopic or '').lower()
    for fam, kws in _FAMILY_KEYWORDS.items():
        if any(kw in name for kw in kws):
            return fam
    return None

def axis2_capability(observed_axis2, presentation_family, fmt):
    """The Axis-2 forms a subtopic may FAITHFULLY take (decisions (b)/(c)):
       observed ∪ family-menu ∪ {DIRECT}, with LINKED added iff the subtopic is
       stimulus-linked (format PASSAGE/DI) and removed otherwise. Forcing a form OUTSIDE
       this set is fabrication (Step 7 decision-(iii) ban)."""
    cap = set(observed_axis2) | {'DIRECT'}
    cap |= FAMILY_AXIS2_MENU.get(presentation_family, set())
    if str(fmt).upper() in ('PASSAGE', 'DI'):
        cap.add('LINKED')
    else:
        cap.discard('LINKED')
    return [c for c in AXIS2_CLASSES if c in cap]   # canonical ladder order

def compute_section_axis_distribution(sec_entries, progress, mocks_per_window=10,
                                      window_years=None):
    """CATEGORY A per-section target. Averages each axis's class counts PER PAPER over the
    N most-recent distinct years, and classifies each Axis-2 class band vs guarantee-only.
    Returns None for a section with no observed questions (all Zero-PYQ scaffolds).

    v2.26 — N is bc.AXIS_WINDOW_YEARS (5), raised from a hardcoded 3. Two separate
    questions must not share one sample:
      • HOW MANY of a class a mock gets  → this window (5 years). A wider window is a
        steadier average; on the reference exam the figural mean moved only 4.33 → 4.40,
        which is the point — the number should not lurch on one unusual paper.
      • WHICH subtopics may carry it     → the FULL corpus (figural_rate, below). Most
        subtopics appear 1-3 times per paper, so a 5-paper sample leaves their denominator
        far too small to rate; the reference exam's Population Genetics reads 0/8 over five
        years and 1/16 over twenty-two.

    The §S1-3 `get_detection_sample()` window stays at 3 ON PURPOSE and is NOT this
    constant: it detects the CURRENT PATTERN (option format, section layout, marking), and
    a layout that changed four years ago must not pollute today's reading of the paper.

    ERA-SCOPING STILL RUNS FIRST (frequency_scope == 'current-era', §14). A wider window
    can therefore never straddle a pattern change — it widens only WITHIN the current era.
    """
    import blueprint_core as bc
    window_years = int(window_years or bc.AXIS_WINDOW_YEARS)
    qs = []
    for e in sec_entries:
        qs.extend(progress.get((e['section'], e['topic'], e['subtopic']), []))
    if not qs:
        return None
    for q in qs:                                   # ensure tagged (idempotent)
        if 'axis2' not in q:
            tag_axes(q)
    years   = sorted({q.get('year') for q in qs if q.get('year')}, reverse=True)
    recentN = set(years[:window_years]) if years else set()
    rq      = [q for q in qs if q.get('year') in recentN] or qs
    n_papers = max(1, len({(q.get('year'), q.get('shift')) for q in rq}))

    def per_paper(axis):
        c = Counter(q.get(axis, '?') for q in rq)
        return {k: round(v / n_papers, 3) for k, v in c.items()}

    axis2 = per_paper('axis2')
    audit_mode = {}
    for cls, avg in axis2.items():
        if cls == 'DIRECT':
            audit_mode[cls] = 'float'              # residual filler — never audited (decision 5/10)
        else:
            audit_mode[cls] = 'band' if avg * mocks_per_window >= 1 else 'guarantee'
    # v2.44 — the raw per-paper figural counts and the per-subtopic totals, carried
    # through so Step 6 can build the target SERIES and the per-subtopic figure quota,
    # and so the auditor's band is this exam's own volatility. Without these the band is
    # a fixed percentage, which rejected four of the reference exam's five real papers.
    # PAPER KEY: str() everything. sorted() over a mixed {2026, None} set raises
    # TypeError in Python 3 ("'<' not supported between int and NoneType"), which is
    # reachable the moment ONE question in the corpus lacks a year or paper_id — a
    # scan gap, a hand-added question, a legacy row. That would take out the whole of
    # Step 5 for an entire exam. A missing key becomes its own bucket rather than a crash.
    def _paper_key(q):
        return str(q.get('paper_id') or q.get('year') or '__unknown__')

    # SUBTOPIC KEY: subtopic_id ONLY, never the display name. Step 7 matches these keys
    # against blueprint subtopic_ids; a display-name key matches NOTHING, so the mock
    # would silently render ZERO figures with every fixture green. Falling back to the
    # display name looks defensive and is the more dangerous failure — a hard skip of the
    # unkeyed question keeps the quota honest and is visible in the totals.
    # v2.46 — MEASURE EVERY STIMULUS CLASS, NOT JUST FIGURAL. v2.44/v2.45 measured
    # figures per paper and per subtopic and left DI and PASSAGE with a rate and nothing
    # else, so Step 6 could build the quota/series chain for FIGURAL only. On a DI-heavy
    # exam that means DI reaches its budget but ignores each subtopic's measured DI
    # frequency. Measuring per class costs one loop and lets a future class inherit the
    # whole chain rather than needing its own release.
    def _class_of(q):
        if q.get('image_role', 'none') != 'none':
            return 'FIGURAL'
        if q.get('linked_group_id'):
            return 'PASSAGE'
        if _looks_like_table_stimulus(q.get('stem', '')) or q.get('has_rendered_table'):
            return 'DI'
        return 'TEXT'

    # GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE (D2). These three lines are
    # MODULE-QUALIFIED — `collections.X`, not the bare `Counter` used everywhere
    # else in §5 — but `import collections` appears NOWHERE in this file. Every
    # other import is the `from collections import ...` form, which binds
    # `Counter` / `defaultdict` and never the module name; the sole
    # `import collections as _c` (§8) is function-local to another function.
    # So this function raised `NameError: name 'collections' is not defined`
    # every time it ran, from v2.46.0 (2026-08-06) onward. write_section_rules
    # calls it, so section_rules.md — the primary Step 5 artefact — was never
    # written.
    # WHY IT SURVIVED: spec_name_audit DID detect it. `collections` sits in
    # spec_name_audit_baseline.json under this file, accepted as a known unbound
    # name, so the ratchet reported OK. A baseline entry cannot distinguish "this
    # name is legitimately bound elsewhere at runtime" from "this name is a
    # guaranteed NameError" — see the typed-reason baseline added in this release.
    # The import is function-local, matching this file's dominant idiom.
    import collections
    _by_paper = collections.defaultdict(collections.Counter)   # cls -> paper -> n
    _by_sub = collections.defaultdict(collections.Counter)     # cls -> subtopic_id -> n
    _unkeyed = collections.Counter()                           # cls -> n
    for _q in rq:
        _c = _class_of(_q)
        if _c == 'TEXT':
            continue
        _by_paper[_c][_paper_key(_q)] += 1
        _sid = _q.get('subtopic_id')
        if _sid:
            _by_sub[_c][str(_sid)] += 1
        else:
            _unkeyed[_c] += 1
    _papers = sorted({_paper_key(_q) for _q in rq})
    _obs_by_class = {c: [_by_paper[c].get(p, 0) for p in _papers] for c in _by_paper}
    _mean_by_class = {c: (sum(v) / len(v) if v else 0.0) for c, v in _obs_by_class.items()}

    _fig_by_paper = _by_paper['FIGURAL']
    _fig_by_sub = _by_sub['FIGURAL']
    _fig_unkeyed = _unkeyed['FIGURAL']
    _per_paper = [_fig_by_paper.get(_p, 0) for _p in _papers]
    return {
        'figural_per_paper_observed' : _per_paper,                       # v2.44
        'figural_per_paper_mean'     : (sum(_per_paper) / len(_per_paper)) if _per_paper else 0.0,
        'figural_count_by_subtopic'  : dict(_fig_by_sub),                # v2.44
        # v2.46 — per-class views, consumed by derive_axis_schedule to build the
        # quota/series chain for EVERY stimulus class rather than FIGURAL alone.
        'per_paper_observed_by_class': {c: list(v) for c, v in _obs_by_class.items()},
        'per_paper_mean_by_class'    : dict(_mean_by_class),
        'count_by_subtopic_by_class' : {c: dict(v) for c, v in _by_sub.items()},
        'unkeyed_questions_by_class' : dict(_unkeyed),
        'figural_unkeyed_questions'  : _fig_unkeyed,   # v2.45 — figural questions with
                                                       # no subtopic_id, excluded from the
                                                       # quota. Non-zero means the corpus
                                                       # lost keys upstream; surfaced so a
                                                       # shortfall has a visible cause
                                                       # instead of looking like a bug here.
        'recent_years'    : sorted(recentN, reverse=True),
        'window_years'    : window_years,     # v2.26 provenance — Step 6 echoes this into
                                              # blueprint.axis_schedule.axis_window_years so
                                              # a target can always be traced to its sample.
        'n_papers_recent' : n_papers,
        'mocks_per_window': mocks_per_window,
        'axis1_per_paper' : per_paper('axis1'),
        'axis2_per_paper' : axis2,
        'axis3_per_paper' : per_paper('axis3'),
        'axis2_audit_mode': audit_mode,
        'negative_rate'   : round(sum(1 for q in rq if q.get('is_negative')) / len(rq), 3),
    }

# v2.44 — plain-majority threshold, both halves of the reducibility test (§S5-FIG).
# MODULE SCOPE deliberately: it is referenced inside synthesise_subtopic() before the
# line it was first written on, which is a NameError waiting for the first exam whose
# corpus reaches that branch. Not tunable per exam ON PURPOSE — a per-exam knob here is
# how an exemption rule gets quietly widened until it swallows the budget again.
FIGURAL_IRREDUCIBLE_RATE = 0.50


def synthesise_subtopic(section, topic, subtopic, questions, progress, figural_data=None,
                        nat_allowed=False):
    """
    BUG-A07 fix: progress added to function signature (cfg parameter removed — no config file).
    BUG-A10 fix: sub_format value ('Cloze' or 'RC') has no leading space.
    BUG-A13 fix: empty patterns from FIGURAL subtopics handled with placeholder.
    BUG-B09 fix: difficulty calibration stores is_inferred bool, not fragile string.
    BUG-B14 fix: figural subtopics with all-empty stems get placeholder pattern.
    BUG-B15 fix: §14 schema documents option_format as dict.
    """
    if not questions:
        return _absent_entry(section, topic, subtopic)

    # ── v2.56 (GAP-2026-08-29-STYLE-FIDELITY §6.1) — MEASURE, NEVER LOOK UP ─────
    # determine_strip_mode (name-table lookup) is RETIRED as an input here (Q9).
    # The content signature is measured from the questions; legacy_mode is DERIVED
    # from that signature — one of the five legacy names for an aptitude-class
    # signature (preserving the legacy masking vocabulary byte-for-byte, EC-26),
    # None for a content-class signature (v2 masking). Every per-question style
    # stamp below is computed on the question's own tokens, never a name.
    stamp_medium(questions)
    sig         = derive_content_signature(questions)
    legacy_mode = derive_legacy_mode(sig, questions=questions,
                                     exam_class=_exam_style_class(progress, sig),
                                     section=section, topic=topic, subtopic=subtopic)
    mode        = legacy_mode                    # None => content-class => v2 masks
    for q in questions:
        # ORDER IS LOAD-BEARING: option_shape must be stamped WITH the figural
        # descriptor BEFORE tag_axes runs, because the Axis-2 ladder reads the
        # shape (_axis2_option_shape) and the descriptor is what refines a bare
        # 'figure' into 'plot_image' / 'structure_image'. Tagging first made the
        # ladder read the UNREFINED shape: a choose-the-graph question with image
        # options classified DIRECT instead of SELECT_PLOT. (Found by reading the
        # diff line by line, 2026-08-31; fixture axis2_needs_figural_shape_first.)
        if 'option_shape' not in q:
            q['option_shape'] = detect_option_shape(q, figural_data)
        if 'axis2' not in q:
            tag_axes(q)
        if q.get('medium', 'en') == 'en':
            if 'mechanic' not in q:
                q['mechanic'] = detect_mechanic(q, figural_descriptor=figural_data)
            q['skeleton'] = strip_variables_v2(q.get('stem') or '', sig, legacy_mode)
        else:
            q.setdefault('mechanic', bc.MECHANIC_UNKNOWN)   # §6.1.9: counted, not guessed
            q.setdefault('skeleton', '')
    patterns = generate_templates(questions, mode, sig=sig)
    _annotate_patterns_with_style(patterns, questions)

    # BUG-A13 / BUG-B14 fix: handle empty patterns for figural or other edge cases
    if not patterns and questions:
        year_set = sorted(set(q.get('year') for q in questions if q.get('year')))
        patterns = [{
            'id':'P1', 'template':'(figural -- no text stem)',
            'frequency':100, 'raw_count':len(questions),
            'confidence':'observed', 'deprecated':False, 'years':year_set,
        }]

    # E-6: NOTE block analysis
    note_count  = sum(1 for q in questions if q.get('has_note'))
    notes_by_yr = {}
    for q in questions:
        if q.get('has_note') and q.get('year'):
            notes_by_yr.setdefault(q['year'], []).append(q.get('note_text',''))
    note_freq  = classify_note_frequency(note_count, len(questions))
    canon_note = canonical_note_text(notes_by_yr)

    for p in patterns:
        p['note_block'] = note_freq
        p['note_text']  = canon_note if note_freq in ('mandatory','conditional') else ''
        # v2.56 (E-3): the aptitude path keeps infer_approach VERBATIM (EC-26
        # byte-identity); the content path derives the approach from the pattern's
        # measured mechanic — never from template keywords.
        p['approach']   = (infer_approach(p['template'], mode, subtopic) if mode
                           else MECHANIC_TO_APPROACH.get(
                                   p.get('mechanic', bc.MECHANIC_UNKNOWN),
                                   MECHANIC_TO_APPROACH[bc.MECHANIC_UNKNOWN]))

    opt_fmt   = subtopic_option_format(questions)
    # PYQ_DIFFICULTY_CALIBRATION RETIRED (GAP-2026-08-27-DIFFICULTY-PROFILE): per-subtopic
    # calibration is read by Step 7 / ScopedBlueprint from the difficulty profile
    # (blueprint_core dp_calibration), measured by PYQExplain with the rubric.
    wrong_opt = classify_wrong_option_structure(questions)

    neg_ct  = sum(1 for q in questions if q.get('is_negative'))
    blank_d = Counter(q.get('blank_pos','none') for q in questions).most_common(1)[0][0]
    sw      = [len(q['stem'].split()) for q in questions if q.get('stem')]
    # ── v2.26 (GAP-2026-08-06-AXIS1) — FIGURAL IS A RATE, NOT A FLAG ────────────────
    # WAS: has_img = any(q.get('image_role','none') != 'none' for q in questions)
    #
    # `any()` is an EXISTENTIAL quantifier with no denominator: ONE figural question
    # anywhere in a 22-year corpus stamped the subtopic FIGURAL permanently. Because
    # Step 7 read `format` as a RENDERING IMPERATIVE ("format==FIGURAL ⇒ draw a
    # picture"), that flag then forced a figure onto EVERY mock question the subtopic
    # ever received. Measured on the reference exam: 46 of 131 subtopics flagged,
    # holding 42.7% of allocation weight, against a true question-level figural rate
    # of 7.3% — and the delivered mocks carried 26 and 30 figures against a blueprint
    # budget of 4. Both passed every gate, because no gate counted.
    #
    # COUNT QUESTIONS, NOT IMAGES. A `stem_and_options` question contributes ONE to
    # figural_q_count and FIVE to images_analysed (one stem figure + four option
    # figures). The two fields answer different questions and must never be confused:
    #   figural_q_count  → HOW MANY QUESTIONS carry a figure   → drives the Axis-1 budget
    #   images_analysed  → HOW MANY IMAGES were inspected      → drives the figure PROFILE
    # Reading images_analysed as a question count inflates the reference corpus from
    # ~118 figural questions to 154 and biases every rate toward option-figure subtopics.
    #
    # `format` SURVIVES UNCHANGED for back-compat (~200 deployed exams read it, and the
    # GOLDEN RULE still forbids it from EXCLUDING anything). What changes is its meaning
    # downstream: from "always draw this" to "CAPABLE of being drawn this way". The new
    # figural_rate decides which capable questions actually claim the scarce budget.
    _fig_qs  = [q for q in questions if q.get('image_role', 'none') != 'none']
    figural_q_count     = len(_fig_qs)
    figural_denominator = len(questions)
    figural_rate = (figural_q_count / float(figural_denominator)) if figural_denominator else 0.0

    # REDUCIBILITY. A question whose OPTIONS are themselves images (organic structures,
    # circuit diagrams, spectra) cannot be rewritten as text without becoming
    # unanswerable, so it is granted a figure even over budget (bc.axis_grant_figural
    # rule 2). A stem-only figure usually CAN be replaced by an attested text question
    # from the same subtopic — which is what REPLACEMENT_RULE below is for.
    # ── v2.44 (GAP-2026-08-06-IRREDUCIBLE) — REDUCIBILITY IS A RATE, NOT AN any() ──
    # WAS: figural_reducible = not any(q.image_role in ('stem_and_options',...))
    #
    # THE SAME EXISTENTIAL DEFECT THIS FILE EXISTS TO FIX, ONE FIELD TO THE LEFT.
    # v2.26 replaced `has_img = any(...)` with a rate, and then decided reducibility
    # with a fresh any(). One question anywhere in a 22-year corpus made a subtopic
    # PERMANENTLY EXEMPT FROM THE FIGURAL BUDGET — and irreducible grants pass even
    # over budget by design, so the exemption silently became the budget.
    #
    # MEASURED ON THE REFERENCE EXAM (2026-08-06, first real PYQExtract run):
    #     21 of 133 subtopics marked irreducible, forcing a mean of 14.3 figures per
    #     mock against a budget of 5; 13 of 15 mocks over budget; worst mock 29.
    #     Complex Formation was exempt on 1 figural question in 32 (3.1%).
    #     Chemical Bonding on 1 in 24. Trigonometry on 1 in 15.
    #     ELEVEN of the 21 carried aggregate image_role 'stem_only' — the figure is in
    #     the stem and the options are TEXT, which is the REDUCIBLE case by definition.
    #     They were exempt anyway, because one minority question had image options.
    #
    # THE CORRECT TEST IS 'ARE THE OPTIONS USUALLY IMAGES', WHICH NEEDS A DENOMINATOR.
    # An exemption must earn itself twice over: the subtopic has to be figure-dominant
    # AT ALL (figural_rate), and its figures have to live in the OPTIONS rather than the
    # stem (option_image_rate). A stem figure with text options can nearly always be
    # replaced by an attested text question from the same subtopic; option figures —
    # organic structures, circuit diagrams, spectra — cannot, because the answer set
    # itself is pictorial.
    #
    # Thresholds tested against the reference exam's real 15-mock allocation:
    #     rule                                   irreducible   mean forced   over budget
    #     any()                     (v2.26)         21 / 133       14.3        13 / 15
    #     aggregate role only                       10 / 133        6.1         7 / 15
    #     rate>=0.50 AND option-majority (THIS)       3 / 133        1.5         0 / 15
    # The three survivors are Alcohols/Aldehydes (71% figural), Aromatic Compounds
    # (71%) and Carboxylic Acids (56%) — every one a case where the options ARE the
    # structures. 0.50 is the plain majority line, deliberately not a tuned constant.
    _opt_img = [q for q in _fig_qs
                if str(q.get('image_role', '')).strip() in ('stem_and_options', 'options_only')]
    option_image_rate = (len(_opt_img) / float(len(_fig_qs))) if _fig_qs else 0.0
    figural_reducible = not (figural_rate >= FIGURAL_IRREDUCIBLE_RATE
                             and option_image_rate >= FIGURAL_IRREDUCIBLE_RATE)

    has_img  = figural_q_count > 0
    has_pass = any(q.get('linked_group_id') for q in questions)

    # ── v2.43 (GAP-2026-08-06-DI) — DI IS A RATE TOO ────────────────────────────
    # `has_tbl` below is the SAME existential quantifier that produced the figural
    # defect: one table-bearing question anywhere in the corpus marks the subtopic DI
    # forever, with no denominator. Measured on the reference exam, Electrochemistry
    # (14 observed questions) and Matrices & Determinants (19) are both flagged DI on
    # roughly one table each. DI is less damaging than FIGURAL was — DI and TEXT share
    # a rendering path, so the flag never FORCED a table the way format=FIGURAL forced
    # an image — but the same rate belongs here for the same reasons: it ranks claims
    # on the Axis-1 DI budget, and it stops a 1-in-19 subtopic outranking a genuine one.
    _di_qs = [q for q in questions
              if _looks_like_table_stimulus(q.get('stem', ''))
              or q.get('has_rendered_table')]
    di_q_count = len(_di_qs)
    di_rate = (di_q_count / float(len(questions))) if questions else 0.0
    # A DI question whose data table IS the question (a computation over supplied
    # values) has no table-free form; one that merely tabulates prose does. Absent a
    # per-question signal the conservative reading is reducible — the stricter of the
    # two, since it lets the budget bind.
    di_reducible = True
    # v2.24.6 FIX C — delegates to the SAME structural/word-boundary table detector used
    # by the canonical classify_axis1() (SHARED AXIS CLASSIFIER v1.0, below) — was a
    # locally-duplicated naive substring match (`'table' in stem.lower()`), which
    # false-positived on "vege**table**", "accep**table**", "no**table**", etc. Single
    # source of truth now; both DI derivations can never drift apart again.
    # v2.45 — THE LAST EXISTENTIAL OF THIS FAMILY. `has_tbl = any(...)` is the same
    # construct that produced GAP-2026-08-06-AXIS1 (format) and GAP-2026-08-06-
    # IRREDUCIBLE (reducibility): ONE table-bearing question anywhere in a 22-year
    # corpus marked the whole subtopic DI. On the reference exam that put
    # Electrochemistry (1 table in 14) and Matrices & Determinants (1 in 19) into DI
    # permanently. DI was less damaging than FIGURAL only because DI and TEXT share a
    # rendering path, so the flag never FORCED a table — an accident of wiring, not a
    # property of the rule, and it would bite on the first DI-heavy exam (banking or
    # CAT-style aptitude) where the class carries real weight.
    #
    # `di_rate` is already computed above from the same predicate. Use it, with the same
    # plain-majority line the reducibility test uses, and require at least 2 observed
    # table questions so a lone outlier can never define a subtopic's format.
    # A subtopic below the line keeps its DI questions in the corpus and its di_rate for
    # ranking; it simply stops being DECLARED a DI subtopic. Nothing is dropped.
    has_tbl = (di_q_count >= 2 and di_rate >= FIGURAL_IRREDUCIBLE_RATE)
    fmt      = ('FIGURAL' if has_img else 'PASSAGE' if has_pass else
                'DI' if has_tbl else 'TEXT')

    # v2.22 — INHERENTLY-VISUAL OVERRIDE (keyword heuristic):
    # Some subtopics are inherently visual by definition (e.g. counting figures,
    # embedded figures, mirror images). If PYQ image extraction failed (scanned
    # PDF, missing media), has_img is False and fmt becomes TEXT — a misclassification.
    # This heuristic checks the subtopic name against a universal visual-keyword set
    # and overrides to FIGURAL when a match is found AND fmt is currently TEXT.
    # Exam-agnostic — keywords cover geometric/spatial/visual terms across all exams.
    _VISUAL_KEYWORDS = re.compile(
        r'(?i)\b(figure[s]?|figural|diagram[s]?|venn|'
        r'mirror\s+image[s]?|water\s+image[s]?|paper\s*fold(ing)?|'
        r'counting\s+(figure|triangle|shape)[s]?|embedded\s+figure[s]?|'
        r'completion\s+of\s+(figure|pattern)[s]?|dice|cube\s+fold(ing)?|'
        r'pattern\s+completion|image\s+series|visual\s+reasoning)\b')
    _inherently_visual = False
    if fmt == 'TEXT' and _VISUAL_KEYWORDS.search(subtopic):
        fmt = 'FIGURAL'
        _inherently_visual = True
        # v2.26 — this branch fires precisely when NO PYQ image was observed (extraction
        # failed, or the corpus is a scanned PDF), so the measured figural_rate is 0.0 and
        # the subtopic would rank LAST for the Axis-1 budget and never be drawn. But a
        # subtopic named "mirror images" or "paper folding" IS the figure — there is no
        # text form of it. Mark it irreducible so bc.axis_grant_figural() grants it
        # regardless of budget, and give it rate 1.0 so ranking treats it as the certainty
        # it is. Non-figural exams are untouched: the keyword set never matches them.
        figural_reducible = False
        figural_rate = 1.0
        if not figural_q_count:
            figural_q_count = figural_denominator
        # Assign a default figural_data with image_role='stem_only' (conservative:
        # most inherently-visual TEXT-classified subtopics have text options).
        # If PYQ data DID have figural info, it would have set has_img=True above,
        # so we only reach this branch when no PYQ images were observed at all.
        if not figural_data:
            figural_data = {
                'image_role': 'stem_only',
                'object_types': {'dominant': [], 'observed': [], 'avoid': []},
                'transformation_types': [],
                'arrangement_types': [],
                'complexity_dist': {},
                'images_analysed': 0,
                'images_unclear': 0,
            }
        print(f"  INHERENTLY-VISUAL override: '{subtopic}' TEXT→FIGURAL "
              f"(keyword match, image_role={figural_data.get('image_role', 'stem_only')})")
    # Allow explicit curator override via entry field (from taxonomy_draft or prior run):
    # figural_override=true forces FIGURAL; figural_override=false suppresses keyword match.
    # Checked via the 'questions' metadata: if ANY question in the PYQ set carries
    # a figural_override annotation, respect it. (Future: read from taxonomy_draft entry.)

    # BUG-A07 fix: progress now properly in scope via parameter
    q_nums  = {q['num'] for q in questions}
    lk_grps = [g for g in progress.get('_linked_groups',{}).values()
               if any(qn in q_nums for qn in g.get('q_numbers',[]))]
    lk_size = round(sum(len(g['q_numbers']) for g in lk_grps)/len(lk_grps)) \
              if lk_grps else 0

    # NEW (v2.3): max_per_paper + typical_per_paper from per-paper question counts
    paper_q_counts = Counter(q.get('paper_id', '?') for q in questions)
    pvals          = list(paper_q_counts.values())
    max_pp     = max(pvals)                           if pvals else 0
    typical_pp = round(sum(pvals) / len(pvals))       if pvals else 0

    # NEW (v2.3): recycled_datasets — stimuli appearing in >=2 different papers
    recycled = _detect_recycled_stimuli(questions)
    # v2.56 (§6.1.5): pools/ranges no longer gated on a name-derived mode. The
    # quantitative aptitude path is byte-preserved; every other signature measures.
    # EC-26: the WHOLE aptitude path is byte-preserved — every derived legacy
    # mode routes through the v2.55 extractors verbatim (extract_number_ranges
    # returns ranges only for quantitative, None otherwise, exactly as before).
    # Only a content-class signature (mode None) measures with the v2 extractors.
    if mode is not None:
        ctx_pool = extract_context_pool(questions, mode)
        _num_ranges, _excl_vals = extract_number_ranges(questions, mode), []
    else:
        ctx_pool = extract_context_pool_v2(questions, sig)
        _num_ranges, _excl_vals = extract_number_ranges_v2(questions, sig)
    if recycled:
        ctx_pool = ctx_pool or {}
        ctx_pool['recycled_datasets'] = recycled
        ctx_pool['ban_recycled']      = True

    # NEW (v2.5): per-subtopic MSQ aggregation → answer_cardinality (Step 7 dispatch unit).
    # Whole-subtopic mode: a subtopic is treated as uniformly single- or multi-answer.
    # A subtopic is 'multi' when a MAJORITY of its observed Qs are MSQ (>50%), so a
    # stray false-detect cannot flip a single-answer subtopic. msq_freq% is recorded
    # for transparency. Inert when multi_select_allowed=false (is_msq always False).
    msq_ct    = sum(1 for q in questions if q.get('is_msq'))
    msq_freq  = round(msq_ct / len(questions) * 100) if questions else 0
    answer_cardinality = 'multi' if msq_freq > 50 else 'single'

    # NEW (v2.8): per-subtopic NAT aggregation → answer_type (the SECOND dispatch axis).
    # A question is NAT when it has NO selectable options at all — neither text option
    # labels NOR option-images. The forgery-resistant signal is option SHAPE (zero options)
    # plus image_role: a NAT carries no option-images (image_role is 'none' for a text NAT
    # or 'stem_only' for a figural NAT — a problem diagram with a typed answer, per ND10),
    # whereas a figural MCQ carries option-images ('options_only'/'stem_and_options') and is
    # therefore NOT NAT. Gated on nat_allowed (PARAMETER 11) so the entire path is inert for
    # non-NAT exams: answer_type is always 'option' unless the exam pattern enables NAT.
    # Whole-subtopic majority (>50%) mirrors the MSQ aggregation, so a stray detect cannot
    # flip a subtopic. answer_type is ORTHOGONAL to answer_cardinality: a 'numerical'
    # subtopic's cardinality is moot (kept 'single' by convention).
    _OPT_IMG_ROLES = ('options_only', 'stem_and_options')
    nat_ct = sum(1 for q in questions
                 if len(q.get('options', [])) == 0
                 and q.get('image_role', 'none') not in _OPT_IMG_ROLES)
    nat_freq    = round(nat_ct / len(questions) * 100) if questions else 0
    answer_type = 'numerical' if (nat_allowed and nat_freq > 50) else 'option'

    # v2.23 THREE-AXIS (CATEGORY B): this subtopic's observed Axis-2 distribution + the
    # capability set Step 6/7 read. Every question is tagged by tag_axes() at extraction;
    # the guard below re-tags defensively if `questions` came from a loaded/pre-v2.23
    # progress json (idempotent — otherwise observed_axis2 would collapse to all-DIRECT).
    for q in questions:
        if 'axis2' not in q:
            tag_axes(q)
    observed_axis2 = dict(Counter(q.get('axis2', 'DIRECT') for q in questions))
    pres_family    = resolve_presentation_family_s5(subtopic, fmt)
    axis2_cap      = axis2_capability(observed_axis2.keys(), pres_family, fmt)

    _en_qs       = [q for q in questions if q.get('medium', 'en') == 'en']
    _mc          = Counter(q.get('mechanic', bc.MECHANIC_UNKNOWN) for q in _en_qs)
    _mech_mix    = ({k: round(v / len(_en_qs), 4) for k, v in sorted(_mc.items())}
                    if _en_qs else {})
    _mech_unknown = _mc.get(bc.MECHANIC_UNKNOWN, 0)
    _stim_stats  = compute_style_cell(questions).get('stimulus_stats')

    return {
        'subtopic'               : subtopic,
        'section'                : section,
        'topic'                  : topic,
        'observed_count'         : len(questions),
        'format'                 : fmt,
        'option_format'          : opt_fmt,   # BUG-B15/C04: full dict, see format_entry
        'OMML_required'          : any(q.get('omml_present') for q in questions),
        'negative_question_freq' : round(neg_ct / len(questions) * 100),
        'answer_type'            : answer_type,   # v2.8: 'option'|'numerical' (NAT axis)
        'nat_freq'               : nat_freq,      # v2.8: % of observed Qs that are NAT
        'answer_cardinality'            : answer_cardinality,   # v2.5: 'single'|'multi' (Step 7 dispatch)
        'msq_freq'               : msq_freq,      # v2.5: % of observed Qs that are MSQ
        'observed_axis2'         : observed_axis2, # v2.23: {AXIS2_CLASS: count} this subtopic
        'presentation_family'    : pres_family,    # v2.23: family key (Step 7 menu source)
        'axis2_capability'       : axis2_cap,      # v2.23: forms this subtopic may faithfully take
        'fill_in_blank'          : blank_d,
        'linked_group_size'      : lk_size,
        'max_per_paper'          : max_pp,
        'typical_per_paper'      : typical_pp,
        'stem_word_count'        : {'min':(min(sw) if sw else 0),
                                     'max':(max(sw) if sw else 0),
                                     'typical':round(sum(sw)/len(sw)) if sw else 0},
        'sub_type_label'         : subtopic,
        'PYQ_STEM_PATTERNS'      : patterns,
        'wrong_option_structure' : wrong_opt,
        'PYQ_NUMBER_RANGES'      : _num_ranges,
        'PYQ_CONTEXT_POOL'       : ctx_pool,
        'PYQ_IMAGE_ANALYSIS'     : figural_data,
        'PYQ_PASSAGE_STRUCTURE'  : extract_passage_structure(questions) if has_pass else None,
        'inherently_visual'      : _inherently_visual,   # v2.22: True if keyword heuristic fired
        # v2.26 (GAP-2026-08-06-AXIS1) — the three fields that turn `format` from an
        # imperative into an eligibility flag. Additive: a consumer that ignores them
        # behaves exactly as before, which is what keeps ~200 exams working un-remeasured.
        'figural_q_count'        : figural_q_count,      # QUESTIONS with a figure (not images)
        'figural_denominator'    : figural_denominator,  # observed questions in this subtopic
        'figural_rate'           : round(figural_rate, 4),  # ranks claims on the Axis-1 budget
        'figural_reducible'      : figural_reducible,    # False ⇒ options ARE images ⇒ never
                                                         # downgrade to text (unanswerable)
        'option_image_rate'      : round(option_image_rate, 4),   # v2.44 — the DENOMINATOR
                                                         # the old any() threw away; makes
                                                         # every exemption auditable
        # v2.43 — DI counterparts. Same contract, same absent-safe defaults.
        'di_q_count'             : di_q_count,           # QUESTIONS with a data table
        'di_rate'                : round(di_rate, 4),    # ranks claims on the DI budget
        'di_reducible'           : di_reducible,
        # ── v2.56 additive style fields (GAP-2026-08-29-STYLE-FIDELITY §6.1.8) ──
        # Additive-only (P-9): a reader that ignores them behaves exactly as before.
        'content_signature'      : {k: v for k, v in sig.items()
                                    if not k.startswith('_')},
        'legacy_mode'            : mode,
        'mechanic_mix'           : _mech_mix,
        'mechanic_unknown_count' : _mech_unknown,
        'pattern_keys'           : build_pattern_keys(questions),
        'distractor_mix'         : mine_distractor_mechanisms(questions, figural_data),
        'low_entropy'            : detect_low_entropy(questions),
        'stimulus_stats'         : _stim_stats,
        'PYQ_EXCLUDE_VALUES'     : _excl_vals,
    }

def _detect_recycled_stimuli(questions):
    """
    NEW (v2.3): Detect linked-group stimuli that appear in >=2 different PYQ papers.
    When the same passage/puzzle/table appears across multiple papers it is a recycled
    dataset — Step 7 must not reproduce it verbatim.
    Returns list of short identifying descriptors (first 12 words of stimulus).
    Called from synthesise_subtopic; result written to PYQ_CONTEXT_POOL.
    """
    stim_papers = {}  # normalised_key → set of paper_ids
    for q in questions:
        if not q.get('linked_group_id'):
            continue
        words = q.get('stem', '').split()
        key   = ' '.join(words[:12]).lower().strip()
        if len(key) < 20:
            continue   # too short to be a meaningful stimulus identifier
        pid = q.get('paper_id', '?')
        stim_papers.setdefault(key, set()).add(pid)
    # Return descriptors for stimuli seen in 2+ distinct papers
    return [key[:80] for key, pids in stim_papers.items() if len(pids) >= 2]


def _absent_entry(section, topic, subtopic):
    _fam = resolve_presentation_family_s5(subtopic, 'TEXT')   # v2.23
    return {
        'subtopic':subtopic,'section':section,'topic':topic,'observed_count':0,
        'format':'TEXT',
        'option_format':{'primary':'single_value','recent_format':'single_value',
                         'changed_recently':False,'all_observed':[]},
        'OMML_required':False,'negative_question_freq':0,'fill_in_blank':'none',
        'answer_type':'option','nat_freq':0,           # v2.8: no PYQ → assume option-type
        'answer_cardinality':'single','msq_freq':0,   # v2.5: no PYQ → assume single-answer
        'linked_group_size':0,'max_per_paper':0,'typical_per_paper':0,
        'stem_word_count':{'min':0,'max':0,'typical':0},
        'sub_type_label':subtopic,
        'concept_group': None,        # v2.24.1 (D8): stamped later by stamp_mechanic_axes()
        'question_mechanic': None,    # v2.24.1 (D8)
        'form_key': None,             # v2.24.1 (D8)
        'collision_domain': None,     # v2.24.1 (D8)
        'PYQ_STEM_PATTERNS':[{'id':'P1','template':'(no PYQ observed)','approach':'(unknown)',
                               'frequency':100,'raw_count':0,'confidence':'absent',
                               'deprecated':False,'years':[],'note_block':'never','note_text':''}],
        'wrong_option_structure':{'type':'varied','description':'No PYQ data'},
        'PYQ_NUMBER_RANGES':None,'PYQ_CONTEXT_POOL':None,
        'PYQ_IMAGE_ANALYSIS':None,'PYQ_PASSAGE_STRUCTURE':None,
        'inherently_visual':False,   # v2.22: no PYQ → cannot determine; default False
        # v2.23: no PYQ → no observed Axis-2; capability = family menu ∪ {DIRECT} so a
        # Zero-PYQ subtopic is still a usable format-elastic filler for Step 7 (decision 11).
        'observed_axis2':{}, 'presentation_family':_fam,
        'axis2_capability':axis2_capability([], _fam, 'TEXT'),
    }


# build_diff_criteria RETIRED — GAP-2026-08-27-DIFFICULTY-PROFILE

def infer_approach(template, mode, subtopic):
    t = template.lower(); u = subtopic.lower()
    if mode == 'quantitative':
        if 'interest' in u or '_p_' in t: return 'Apply SI/CI formula'
        if 'profit' in u:                  return 'Apply profit/loss formula'
        if 'speed' in t or 'distance' in t:return 'Apply speed-distance-time formula'
        return 'Solve using relevant arithmetic formula'
    elif mode == 'reasoning':
        if 'coded as' in t or 'written as' in t: return 'Decode substitution pattern'
        if 'related to' in t: return 'Find operation A->B, apply same to C'
        return 'Apply logical reasoning pattern'
    elif mode == 'factual':
        if 'who' in t:    return 'Identify person from contextual clues'
        if '_year_' in t: return 'Recall year of event'
        return 'Recall factual information'
    elif mode == 'english':
        if 'synonym' in u: return 'Select semantically equivalent word'
        if 'antonym' in u: return 'Select semantically opposite word'
        return 'Apply English language rule'
    elif mode == 'logical':
        return 'Evaluate statement-conclusion pairs using syllogism rules'
    return 'Apply appropriate strategy'

def extract_number_ranges(questions, mode):
    if mode != 'quantitative': return None
    # BUG-A14 fix: non-raw string so \u20b9 is actual ₹ character
    VAR_PATS = [
        ('P',   '\u20b9\\s*([\\d,]+)'),
        ('R',   r'(\d+(?:\.\d+)?)%'),
        ('T',   r'(\d+)\s*(?:years?|months?)'),
        ('NUM', r'\b(\d+)\b'),
    ]
    var_vals = {}
    for q in questions:
        for var, pat in VAR_PATS:
            for m in re.finditer(pat, q.get('stem',''), re.IGNORECASE):
                try: var_vals.setdefault(var,[]).append(int(m.group(1).replace(',','')))
                except: pass
    if not var_vals: return None
    from math import gcd
    from functools import reduce
    result = {}
    for var, vals in var_vals.items():
        gcf = reduce(gcd, vals[:10]) if len(vals) >= 2 else vals[0]
        result[var] = {'min':min(vals),'max':max(vals),'multiples_of':gcf,
                       'notes':f'n={len(vals)} observed'}
    return result

def extract_context_pool(questions, mode):
    if mode != 'quantitative': return None
    CTXS = [
        (r'\bborrow|lend|loan\b','loan'),
        (r'\bbank|deposit|savings\b','bank_deposit'),
        (r'\bshop|sell|buy|profit\b','retail_trade'),
        (r'\btrain|speed|distance\b','speed_distance'),
        (r'\bpipe|cistern|tank\b','pipes_cisterns'),
        (r'\bwork|days|complete\b','work_time'),
        (r'\bscheme|invest\b','investment'),
    ]
    counts = Counter()
    for q in questions:
        for pat, label in CTXS:
            if re.search(pat, q.get('stem',''), re.IGNORECASE): counts[label] += 1
    if not counts: return None
    tot = len(questions)
    return {'dominant':[c for c,n in counts.items() if n/tot>0.20],
            'common'  :[c for c,n in counts.items() if 0.05<=n/tot<=0.20],
            'rare'    :[c for c,n in counts.items() if 0<n/tot<0.05],
            'avoid'   :[]}

def extract_passage_structure(questions):
    """
    BUG-A10 fix: 'Cloze' without leading space.
    BUG-C05/v2.3: paragraph_count estimated from avg word count;
                  topic_domains populated from content keyword matching.
    """
    linked = [q for q in questions if q.get('linked_group_id')]
    if not linked: return None
    words  = [len(q.get('stem','').split()) for q in linked]
    qtypes = Counter()
    for q in linked:
        s = q.get('stem','').lower()
        if any(kw in s for kw in ['suggest','imply','infer','conclude']): qtypes['inference'] += 1
        elif any(kw in s for kw in ['according','states','passage']):      qtypes['direct']    += 1
        elif any(kw in s for kw in ['meaning','synonym','vocabulary']):    qtypes['vocab']     += 1
        elif any(kw in s for kw in ['blank','fill','appropriate']):        qtypes['grammar']   += 1
        else:                                                                qtypes['direct']    += 1
    tot  = sum(qtypes.values()) or 1
    dist = {k:round(v/tot*100) for k,v in qtypes.items()}
    has_cloze = any('blank' in q.get('stem','').lower() for q in linked)

    # Estimate paragraph_count from average stimulus word count
    # (typical prose: ~80 words/paragraph; Cloze: 1-2 paragraphs)
    avg_words = round(sum(words) / len(words)) if words else 0
    est_paras = max(1, round(avg_words / 80)) if not has_cloze else max(1, round(avg_words / 120))

    # Detect topic domains from stimulus content (exam-agnostic keyword matching)
    # These are the most common passage domain categories across competitive exams.
    DOMAIN_MAP = [
        (['science','technology','research','experiment','discovery','innovation'],'science_technology'),
        (['social','society','community','culture','tradition','diversity'],'social_issues'),
        (['environment','ecology','climate','biodiversity','conservation','pollution'],'environment'),
        (['economy','economic','finance','trade','market','gdp','inflation'],'economy'),
        (['history','ancient','medieval','civilization','empire','dynasty'],'history'),
        (['health','medicine','disease','body','nutrition','mental','therapy'],'health_medicine'),
        (['education','school','learning','student','knowledge','pedagogy'],'education'),
        (['philosophy','ethics','morality','consciousness','value','virtue'],'philosophy'),
        (['governance','policy','government','law','constitution','democracy'],'governance'),
        (['sport','game','athletic','champion','competition','tournament'],'sports'),
        (['literature','poetry','novel','author','narrative','character'],'literature'),
        (['art','music','painting','sculpture','aesthetic','creative'],'arts'),
    ]
    all_stems = ' '.join(q.get('stem','').lower() for q in linked[:5])
    observed_domains = [label for keywords, label in DOMAIN_MAP
                        if any(kw in all_stems for kw in keywords)]

    return {'sub_format'         : 'Cloze' if has_cloze else 'RC',
            'word_range'         : {'min':(min(words) if words else 0),'max':(max(words) if words else 0)},
            'paragraph_count'    : {'typical': est_paras},
            'topic_domains'      : {'observed': observed_domains, 'avoid': []},
            'q_type_distribution': dist}

# ═══ FROM Framework_MockTestAnalyse.md §5, fence L3618-5667 (v2.53.2) — VERBATIM ═══
def _compute_structural_changes(entries):
    """
    NEW (v2.3): Compute observable year-over-year structural changes from PYQ data.
    Returns list of strings for STRUCTURAL_CHANGES_BY_YEAR block.
    All conclusions are DATA-DRIVEN — zero hardcoding.
    Detects: subtopics removed in recent years, new subtopics, DI format shifts,
             FIGURAL elimination, format type changes.
    Called by write_section_rules.
    """
    changes = []

    # Collect global year range from all pattern years lists
    all_years = sorted(set(
        y for e in entries
        for p in e.get('PYQ_STEM_PATTERNS', [])
        for y in p.get('years', [])
        if isinstance(y, int)
    ), reverse=True)

    if not all_years:
        return changes   # no year data — cannot compute changes

    max_year = all_years[0]

    for e in entries:
        subtopic = e['subtopic']
        # PROVEN EQUIVALENT MUTANT (B7, 2026-08-20). `sorted -> list` here is the one
        # surviving pipeline mutant in this engine, and it survives because it CANNOT
        # change behaviour: `years_seen` is read on exactly four lines below and every
        # one of them is order-invariant — truthiness, max(), min(), len(). There is no
        # fixture that could kill it, so none is written; inventing an assertion that
        # only appears to cover it would be worse than the honest entry in
        # MUTATION_BUDGETS.json.
        #
        # The sorted() STAYS. It costs nothing, and the day someone adds an
        # order-dependent read (a first/last, a slice, an emitted list) the ordering is
        # already correct rather than newly broken. If that day comes, this comment is
        # the thing to delete — and the mutant becomes killable, which is the point.
        years_seen = sorted(set(
            y for p in e.get('PYQ_STEM_PATTERNS', [])
            for y in p.get('years', []) if isinstance(y, int)
        ))
        if not years_seen:
            continue

        last_seen = max(years_seen)
        first_seen = min(years_seen)

        # Subtopic absent from most recent year and last seen 2+ years ago → REMOVED
        if last_seen < max_year - 1:
            changes.append(
                f'  {subtopic} (format={e["format"]}): '
                f'last seen {last_seen} — likely REMOVED after {last_seen}'
            )

        # Subtopic appears only in last 1-2 years → NEW
        elif first_seen >= max_year - 1 and len(years_seen) <= 2:
            changes.append(
                f'  {subtopic} (format={e["format"]}): '
                f'first seen {first_seen} — NEW subtopic'
            )

        # v2.15 BUG-D10 fix: only emit FIGURAL-eliminated if not already flagged as REMOVED
        # (a REMOVED FIGURAL subtopic would otherwise produce two entries).
        if (e.get('format') == 'FIGURAL' and e.get('observed_count', 0) == 0
                and not (last_seen < max_year - 1)):
            changes.append(
                f'  {subtopic} (FIGURAL): observed_count=0 '
                f'— FIGURAL format appears eliminated for this subtopic'
            )

        # DI: single-Q format (linked_group_size <= 1) — may indicate format shift
        if e.get('format') == 'DI' and e.get('linked_group_size', 0) <= 1 and e.get('observed_count', 0) > 0:
            changes.append(
                f'  {subtopic} (DI): linked_group_size={e["linked_group_size"]} '
                f'— single-Q DI format observed (previously multi-Q)'
            )

    return changes


# ══════════════════════════════════════════════════════════════════════════════
# v2.24.1 MECHANIC / FORM-KEY ENGINE  (permanent, EXAM-INDEPENDENT fix for the
#                                      BV-10a form_key-collision HALT class)
# ------------------------------------------------------------------------------
# THREE AXES, THREE GRANULARITIES. They are NOT the same variable.
#   family / concept_group  — COARSE. May be shared. Feeds BV-10b (SOFT cap).
#   question_mechanic       — SEMANTIC FORM. May be shared. Feeds Step 7 CHECK D.
#   form_key                — IDENTITY. MUST be unique per collision_domain.
#                             Feeds BV-10a (HARD cap = 1, not configurable).
#
# v2.24 collapsed all three onto `family` whenever no qualifier was found — which,
# because _QUALIFIERS is a reasoning-domain vocabulary, is ALWAYS the case on a
# subject-knowledge exam. Any two subtopics sharing a _FAMILY_MAP keyword then got
# an IDENTICAL form_key, and Step 6 BV-10a HALTed two steps later. This is not
# exam-specific: it is latent in EVERY single-section subject exam. See the
# v2.24.1 defect report (D1..D9).
#
# v2.24.1 relocates the uniqueness guarantee from an ACCIDENT (a reasoning
# qualifier happening to match the name) to a CONSTRUCTION: form_key derives from
# the subtopic's own identity base, which bottoms out at the unique subtopic_id.
# It scopes the keyword table by a per-exam `template_sets` declaration, derives +
# stamps every axis exactly ONCE after id minting, and asserts form_key uniqueness
# BEFORE any artifact is written. The failure mode (defective manifest promoted,
# HALT two steps later) is therefore impossible for any exam: the worst case is a
# LOUD Step-5 FAIL naming the offending ids, never a silent downstream HALT.
#
# EXAM-INDEPENDENCE: the engine is identical for every exam. The ONLY per-exam
# input is [ExamCode]_mechanic_overrides.json. Absent file ⇒ legacy family
# selection (REGR-1) PLUS the SPEC-1 uniqueness improvement, which is always on.

def canon_text(s):
    """NFC + casefold + hyphen/slash/ampersand->space + collapse whitespace. Never raises."""
    import re, unicodedata
    s = unicodedata.normalize('NFC', (s or ''))
    for ch in ('—','–','-','/','&'):
        s = s.replace(ch, ' ')
    s = s.casefold().strip()
    return re.sub(r'\s+', ' ', s)

def _has_word(text, kw):
    """Word-boundary containment; allows only simple plural (s/es), never arbitrary
    continuation. 'voice' does NOT match 'invoice'; 'clock' does NOT match 'clockwise';
    but 'antonym' DOES match 'antonyms'. EC-M12: plural suffix applies to the FINAL
    token of a multi-word keyword too."""
    import re
    if ' ' in kw:
        head, _, tail = kw.rpartition(' ')
        pat = r'(?<!\w)' + re.escape(head) + r'\s+' + re.escape(tail) + r'(?:e?s)?(?!\w)'
        return re.search(pat, text) is not None
    return re.search(r'\b'+re.escape(kw)+r'(?:e?s)?\b', text) is not None

# Minimal, extensible transliteration for common Hindi verbal/reasoning terms so a
# pure-Devanagari name never collapses to '' for the FAMILY axis (never for form_key).
_HI_MAP = {
    'पर्यायवाची':'synonym',
    'विलोम':'antonym',
    'मुहावरे':'idiom','मुहावरा':'idiom',
    'श्रृंखला':'series','शृंखला':'series',
    'सादृश्य':'analogy','वर्गीकरण':'classification',
    'कोडिंग':'coding_decoding','दिशा':'direction_sense',
    'वर्तनी':'spelling',
}
def _translit_hint(raw):
    for k,v in _HI_MAP.items():
        if k in (raw or ''):
            return v
    return None

_VERBAL_SECTION_HINTS = ('english','verbal','language','comprehension','hindi',
                         'हिंदी','भाषा')
def _is_verbal(section, fmt):
    """v2.24.1: DEMOTED. It may only NARROW a declared template set (advisory), never
    WIDEN one. Its old ability to enable the verbal table for any TEXT exam (the D2
    defect) is gone — the verbal table now fires only when the exam DECLARES 'verbal'
    in template_sets AND this returns True."""
    sec = canon_text(section)
    if any(h in sec for h in _VERBAL_SECTION_HINTS):
        return True
    return (fmt or 'TEXT').upper() in ('TEXT','PASSAGE')

# ── Per-exam overrides loader (SPEC-2 / SPEC-6; EC-M17 exam_code, EC-M18 malformed) ──
_OVERRIDES = None
_MERGE_LOG = []          # populated by apply_subtopic_merges(); read by write_analysis_summary()

OVERRIDE_SEARCH_DIRS = ('/mnt/project/', '/mnt/user-data/uploads/')

# ── ARTEFACT OUTPUT DIRECTORY (Wave 2 Part C B7) ─────────────────────────────
# The four writers downstream hardcoded '/mnt/user-data/outputs/'. That is correct in the
# session sandbox and ABSENT on a CI runner, which is why every one of them sat outside
# the mutation gate: a fixture that calls them cannot open the file, so no fixture called
# them, so nine surviving mutants lived in code no test had ever executed.
#
# GAP-2026-08-17-B4-ENV-SKEW is the precedent and the warning. There the fix was
# `search_dirs=None` on load_mechanic_overrides — additive, default unchanged, and the
# assertion moved from "silently skipped when /mnt is unwritable" to "runs everywhere".
# The same shape applies here. An `if os.path.isdir(...)` guard around the fixture would
# NOT do: a guarded assertion is a dormant assertion, and a dormant assertion is exactly
# where a mutation survivor hides while CI reports green.
#
# DEFAULT IS UNCHANGED. Every production call site omits out_dir and writes precisely
# where it wrote before, so no artefact moves and the golden set is untouched.
OUTPUT_DIR = '/mnt/user-data/outputs'


def _artefact_path(out_dir, filename):
    """Resolve an artefact path under out_dir, falling back to the session sandbox."""
    return os.path.join(out_dir or OUTPUT_DIR, filename)


def load_mechanic_overrides(exam_code, search_dirs=None):
    """Discovery: /mnt/project/, then /mnt/user-data/uploads/.
    Absent               → all defaults (legacy family selection; SPEC-1 still applies).
    Present + malformed  → FAIL (EC-M18).  exam_code mismatch → FAIL (EC-M17).

    search_dirs is ADDITIVE AND DEFAULTS TO THE DOCUMENTED ORDER, so production
    behaviour is byte-for-byte unchanged. It exists because the EC-M18 hard stop was
    UNTESTABLE without it: the only way to reach that branch is to place a malformed
    file in one of two absolute paths, and on a GitHub runner neither is writable. The
    fixture silently skipped, the guard's mutant survived there and died here, and the
    mutation budget measured 28 locally against 29 in CI — a gate that disagreed with
    itself across environments (GAP-2026-08-17-B4-ENV-SKEW).
    A guard that cannot be exercised in the environment that gates the build is not
    guarded. Injection is the same remedy applied to session_re in B3."""
    global _OVERRIDES
    if _OVERRIDES is not None:
        return _OVERRIDES
    _OVERRIDES = {'template_sets': None, 'template_sets_by_section': {},
                  'subtopic_overrides': {}, 'subtopic_merges': []}
    for d in (search_dirs or OVERRIDE_SEARCH_DIRS):
        path = os.path.join(d, f'{exam_code}_mechanic_overrides.json')
        if not os.path.exists(path):
            continue
        try:
            with io_open_utf8(path) as fh:
                data = json.load(fh)
        except json.JSONDecodeError as ex:
            raise SystemExit(f'FAIL: {path} is not valid JSON — {ex}')                 # EC-M18
        if data.get('exam_code') != exam_code:                                          # EC-M17
            raise SystemExit(f"FAIL: {path} declares exam_code={data.get('exam_code')!r}, "
                             f"expected {exam_code!r}")
        _valid = {'verbal','reasoning'}
        bad = set(data.get('template_sets') or []) - _valid                             # OV-5
        if bad:
            raise SystemExit(f"FAIL: {path} template_sets has unknown values {sorted(bad)}")
        for _sec, _sets in (data.get('template_sets_by_section') or {}).items():         # OV-5b
            _badsec = set(_sets or []) - _valid
            if _badsec:
                raise SystemExit(f"FAIL: {path} template_sets_by_section[{_sec!r}] has "
                                 f"unknown values {sorted(_badsec)} (check case/whitespace)")
        _OVERRIDES.update(data)
        break
    return _OVERRIDES

def io_open_utf8(path):
    import io as _io
    return _io.open(path, encoding='utf-8')

# ── The keyword table. Rows are 3-tuples now: (keywords, family, template_set). ──
#    Every former verbal_only=True row is 'verbal'; every former False row is
#    'reasoning'. A row fires only if its template_set is declared for the exam.
_FAMILY_MAP = [
    (['synonym','similar in meaning','nearest in meaning'], 'synonym', 'verbal'),
    (['antonym','opposite in meaning','opposite meaning'],  'antonym', 'verbal'),
    (['one word substitution','one word substitute'],       'one_word_substitution', 'verbal'),
    (['idiom','phrasal verb'],                              'idiom', 'verbal'),
    (['cloze'],                                             'cloze_test', 'verbal'),
    (['fill in the blank','fill in the blanks','sentence completion'], 'fill_in_blank', 'verbal'),
    (['spelling','correctly spelt','misspelt'],             'spelling', 'verbal'),
    (['grammatical error','error detection','spotting error','find the error'], 'error_detection', 'verbal'),
    (['sentence improvement','sentence correction','best improves'], 'sentence_improvement', 'verbal'),
    (['voice','active voice','passive voice'],              'voice', 'verbal'),
    (['narration','direct speech','indirect speech','reported speech'], 'narration', 'verbal'),
    (['para jumble','sentence rearrangement','sentence order'], 'para_jumble', 'verbal'),
    (['reading comprehension'],                            'reading_comprehension', 'verbal'),
    (['series'],                                           'series', 'reasoning'),
    (['analogy'],                                          'analogy', 'reasoning'),
    (['classification','odd one out','odd pair','does not belong'], 'classification', 'reasoning'),
    (['coding','decoding'],                                'coding_decoding', 'reasoning'),
    (['blood relation','family relation'],                 'blood_relation', 'reasoning'),
    (['direction'],                                        'direction_sense', 'reasoning'),
    (['seating','arrangement','puzzle'],                   'arrangement', 'reasoning'),
    (['syllogism','statement conclusion','statement assumption','course of action','logical deduction'], 'syllogism', 'reasoning'),
    (['mirror image'],                                     'mirror_image', 'reasoning'),
    (['water image'],                                      'water_image', 'reasoning'),
    (['paper folding','paper cutting'],                    'paper_folding', 'reasoning'),
    (['embedded figure','hidden figure'],                  'embedded_figure', 'reasoning'),
    (['venn diagram'],                                     'venn_diagram', 'reasoning'),
    (['dice','cube'],                                      'dice', 'reasoning'),
    (['missing number','number matrix','number grid'],     'missing_number', 'reasoning'),
    (['calendar'],                                         'calendar', 'reasoning'),
    (['clock'],                                            'clock', 'reasoning'),
    (['data interpretation','bar graph','pie chart','line graph','tabulation'], 'data_interpretation', 'reasoning'),
    (['current affairs'],                                  'current_affairs', 'reasoning'),
    (['static gk','static general knowledge'],             'static_gk', 'reasoning'),
]
_QUALIFIERS = ['alphanumeric','symbolic','semantic','number','numeric','numerical',
               'letter','alphabet','word','figural','figure','spatial',
               'linear','circular','floor','matrix','wheel','triangle','substitution']
_QUALIFIABLE = {'series','analogy','classification','coding_decoding','missing_number',
                'arrangement','mirror_image','paper_folding','embedded_figure'}
_ALL_FAMILY_NAMES = {fam for _, fam, _ in _FAMILY_MAP}
_TEMPLATE_SET     = {fam: ts for _, fam, ts in _FAMILY_MAP}

def _identity_base(display_name, subtopic_id):
    """SPEC-1 / EC-M2. The collision-safe identity root. ONE recipe, ONE call site.
    A fully non-Latin name slugifies to '' and must still yield a stable, unique,
    call-site-independent value. Ultimately bottoms out at the unique subtopic_id."""
    import hashlib as _hl
    return (slugify(display_name)
            or slugify(subtopic_id)
            or ('u_' + _hl.md5((display_name or '').encode('utf-8')).hexdigest()[:8]))

def _redundant(qual, base):
    """EC-M13: word-boundary test, not substring. 'number' is redundant against base
    'missing_number' only if it appears there as a whole token."""
    return _has_word(base.replace('_', ' '), qual)

def _extract_qualifiers(name_c):
    """EC-M14: ALL matching qualifiers, canonical (alphabetical) order — order-independent."""
    import re
    return tuple(sorted(re.sub(r'\s+', '_', q) for q in _QUALIFIERS if _has_word(name_c, q)))

def _allowed_template_sets(section, ov):
    allowed = ov.get('template_sets_by_section', {}).get(section, ov.get('template_sets'))
    if allowed is None:
        allowed = ['verbal', 'reasoning']            # ABSENT ⇒ legacy behaviour (REGR-1)
    return allowed

def derive_mechanic(section, subtopic, sub_type_label=None, templates='',
                    fmt='TEXT', subtopic_id=None, prefix_overrides=None):
    """Returns {family, mechanic, form_key, collision_domain}.
    PRECONDITION: subtopic_id is the FINAL, de-duplicated, minted id (EC-M3)."""
    import re
    ov       = _OVERRIDES or {}
    raw_name = subtopic or sub_type_label or ''
    name_c   = canon_text(raw_name + ' ' + (sub_type_label or ''))
    base     = _identity_base(raw_name, subtopic_id)
    domain   = section_prefix(section, prefix_overrides) or 'default'    # SPEC-3 / EC-M4
    allowed  = _allowed_template_sets(section, ov)                       # SPEC-2 / EC-M6

    def _match(hay):
        for kws, fam, tset in _FAMILY_MAP:
            if tset not in allowed:                                 continue
            if tset == 'verbal' and not _is_verbal(section, fmt):   continue   # advisory NARROW only
            if any(_has_word(hay, kw) for kw in kws):
                return fam
        return None

    # COARSE axis. NAME is authoritative. Templates rescue ONLY a name with no
    # alphanumeric content at all (never merely "a name that matched nothing").
    family = _match(name_c) if name_c.strip() else None
    if family is None and not re.search(r'[a-z0-9]', name_c):
        family = _match(canon_text(templates))
    if family is None:
        family = _translit_hint(raw_name)                           # EC-M2: family ONLY
    if family is None:
        family = base                                               # coarse identity == the name

    # FINE axis. SPEC-1: starts from the NAME/id base, NEVER from the family token.
    quals    = _extract_qualifiers(name_c) if family in _QUALIFIABLE else ()
    suffix   = '__'.join(q for q in quals if not _redundant(q, base))   # EC-M13 / EC-M14
    form_key = f'{base}__{suffix}' if suffix else base

    # Explicit curator override wins over everything (applied per subtopic_id).
    per      = ov.get('subtopic_overrides', {}).get(subtopic_id or '', {})
    family   = per.get('concept_group',     family)
    form_key = per.get('form_key',          form_key)
    mechanic = per.get('question_mechanic', form_key)               # v2.24.1: mechanic == form_key

    return {'family': family, 'mechanic': mechanic,
            'form_key': form_key, 'collision_domain': domain}

def mint_subtopic_ids(entries, exam_meta=None):
    """v2.24.1: mint the canonical, collision-safe subtopic_id for every entry, IDEMPOTENTLY
    (an entry that already carries an id is left untouched). Called from run_synthesise()
    BEFORE merges/stamp/QV so the gate and every writer read the SAME id. write_section_rules()
    also calls it as a no-op guard."""
    pov  = (exam_meta or {}).get('section_prefix_overrides', {})
    pmap = build_section_prefix_map(sorted({e['section'] for e in entries}), pov)
    seen = {}
    for e in entries:
        if e.get('subtopic_id'):
            seen[e['subtopic_id']] = (e['section'], e['topic'], e['subtopic'])
    for e in entries:
        if e.get('subtopic_id'):
            continue
        sid = make_subtopic_id(e['section'], e['topic'], e['subtopic'], pmap)
        base = sid; n = 2
        key = (e['section'], e['topic'], e['subtopic'])
        while sid in seen and seen[sid] != key:
            sid = f'{base}_{n}'; n += 1
        seen[sid] = key
        e['subtopic_id'] = sid
    return entries

def apply_subtopic_merges(entries, exam_code):
    """D7 / SPEC-7. TRUE-duplicate merge (NOT an allowlist). Each group of >=2 subtopic_ids
    is collapsed into the FIRST member; the others are dropped. Members must share a section
    (OV-3). Runs AFTER id minting, BEFORE stamp/QV. Records every drop in _MERGE_LOG so
    write_analysis_summary() can emit the mandatory '## MERGED SUBTOPICS' section (a merge
    removes a subtopic_id that Steps 6/7 join on — the curator must see what disappeared)."""
    global _MERGE_LOG
    _MERGE_LOG = []
    ov = load_mechanic_overrides(exam_code)
    groups = ov.get('subtopic_merges', []) or []
    if not groups:
        return entries
    by_id = {e.get('subtopic_id'): e for e in entries}
    drop  = set()
    _across = {}                                                                     # id -> group index
    for _gi, _grp in enumerate(groups):                                              # OV-4b: groups disjoint
        for _g in _grp:
            if _g in _across:
                raise SystemExit(f"FAIL: subtopic_merges: {_g!r} appears in two groups "
                                 f"({groups[_across[_g]]} and {_grp}); groups must be disjoint")
            _across[_g] = _gi
    for grp in groups:
        if len(grp) < 2:                                                             # OV-4
            raise SystemExit(f"FAIL: subtopic_merges group {grp} has <2 members")
        missing = [g for g in grp if g not in by_id]
        if missing:                                                                  # OV-3
            raise SystemExit(f"FAIL: subtopic_merges references unknown subtopic_id(s) {missing}")
        secs = {by_id[g]['section'] for g in grp}
        if len(secs) > 1:                                                            # OV-3
            raise SystemExit(f"FAIL: subtopic_merges group {grp} spans sections {sorted(secs)}")
        survivor = by_id[grp[0]]
        members  = [by_id[g] for g in grp]
        survivor['observed_count'] = sum(m.get('observed_count', 0) for m in members)
        survivor['max_per_paper']  = max(m.get('max_per_paper', 0) for m in members)
        tpp = [m.get('typical_per_paper', 0) for m in members]
        survivor['typical_per_paper'] = round(sum(tpp) / len(tpp)) if tpp else 0
        for g in grp[1:]:
            drop.add(g)
            _MERGE_LOG.append((g, grp[0]))
    kept = [e for e in entries if e.get('subtopic_id') not in drop]
    for g, keep in _MERGE_LOG:
        print(f"  MERGE: {g} → {keep} (observed_count summed)")
    return kept

def stamp_mechanic_axes(entries, exam_code, exam_meta=None):
    """SPEC-5 / D8. Derive ONCE, stamp, then assert. Called from run_synthesise() AFTER
    taxonomy sync, AFTER merges, AFTER subtopic_id minting — and BEFORE run_qv() and any
    writer. Every downstream consumer reads the stamped field; NONE recompute."""
    import difflib
    ov = load_mechanic_overrides(exam_code)
    pov = (exam_meta or {}).get('section_prefix_overrides', {})
    po  = build_section_prefix_map(sorted({e['section'] for e in entries}), pov)

    # PRECONDITION (§8-4): every entry must already carry a minted subtopic_id. mint runs
    # before stamp in run_synthesise; guard here so a misordered caller fails clearly, not
    # with a bare KeyError.
    _noid = [i for i, e in enumerate(entries) if not e.get('subtopic_id')]
    if _noid:
        raise SystemExit(f"stamp_mechanic_axes precondition violated: {len(_noid)} entr(y/ies) "
                         f"have no subtopic_id — mint_subtopic_ids() must run first.")

    known   = {e['subtopic_id'] for e in entries}
    unknown = set(ov.get('subtopic_overrides', {})) - known                          # OV-1 / EC-M7
    if unknown:
        for u in sorted(unknown):
            near = difflib.get_close_matches(u, list(known), n=3)
            print(f'  FAIL: override key {u!r} matches no subtopic_id. Did you mean: {near}')
        raise SystemExit(f'OV-1: {len(unknown)} override key(s) match no subtopic_id')
    for sec in ov.get('template_sets_by_section', {}):                               # OV-6
        if sec not in {e['section'] for e in entries}:
            raise SystemExit(f"OV-6: template_sets_by_section names unknown section {sec!r}")

    _ovsub = ov.get('subtopic_overrides', {})
    _forced_fk = set()          # ids whose form_key was set by an EXPLICIT curator override
    for e in entries:
        templates = ' '.join(p.get('template', '') for p in e.get('PYQ_STEM_PATTERNS', []))
        ax = derive_mechanic(e['section'], e.get('subtopic') or e.get('sub_type_label',''),
                             e.get('sub_type_label'), templates,
                             e.get('format', 'TEXT'), e['subtopic_id'], po)
        e['concept_group']     = ax['family']
        e['question_mechanic'] = ax['mechanic']
        e['form_key']          = ax['form_key']
        e['collision_domain']  = ax['collision_domain']
        if 'form_key' in _ovsub.get(e['subtopic_id'], {}):
            _forced_fk.add(e['subtopic_id'])

    # EC-M1: two subtopics whose DISPLAY_NAME yields the SAME derived base in one
    # collision_domain both fall back to slugify(subtopic_id) — deterministic, unique.
    # This applies ONLY to derived collisions. A collision in which ANY member carries an
    # EXPLICIT override form_key is NOT auto-resolved: that is an OV-2 curator error and
    # must FAIL loudly below (silently rewriting a curator's declared value would hide a
    # real mistake). See EC-M8.
    import hashlib as _hl
    claims = {}
    for e in entries:
        claims.setdefault((e['collision_domain'], e['form_key']), []).append(e)
    for (_dom, _fk), grp in claims.items():
        if len({e['subtopic_id'] for e in grp}) > 1 and not any(e['subtopic_id'] in _forced_fk for e in grp):
            for e in grp:
                e['form_key'] = slugify(e['subtopic_id'])
    # Residual guard: subtopic_ids are unique, but slugify() maps '.' and '_' alike, so two
    # DISTINCT ids can slugify to the same string (e.g. 'p.a_b.c' and 'p.a.b_c'). Any residual
    # (domain, form_key) collision among AUTO-RESOLVED entries (no forced override) is broken
    # deterministically with a short id-hash suffix, so genuinely-distinct subtopics never
    # provoke a false HALT. Override-induced collisions are left for the invariant to reject.
    _res = {}
    for e in entries:
        _res.setdefault((e['collision_domain'], e['form_key']), []).append(e)
    for (_dom, _fk), grp in _res.items():
        if len({e['subtopic_id'] for e in grp}) > 1 and not any(e['subtopic_id'] in _forced_fk for e in grp):
            for e in grp:
                _h = _hl.md5(e['subtopic_id'].encode('utf-8')).hexdigest()[:6]
                e['form_key'] = f"{e['form_key']}_{_h}"

    # OV-2 / D7 / EC-M8 / EC-M15: uniqueness is an INVARIANT, asserted before any write.
    seen, violations = {}, []
    for e in entries:
        k = (e['collision_domain'], e['form_key'])
        if not e['form_key']:
            violations.append(f"{e['subtopic_id']}: empty form_key")
        if e['form_key'] == e['collision_domain']:
            violations.append(f"{e['subtopic_id']}: form_key equals collision_domain")
        if k in seen and seen[k] != e['subtopic_id']:
            violations.append(f"{k[0]}:{k[1]} shared by {seen[k]} and {e['subtopic_id']}")
        seen[k] = e['subtopic_id']
    if violations:
        raise SystemExit(
            'form_key uniqueness invariant violated. A manifest with a shared form_key is a '
            'latent Step 6 BV-10a HALT whose triggering depends on N_mocks and batch_size, '
            'which Step 5 cannot know. Merge the subtopics via subtopic_merges, or give them '
            'distinct form_keys via subtopic_overrides.\n  ' + '\n  '.join(violations))
    print(f'  form_key invariant: PASS — {len(entries)} entries, '
          f'{len(seen)} distinct (collision_domain, form_key) pairs')
    return entries

# ── Back-compat wrappers. They read the STAMPED field first and only fall back to a
#    fresh derivation for a legacy caller that never ran stamp_mechanic_axes(). After
#    the v2.24.1 run_synthesise() path they always return the stamped value. ──
def _derive_axes(entry):
    templates = ' '.join(p.get('template','') for p in entry.get('PYQ_STEM_PATTERNS', []))
    return derive_mechanic(entry.get('section',''),
                           entry.get('subtopic') or entry.get('sub_type_label',''),
                           entry.get('sub_type_label'), templates,
                           entry.get('format','TEXT'), entry.get('subtopic_id'))

def _derive_concept_group(entry):
    return entry.get('concept_group') or _derive_axes(entry)['family']

def _derive_question_mechanic(entry):
    return entry.get('question_mechanic') or _derive_axes(entry)['mechanic']

def _derive_form_key(entry):
    return entry.get('form_key') or _derive_axes(entry)['form_key']

def _derive_collision_domain(entry):
    return entry.get('collision_domain') or _derive_axes(entry)['collision_domain']


# ── PROVENANCE STAMPS — ONE DEFINITION (v2.50.0) ─────────────────────────────
# These three strings are the ONLY place this spec's version appears in emitted
# output. Until v2.50.0 the same number was hand-copied into five scattered lines,
# and it drifted at four separate releases (CHANGELOG v2.14/v2.15/v2.17/v2.47) and
# again at the v2.49.1 -> v2.50.0 bump, where it blocked release 2026.08.15.8: the
# header said v2.50 while every section_rules.md and subtopic_manifest.json would
# have been stamped "produced by v2.49", on every run of every exam, with the
# artefacts persisting into Steps 6 and 7.
#
# THEY ARE DELIBERATELY LITERAL, NOT DERIVED. mock_sync_audit MS-3 STAMP-PARITY
# verifies them by scanning this file's fences for the three patterns below and
# comparing each against the header's major.minor. Building them from an f-string
# would leave MS-3 with nothing to match: the check would report 0 issues while
# verifying nothing — a check-shaped hole of exactly the kind C6-PRE was added to
# close one level up. One literal per pattern is the smallest form MS-3 can still
# police, and MS-3 now also fails if any of the three disappears.
#
# ON A VERSION BUMP: change the major.minor here and nowhere else in emitted code.
# The illustrative copy inside write_subtopic_manifest's docstring documents the
# OUTPUT shape and is checked by MS-3 too, so keep it in step.
FRAMEWORK_STAMP         = 'Framework_MockTestAnalyse v2.57'
GENERATED_BY_STAMP      = 'Generated by Framework_MockTestAnalyse v2.57'
FRAMEWORK_VERSION_STAMP = 'framework_version: v2.57'


def write_section_rules(entries, exam_code, exam_meta=None, progress=None, out_dir=None):
    """
    BUG-A25 fix: output to /mnt/user-data/outputs/ for present_files access.
    v2.3 NEW: writes EXAM_STRUCTURE header block (CATEGORY C) when exam_meta provided.
    v2.3 NEW: writes STRUCTURAL_CHANGES_BY_YEAR block computed from PYQ data.
    v2.3 NEW: writes figural_banned flag per section header.
    v2.15 BUG-D03: accepts progress dict for per-section option_label_format aggregation.

    exam_meta: dict built by run_synthesise from progress['_meta'].
      Keys: time_per_q_sec, language, q_types, marks_per_q, negative_marking,
            options_count, multi_select_allowed, papers_analysed, questions_analysed,
            years_covered, generation_date, option_label_format,
            marking_scheme, level, medium (v2.18 additions).
    progress: the full progress dict (needed for per-question option_label aggregation).
    """
    from datetime import datetime
    out  = _artefact_path(out_dir, f'{exam_code}_section_rules.md')
    meta = exam_meta or {}
    progress = progress or {}   # v2.15: safe fallback for label aggregation

    # ── CATEGORY C: exam-level header (auto-detected — not hardcoded) ─────────
    lines = [
        f'# {exam_code}_section_rules.md',
        f'# {GENERATED_BY_STAMP}',   # v2.54: stamp already begins 'Generated by' — the prefix here doubled it since v2.3
        f'# DO NOT edit manually -- regenerate via: PYQExtract {exam_code} --synthesise ALL',
        f'# Download this file from chat → upload to {exam_code} project Files/Knowledge section.',
        '',
        '=== EXAM_STRUCTURE ===',
        f'exam_code: {exam_code}',
        f'total_papers_analysed: {meta.get("papers_analysed", 0)}',
        f'total_questions_analysed: {meta.get("questions_analysed", 0)}',
        f'years_covered: {meta.get("years_covered", [])}',
        f'generation_date: {meta.get("generation_date", datetime.now().isoformat()[:10])}',
        f'time_per_q_sec: {meta.get("time_per_q_sec", "unknown")}',
        f'language: {meta.get("language", "unknown")}',
        # v2.18: new fields from Step 2a v2.5 exam_config contract
        f'medium: {meta.get("medium", "unknown")}',
        f'level: {meta.get("level", "unknown")}',
        f'q_types: {meta.get("q_types", ["MCQ"])}',
        f'marks_per_q: {meta.get("marks_per_q", {"MCQ": 1})}',
        f'negative_marking: {meta.get("negative_marking", 0)}',
        # v2.18: full per-range marking_scheme from exam_config. Steps 7/8/9 can use this
        # for exact per-Q-position marks lookup (e.g., CSIR NET Q.72 → 4 marks, Q.25 → 2 marks).
        # marks_per_q and negative_marking above remain as summary scalars for backward compat.
        f'marking_scheme: {meta.get("marking_scheme", [])}',
        f'options_count: {meta.get("options_count", 4)}',
        # options_count → S1 reads as total_options and writes to blueprint.json Step 7.
        # S2 reads as num_options via bp.get('total_options', 4). SYNC CHAIN:
        #   S0 writes options_count → S1 reads → writes total_options → S2 reads.
        f'option_label_format: {meta.get("option_label_format", "1/2/3/4")}',
        # option_label_format → S1 reads and writes as option_label to blueprint.json Step 7.
        # S2 reads via bp.get('option_label','1/2/3/4') AND re-reads per section from
        # section_rules.md option_label_format field below (per-section override).
        # HOW TO DETECT option_label_format from PYQ: scan PYQ option lines for pattern:
        #   "(1)" / "(A)" / "1." / "A." → set option_label_format accordingly.
        #   Most SSC/banking exams: "1/2/3/4". UPSC/GATE: "A/B/C/D". Regional: varies.
        f'multi_select_allowed: {str(meta.get("multi_select_allowed", False)).lower()}',
        # v2.5 MSQ contract fields (consumed by Step 6/7/9). Inert when multi_select_allowed=false.
        f'msq_k_mode: {meta.get("msq_k_mode", "n/a")}',
        f'msq_k: {meta.get("msq_k", "none")}',
        # v2.6 D5: AOTA policy under MSQ. Step 7 (R-MSQ-ESCAPE/G-MSQ-SET) and Step 8
        # (A-MSQ-KEY) read this directly from section_rules. Default false.
        f'msq_allow_aota: {str(meta.get("msq_allow_aota", False)).lower()}',
        # v2.9 (contract-sync fix): the localized MSQ select-instruction, the EXACT MSQ
        # analogue of nat_instruction below. Step 7 (msq_instruction_for) and Step 8
        # (msq_instruction_phrases) READ this from section_rules; before v2.9 no producer
        # emitted it, so the instruction was silently locked to a hardcoded fallback and was
        # not exam-configurable/localizable the way NAT's is. Now authoritative + overridable
        # per exam; the *_hi variant carries the Hindi/bilingual phrasing. Parenthesised so it
        # reads as an in-stem instruction. Inert when multi_select_allowed=false.
        f'msq_instruction: {meta.get("msq_instruction", "(One or more options may be correct)")}',
        f'msq_instruction_hi: {meta.get("msq_instruction_hi", "(एक या अधिक विकल्प सही हो सकते हैं)")}',
        f'negative_marking_by_type: {meta.get("negative_marking_by_type", {})}',
        f'partial_credit: {str(meta.get("partial_credit", False)).lower()}',
        # v2.12 difficulty_labels — the CANONICAL, exam-overridable difficulty vocabulary
        # used as the RENDERED/stored Complexity value in the per-question index
        # (registry.question_index: Step 6 seeds, Step 7 fills, Step 8 certifies, Step 11
        # renders). Default ['Easy','Medium','Hard']. ALIAS CONTRACT — do NOT conflate the
        # three pre-existing internal spellings:
        #   • (Step 5 keyword difficulty levels RETIRED — GAP-2026-08-27-DIFFICULTY-PROFILE)
        #   • Step 6 difficulty_schedule COUNT keys     : simple / medium / hard  (per-mock counts)
        #   • canonical LABEL (this field, index/tags)  : Easy  / Medium / Hard
        # Fixed alias: simple→Easy, medium→Medium, hard→Hard. An exam may override this list
        # (e.g. a 2- or 5-band set); the index value and the schedule bands then draw from it.
        f'difficulty_labels: {meta.get("difficulty_labels", ["Easy", "Medium", "Hard"])}',
        # v2.8 NAT contract fields (consumed by Step 6/7/9). nat_allowed is the capability
        # gate (mirrors multi_select_allowed); nat_present is a rollup of THIS analysis
        # (true iff any subtopic resolved to answer_type=='numerical'). All inert/absent-safe
        # when nat_allowed=false. nat_answer_type/tolerance/instruction define the answer model.
        f'nat_allowed: {str(meta.get("nat_allowed", False)).lower()}',
        f'nat_present: {str(any(e.get("answer_type") == "numerical" for e in entries)).lower()}',
        f'nat_answer_type: {meta.get("nat_answer_type", "real")}',
        f'nat_tolerance: {meta.get("nat_tolerance", "0")}',
        f'nat_instruction: {meta.get("nat_instruction", "Enter your answer as a numerical value.")}',
        f'total_sections: {len(set(e["section"] for e in entries))}',
        f'{FRAMEWORK_VERSION_STAMP}',  # v2.50.0: was five hand-copied literals — a five-times-recurred stamp-drift class (CHANGELOG v2.14/v2.15/v2.17/v2.47, and again at the v2.49.1->v2.50.0 bump, which blocked release 2026.08.15.8). Now ONE definition; no consumer parses it (verified), keep honest anyway
        '',
    ]

    # ── STRUCTURAL_CHANGES_BY_YEAR (observed from data — not hardcoded) ───────
    year_changes = _compute_structural_changes(entries)
    if year_changes:
        lines += [
            '=== STRUCTURAL_CHANGES_BY_YEAR ===',
            '# Observed from PYQ data — NOT hardcoded. Shows removed/new subtopics,',
            '# format changes, DI shifts, FIGURAL elimination across years.',
            '# Step 7 reads these to understand recent exam structural changes.',
        ]
        lines += year_changes
        lines.append('')

    by_sec = {}
    for e in entries: by_sec.setdefault(e['section'], []).append(e)

    # ── v2.4 / v2.24.1 SUBTOPIC_ID: mint once, collision-safe, IDEMPOTENT ──
    # v2.24.1: run_synthesise() now mints ids BEFORE run_qv() (so the gate and every
    # writer read the SAME id). This is a no-op guard for entries already stamped.
    mint_subtopic_ids(entries, exam_meta)

    for section, sec_entries in by_sec.items():
        # v2.15 BUG-D03 fix: derive option LABEL style from per-question option_label
        # stored during extraction — NOT from option_format.primary (which is the
        # content FORMAT type like 'single_value', a completely different domain).
        # Aggregation: collect option_label from all observed Qs in this section via
        # progress, then majority-vote. Fallback to exam-level option_label_format.
        sec_keys = [(e['section'], e['topic'], e['subtopic']) for e in sec_entries]
        all_labels = []
        for sk in sec_keys:
            for q in progress.get(sk, []):
                lbl = q.get('option_label', 'unknown')
                if lbl != 'unknown':
                    all_labels.append(lbl)
        if all_labels:
            fmt_label = Counter(all_labels).most_common(1)[0][0]
        else:
            fmt_label = meta.get('option_label_format', '1/2/3/4')

        # figural_banned: True when ALL FIGURAL subtopics in section are deprecated/absent
        # Computed from observed data — NOT hardcoded per exam.
        figural_entries = [e for e in sec_entries if e.get('format') == 'FIGURAL']
        figural_banned  = bool(figural_entries) and all(
            e.get('observed_count', 0) == 0
            or all(p.get('deprecated', False)
                   for p in e.get('PYQ_STEM_PATTERNS', []))
            for e in figural_entries
        )

        lines += ['', f'=== SECTION: {section} ===',
                  f'option_label_format: {fmt_label}',
                  f'figural_banned: {str(figural_banned).lower()}',
                  'sub_types_observed:']
        for e in sorted(sec_entries, key=lambda x: -x['observed_count']):
            lines.append(f'  - {e["sub_type_label"]} (n={e["observed_count"]})')
        # v2.23 THREE-AXIS (CATEGORY A): the per-section 3-year format-distribution TARGET
        # (per-paper averages) that Step 6 (allocation: Axis-1/3 + LINKED) and Step 7
        # (generation: the other 7 Axis-2 classes, joint-solved with difficulty) enforce,
        # and Step 8 audits. Omitted for all-Zero-PYQ sections (compute returns None).
        axdist = compute_section_axis_distribution(sec_entries, progress)
        if axdist:
            lines.append('axis_distribution:')
            lines.append(f'  recent_years: {axdist["recent_years"]}')
            lines.append(f'  n_papers_recent: {axdist["n_papers_recent"]}')
            lines.append(f'  mocks_per_window: {axdist["mocks_per_window"]}')
            lines.append(f'  negative_rate: {axdist["negative_rate"]}')
            lines.append(f'  axis1_per_paper: {axdist["axis1_per_paper"]}')
            lines.append(f'  axis2_per_paper: {axdist["axis2_per_paper"]}')
            lines.append(f'  axis3_per_paper: {axdist["axis3_per_paper"]}')
            lines.append(f'  axis2_audit_mode: {axdist["axis2_audit_mode"]}')
        for e in sorted(sec_entries, key=lambda x: -x['observed_count']):
            lines += format_entry(e)

    with open(out, 'w', encoding='utf-8') as f: f.write('\n'.join(lines))
    print(f'Written: {out} ({len(lines)} lines)')
    return out

slugify = bc.slugify   # DELEGATED (Cluster D) — one definition corpus-wide


# Section-prefix recipe (v2.4): readable, stable, deterministic.
# Uses the INITIALS of significant section words (skipping and/of/the/...), so
# 'General Intelligence & Reasoning' -> 'gir', 'Quantitative Aptitude' -> 'qa'.
# Single-word sections fall back to first 4 chars. Prefix collisions between two
# DISTINCT sections are resolved by build_section_prefix_map() with a numeric
# suffix (gs, gs2, ...). exam_meta['section_prefix_overrides'] can pin a prefix.
_PREFIX_STOPWORDS = ('and', 'of', 'the', 'for', 'in', 'to', 'a', 'an')

def section_prefix(section_name, overrides=None):
    overrides = overrides or {}
    if section_name in overrides:
        return overrides[section_name]
    words = [w for w in slugify(section_name).split('_') if w not in _PREFIX_STOPWORDS]
    if len(words) >= 2:
        return ''.join(w[0] for w in words)   # word-initials, e.g. gir / qa / ec
    return (words[0][:4] if words else 'sec')

def build_section_prefix_map(sections, overrides=None):
    """
    Deterministic, collision-safe section→prefix map. Two DISTINCT sections that
    would collapse to the same prefix get numeric suffixes (gs, gs2, ...).
    Build this ONCE from the full section list, then pass the resulting overrides
    into make_subtopic_id so every subtopic uses the same stable prefix.
    """
    seen = {}; result = {}
    for s in sections:
        p = section_prefix(s, overrides)
        if p in seen and seen[p] != s:
            base = p; n = 2
            while p in seen and seen[p] != s:
                p = f'{base}{n}'; n += 1
        seen[p] = s; result[s] = p
    return result

def _as_mandate_int(v):
    """v2.30 — renamed from _as_int (GAP-2026-07-25-002): blueprint_core defines a
    DIFFERENT _as_int with different semantics (returns a 0 default, not None). Two
    private helpers sharing a name across an engine and a spec that imports it is a
    trap waiting for the first `from blueprint_core import *`.
    v2.11 — coerce a mandate integer field (may arrive as a str from section_rules,
    or as 'none'/None/'' when unset) to int or None. Used for min_per_series_window,
    min_count, and any future numeric mandate field."""
    if v in (None, '', 'none', 'None'):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _mandate_from_note(note):
    """v2.10 — Robust, sentence-independent 'mandatory every mock' detector.
    Returns True iff the NOTE both (a) mentions MANDATORY and (b) contains an
    every/per mock|paper phrase — anywhere, not necessarily in the same sentence.
    Replaces the v2.4 regex MANDATORY[^.\\n]*\\b(every mock|per mock)\\b, which was
    anchored to a single sentence and accepted only "mock", so it silently missed
    real wording such as "MANDATORY SUBTOPIC ... 1Q per every paper". An explicit
    mandate_every_mock line (emitted by format_entry / editable in section_rules)
    always takes precedence over this fallback. Exam-agnostic."""
    import re as _re
    if not note:
        return False
    return bool(_re.search(r'\bMANDATORY\b', note, _re.I)) and \
           bool(_re.search(r'\b(?:every|per)\s+(?:mock|paper)\b', note, _re.I))


def make_subtopic_id(section, topic, subtopic, prefix_overrides=None):
    """
    Mint the canonical subtopic_id: <section_prefix>.<topic_slug>.<subtopic_slug>
    - Deterministic: same (section, topic, subtopic) → same id, always.
    - Independent of concept_group (a separate concern — DOUBT-3 uniqueness).
    - Human-debuggable: e.g. 'qa.mensuration.mensuration_3d_combined',
      'gir.analogy.mixed_number_letter_analogy'.
    - prefix_overrides: the map from build_section_prefix_map() (collision-safe).
    Collisions (two different subtopics producing the same id) are detected and
    de-duplicated by write_subtopic_manifest() with a numeric suffix _2, _3, ...
    """
    return f'{section_prefix(section, prefix_overrides)}.{slugify(topic)}.{slugify(subtopic)}'


# ═══════════════════════════════════════════════════════════════════════════
# v2.20 — ZERO-PYQ SCAFFOLD ENTRY + TAXONOMY SYNC
# ═══════════════════════════════════════════════════════════════════════════
# These two functions implement Fix A (taxonomy sync) and Fix C (scaffold
# section_rules blocks) from the BUGFIX_Zero_PYQ_Manifest_Gap report.
# Called by run_synthesise() AFTER PYQ-based entries are built and BEFORE
# write_section_rules() / write_subtopic_manifest() run.

# ── v2.24.5: AUTOMATIC ZERO-PYQ FORMAT INFERENCE ────────────────────────────────
# Mirrors the _VISUAL_KEYWORDS set used in synthesise() (the PYQ path). KEEP IN SYNC — any
# keyword added there should be added here so PYQ and zero-PYQ agree on what "looks visual".
_ZP_VISUAL_KEYWORDS = re.compile(
    r'(?i)\b(figure[s]?|figural|diagram[s]?|venn|'
    r'mirror\s+image[s]?|water\s+image[s]?|paper\s*fold(ing)?|'
    r'counting\s+(figure|triangle|shape)[s]?|embedded\s+figure[s]?|'
    r'completion\s+of\s+(figure|pattern)[s]?|dice|cube\s+fold(ing)?|'
    r'pattern\s+completion|image\s+series|visual\s+reasoning)\b')


def infer_zero_pyq_axes(name, sibling_formats, sibling_answer_types, sibling_cardinalities):
    """v2.24.5 — PURE. Automatically infer a zero-PYQ subtopic's Axis-1 format + answer_type +
    answer_cardinality from evidence available WITHOUT any curator input:
      (1) NAME keyword  -> FIGURAL  (per-subtopic; the same heuristic trusted for PYQ subtopics);
      (2) same-topic PYQ SIBLINGS   -> inherit a UNANIMOUS non-TEXT format (>=2 siblings, all
          identical); inherit 'numerical' / 'multi' when >= two-thirds of (>=2) siblings are.
    Falls back to TEXT / option / single when there is no strong signal. Deterministic; no I/O.
    Precedence: name keyword (1) beats sibling inheritance (2). Returns
      {format, answer_type, answer_cardinality, inherently_visual, reason}.
    """
    fmt, inh_vis, reason = 'TEXT', False, 'default'
    if _ZP_VISUAL_KEYWORDS.search(name or ''):                      # (1) name -> FIGURAL
        fmt, inh_vis, reason = 'FIGURAL', True, 'name_keyword'
    else:                                                           # (2) UNANIMOUS topic siblings
        sibs = [f for f in (sibling_formats or []) if f]
        if len(sibs) >= 2 and len(set(sibs)) == 1 and sibs[0] != 'TEXT':
            fmt, reason = sibs[0], 'topic_unanimous'

    ans_type = 'option'                                             # NAT: >=2/3 of >=2 siblings
    at = [a for a in (sibling_answer_types or []) if a]
    if len(at) >= 2 and at.count('numerical') * 3 >= len(at) * 2:
        ans_type = 'numerical'

    ans_card = 'single'                                             # MSQ: >=2/3 of >=2 siblings
    ac = [c for c in (sibling_cardinalities or []) if c]
    if len(ac) >= 2 and ac.count('multi') * 3 >= len(ac) * 2:
        ans_card = 'multi'

    return {'format': fmt, 'answer_type': ans_type, 'answer_cardinality': ans_card,
            'inherently_visual': inh_vis, 'reason': reason}


def apply_zero_pyq_format_inference(entries):
    """v2.24.5 — In-place post-pass over the full entry set. For each ZERO-PYQ entry
    (observed_count == 0), infer format/answer_type/answer_cardinality from its NAME and its
    same-(section, topic) PYQ siblings (observed_count > 0), via infer_zero_pyq_axes(). PYQ
    entries are NEVER touched. Every change is logged (audit trail; no prompt). Runs BEFORE
    id-mint / stamp / QV / writers, so both section_rules and the manifest see the inferred axes.
    """
    from collections import defaultdict
    sib_fmt, sib_at, sib_ac = defaultdict(list), defaultdict(list), defaultdict(list)
    for e in entries:
        if e.get('observed_count', 0) > 0:                          # PYQ sibling
            k = (e.get('section'), e.get('topic'))
            sib_fmt[k].append(e.get('format', 'TEXT'))
            sib_at[k].append(e.get('answer_type', 'option'))
            sib_ac[k].append(e.get('answer_cardinality', 'single'))

    changed = 0
    for e in entries:
        if e.get('observed_count', 0) != 0:                         # PYQ subtopic — untouched
            continue
        k = (e.get('section'), e.get('topic'))
        res = infer_zero_pyq_axes(e.get('subtopic', ''), sib_fmt.get(k, []),
                                  sib_at.get(k, []), sib_ac.get(k, []))
        if res['reason'] == 'default' and res['answer_type'] == 'option' \
                and res['answer_cardinality'] == 'single':
            continue                                                # no signal → keep safe defaults
        before = (e.get('format'), e.get('answer_type'), e.get('answer_cardinality'))
        e['format'] = res['format']
        e['answer_type'] = res['answer_type']
        e['answer_cardinality'] = res['answer_cardinality']
        e['inherently_visual'] = bool(res['inherently_visual']) or e.get('inherently_visual', False)
        # keep the freq fields consistent with the inferred string type (some derivations read them)
        if res['answer_type'] == 'numerical':
            e['nat_freq'] = max(e.get('nat_freq', 0), 100)
        if res['answer_cardinality'] == 'multi':
            e['msq_freq'] = max(e.get('msq_freq', 0), 100)
        if res['format'] == 'FIGURAL' and not e.get('figural_data'):
            e['figural_data'] = {                                   # same conservative default as PYQ path
                'image_role': 'stem_only',
                'object_types': {'dominant': [], 'observed': [], 'avoid': []},
                'transformation_types': [], 'arrangement_types': [],
                'complexity_dist': {}, 'images_analysed': 0, 'images_unclear': 0}
        changed += 1
        print(f"  ZERO-PYQ INFERENCE: '{e.get('subtopic')}' "
              f"[{e.get('section')} > {e.get('topic')}] {before} -> "
              f"({res['format']}, {res['answer_type']}, {res['answer_cardinality']}) "
              f"[{res['reason']}]")
    if changed:
        print(f"Zero-PYQ format inference: {changed} scaffold entr"
              f"{'y' if changed == 1 else 'ies'} refined from name/topic evidence.")
    return entries


def make_zero_pyq_scaffold_entry(section, topic, subtopic):
    """
    v2.20 Fix C — Create a COMPLETE entry dict for a zero-PYQ subtopic.
    This entry has every field that format_entry() expects, filled with
    zero-PYQ defaults. The resulting section_rules block gives Step 7
    enough structure to generate questions using training knowledge +
    format guidance, even though no PYQ data exists for this subtopic.

    Returns: dict matching the synthesise_subtopic() output schema.
    """
    return {
        'section': section,
        'topic': topic,
        'subtopic': subtopic,
        'observed_count': 0,
        'format': 'TEXT',
        'option_format': {
            'primary': 'single_value',
            'recent_format': 'single_value',
            'changed_recently': False,
            'all_observed': ['single_value']
        },
        'OMML_required': False,
        'negative_question_freq': 0,
        'answer_type': 'option',
        'nat_freq': 0,
        'answer_cardinality': 'single',
        'msq_freq': 0,
        # v2.23: Zero-PYQ scaffold — no observed Axis-2; capability = family ∪ {DIRECT}
        # so Step 7 can use it as a format-elastic filler (decision 11).
        'observed_axis2': {},
        'presentation_family': resolve_presentation_family_s5(subtopic, 'TEXT'),
        'axis2_capability': axis2_capability(
            [], resolve_presentation_family_s5(subtopic, 'TEXT'), 'TEXT'),
        'fill_in_blank': 'none',
        'linked_group_size': 0,
        'max_per_paper': 2,
        'typical_per_paper': 1,
        'stem_word_count': {'min': 10, 'max': 50, 'typical': 25},
        'sub_type_label': subtopic,
        'concept_group': None,        # v2.24.1 (D8): stamped later by stamp_mechanic_axes()
        'question_mechanic': None,    # v2.24.1 (D8)
        'form_key': None,             # v2.24.1 (D8)
        'collision_domain': None,     # v2.24.1 (D8)
        'mandate_every_mock': False,
        'alternation_group': None,
        'min_per_series_window': None,
        'mandatory_group': None,
        'min_count': None,
        'PYQ_STEM_PATTERNS': [
            {
                'id': 'P1',
                'template': f'Standard {subtopic} question',
                'approach': 'direct',
                'frequency': 100,
                'confidence': 'absent',
                'raw_count': 0,
                'years': [],
                'note_block': 'never',
                'deprecated': False
            }
        ],
        'wrong_option_structure': {
            'type': 'same_category',
            'description': 'Options from the same conceptual category as the correct answer.'
        },
        'NOTE': (f'Zero-PYQ subtopic. No PYQ observations available. '
                 f'Added from exam syllabus/taxonomy via v2.20 taxonomy sync. '
                 f'Step 7 generates questions using exam-level knowledge and '
                 f'the format/option guidance above.'),
        'note_text': (f'Zero-PYQ subtopic. No PYQ observations available. '
                      f'Added from exam syllabus/taxonomy via v2.20 taxonomy sync. '
                      f'Step 7 generates questions using exam-level knowledge and '
                      f'the format/option guidance above.')
    }


def taxonomy_sync_entries(existing_entries, exam_code):
    """
    v2.20 Fix A — TAXONOMY SYNC PROTOCOL.

    After PYQ-based synthesis produces the entry list, this function
    synchronises it with the exam's approved taxonomy to ensure the
    manifest covers the COMPLETE subtopic vocabulary, not just the
    PYQ-observed subset.

    For every taxonomy-defined subtopic NOT already in existing_entries,
    creates a scaffold entry via make_zero_pyq_scaffold_entry().

    Taxonomy sources (UNION — both loaded, not primary/fallback):
      1. taxonomy_draft.json: [ExamCode]_taxonomy_draft.json in project
         knowledge — the syllabus-faithful taxonomy from Step 2a. This is
         the source that contains zero-PYQ subtopics (new syllabus entries
         that have never appeared in any PYQ paper). PRIMARY source.
      2. Approved Analysis doc: [ExamCode]_PYQ_Analysis.docx in project
         knowledge — ADDITIONAL source. The Analysis doc typically contains
         only PYQ-observed subtopics (which are already in existing_entries),
         but is loaded as a safety net in case taxonomy_draft is absent.

    WHY taxonomy_draft IS PRIMARY (not Analysis doc):
      The Analysis doc (Step 2c output) contains ~PYQ-observed subtopics.
      Zero-PYQ subtopics (the ones this function exists to find) are NOT
      in the Analysis doc — they have zero PYQ observations so they were
      never processed. The taxonomy_draft.json (Step 2a output) IS the
      syllabus-faithful source that contains ALL subtopics including
      zero-PYQ ones. Evidence: SSC CGL Tier 2 — taxonomy_draft had 93
      subtopics (all 7 orphans present), Analysis doc had ~96 (orphans
      absent, but finer PYQ-discovered splits present).

    Returns: (new_entries_list, sync_log_lines)
      new_entries_list: scaffold entries to APPEND to existing_entries.
      sync_log_lines:   human-readable log of what was added/skipped.

    EDGE CASES (ref: BUGFIX_Zero_PYQ_Manifest_Gap.md):
      EC-ZP-1: Taxonomy subtopic name COLLIDES with existing entry
               (same slugified name) → use existing, do not duplicate.
      EC-ZP-2: Taxonomy subtopic is COVERED by finer-grained existing
               entries (e.g., taxonomy "Geometry" but entries have
               "Triangles", "Circles") → do not create; log coverage.
      EC-ZP-3: Taxonomy subtopic under DIFFERENT section than existing
               → section mismatch is a HARD STOP (taxonomy must align
               with exam_config sections).
      EC-ZP-4: Exam with ZERO PYQ papers (100% zero-PYQ) → all
               subtopics become scaffold entries. Supported.
      EC-ZP-5: taxonomy_draft.json absent AND Analysis doc absent
               → sync SKIPS (cannot add what doesn't exist). Log warning.
      EC-ZP-6: PYQ-discovered subtopic NOT in taxonomy → already in
               entries. Taxonomy sync only ADDS; it never removes.
      EC-ZP-7: Same subtopic in MULTIPLE sections in taxonomy
               → each gets its own entry (slugify includes section prefix).
      EC-ZP-8: Subtopic in taxonomy_draft but REMOVED from Analysis doc
               → taxonomy_draft is the syllabus-faithful source; if a
               subtopic is in taxonomy_draft it represents the exam
               authority's syllabus. If it was intentionally excluded,
               it should be removed from taxonomy_draft.json itself.
      EC-ZP-9: Subtopic in section_rules but not manifest → handled by
               rebuild_subtopic_manifest_from_section_rules() (separate).
      EC-ZP-10: Large exam (300+ taxonomy, 50 in PYQs) → adds 250+
                scaffold entries. No performance issue (flat list).
    """
    import json, os, glob, re

    sync_log = []
    new_entries = []

    # ── Step 1: Build the EXISTING subtopic index from PYQ-derived entries ──
    # Key: (section_slug, topic_slug, subtopic_slug) for collision detection.
    # Also build a set of subtopic_slugs per section for EC-ZP-2 coverage check.
    existing_keys = set()
    existing_slugs_by_section = {}   # section -> set of subtopic slugs
    existing_topic_slugs_by_section = {}  # section -> set of topic slugs
    for e in existing_entries:
        sec_s = slugify(e['section'])
        top_s = slugify(e['topic'])
        sub_s = slugify(e['subtopic'])
        existing_keys.add((sec_s, top_s, sub_s))
        existing_slugs_by_section.setdefault(sec_s, set()).add(sub_s)
        existing_topic_slugs_by_section.setdefault(sec_s, set()).add(top_s)

    # ── Step 2: Load taxonomy sources (UNION — both tried) ──
    taxonomy_tuples = []   # list of (section, topic, subtopic) raw strings
    seen_tuples = set()    # dedup across sources

    # Source 1 (PRIMARY): taxonomy_draft.json — the syllabus-faithful taxonomy
    # This is the source that contains zero-PYQ subtopics (the orphans).
    for search_dir in ['/mnt/project/', '/mnt/user-data/uploads/']:
        for f in glob.glob(os.path.join(search_dir, f'*taxonomy_draft*.json')):
            try:
                with open(f, encoding='utf-8') as fp:
                    tax_data = json.load(fp)
                # v2.17 BUGFIX — the taxonomy lives under ['sections'], not at
                # top level. Reading the top level made the isinstance guards
                # skip everything: this PRIMARY source silently yielded ~0
                # subtopics and fell through to Source 2, losing exactly the
                # zero-PYQ orphan subtopics it exists to supply. Consistent with
                # the SSC CGL Tier 2 Mock 1 incident (7 syllabus-only subtopics
                # first surfacing at Step 6/7). Pre-existing, not v2.17.
                # CANONICAL READER — do not hand-roll another copy.
                from syllabus_provenance import read_taxonomy_draft
                for (sec, top, sub) in read_taxonomy_draft(tax_data):
                    tup_key = (slugify(sec), slugify(top), slugify(sub))
                    if tup_key not in seen_tuples:
                        taxonomy_tuples.append((sec, top, sub))
                        seen_tuples.add(tup_key)
                if taxonomy_tuples:
                    sync_log.append(f'  Taxonomy source 1 (PRIMARY): taxonomy_draft.json ({os.path.basename(f)}) — {len(taxonomy_tuples)} subtopics')
                    break
            except Exception as ex:
                sync_log.append(f'  WARN: taxonomy_draft.json unreadable: {ex}')

    # Source 2 (ADDITIONAL): approved Analysis doc — safety net for when
    # taxonomy_draft is absent. Read through corpus_io Cluster K (v2.30) and
    # asserted against the approval record's taxonomy_fingerprint (v2.31).
    analysis_added = 0
    for search_dir in ['/mnt/project/', '/mnt/user-data/uploads/']:
        for f in glob.glob(os.path.join(search_dir, '*.docx')):
            bn = os.path.basename(f).lower()
            if 'analysis' in bn and exam_code.lower() in bn:
                try:
                    doc_tuples = _extract_taxonomy_tuples_from_analysis_doc(f)
                    for tup in doc_tuples:
                        tup_key = (slugify(tup[0]), slugify(tup[1]), slugify(tup[2]))
                        if tup_key not in seen_tuples:
                            taxonomy_tuples.append(tup)
                            seen_tuples.add(tup_key)
                            analysis_added += 1
                    if doc_tuples:
                        sync_log.append(
                            f'  Taxonomy source 2 (ADDITIONAL): Analysis doc ({os.path.basename(f)}) — '
                            f'{analysis_added} new subtopics added beyond taxonomy_draft')
                except corpus_io.AnalysisDocError:
                    # v2.31 — NEVER a WARN. An unreadable or unapproved Analysis doc
                    # is not a source that happened to contribute nothing; it is the
                    # wrong taxonomy, and every id minted from it is wrong. The old
                    # code caught this here AND inside the extractor, so the fault
                    # had to survive two independent downgrades to be seen at all.
                    raise
                except Exception as ex:
                    sync_log.append(f'  WARN: Analysis doc {os.path.basename(f)} unreadable: {ex}')

    # EC-ZP-5: no taxonomy source found → skip sync
    if not taxonomy_tuples:
        sync_log.append(
            '  Taxonomy sync SKIPPED: no taxonomy_draft.json or Analysis doc found. '
            'If Step 6 encounters unresolvable subtopics, re-run Step 2a (PYQDraft) '
            'to generate taxonomy_draft.json, then re-run Step 5.')
        return new_entries, sync_log

    sync_log.append(f'  Total taxonomy subtopics (union): {len(taxonomy_tuples)}')

    # ── Step 3: Diff taxonomy against existing entries ──
    added_count = 0
    skipped_existing = 0
    skipped_covered = 0

    for (tax_sec, tax_top, tax_sub) in taxonomy_tuples:
        sec_s = slugify(tax_sec)
        top_s = slugify(tax_top)
        sub_s = slugify(tax_sub)

        # EC-ZP-1: exact match (same section+topic+subtopic slug) → already exists
        if (sec_s, top_s, sub_s) in existing_keys:
            skipped_existing += 1
            continue

        # EC-ZP-1 variant: same subtopic slug under same section (different topic slug)
        # Check if ANY existing entry in this section has the same subtopic slug.
        # This catches renames/reclassifications at the topic level.
        if sec_s in existing_slugs_by_section and sub_s in existing_slugs_by_section[sec_s]:
            skipped_existing += 1
            sync_log.append(
                f'  SKIP (EC-ZP-1): "{tax_sub}" in [{tax_sec}] — slug matches existing entry.')
            continue

        # EC-ZP-2: coverage check — is this taxonomy subtopic COVERED by finer-grained
        # existing entries? If the taxonomy subtopic slug is a PREFIX of any existing
        # subtopic's topic_slug or subtopic_slug within the same section, the finer
        # entries likely cover the taxonomy scope.
        if sec_s in existing_slugs_by_section:
            covered_by = [s for s in existing_slugs_by_section[sec_s]
                          if sub_s in s and sub_s != s]
            if covered_by:
                skipped_covered += 1
                sync_log.append(
                    f'  SKIP (EC-ZP-2): "{tax_sub}" in [{tax_sec}] — covered by '
                    f'finer entries: {covered_by[:5]}')
                continue

        # Not in existing → create scaffold entry
        scaffold = make_zero_pyq_scaffold_entry(tax_sec, tax_top, tax_sub)
        new_entries.append(scaffold)
        existing_keys.add((sec_s, top_s, sub_s))
        existing_slugs_by_section.setdefault(sec_s, set()).add(sub_s)
        added_count += 1
        sync_log.append(f'  ADDED: "{tax_sub}" [{tax_sec} > {tax_top}] — zero-PYQ scaffold')

    sync_log.append(
        f'  Taxonomy sync complete: {added_count} zero-PYQ scaffolds added, '
        f'{skipped_existing} already existed, {skipped_covered} covered by finer entries.')

    return new_entries, sync_log


def _extract_taxonomy_tuples_from_analysis_doc(docx_path):
    """(section, topic, subtopic) tuples from the approved Analysis doc.

    v2.30 (GAP-2026-07-25-002) — DELEGATED to corpus_io Cluster K, THE reader for
    this artefact. The previous implementation was the tuple-returning twin of
    extract_taxonomy_from_analysis_doc() and shared all of its defects: paragraphs
    only (subtopics are in tables), Heading-style detection against a generator
    that emits no styles, and a first-value latch. It returned an empty list from
    every real Analysis doc, and — because both its except branches swallow
    silently — that emptiness was indistinguishable from "the doc had nothing to
    add". Measured on the first exam's live doc: 0 tuples against a truth of 131.

    Returns [] when the doc is ABSENT, which is a supported configuration: this is
    an ADDITIONAL source and Step 5 must still run without it.

    v2.31 — WHAT IT NO LONGER SWALLOWS. The old body was `except Exception: return
    []`, which made three different situations indistinguishable: "no Analysis doc,
    by design", "the Analysis doc is unreadable", and "the Analysis doc is not the
    approved one". Only the first is benign. The other two returned [] and the
    caller logged a WARN, so the safety net that mints ids for zero-PYQ subtopics
    quietly contributed nothing — exactly the shape of the v2.30 defect, one layer
    up. Absence is still tolerated; a doc that is PRESENT and wrong is now loud.
    """
    import corpus_io
    try:
        doc = corpus_io.load_taxonomy(path=docx_path, step='PYQExtract')
    except corpus_io.AnalysisDocError as ex:
        if 'no Analysis doc found' in str(ex):
            return []                      # absent by design — the only benign case
        raise                              # present but unreadable — never a WARN
    return list(doc['triples'])


def write_taxonomy_xlsx(manifest, exam_code, out_dir=None):
    """v2.24 — Emit [ExamCode]_taxonomy.xlsx: a plain, human-readable list of
    Subject | Topic | Sub Topic Name | Sub Topic Id | Upload Order (one row per
    sub-topic, in MANIFEST ORDER), so the Step-6 operator can pick scope values
    WITHOUT reading the manifest JSON. It is a companion to
    subtopic_manifest.json, generated from the SAME dict — the JSON stays
    authoritative. Failure to write (e.g. openpyxl absent) is a WARN, never a
    hard stop.

    v2.48.0 — ROW ORDER IS THE TEACHING ORDER, NEVER RE-SORTED
    (GAP-2026-08-14-TAXONOMY-ORDER). Through v2.47.x this function sorted rows
    alphabetically "for readability" — which silently DISCARDED the manifest's
    curated order. That order is load-bearing three times over:
    notes_core.assign_numbering freezes first-seen manifest order into the
    permanent S/T/ST unit numbers (verified against delivered filenames:
    Vector Calculus is ST02 by syllabus order, ST06 alphabetically); the Notes
    portal upload sequence follows this sheet top-to-bottom; and the Notes
    integration sections' backward-only rule runs on the same order. An
    alphabetical resort therefore showed the operator a DIFFERENT order from
    the one every filename and title number already carried. Rows now emit in
    manifest insertion order with an explicit Upload Order column (E). Columns
    A-D are UNCHANGED in position and meaning — every downstream instruction
    that says "column D" still holds.

    Column -> Step-6 use:  Subject = subject scope · "Subject::Topic" = topic scope ·
    "Subject::Topic::Sub Topic Name" = subtopic scope (or the Sub Topic Id if that name repeats
    under the same topic). Upload Order (column E) is for the PORTAL: upload
    sub-topics in this sequence, top to bottom.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARN: openpyxl unavailable — skipped taxonomy.xlsx (manifest.json is authoritative).")
        return None

    # MANIFEST ORDER, verbatim — dict insertion order IS the taxonomy order
    # (the same order notes_core.assign_numbering freezes into unit numbers).
    rows = [[v.get('section', ''), v.get('topic', ''), v.get('display_name', ''),
             sid, i]
            for i, (sid, v) in enumerate(manifest.get('subtopics', {}).items(),
                                         start=1)]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Taxonomy'
    head_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='1F4E79')
    arial = Font(name='Arial', size=11)
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, name in enumerate(['Subject', 'Topic', 'Sub Topic Name', 'Sub Topic Id',
                              'Upload Order'], start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
    band = PatternFill('solid', fgColor='F2F6FB')
    prev_subject, shade = None, False
    for i, row in enumerate(rows, start=2):
        if row[0] != prev_subject:
            shade = not shade
            prev_subject = row[0]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = arial
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if shade:
                cell.fill = band
    for c, width in enumerate([26, 30, 46, 30, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{len(rows) + 1}'

    hw = wb.create_sheet('How to use')
    guide = [
        (f'{exam_code} — how to use this list in Step 6', True),
        ('', False),
        ('This lists every Subject, Topic and Sub-Topic in your exam. Pick from the', False),
        ('"Taxonomy" tab, then copy the value into the Step-6 trigger (match text EXACTLY —', False),
        ('spelling and capital letters matter). Use the header filter arrows to narrow down.', False),
        ('', False),
        ('PORTAL UPLOAD ORDER — the rows are already in the correct teaching sequence', True),
        ('(basics first). Upload sub-topics to the portal from TOP TO BOTTOM, exactly as', False),
        ('listed; the "Upload Order" column (E) numbers that sequence. Do NOT re-sort', False),
        ('this sheet before uploading — the row order matches the numbering printed in', False),
        ('every delivered file.', False),
        ('', False),
        ('SUBJECT test   -> copy the Subject cell (column A):', True),
        ('    ScopedBlueprint --level subject  --scope "<Subject>"          --count N --qs_per_paper Q', False),
        ('', False),
        ('TOPIC test     -> join Subject and Topic with "::" (columns A and B):', True),
        ('    ScopedBlueprint --level topic    --scope "<Subject>::<Topic>" --count N --qs_per_paper Q', False),
        ('', False),
        ('SUB-TOPIC test -> join Subject + Topic + Sub Topic Name with "::" (columns A, B, C):', True),
        ('    ScopedBlueprint --level subtopic --scope "<Subject>::<Topic>::<Sub Topic Name>" --count N --qs_per_paper Q', False),
        ('    The Topic in the middle keeps same-named sub-topics apart — e.g. "Kinematics"', False),
        ('    under Mechanics vs under Rotational Motion resolve to two DIFFERENT sub-topics.', False),
        ('    Use the Sub Topic Id (column D) ONLY if the same name repeats under the same Topic:', False),
        ('        --scope <Sub Topic Id>', False),
    ]
    for i, (text, bold) in enumerate(guide, start=1):
        hw.cell(row=i, column=1, value=text).font = Font(name='Arial', size=11, bold=bold)
    hw.column_dimensions['A'].width = 100

    out = _artefact_path(out_dir, f'{exam_code}_taxonomy.xlsx')
    wb.save(out)
    print(f'Written: {out} ({len(rows)} sub-topics — human-readable taxonomy for Step 6)')
    return out


# ═══ SUBTOPIC ORDERING RULE (v2.48.0 — GAP-2026-08-14-TAXONOMY-ORDER) ═══
# The ORDER in which subtopics enter this manifest is the exam's PERMANENT
# order: notes_core.assign_numbering derives the S/T/ST unit numbers from
# manifest insertion order and preserves them verbatim forever after (a
# re-run never renumbers a delivered unit). The taxonomy xlsx above emits
# rows in this same order, and the operator uploads to the portal top to
# bottom. Three consequences, all binding:
#   (1) FIRST EXTRACTION OF A NEW EXAM: enter each topic's subtopics in
#       TEACHING ORDER — the official syllabus order by default (syllabi are
#       written basics-first), adjusted by SME judgement where the syllabus
#       order is not pedagogical. Alphabetical entry is a DEFECT: it freezes
#       an arbitrary order into every filename, title number and portal link
#       the exam will ever ship.
#   (2) RE-RUNS APPEND, NEVER INSERT: a Step-5 re-run that discovers a new
#       subtopic adds it at the END of its topic's block, so manifest order
#       and the persisted numbering can never disagree mid-list.
#   (3) EXISTING EXAMS ARE NEVER REORDERED: persisted numbering makes a
#       reorder cosmetic at best and a filename churn at worst. The fix for
#       an exam whose order is already frozen is none — its numbering was
#       taken from syllabus order at first extraction and stands.
def write_subtopic_manifest(entries, exam_code, exam_meta=None, progress=None,
                            frequency_scope='all', exam_config=None, out_dir=None):
    """
    v2.4 — Write [ExamCode]_subtopic_manifest.json: the AUTHORITATIVE id↔name
    registry and the formal cross-step contract artifact.
    v2.23 — also carries the THREE-AXIS machine data Step 6 reads without re-parsing PYQ:
      • per subtopic: observed_axis2, presentation_family, axis2_capability
      • top-level 'axis_distribution': {SUBJECT: <compute_section_axis_distribution()>}
      • top-level 'axis_distribution_by_exam_section' (v2.26): the same statistic counted
        per EXAM SECTION via exam_config q_ranges. Step 6 PREFERS this; without it Step 6
        must sum every subject and split by section SIZE, which mis-assigns the mix
        (reference exam: measured A 1.4 / B 1.0 / C 2.0 vs apportioned A 2.2 / B 0.7 / C 1.4).
        (needs `progress`; omitted per-section when progress is absent — section_rules.md
        remains the authoritative human-readable copy either way).

    This manifest is the SINGLE SOURCE OF TRUTH for the subtopic vocabulary.
    Step 6 (Blueprint) and Step 7 (Create) MUST read it and reference subtopics
    by subtopic_id. They never invent ids; an unknown id is a HARD STOP.

    Structure:
    {
      "exam_code": "...",
      "manifest_version": "1.0",
      "generated_by": "Framework_MockTestAnalyse v2.53",
      "id_recipe": "<section_prefix>.<topic_slug>.<subtopic_slug> via slugify v2.4",
      "subtopics": {
         "<subtopic_id>": {
            "display_name": "...",      # decorative; may change without breaking joins
            "section": "...",
            "topic": "...",
            "concept_group": "...",     # carried for convenience; NOT part of the id
            "format": "TEXT|PASSAGE|FIGURAL|DI|...",
            "inherently_visual": false,   // v2.22: true if keyword heuristic fired
            "mandates": {               # structured MANDATE data (not prose)
               "mandatory_every_mock": false,
               "alternation_group": null,   # e.g. "ci_si" ; members alternate, never co-occur
               "min_per_series_window": null
            }
         }, ...
      },
      "alternation_groups": {           # groups whose members must NOT co-occur in one mock
         "ci_si": ["qa.interest.simple_interest", "qa.interest.compound_interest"]
      },
      "mandatory_every_mock": [ "<id>", ... ]   # flat list for fast Step-1/2 checks
    }

    MANDATE EXTRACTION (exam-agnostic, v2.10). Three round-trippable fields per
    subtopic, emitted by format_entry and read here AND by
    rebuild_subtopic_manifest_from_section_rules():
      • mandatory_every_mock — an explicit entry field mandate_every_mock wins;
        else _mandate_from_note() (NOTE mentions MANDATORY *and* an every/per
        mock|paper phrase, sentence-independent). This replaces the pre-v2.10
        single-sentence, mock-only regex that missed wording like "1Q per every paper".
      • alternation_group — the authored group name on BOTH members (e.g. set
        alternation_group='ci_si' on Simple Interest AND Compound Interest); the
        manifest groups members by that name. Members must NOT co-occur in one mock.
      • min_per_series_window — authored cadence integer (reserved for the Issue-2b
        window/cadence enforcement; carried through as data, inert until then).
    Because format_entry now writes all three into section_rules, the manifest is
    fully reproducible from the section_rules FILE alone — no ephemeral entry state.
    """
    import json, re
    out = _artefact_path(out_dir, f'{exam_code}_subtopic_manifest.json')
    meta = exam_meta or {}
    overrides = meta.get('section_prefix_overrides', {})

    manifest = {
        'exam_code': exam_code,
        'manifest_version': '1.0',
        'generated_by': FRAMEWORK_STAMP,
        'id_recipe': '<section_prefix>.<topic_slug>.<subtopic_slug>; section_prefix=word-initials (gir/ga/qa/ec), slugify v2.4',
        'subtopics': {},
        'alternation_groups': {},
        'mandatory_every_mock': [],
        # v2.11 — the three mandate types a flat per-id list cannot express (Issue 2b):
        'mandatory_groups': {},   # {group: {members:[ids], min:int}} — >=min members present per mock
        'cadence_windows': {},    # {id: N}  — subtopic must appear >=1 in every N-mock window
        'min_counts': {},         # {id: k}  — subtopic must have >=k Q per mock
        'axis_distribution': {},  # v2.23 {SUBJECT: per-subject format-distribution target}
        'axis_distribution_by_exam_section': {},  # v2.26 {EXAM SECTION: same, measured directly}
        'pattern_eras': {}        # v2.25 per-paper era + the scope used (audit trail)
    }

    # v2.23 per-section axis distribution (needs the question lists in `progress`).
    # v2.25 — the axis distribution is a MIX quantity, so it obeys frequency_scope exactly
    # as the Frequency xlsx does. Two protections already existed and both are kept:
    #   (a) compute_section_axis_distribution() windows to the 3 most recent distinct years,
    #       which for most pattern changes is already entirely current-era; and
    #   (b) bc.derive_axis_schedule() (Framework_Blueprint v1.36) rescales every axis to
    #       sec_qs, so the SIZE is right even when the measurement era is not.
    # Neither fixes the residual case where the pattern changed WITHIN the last 3 years:
    # there the 3-year window straddles two patterns and the class PROPORTIONS are blended.
    # Era-scoping closes that gap at the source. `progress` is never mutated — the filtered
    # structure is a counting view (bc.filter_progress_to_eras returns a copy), so §14
    # pattern synthesis below still sees every era.
    if progress:
        _axis_progress = progress
        if frequency_scope == 'current-era' and exam_config:
            import blueprint_core as bc
            _eras = bc.paper_eras_from_progress(progress, exam_config)
            _filtered, _st = bc.filter_progress_to_eras(progress, _eras,
                                                        keep=('current',))
            # Never trade a real distribution for an empty one: if era-scoping would leave
            # nothing to measure, keep the full-corpus axis and say so. A blended axis that
            # is rescaled to the right size beats no axis at all (status='no_pyq' would
            # switch the whole three-axis feature off for this section).
            if _st['kept_questions'] > 0:
                _axis_progress = _filtered
            else:
                print("  NOTE: era-scoped axis skipped — no current-era questions; "
                      "using full-corpus axis (still rescaled to sec_qs at Step 6).")
        _by_sec = {}
        for e in entries:
            _by_sec.setdefault(e['section'], []).append(e)
        for _sec, _ents in _by_sec.items():
            _ax = compute_section_axis_distribution(_ents, _axis_progress)
            if _ax:
                manifest['axis_distribution'][_sec] = _ax

        # ── v2.26 (GAP-2026-08-06-AXIS1) — MEASURE THE EXAM SECTIONS DIRECTLY ──────
        # The map above is keyed by SUBJECT (the taxonomy's top level). Step 6 needs it
        # per EXAM SECTION (A/B/C), and until now bridged that gap by SUMMING every
        # subject and splitting the total by section SIZE. That is a real distortion,
        # not a rounding detail — on the reference exam:
        #
        #     measured per section   : A 1.4   B 1.0   C 2.0   (figures/paper)
        #     size-apportioned       : A 2.2   B 0.7   C 1.4
        #
        # It hands the FEWEST figures to Section C, which actually carries the MOST
        # (5 of the 8 figures in the most recent paper). Sections are format bands, not
        # content domains; their format mix has no reason to track their question count.
        #
        # Counting them directly is possible because every question already carries its
        # q_num, and exam_config already carries each section's q_range. Emitted under a
        # SEPARATE key so a consumer that wants the subject view still has it, and a
        # pre-v2.26 blueprint that never looks here is completely unaffected.
        _ranges = []
        for _s in (exam_config or {}).get('sections', []) or []:
            _r = _s.get('q_range') or []
            if len(_r) >= 2:
                _ranges.append((_s.get('section_name') or _s.get('name'),
                                int(_r[0]), int(_r[1])))
        if _ranges:
            manifest['axis_distribution_by_exam_section'] = {}
            for _nm, _lo, _hi in _ranges:
                # A counting VIEW over the same question lists — `progress` is never
                # mutated, exactly as the era filter above leaves it untouched.
                _sec_view, _seen = {}, False
                for _k, _qs in _axis_progress.items():
                    _keep = [q for q in _qs
                             if q.get('q_num') is not None and _lo <= int(q['q_num']) <= _hi]
                    if _keep:
                        _sec_view[_k] = _keep
                        _seen = True
                if not _seen:
                    continue      # section not represented (out-of-pattern era) — skip
                _ents_all = [{'section': k[0], 'topic': k[1], 'subtopic': k[2]}
                             for k in _sec_view]
                _ax = compute_section_axis_distribution(_ents_all, _sec_view)
                if _ax:
                    _ax['measured_by'] = 'measured'   # provenance → blueprint.axis_measured_by
                    manifest['axis_distribution_by_exam_section'][_nm] = _ax

    # v2.25 — record every paper's pattern era in the manifest so the decision is auditable
    # downstream and after the fact. Purely additive: absent on a pre-v2.25 manifest, and no
    # consumer is required to read it.
    if progress and exam_config:
        import blueprint_core as bc
        try:
            _pe = bc.paper_eras_from_progress(progress, exam_config)
            manifest['pattern_eras'] = {
                'frequency_scope': frequency_scope,
                'papers': {f'{y}|{sh}': info for (y, sh), info in sorted(
                    _pe.items(), key=lambda kv: (str(kv[0][0]), str(kv[0][1])))},
                'era_counts': {e: sum(1 for i in _pe.values() if i['era'] == e)
                               for e in bc.PATTERN_ERAS},
            }
        except ValueError:
            pass          # exam_config without sections[] — nothing to classify against

    seen_ids = {}
    for e in entries:
        # Reuse the id stamped by write_section_rules (e['subtopic_id']) so the
        # manifest id is IDENTICAL to the section_rules block id. Only recompute as
        # a fallback if write_subtopic_manifest is ever called before stamping.
        sid = e.get('subtopic_id')
        if not sid:
            sid = make_subtopic_id(e['section'], e['topic'], e['subtopic'],
                                   build_section_prefix_map(
                                       sorted({x['section'] for x in entries}), overrides))
            base = sid; n = 2
            key = (e['section'], e['topic'], e['subtopic'])
            while sid in seen_ids and seen_ids[sid] != key:
                sid = f'{base}_{n}'; n += 1
        seen_ids[sid] = (e['section'], e['topic'], e['subtopic'])

        note = (e.get('NOTE') or e.get('note') or e.get('note_text') or '')
        # v2.10: an explicit mandate_every_mock (from the entry / round-tripped from
        # section_rules) wins; otherwise the robust sentence-independent detector
        # (_mandate_from_note) replaces the brittle single-sentence, mock-only regex
        # that missed wording like "MANDATORY ... 1Q per every paper".
        mand_every = bool(e['mandate_every_mock']) if 'mandate_every_mock' in e \
                     else _mandate_from_note(note)
        manifest['subtopics'][sid] = {
            'display_name': e['subtopic'],
            'section': e['section'],
            'topic': e['topic'],
            'concept_group': e.get('concept_group') or _derive_concept_group(e),
            'question_mechanic': e.get('question_mechanic') or _derive_question_mechanic(e),  # v2.24
            'form_key': e.get('form_key') or _derive_form_key(e),                             # v2.24
            'collision_domain': e.get('collision_domain') or _derive_collision_domain(e),     # v2.24
            'format': e.get('format', 'TEXT'),
            'inherently_visual': bool(e.get('inherently_visual', False)),   # v2.22
            # v2.26 — Step 6 copies these into blueprint.subtopic_list and Step 7 ranks
            # the Axis-1 budget by them. Absent on a pre-v2.26 manifest, in which case
            # bc.rank_figural_candidates() degrades to irreducible-first ordering and the
            # budget is still enforced — the cap is what matters; the ranking is what
            # makes the capped set faithful.
            'figural_q_count'  : int(e.get('figural_q_count', 0)),
            'figural_rate'     : float(e.get('figural_rate', 0.0)),
            'figural_reducible': bool(e.get('figural_reducible', True)),
            'di_q_count'       : int(e.get('di_q_count', 0)),          # v2.43
            'di_rate'          : float(e.get('di_rate', 0.0)),         # v2.43
            'di_reducible'     : bool(e.get('di_reducible', True)),    # v2.43
            # v2.23 THREE-AXIS per-subtopic capability (Step 6 rare-format reachability;
            # Step 7 renders only within axis2_capability — fabrication banned).
            'observed_axis2': e.get('observed_axis2', {}),
            'presentation_family': e.get('presentation_family'),
            'axis2_capability': e.get('axis2_capability', ['DIRECT']),
            'mandates': {
                'mandatory_every_mock': mand_every,
                'alternation_group': e.get('alternation_group'),
                'min_per_series_window': _as_mandate_int(e.get('min_per_series_window')),
                'mandatory_group': e.get('mandatory_group'),          # v2.11 group-presence
                'min_count': _as_mandate_int(e.get('min_count'))              # v2.11 min Q per mock
            }
        }
        if mand_every:
            manifest['mandatory_every_mock'].append(sid)
        ag = e.get('alternation_group')
        if ag:
            manifest['alternation_groups'].setdefault(ag, [])
            if sid not in manifest['alternation_groups'][ag]:
                manifest['alternation_groups'][ag].append(sid)
        # v2.11 rollups (Issue 2b) — group-presence, cadence, min-count:
        mg = e.get('mandatory_group')
        if mg and mg != 'none':
            g = manifest['mandatory_groups'].setdefault(mg, {'members': [], 'min': 1})
            if sid not in g['members']:
                g['members'].append(sid)
        _win = _as_mandate_int(e.get('min_per_series_window'))
        if _win:
            manifest['cadence_windows'][sid] = _win
        _mc = _as_mandate_int(e.get('min_count'))
        if _mc:
            manifest['min_counts'][sid] = _mc

    with open(out, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f'Written: {out} ({len(manifest["subtopics"])} subtopic ids, '
          f'{len(manifest["mandatory_every_mock"])} mandatory-every-mock, '
          f'{len(manifest["alternation_groups"])} alternation groups, '
          f'{len(manifest["mandatory_groups"])} group-presence, '
          f'{len(manifest["cadence_windows"])} cadence, '
          f'{len(manifest["min_counts"])} min-count)')
    write_taxonomy_xlsx(manifest, exam_code, out_dir)   # v2.24 — human-readable companion
    return out


def rebuild_subtopic_manifest_from_section_rules(section_rules_path, exam_code, out_dir=None):
    """v2.10 — Reconstruct a COMPLETE subtopic_manifest.json from an existing
    section_rules.md ALONE (no source PYQ, no in-memory entries). This is the
    supported path to (re)generate a MISSING or INCOMPLETE manifest — e.g. a project
    that has section_rules + blueprint but lost the manifest, or one whose mandate
    markers were edited by hand and must be re-minted.

    It parses every '--- Subtopic: --- block and reads the round-trippable mandate
    lines emitted by format_entry (mandate_every_mock / alternation_group /
    min_per_series_window), falling back to the robust NOTE detector
    (_mandate_from_note) only when an explicit mandate_every_mock line is absent
    (legacy blocks written before v2.10). ONLY id-carrying blocks enter the manifest
    join; a block without subtopic_id cannot be joined by Step 6/7, so it is skipped
    and surfaced as a WARN (mint its id in Step 5, or add one, then re-run).

    Output schema is byte-identical to write_subtopic_manifest(). Exam-agnostic —
    zero subtopic names hardcoded.
    """
    import json, re, ast
    text = open(section_rules_path, encoding='utf-8').read()
    manifest = {
        'exam_code': exam_code,
        'manifest_version': '1.0',
        'generated_by': f'{FRAMEWORK_STAMP} (rebuild_from_section_rules)',
        'id_recipe': '<section_prefix>.<topic_slug>.<subtopic_slug>; slugify v2.4',
        'subtopics': {},
        'alternation_groups': {},
        'mandatory_every_mock': [],
        'mandatory_groups': {},   # v2.11 group-presence
        'cadence_windows': {},    # v2.11 cadence
        'min_counts': {},         # v2.11 min Q per mock
        'axis_distribution': {}   # v2.23 (round-tripped from the AXIS_DISTRIBUTION block below)
    }

    def _lit(s, default):
        """Safe Python-literal parse for round-tripped axis fields ({...}/[...]/floats)."""
        try:
            return ast.literal_eval(s.strip())
        except Exception:
            return default
    EXPL_MAND = re.compile(r'^\s*mandate_every_mock:\s*(true|false)\s*$', re.I | re.M)
    EXPL_ALT  = re.compile(r'^\s*alternation_group:\s*(\S+)\s*$', re.M)
    EXPL_WIN  = re.compile(r'^\s*min_per_series_window:\s*(\S+)\s*$', re.M)
    EXPL_GRP  = re.compile(r'^\s*mandatory_group:\s*(\S+)\s*$', re.M)          # v2.11
    EXPL_MINC = re.compile(r'^\s*min_count:\s*(\S+)\s*$', re.M)               # v2.11
    ID_RE     = re.compile(r'^\s*subtopic_id:\s*(\S+)\s*$', re.M)
    SEC_RE    = re.compile(r'^\s*section:\s*(.+)$', re.M)
    TOP_RE    = re.compile(r'^\s*topic:\s*(.+)$', re.M)
    FMT_RE    = re.compile(r'^\s*format:\s*(\S+)\s*$', re.M)
    INHERENT_RE = re.compile(r'^\s*inherently_visual:\s*(true|false)\s*$', re.I | re.M)  # v2.22
    # v2.26 (GAP-2026-08-06-AXIS1) — round-trip the figural rate fields. Tolerant by
    # construction: a block written before v2.26 carries none of them, and every
    # consumer defaults (rate 0.0, reducible True) to the pre-v2.26 reading. A parser
    # that HALTED on the older shape would strand ~200 deployed exams on re-read.
    FIGRATE_RE = re.compile(r'^\s*figural_rate:\s*([0-9]*\.?[0-9]+)\s*$', re.I | re.M)
    FIGQC_RE   = re.compile(r'^\s*figural_q_count:\s*(\d+)\s*$', re.I | re.M)
    FIGRED_RE  = re.compile(r'^\s*figural_reducible:\s*(true|false)\s*$', re.I | re.M)
    OPTIMG_RE  = re.compile(r'^\s*option_image_rate:\s*([0-9]*\.?[0-9]+)\s*$', re.I | re.M)
    DIRATE_RE  = re.compile(r'^\s*di_rate:\s*([0-9]*\.?[0-9]+)\s*$', re.I | re.M)   # v2.43
    DIQC_RE    = re.compile(r'^\s*di_q_count:\s*(\d+)\s*$', re.I | re.M)
    DIRED_RE   = re.compile(r'^\s*di_reducible:\s*(true|false)\s*$', re.I | re.M)
    CG_RE     = re.compile(r'^\s*concept_group:\s*(.+)$', re.M)
    QM_RE      = re.compile(r'^\s*question_mechanic:\s*(.+)$', re.M)    # v2.24
    FORMKEY_RE = re.compile(r'^\s*form_key:\s*(.+)$', re.M)             # v2.24
    CDOM_RE    = re.compile(r'^\s*collision_domain:\s*(.+)$', re.M)     # v2.24
    NOTE_RE   = re.compile(r'^\s*NOTE\s*:\s*(.+)$', re.M | re.I)
    OBSAX_RE  = re.compile(r'^\s*observed_axis2:\s*(.+)$', re.M)          # v2.23
    PFAM_RE   = re.compile(r'^\s*presentation_family:\s*(.+)$', re.M)     # v2.23
    AX2CAP_RE = re.compile(r'^\s*axis2_capability:\s*(.+)$', re.M)        # v2.23

    id_less = []
    for raw in re.split(r'\n--- Subtopic:', text)[1:]:
        disp  = raw.split('---', 1)[0].strip()
        sid_m = ID_RE.search(raw)
        if not sid_m:
            id_less.append(disp)                 # cannot be joined — skip, warn below
            continue
        sid    = sid_m.group(1)
        note_m = NOTE_RE.search(raw)
        note   = note_m.group(1) if note_m else ''
        em = EXPL_MAND.search(raw)
        mand = (em.group(1).lower() == 'true') if em else _mandate_from_note(note)
        av = EXPL_ALT.search(raw)
        ag = av.group(1) if (av and av.group(1).lower() != 'none') else None
        wv  = EXPL_WIN.search(raw)
        win = _as_mandate_int(wv.group(1)) if wv else None
        gv  = EXPL_GRP.search(raw)                                   # v2.11
        grp = gv.group(1) if (gv and gv.group(1).lower() != 'none') else None
        mcv = EXPL_MINC.search(raw)                                  # v2.11
        mc  = _as_mandate_int(mcv.group(1)) if mcv else None
        _sec_r = SEC_RE.search(raw).group(1).strip() if SEC_RE.search(raw) else ''   # v2.24
        _fmt_r = FMT_RE.search(raw).group(1) if FMT_RE.search(raw) else 'TEXT'        # v2.24
        _ax_r  = derive_mechanic(_sec_r, disp, None, '', _fmt_r, sid)                 # v2.24
        manifest['subtopics'][sid] = {
            'display_name':  disp,
            'section':       (SEC_RE.search(raw).group(1).strip() if SEC_RE.search(raw) else ''),
            'topic':         (TOP_RE.search(raw).group(1).strip() if TOP_RE.search(raw) else ''),
            'concept_group': (CG_RE.search(raw).group(1).strip()  if CG_RE.search(raw)  else _ax_r['family']),
            'question_mechanic': (QM_RE.search(raw).group(1).strip() if QM_RE.search(raw) else _ax_r['mechanic']),   # v2.24
            'form_key': (FORMKEY_RE.search(raw).group(1).strip() if FORMKEY_RE.search(raw) else _ax_r['form_key']),  # v2.24
            'collision_domain': (CDOM_RE.search(raw).group(1).strip() if CDOM_RE.search(raw) else _ax_r['collision_domain']),  # v2.24
            'format':        (FMT_RE.search(raw).group(1)         if FMT_RE.search(raw) else 'TEXT'),
            'inherently_visual': (INHERENT_RE.search(raw).group(1).lower() == 'true'
                                  if INHERENT_RE.search(raw) else False),   # v2.22
            'figural_rate': (float(FIGRATE_RE.search(raw).group(1))
                             if FIGRATE_RE.search(raw) else 0.0),            # v2.26
            'figural_q_count': (int(FIGQC_RE.search(raw).group(1))
                                if FIGQC_RE.search(raw) else 0),             # v2.26
            'figural_reducible': (FIGRED_RE.search(raw).group(1).lower() == 'true'
                                  if FIGRED_RE.search(raw) else True),       # v2.26
            'option_image_rate': (float(OPTIMG_RE.search(raw).group(1))
                                  if OPTIMG_RE.search(raw) else 0.0),        # v2.44
            'di_rate': (float(DIRATE_RE.search(raw).group(1))
                        if DIRATE_RE.search(raw) else 0.0),                  # v2.43
            'di_q_count': (int(DIQC_RE.search(raw).group(1))
                           if DIQC_RE.search(raw) else 0),                   # v2.43
            'di_reducible': (DIRED_RE.search(raw).group(1).lower() == 'true'
                             if DIRED_RE.search(raw) else True),             # v2.43
            # v2.23 THREE-AXIS per-subtopic (round-tripped; safe defaults for legacy blocks)
            'observed_axis2':      (_lit(OBSAX_RE.search(raw).group(1), {})
                                    if OBSAX_RE.search(raw) else {}),
            'presentation_family': ((lambda v: None if v in ('None', '') else v)
                                    (PFAM_RE.search(raw).group(1).strip())
                                    if PFAM_RE.search(raw) else None),
            'axis2_capability':    (_lit(AX2CAP_RE.search(raw).group(1), ['DIRECT'])
                                    if AX2CAP_RE.search(raw) else ['DIRECT']),
            'mandates': {
                'mandatory_every_mock': mand,
                'alternation_group':    ag,
                'min_per_series_window': win,
                'mandatory_group':      grp,     # v2.11 group-presence
                'min_count':            mc       # v2.11 min Q per mock
            }
        }
        if mand and sid not in manifest['mandatory_every_mock']:
            manifest['mandatory_every_mock'].append(sid)
        if ag:
            manifest['alternation_groups'].setdefault(ag, [])
            if sid not in manifest['alternation_groups'][ag]:
                manifest['alternation_groups'][ag].append(sid)
        # v2.11 rollups (Issue 2b) — group-presence, cadence, min-count:
        if grp:
            g = manifest['mandatory_groups'].setdefault(grp, {'members': [], 'min': 1})
            if sid not in g['members']:
                g['members'].append(sid)
        if win:
            manifest['cadence_windows'][sid] = win
        if mc:
            manifest['min_counts'][sid] = mc

    # v2.23 — round-trip the per-section AXIS_DISTRIBUTION block so a rebuilt manifest is
    # schema-identical to the write path. Each '=== SECTION: X ===' chunk may carry an
    # 'axis_distribution:' sub-block of indented 'key: <python-literal>' lines.
    _AXKEYS = {'recent_years', 'n_papers_recent', 'mocks_per_window', 'negative_rate',
               'axis1_per_paper', 'axis2_per_paper', 'axis3_per_paper', 'axis2_audit_mode'}
    for chunk in re.split(r'^=== SECTION:\s*', text, flags=re.M)[1:]:
        sec_name = chunk.split('===', 1)[0].strip()
        m = re.search(r'^axis_distribution:\s*$', chunk, flags=re.M)
        if not m:
            continue
        block = {}
        for ln in chunk[m.end():].splitlines():
            if not ln.startswith('  '):           # indented sub-lines only; stop at dedent
                if ln.strip() == '':
                    continue
                break
            km = re.match(r'\s+(\w+):\s*(.+)$', ln)
            if km and km.group(1) in _AXKEYS:
                block[km.group(1)] = _lit(km.group(2), km.group(2).strip())
        if block:
            manifest['axis_distribution'][sec_name] = block

    out = _artefact_path(out_dir, f'{exam_code}_subtopic_manifest.json')
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f'Rebuilt: {out} ({len(manifest["subtopics"])} ids, '
          f'{len(manifest["mandatory_every_mock"])} mandatory-every-mock, '
          f'{len(manifest["alternation_groups"])} alternation groups, '
          f'{len(manifest["mandatory_groups"])} group-presence, '
          f'{len(manifest["cadence_windows"])} cadence, '
          f'{len(manifest["min_counts"])} min-count)')
    if id_less:
        print(f'WARN: {len(id_less)} subtopic block(s) had NO subtopic_id and were '
              f'SKIPPED (cannot be joined by Step 6/7): {id_less[:8]}'
              f'{" ..." if len(id_less) > 8 else ""}. Mint ids in Step 5 or add them, '
              f'then re-run rebuild.')
    write_taxonomy_xlsx(manifest, exam_code, out_dir)   # v2.24 — human-readable companion
    return out


def format_entry(e):
    # BUG-C04 fix (v2.3): option_format written as full dict per §14 BUG-B15 spec
    ofmt = e['option_format']
    if isinstance(ofmt, dict):
        ofmt_primary  = ofmt.get('primary', 'single_value')
        ofmt_recent   = ofmt.get('recent_format', ofmt_primary)
        ofmt_changed  = str(ofmt.get('changed_recently', False)).lower()
        ofmt_all      = ofmt.get('all_observed', [ofmt_primary])
    else:
        ofmt_primary = ofmt_recent = str(ofmt)
        ofmt_changed = 'false'
        ofmt_all     = [str(ofmt)]
    lines = [
        '', f'--- Subtopic: {e["subtopic"]} ---',
        # ── subtopic_id (v2.4 — CROSS-STEP JOIN KEY, the single source of truth) ──
        # subtopic_id is the STABLE machine identifier that Step 6 and Step 7 use to
        # join blueprint.json ↔ section_rules.md. It is minted HERE (Step 5) and
        # ONLY here. Downstream steps copy it verbatim and NEVER invent their own.
        # The display name ("--- Subtopic: X ---") is decorative and may be reworded
        # freely WITHOUT breaking the pipeline, because the id (not the name) is the
        # load-bearing join key. See §15 SUBTOPIC_ID CONTRACT for the full recipe.
        f'subtopic_id: {e["subtopic_id"]}',
        f'section: {e["section"]}', f'topic: {e["topic"]}',
        f'observed_count: {e["observed_count"]}', f'format: {e["format"]}',
        f'inherently_visual: {str(e.get("inherently_visual", False)).lower()}',
        # ── v2.26 (GAP-2026-08-06-AXIS1) ────────────────────────────────────────────
        # figural_rate is the SHARE OF THIS SUBTOPIC'S OBSERVED QUESTIONS that carried a
        # figure — read `4/39` not `10%` when you are debugging, because the denominator
        # is what the old boolean threw away. Step 7 ranks claims on the scarce Axis-1
        # budget by this number, so a subtopic the exam illustrates 79% of the time
        # (organic stereochemistry) outranks one it illustrates 3% of the time (complex
        # formation) — both of which the pre-v2.26 flag marked identically FIGURAL.
        f'figural_q_count: {int(e.get("figural_q_count", 0))}',
        f'figural_rate: {float(e.get("figural_rate", 0.0)):.4f}',
        # figural_reducible: false ⇒ the OPTIONS are images, so there is no text form of
        # this question. Granted a figure even over budget (bc.axis_grant_figural rule 2);
        # the audit raises its expectation by the same count so the overage is silent.
        f'figural_reducible: {str(e.get("figural_reducible", True)).lower()}',
        # v2.44 — share of this subtopic's FIGURAL questions whose figure is in the
        # OPTIONS. Read '2/15' not '13%' when debugging: the denominator is exactly what
        # the pre-v2.44 any() discarded, and discarding it made 21 subtopics exempt from
        # a budget of 5 on the strength of one question each.
        f'option_image_rate: {float(e.get("option_image_rate", 0.0)):.4f}',
        # v2.43 — DI rate, read by Step 7 to rank claims on the Axis-1 DI budget and by
        # A-AXIS1 to know whether a DI target is reachable at all.
        f'di_q_count: {int(e.get("di_q_count", 0))}',
        f'di_rate: {float(e.get("di_rate", 0.0)):.4f}',
        f'di_reducible: {str(e.get("di_reducible", True)).lower()}',
        # REPLACEMENT_RULE — what to render when a FIGURAL-CAPABLE question does NOT win
        # a budget slot. It is NOT a downgrade to a worse question: it draws from this
        # subtopic's OWN observed text patterns (PYQ_STEM_PATTERNS below), which is the
        # majority shape for every subtopic whose figural_rate is under 50%. Required by
        # HS-16 for any subtopic that can be denied a figure — which, post-v2.26, is every
        # reducible FIGURAL subtopic, so it is emitted unconditionally rather than left to
        # a hand-authored ban list (the reference exam had 46 FIGURAL subtopics and ZERO
        # REPLACEMENT_RULE blocks, so nothing was ever downgradable and the cap could not
        # have been honoured even if Step 7 had tried).
        ('REPLACEMENT_RULE: render as TEXT using this subtopic\'s observed '
         'PYQ_STEM_PATTERNS; preserve subtopic, difficulty band and answer_cardinality; '
         'never substitute another subtopic'
         if str(e.get('format', 'TEXT')).upper() == 'FIGURAL'
            and bool(e.get('figural_reducible', True))
         else 'REPLACEMENT_RULE: none (not reducible)'),
        f'option_format_primary: {ofmt_primary}',
        f'option_format_recent: {ofmt_recent}',
        f'option_format_changed_recently: {ofmt_changed}',
        f'option_format_all_observed: {ofmt_all}',
        f'OMML_required: {str(e["OMML_required"]).lower()}',
        f'negative_question_freq: {e["negative_question_freq"]}%',
        f'answer_type: {e.get("answer_type", "option")}',
        f'nat_freq: {e.get("nat_freq", 0)}%',
        f'answer_cardinality: {e.get("answer_cardinality", "single")}',
        f'msq_freq: {e.get("msq_freq", 0)}%',
        # v2.23 THREE-AXIS (CATEGORY B). observed_axis2 = PYQ-observed counts; axis2_capability
        # = faithful forms Step 6/7 may use; presentation_family = Step 7 menu key. Single-line
        # so rebuild_subtopic_manifest_from_section_rules() can round-trip them.
        f'observed_axis2: {e.get("observed_axis2", {})}',
        f'presentation_family: {e.get("presentation_family")}',
        f'axis2_capability: {e.get("axis2_capability", [])}',
        f'fill_in_blank: {e["fill_in_blank"]}',
        f'linked_group_size: {e["linked_group_size"]}',
        f'max_per_paper: {e.get("max_per_paper", 0)}',
        f'typical_per_paper: {e.get("typical_per_paper", 0)}',
        f'stem_word_count: min={e["stem_word_count"]["min"]} '
          f'max={e["stem_word_count"]["max"]} typical={e["stem_word_count"]["typical"]}',
        f'sub_type_label: {e["sub_type_label"]}',
        # ── CONCEPT_GROUP and QUESTION_MECHANIC (MANDATORY — Step 7 relies on these) ──
        # CONCEPT_GROUP: broad theme grouping (Step 7 resolve_concept_group priority #1).
        # QUESTION_MECHANIC: the normalised student task for CHECK D mechanic-collision.
        # Both MUST be written here. If absent → Step 7 falls to keyword-matching
        # fallback which FAILS for non-standard subtopic names → mechanic collisions
        # like synonym×2 or antonym×2 in the same mock go undetected.
        #
        # HOW TO DERIVE concept_group (exam-agnostic):
        #   Read the PYQ_STEM_PATTERNS template for this subtopic.
        #   Group all subtopics that share the same broad topic theme.
        #   Example: all "find missing number" variants → concept_group: missing_number
        #   Example: all "series" variants              → concept_group: series
        #   Use the SAME string for every subtopic in the same theme group.
        #
        # HOW TO DERIVE question_mechanic (exam-agnostic):
        #   Normalise the stem instruction to its atomic task:
        #     "opposite in meaning" / "ANTONYM"       → antonym
        #     "similar in meaning"  / "SYNONYM"       → synonym
        #     "find the error"                        → error_detection
        #     "improve the sentence" / "best replaces"→ sentence_improvement
        #     "one word substitution"                 → one_word_substitution
        #     "fill in the blank"                     → fill_in_blank
        #     "meaning of idiom/phrase"               → idiom
        #     "select the correctly spelt"            → spelling
        #     "active/passive voice"                  → voice
        #     "direct/indirect speech"                → narration
        #     (derive from PYQ_STEM_PATTERNS template — never hardcode exam names)
        #   Two subtopics with the SAME question_mechanic MUST NOT appear in
        #   the same mock (MockCreate mandate-7 CHECK D).
        #
        # SYNONYM vs ANTONYM: these are DIFFERENT mechanics. Both CAN appear in
        # one mock (one synonym Q + one antonym Q = different mechanics = allowed).
        # But two synonym Qs = SAME mechanic = BLOCKED by CHECK D. Always separate.
        f'concept_group: {e.get("concept_group") or _derive_concept_group(e)}',
        f'question_mechanic: {e.get("question_mechanic") or _derive_question_mechanic(e)}',
        f'form_key: {e.get("form_key") or _derive_form_key(e)}',                     # v2.24 (BV-10a HARD identity)
        f'collision_domain: {e.get("collision_domain") or _derive_collision_domain(e)}',  # v2.24 (default=section)
        # ── v2.10 MANDATE ROUND-TRIP: emit the three mandate fields as parseable lines
        #    so the manifest is reproducible from THIS FILE (not only from in-memory
        #    entries). mandate_every_mock defaults to the robust NOTE detector; an explicit
        #    e['mandate_every_mock'] overrides. alternation_group + min_per_series_window
        #    are authored fields ('none' when unset). write_subtopic_manifest AND
        #    rebuild_subtopic_manifest_from_section_rules read these back; Step 6 consumes
        #    the manifest (RULE M1/M2); Step 7 reads it (S3-17). These lines are also the
        #    HAND-EDIT point: set 'mandate_every_mock: true' / 'alternation_group: <name>'
        #    on a block and regenerate the manifest to change policy — no re-analysis.
        f'mandate_every_mock: {str(e.get("mandate_every_mock", _mandate_from_note(e.get("note_text") or e.get("NOTE") or ""))).lower()}',
        f'alternation_group: {e.get("alternation_group") or "none"}',
        f'min_per_series_window: {e.get("min_per_series_window") if e.get("min_per_series_window") not in (None, "") else "none"}',
        # v2.11 (Issue 2b) — two more round-trippable mandate fields:
        #   mandatory_group : a group name; a mock must contain >=1 member of the group
        #                     (group-presence, e.g. any one member of a 3D-solids group).
        #   min_count       : minimum questions of THIS subtopic per mock (>=k).
        f'mandatory_group: {e.get("mandatory_group") or "none"}',
        f'min_count: {e.get("min_count") if e.get("min_count") not in (None, "") else "none"}'
        # _derive_concept_group() and _derive_question_mechanic() are defined below.
        # They extract from PYQ_STEM_PATTERNS when not explicitly set in input data.
    ]
    for p in e.get('PYQ_STEM_PATTERNS', []):
        # BUG-C02 fix (v2.3): raw_count and years written — QV-11 uses years
        dep_tag = ' [DEPRECATED]' if p.get('deprecated') else ''
        lines += [
            f'  {p["id"]}: "{p["template"]}"',
            f'      approach: "{p.get("approach","")}"',
            f'      frequency: {p["frequency"]}%  confidence: {p["confidence"]}{dep_tag}',
            f'      raw_count: {p.get("raw_count", 0)}',
            f'      years: {p.get("years", [])}',
            # v2.56 additive (§6.1.8): the measured pattern abstraction. S7-STYLE
            # composes from THESE, never from the template text (RC-3).
            f'      pattern_key: {p.get("pattern_key", "")}',
            (f'      mechanic: {p.get("mechanic", "unknown")}  '
             f'polarity: {p.get("polarity", "positive")}  '
             f'option_shape: {p.get("option_shape", "none")}'),
            f'      note_block: {p.get("note_block","never")}',
        ]
        if p.get('note_text'): lines.append(f'      note_text: "{p["note_text"][:120]}"')
        # ── SYNC fields: read by Step 7 sub-step 4b (UL-GATE) and sub-step 4c (ANCHOR-LOCK) ──
        # underline_required: true/false — Step 7 sub-step 4b reads this to determine
        #   whether stem_underline_ranges must be computed. Set true for subtopics
        #   where the stem instruction says "underlined" (idiom, sentence_improvement,
        #   error_detection, one_word_substitution, etc.).
        #   HOW TO DETECT from PYQ: if p['template'] contains 'underlined' or 'underline'
        #   → set underline_required: true. Otherwise false.
        _ul_req = 'true' if any(kw in p.get('template','').lower()
                                for kw in ['underlined','underline']) else 'false'
        lines.append(f'      underline_required: {_ul_req}')
        # anchor_option_required: true/false — Step 7 sub-step 4c reads this to determine
        #   whether an anchor option (e.g. "No error", "No improvement required") must
        #   be pinned to a fixed position during shuffle.
        #   HOW TO DETECT from PYQ: if p['template'] contains a phrase like
        #   "select option (N) if there is no error" or "if no improvement" + option number
        #   → set anchor_option_required: true and record anchor_position.
        import re as _re
        _anch_m = _re.search(
            r'select\s+option\s*\(?(\d+|[A-Da-d])\)?.*?(?:no error|no improvement|no change)',
            p.get('template',''), _re.IGNORECASE)
        _anch_req = 'true' if _anch_m else 'false'
        lines.append(f'      anchor_option_required: {_anch_req}')
        if _anch_m:
            # anchor_position: the 1-indexed position stated in the stem
            _pos_str = _anch_m.group(1)
            _pos_int = int(_pos_str) if _pos_str.isdigit() else (ord(_pos_str.upper())-64)
            lines.append(f'      anchor_position: {_pos_int}')
        lines.append('')
    # PYQ_DIFFICULTY_CALIBRATION block RETIRED (GAP-2026-08-27-DIFFICULTY-PROFILE) — never written
    # ── v2.56 additive STYLE block (GAP-2026-08-29-STYLE-FIDELITY §6.1.8) ──────
    if e.get('content_signature') is not None:
        lines += ['STYLE (v2.56):',
                  f'  legacy_mode: {e.get("legacy_mode")}',
                  f'  content_signature: {json.dumps(e.get("content_signature", {}), sort_keys=True)}',
                  f'  mechanic_mix: {json.dumps(e.get("mechanic_mix", {}), sort_keys=True)}',
                  f'  mechanic_unknown_count: {int(e.get("mechanic_unknown_count", 0))}',
                  f'  low_entropy: {str(bool(e.get("low_entropy", False))).lower()}',
                  f'  distractor_mix: {json.dumps(e.get("distractor_mix"), sort_keys=True) if isinstance(e.get("distractor_mix"), dict) else e.get("distractor_mix")}',
                  f'  stimulus_stats: {json.dumps(e.get("stimulus_stats"), sort_keys=True)}',
                  f'  PYQ_EXCLUDE_VALUES: {e.get("PYQ_EXCLUDE_VALUES", [])}',
                  'PATTERN_KEYS (v2.56):']
        for pk in e.get('pattern_keys', []):
            lines.append('  ' + json.dumps(pk, sort_keys=True))
        lines.append('')
    wo = e.get('wrong_option_structure', {})
    lines += ['wrong_option_structure:',
              f'  type: {wo.get("type","varied")}',
              f'  description: "{wo.get("description","")}"']
    if wo.get('fixed_option_texts'): lines.append(f'  fixed_option_texts: {wo["fixed_option_texts"]}')
    if wo.get('shared_pool_words'):  lines.append(f'  shared_pool_words: {wo["shared_pool_words"]}')
    lines.append('')
    if e.get('PYQ_NUMBER_RANGES'):
        lines.append('PYQ_NUMBER_RANGES:')
        for var,rng in e['PYQ_NUMBER_RANGES'].items():
            if 'p50' in rng:   # v2.56 per-unit percentile shape (§6.1.5)
                lines.append(f'  {var}: {{min:{rng["min"]},max:{rng["max"]},'
                             f'p10:{rng["p10"]},p50:{rng["p50"]},p90:{rng["p90"]},'
                             f'n:{rng["n"]}}}')
            else:
                lines.append(f'  {var}: {{min:{rng["min"]},max:{rng["max"]},'
                             f'multiples_of:{rng.get("multiples_of","N/A")}}}')
        lines.append('')
    if e.get('PYQ_CONTEXT_POOL'):
        cp = e['PYQ_CONTEXT_POOL']
        cp_lines = ['PYQ_CONTEXT_POOL:',
                    f'  dominant: {cp.get("dominant",[])}',
                    f'  common: {cp.get("common",[])}',
                    f'  avoid: {cp.get("avoid",[])}',
                   ]
        # NEW v2.3: write recycled_datasets and ban_recycled if present
        if cp.get('recycled_datasets'):
            cp_lines.append(f'  recycled_datasets: {cp["recycled_datasets"]}')
            cp_lines.append(f'  ban_recycled: {str(cp.get("ban_recycled", True)).lower()}')
        cp_lines.append('')
        lines += cp_lines
    if e.get('PYQ_IMAGE_ANALYSIS'):
        ia = e['PYQ_IMAGE_ANALYSIS']
        # v2.37 (GAP-2026-07-26-003 D3/D4). This emitter used to write FOUR fields.
        # aggregate_figural built NINE, so arrangement_types, complexity_dist,
        # object_types.avoid, images_analysed and images_unclear were computed and then
        # DROPPED at the artefact boundary — Step 7 could never have seen complexity or
        # arrangement even with vision working perfectly.
        #
        # vision_status is the D3 fix. In the reference run 'vision_unavailable'
        # appeared 153 times in progress.json and ZERO times in section_rules.md, so a
        # consumer reading `dominant: []` had no way to tell "this exam has no figures"
        # from "vision failed and this profile is empty". The artefact now carries its
        # own provenance.
        _ot = ia.get('object_types', {})
        lines += ['PYQ_IMAGE_ANALYSIS:',
                  f'  image_role: {ia.get("image_role","stem_only")}',
                  f'  vision_status: {ia.get("vision_status","not_applicable")}',
                  '  object_types:',
                  f'    dominant: {_ot.get("dominant",[])}',
                  f'    observed: {_ot.get("observed",[])}',
                  f'    avoid: {_ot.get("avoid",[])}',
                  f'  transformation_types: {ia.get("transformation_types",[])}',
                  f'  arrangement_types: {ia.get("arrangement_types",[])}',
                  f'  complexity_dist: {ia.get("complexity_dist",{})}',
                  f'  images_analysed: {ia.get("images_analysed",0)}',
                  f'  images_unclear: {ia.get("images_unclear",0)}',
                  f'  images_unobserved: {ia.get("images_unobserved",0)}']
        if ia.get('dominant_suppressed'):
            lines += [f'  dominant_suppressed: "{ia["dominant_suppressed"]}"']
        lines += ['']
    # BUG-C05 fix (v2.3): paragraph_count and topic_domains now written
    if e.get('PYQ_PASSAGE_STRUCTURE'):
        ps = e['PYQ_PASSAGE_STRUCTURE']
        lines += ['PYQ_PASSAGE_STRUCTURE:',
                  f'  sub_format: {ps.get("sub_format","RC")}',
                  f'  word_range: {{min:{ps.get("word_range",{}).get("min",0)},'
                    f'max:{ps.get("word_range",{}).get("max",0)}}}',
                  f'  paragraph_count: {{typical:{ps.get("paragraph_count",{}).get("typical",1)}}}',
                  f'  topic_domains_observed: {ps.get("topic_domains",{}).get("observed",[])}',
                  f'  topic_domains_avoid: {ps.get("topic_domains",{}).get("avoid",[])}',
                  f'  q_type_distribution: {ps.get("q_type_distribution",{})}', '']
    return lines

# ═══ FROM Framework_MockTestAnalyse.md §6, fence L5673-6069 (v2.53.2) — VERBATIM ═══
def run_qv(entries, taxonomy, progress):
    """
    BUG-A09 fix: QV-11 now uses global_max_year across all entries.
    BUG-B09 fix: QV-3 uses is_inferred bool from calibration dict.
    """
    results = {}
    ekeys   = {(e['section'],e['topic'],e['subtopic']):e for e in entries}

    # QV-1: Every taxonomy subtopic has an entry
    missing = [st['subtopic'] for sec,sts in taxonomy.items()
               for st in sts if (sec,st['topic'],st['subtopic']) not in ekeys]
    results['QV-1'] = ('WARN' if missing else 'PASS',
                        f'{len(missing)} missing: {missing[:5]}' if missing else 'All covered')

    # QV-1a: PYQ subtopics are in taxonomy
    pyq_subs = set(k[2] for k in ekeys)
    tax_subs = set(st['subtopic'] for sts in taxonomy.values() for st in sts)
    extra    = pyq_subs - tax_subs
    # GAP-2026-08-05-001 (SG-6). Severity RAISED from WARN to FAIL. The IDENTICAL
    # condition — a parsed subtopic that is not in the locked taxonomy — is a HARD STOP
    # at Step 4 Task 2.5 and was only a WARN here. Per Framework_DeliveryFooter §5 Q0 a
    # WARN does not force amber, so a phantom subtopic passed SILENTLY, rendered a green
    # "Step Complete" footer, and flowed into subtopic_manifest.json -> Step 6 allocation
    # -> Step 7 generation, carrying a truncated stem and dropped options with it.
    # FAIL renders F1 amber and names the check. It MUST NOT halt Step 5: per CLAUDE.md
    # a CLASS T failure must be LOUD and must NOT halt. Loud, not fatal.
    results['QV-1a'] = ('FAIL' if extra else 'PASS',
                          f'In PYQ not taxonomy: {list(extra)[:5]}' if extra else 'OK')

    # QV-15 — BODY TERMINATION SANITY (GAP-2026-08-05-001, SG-7)
    # QV-1a only fires when the phantom's TEXT is absent from the taxonomy. If a misread
    # continuation happens to canon-match a real subtopic name, QV-1a is silent, Step 4's
    # totals still reconcile, and the questions after it are attributed to the WRONG
    # subtopic with no signal anywhere. QV-15 tests the STRUCTURE instead of the name, so
    # it catches that branch too.
    # NAT IS TESTED DIFFERENTLY AND DELIBERATELY. An option-count threshold is meaningless
    # for a question that has no options, and NAT is the shape MOST exposed to a heading
    # misread (no options means its last stem paragraph sits in the same slot as a genuine
    # heading). So for NAT the assertion is on COLOUR: in a file whose date-label probe
    # passed, a body terminated by a non-navy inferred heading is a misread by definition.
    _tbh   = [e for e in entries if e.get('terminated_by_heading')]
    # GAP-2026-08-16-BASELINE-SUPPRESSED-NAMEERRORS (D6). `options_count` is a SESSION
    # parameter detected at S1-3; it is bound nowhere at module scope, so run_qv read a
    # free name. The comprehension only EVALUATES it when _tbh is non-empty — i.e. when
    # at least one question was terminated by an inferred heading — so QV-15 raised
    # NameError EXACTLY when it had something to report, and passed silently when it
    # did not. A check that crashes only on the defect it exists to detect is worse
    # than absent: it reads as a clean PASS on every corpus that does not trigger it.
    # IIT_JAM_MATHEMATICS has 0 such questions, which is the only reason the v2.53.0
    # certification run survived this line.
    # progress._meta.options_count is the parameter's persisted home (written at S1-3,
    # present in every schema_version >= 2.0 corpus). Default 4 matches the documented
    # safe default at §1 ("mixed or ambiguous -> options_count = 4").
    # Coerced, not merely defaulted. `<` between int and str raises TypeError, so a
    # hand-edited exam_config or a pre-2.0 corpus carrying "4" would reproduce the exact
    # crash this fix removes. Every non-numeric or absent value falls back to the
    # documented safe default of 4 (§1: "mixed or ambiguous -> options_count = 4").
    try:
        _options_count = int((progress.get('_meta') or {}).get('options_count') or 4)
    except (TypeError, ValueError):
        _options_count = 4
    if _options_count < 1:
        _options_count = 4
    _short = [e for e in _tbh
              if e.get('has_options', True)
              and len(e.get('options') or []) < _options_count]
    _nat_bad = [e for e in _tbh
                if not e.get('has_options', True)
                and e.get('colour_available')
                and e.get('terminating_heading_colour') != bc.HEADING_NAVY]
    _bad = _short + _nat_bad
    results['QV-15'] = ('FAIL' if _bad else 'PASS',
                        f'{len(_bad)} question(s) terminated by an inferred heading '
                        f'before their options were collected: '
                        f'{[e.get("num") for e in _bad][:5]}' if _bad else
                        f'OK ({len(_tbh)} bodies ended on a heading, all well-formed)')
    # Report the raw counter in the batch summary even when it is 0 — an unexplained
    # jump between papers of the same exam is the earliest visible symptom of a Step-1
    # rendering change that newly exposes this class.
    results['_counter_questions_terminated_by_heading'] = len(_tbh)

    # ── QV-16 — POSITION RESOLUTION INTEGRITY ────────────────────────────────
    # GAP-2026-08-16-PYQEXTRACT-DATE-LABEL-POSITION. The defect was invisible for its
    # whole life because NO check compared the resolved positions against the paper
    # that produced them. Every QV passed, nothing raised, nothing counted it, and the
    # one artefact a reviewer inspects by eye — the template — had already been
    # scrubbed of date labels by E-10 strip_variables, hiding the leak.
    #
    # THE INVARIANT: a v1.18+ sorted paper stamps EXACTLY ONE original position per
    # question, so within one paper the resolved original_q_num values must be
    # DISTINCT. Under the defect every question in a taxonomy block collapses onto the
    # block's first position, and the collision count is the number of questions that
    # inherited a stale label. This is a property of the OUTPUT, not of the parser, so
    # it cannot share the parser's blind spot — which is the whole point (the same
    # lesson as GAP-2026-08-15-BAREQ: input and output counted with one blind detector).
    #
    # VACUOUS PASS on a pre-v1.18 corpus, where parse_original_q_num returns None for
    # every unstamped label and there is nothing to collide. That is correct, not a
    # gap: such a corpus never had positional data to corrupt.
    #
    # FAIL, NOT HALT — the CLASS T convention. The papers are still extracted and the
    # progress file is still written; the operator is told loudly that every positional
    # consumer downstream (is_msq -> answer_cardinality -> Step 7's answer mechanism)
    # is reading stale values, and can re-run from a clean progress file.
    from collections import Counter as _Counter16   # local: each fence binds its own
    _pos_by_paper = {}
    for _k, _v in progress.items():
        if not isinstance(_k, tuple):
            continue
        for _q in _v:
            if _q.get('original_q_num') is not None:
                _pos_by_paper.setdefault(_q.get('paper_id'), []).append(
                    _q['original_q_num'])
    _dupes = []
    for _pid, _nums in _pos_by_paper.items():
        if len(set(_nums)) != len(_nums):
            _rep = [n for n, c in _Counter16(_nums).items() if c > 1]
            _dupes.append(f'{_pid}: {len(_nums) - len(set(_nums))} collision(s), '
                          f'e.g. {_rep[:3]}')
    results['QV-16'] = (('FAIL', f'position collisions — every question in a taxonomy '
                                 f'block inherited the block\'s first exam position: '
                                 f'{_dupes[:3]}')
                        if _dupes else
                        ('PASS', f'{len(_pos_by_paper)} paper(s): one distinct '
                                 f'original_q_num per question'
                                 if _pos_by_paper else
                                 'vacuous — pre-v1.18 corpus, no stamped positions'))

    # QV-2: Frequency% sums to 100 per subtopic
    bad = [e['subtopic'] for e in entries if e.get('PYQ_STEM_PATTERNS') and
           abs(sum(p['frequency'] for p in e['PYQ_STEM_PATTERNS'])-100) > 1]
    results['QV-2'] = ('FAIL' if bad else 'PASS',
                        f'Sums!=100: {bad[:3]}' if bad else 'All=100%')

    # QV-3 RETIRED (GAP-2026-08-27-DIFFICULTY-PROFILE): difficulty calibration is no longer a
    # Step-5 product; the key is kept so QV readers see a stable list.
    results['QV-3'] = ('PASS', 'retired — difficulty calibration now lives in the difficulty profile')

    # QV-4/5: option_format and wrong_option_structure classified
    no_fmt = [e['subtopic'] for e in entries if not e.get('option_format')]
    results['QV-4'] = ('FAIL' if no_fmt else 'PASS',
                        f'Missing: {no_fmt[:3]}' if no_fmt else 'OK')
    no_wo  = [e['subtopic'] for e in entries if not e.get('wrong_option_structure')]
    results['QV-5'] = ('FAIL' if no_wo else 'PASS',
                        f'Missing: {no_wo[:3]}' if no_wo else 'OK')

    # QV-5b: BUG-C07 fix (v2.3) — fixed_set must have non-empty fixed_option_texts
    bad_fixed = [e['subtopic'] for e in entries
                 if e.get('wrong_option_structure', {}).get('type') == 'fixed_set'
                 and not e.get('wrong_option_structure', {}).get('fixed_option_texts')]
    results['QV-5b'] = ('FAIL' if bad_fixed else 'PASS',
                         f'fixed_set missing texts: {bad_fixed[:3]}' if bad_fixed else 'OK')

    # QV-6: confidence accuracy
    bad_c = [f'{e["subtopic"]}.{p["id"]}' for e in entries
             for p in e.get('PYQ_STEM_PATTERNS',[])
             if p.get('confidence')=='observed' and p.get('raw_count',0)<3]
    results['QV-6'] = ('WARN' if bad_c else 'PASS',
                        f'Observed<3: {bad_c[:3]}' if bad_c else 'OK')

    # QV-7: templates have _VAR_ placeholders
    VAR_TOKENS = {'_NUM_','_P_','_R_%','_T_','_WORD_','_CODE_','_NOUN_','_LCLUS_',
                  '_YEAR_','_ORG_','_BLANK_','_NAME_','_SERIES_','_STMT_','_ADJ_'}
    no_var = [f'{e["subtopic"]}.{p["id"]}' for e in entries
              for p in e.get('PYQ_STEM_PATTERNS',[])
              if p['template'] not in ('(no PYQ observed)','(figural -- no text stem)')
              and not any(v in p['template'] for v in VAR_TOKENS)]
    results['QV-7'] = ('WARN' if no_var else 'PASS',
                        f'No _VAR_: {no_var[:3]}' if no_var else 'OK')

    # QV-8: OMML recovery >= 80%
    omml_iss = []
    for e in entries:
        if not e.get('OMML_required'): continue
        qs = progress.get((e['section'],e['topic'],e['subtopic']),[])
        om = [q for q in qs if q.get('omml_present')]
        ok = [q for q in om if not q.get('omml_failed')]
        if om and len(ok)/len(om) < 0.80:
            omml_iss.append(f'{e["subtopic"]}: {len(ok)}/{len(om)}')
    results['QV-8'] = ('WARN' if omml_iss else 'PASS',
                        f'OMML<80%: {omml_iss[:3]}' if omml_iss else 'OK')

    # QV-9: image clarity >= 80%
    img_iss = []
    for e in entries:
        if e.get('format')!='FIGURAL': continue
        ia  = e.get('PYQ_IMAGE_ANALYSIS') or {}
        tot = ia.get('images_analysed',0) + ia.get('images_unclear',0)
        if tot>0 and ia.get('images_unclear',0)/tot>0.20:
            img_iss.append(f'{e["subtopic"]}: {ia["images_unclear"]}/{tot}')
    results['QV-9'] = ('WARN' if img_iss else 'PASS',
                        f'Image<80%: {img_iss[:3]}' if img_iss else 'OK')

    # QV-10: PASSAGE subtopics have linked_group_size > 0
    no_grp = [e['subtopic'] for e in entries
               if e.get('format')=='PASSAGE' and e.get('linked_group_size',0)==0]
    results['QV-10'] = ('WARN' if no_grp else 'PASS',
                          f'Passage no groups: {no_grp}' if no_grp else 'OK')

    # QV-11: BUG-A09 fix -- global_max_year across all entries, not per-entry max
    all_years_global = [y for e in entries
                        for p in e.get('PYQ_STEM_PATTERNS',[])
                        for y in p.get('years',[]) if isinstance(y,int)]
    global_max = max(all_years_global) if all_years_global else None
    old_only = []
    if global_max:
        for e in entries:
            entry_yrs = [y for p in e.get('PYQ_STEM_PATTERNS',[])
                         for y in p.get('years',[]) if isinstance(y,int)]
            if entry_yrs and max(entry_yrs) < global_max - 1:
                old_only.append(e['subtopic'])
    results['QV-11'] = ('WARN' if old_only else 'PASS',
                          f'No recent data: {old_only[:3]}' if old_only else 'OK')

    # QV-12: no near-duplicate templates within a subtopic
    dups = []
    for e in entries:
        pats = [p['template'] for p in e.get('PYQ_STEM_PATTERNS',[])]
        for i in range(len(pats)):
            for j in range(i+1,len(pats)):
                if SequenceMatcher(None,pats[i],pats[j]).ratio()>0.90:
                    dups.append(f'{e["subtopic"]}: P{i+1}~P{j+1}')
    results['QV-12'] = ('WARN' if dups else 'PASS',
                          f'Near-dups: {dups[:3]}' if dups else 'OK')

    # QV-13 (v2.24) — MECHANIC INTEGRITY & COLLISION GUARD. Stops Step 5 from ever
    # QV-13 (v2.24.1) — MECHANIC IDENTITY INTEGRITY. FAIL severity, NO allowlist.
    # Reads the STAMPED fields (D5); never recomputes for the collision test. Step 5
    # cannot know N_mocks/batch_size, so it must reject ALL shared form_keys (D4/D7).
    from collections import defaultdict as _dd
    empty     = [e['subtopic_id'] for e in entries if not e.get('form_key')]
    unstamped = [e['subtopic_id'] for e in entries if 'form_key' not in e or e['form_key'] is None]
    # a REAL determinism check (D5): does a fresh derivation reproduce the stamped value?
    _pov = (progress.get('_meta', {}) or {}).get('section_prefix_overrides', {})
    _po  = build_section_prefix_map(sorted({e['section'] for e in entries}), _pov)
    nondet = []
    for e in entries:
        if not e.get('subtopic_id') or e.get('form_key') is None:
            continue
        _tpl = ' '.join(p.get('template','') for p in e.get('PYQ_STEM_PATTERNS', []))
        _mech = derive_mechanic(e['section'], e.get('subtopic') or e.get('sub_type_label',''),
                                e.get('sub_type_label'), _tpl, e.get('format','TEXT'),
                                e['subtopic_id'], _po)
        _fk  = _mech['form_key']
        # a curator override legitimately makes the stamped value differ from a bare
        # derivation; only flag when NO override explains it.
        _ov  = (_OVERRIDES or {}).get('subtopic_overrides', {}).get(e['subtopic_id'], {})
        if _fk != e['form_key'] and 'form_key' not in _ov and e['form_key'] != slugify(e['subtopic_id']):
            nondet.append(e['subtopic_id'])
        # ── B7: COLLISION_DOMAIN IS THE OTHER HALF OF THE IDENTITY ───────────
        # This gate is named MECHANIC IDENTITY INTEGRITY, and identity is the PAIR
        # (collision_domain, form_key) — the uniqueness test three lines below buckets
        # by exactly that pair. Until now the gate re-derived form_key and threw the
        # freshly-derived collision_domain away, so a drifted domain passed unnoticed
        # while the check that was supposed to catch drift reported PASS.
        #
        # That is not hypothetical. build_section_prefix_map assigns colliding prefixes
        # (ra, ra2, ra3) by iteration order, which is the exact defect B6 fixed in
        # mint_subtopic_ids and stamp_mechanic_axes. Those two were pinned; the GATE
        # that exists to notice when they drift was not. It also made `_po` above dead
        # computation, which is how the mutation run found this: the mutant that
        # hash-orders `_po` survived, because nothing downstream read the result.
        #
        # Same override escape hatch as form_key: a curator who sets collision_domain
        # deliberately is not drift.
        if (_mech['collision_domain'] != e.get('collision_domain')
                and 'collision_domain' not in _ov
                and e.get('collision_domain') is not None):
            nondet.append(e['subtopic_id'])
    dom = _dd(lambda: _dd(list))
    for e in entries:
        dom[e.get('collision_domain')][e.get('form_key')].append(e['subtopic_id'])
    collisions = [f'{d}:{fk}={sorted(ids)}'
                  for d, fks in dom.items() for fk, ids in fks.items() if len(ids) > 1]
    # a bare family token is illegal unless this exam declared the owning template set
    def _decl(e):
        _ov = _OVERRIDES or {}
        a = _ov.get('template_sets_by_section', {}).get(e['section'], _ov.get('template_sets'))
        return ['verbal','reasoning'] if a is None else a
    # A family-named form_key is illegal ONLY when it is a FOREIGN token, i.e. it did NOT
    # come from the subtopic's own identity base. Under v2.24.1 form_key == base for a
    # subtopic legitimately named like a family (e.g. a real verbal "Analogy"); that is not
    # contamination and must NOT FAIL. A foreign token can only arrive via a curator override.
    illegal = [e['subtopic_id'] for e in entries
               if e.get('form_key') in _ALL_FAMILY_NAMES
               and _TEMPLATE_SET.get(e.get('form_key')) not in _decl(e)
               and e['form_key'] != _identity_base(e.get('subtopic') or e.get('sub_type_label',''),
                                                    e.get('subtopic_id'))]
    same_as_domain = [e['subtopic_id'] for e in entries
                      if e.get('form_key') and e['form_key'] == e.get('collision_domain')]
    import re as _re
    bad_shape = [e['subtopic'] for e in entries
                 if len(e['subtopic']) > 60 or e['subtopic'].strip().endswith('?')
                 or _re.match(r'^\s*(what|which|how|why|find|choose|select|the average)\b',
                              e['subtopic'].strip(), _re.I)]
    if empty or unstamped or nondet or collisions or illegal or same_as_domain:
        results['QV-13'] = ('FAIL',
            f'empty={empty[:3]} unstamped={unstamped[:3]} nondeterministic={nondet[:3]} '
            f'collisions={collisions[:3]} illegal_family_token={illegal[:3]} '
            f'equals_domain={same_as_domain[:3]}')
    else:
        results['QV-13'] = ('PASS', f'{len(entries)} form_keys: non-empty, deterministic, '
                                    f'unique per collision_domain, no foreign family tokens')
    # QV-13a — NAME SHAPE. Advisory only; split so an editorial nit can neither mask nor
    # be masked by an identity failure.
    results['QV-13a'] = (('WARN', f'{len(bad_shape)} long/question-shaped names: {bad_shape[:3]}')
                         if bad_shape else ('PASS', 'OK'))

    # ── QV-14 — VISION COVERAGE (v2.37, GAP-2026-07-26-003) ─────────────────
    # FAIL, not WARN. The reference run scored QV-9 PASS with 153/153 figural
    # questions unobserved and 45/45 FIGURAL subtopics shipping an empty profile,
    # because QV-9 computes images_analysed + images_unclear and BOTH are zero when
    # nothing was observed — so its `if tot > 0` branch never fires. A WARN would not
    # have stopped that run either. This is the check that makes the failure impossible
    # to ship silently.
    #
    # QV-14 DOES NOT HALT. It reports. Framework_DeliveryFooter renders F1 amber on any
    # FAIL, and the operator re-runs PHASE B ONLY.
    #
    # DENOMINATOR. Only questions that actually carry a figure. EC-V13 (zero-PYQ
    # inferred FIGURAL subtopic) and EC-V14 (INHERENTLY-VISUAL keyword override) are
    # legitimately FIGURAL with no embedded figure; they contribute no queue item by
    # construction and are excluded here for the same reason. EC-V1: an exam with no
    # figural content at all has an empty denominator and PASSes vacuously — a
    # text-only exam must never fail this check.
    fig_qs = [q for k, v in progress.items() if isinstance(k, tuple) for q in v
              if q.get('image_role', 'none') != 'none']
    if not fig_qs:
        results['QV-14'] = ('PASS', 'no figural questions in this corpus — '
                                    'vision not applicable (EC-V1)')
    else:
        unobs = sum(1 for q in fig_qs if q.get('image_clarity') == 'vision_unavailable')
        noty = sum(1 for q in fig_qs if not q.get('object_type')
                   and q.get('image_clarity') == 'clear')
        seen = len(fig_qs) - unobs
        if unobs == len(fig_qs):
            results['QV-14'] = ('FAIL',
                f'0/{len(fig_qs)} figures observed — PHASE B did not run. '
                f'PYQ_IMAGE_ANALYSIS is empty for every FIGURAL subtopic and Step 7 '
                f'has no object-type, transformation, arrangement or complexity '
                f'guidance. The run is COMPLETE and resumable: re-run PHASE B ONLY '
                f'(S4-2b) against the queue already on disk, then re-run synthesis.')
        elif unobs / len(fig_qs) > 0.50:
            results['QV-14'] = ('FAIL',
                f'only {seen}/{len(fig_qs)} figures observed ({unobs} unobserved) — '
                f'more than half the figural corpus is missing. Re-run PHASE B; it is '
                f'idempotent and fills only the gaps.')
        elif unobs:
            results['QV-14'] = ('WARN',
                f'{seen}/{len(fig_qs)} figures observed, {unobs} unobserved — '
                f're-running PHASE B is idempotent and fills only the gaps.')
        elif noty:
            results['QV-14'] = ('WARN',
                f'{seen}/{len(fig_qs)} observed but {noty} carry no object_type — '
                f'check the Phase B protocol was followed for every labelled cell.')
        else:
            results['QV-14'] = ('PASS', f'{seen}/{len(fig_qs)} figures observed')

    return results

def print_qv(results):
    """Render run_qv()'s result mapping.

    GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE (D1). run_qv returns ONE dict
    carrying TWO kinds of entry, and this function used to unpack every value as
    a (status, detail) pair:

        for check,(status,detail) in results.items():

    v2.41.0 added `results['_counter_questions_terminated_by_heading'] = len(_tbh)`
    — an int, written unconditionally, "reported even when 0". From that release
    this loop raised `TypeError: cannot unpack non-iterable int object` on EVERY
    exam, and because run_synthesise calls print_qv BEFORE write_section_rules,
    Step 5 emitted NOTHING for eleven days and eleven minor versions.

    The counter was written once and read nowhere. The producer changed the dict's
    contract and the consumer was never told — the same producer/consumer blind
    spot as GAP-2026-08-15-BAREQ and GAP-2026-08-16-PYQEXTRACT-DATE-LABEL-POSITION.

    THE CONTRACT IS NOW EXPLICIT AND ENFORCED, not merely tolerated:
      * `QV-*` keys are CHECKS   -> value MUST be a (status, detail) 2-tuple.
      * `_`-prefixed keys are COUNTERS -> scalar, reported on their own line, and
        never mistaken for a check. This preserves v2.41.0's stated intent, which
        was to make the counter VISIBLE, not to make it a check.
      * Anything else raises immediately, naming the offending key, so a future
        malformed check fails loudly at its source instead of presenting as an
        unpackable int three frames from whatever produced it.
    """
    print('\n=== Quality Verification Results ===')
    icons = {'PASS':'v','WARN':'!','FAIL':'X'}
    all_ok = True
    counters = {}
    for check, value in results.items():
        if check.startswith('_'):
            counters[check] = value
            continue
        if not (isinstance(value, tuple) and len(value) == 2):
            raise TypeError(
                f"print_qv: check {check!r} must map to a (status, detail) "
                f"2-tuple, got {type(value).__name__} {value!r}. A counter or "
                f"other bookkeeping value must use a '_'-prefixed key so it is "
                f"reported rather than unpacked (D1).")
        status, detail = value
        if status not in icons:
            raise ValueError(
                f"print_qv: check {check!r} has status {status!r}; expected one "
                f"of {sorted(icons)}.")
        print(f'  {icons[status]} {check}: {status} -- {detail}')
        if status=='FAIL': all_ok = False
    for name, value in counters.items():
        print(f'  . {name.lstrip("_")}: {value}')
    print(f'\n{"All checks passed." if all_ok else "FAIL checks must be resolved."}')
    return all_ok

# ────────────────────────────────────────────────────────────── self-test
def self_test():
    """python3 analyse_engine.py --self-test

    Every assertion below pins a behaviour some caller in Framework_MockTestAnalyse.md
    depends on. A fixture that merely restates the implementation proves nothing, so
    each one names a real input shape and the answer the pipeline needs.
    """
    passed, fails = 0, []

    def check(name, cond):
        nonlocal passed
        if cond:
            passed += 1
        else:
            fails.append(name)

    # ── E-1 question-start detection ──────────────────────────────────────
    # Returns the question NUMBER, not a bool. Q_PATTERNS accepts ONLY `Q.N` and
    # `QN.` forms; a bare `1.` is deliberately NOT a question start, which is what
    # stops an ordinary numbered list inside a stem from splitting the paper.
    check('qstart_Q.1', detect_question_start('Q.1 What is x?') == 1)
    check('qstart_Q.10_two_digits', detect_question_start('Q.10') == 10)
    check('qstart_Q1.', detect_question_start('Q1.') == 1)
    for s_ in ('1. What is x?', '12.', '(1) x', 'Q 1 What is x?'):
        check(f'qstart_rejects_non_Q_form[{s_!r}]', detect_question_start(s_) is None)
    check('qstart_rejects_mid_sentence',
          detect_question_start('The value of Q.1 is') is None)
    check('qstart_empty_is_safe', detect_question_start('') is None)

    # ── E-2 option recognition ────────────────────────────────────────────
    for s_ in ('(A) first', 'A. first', '(a) first'):
        check(f'is_option[{s_!r}]', bool(is_option(s_)))
    for s_ in ('A group of students', ''):
        check(f'not_option[{s_!r}]', not is_option(s_))
    check('clean_option_text_strips_label', clean_option_text('(A) 42').strip() == '42')

    # ── E-8 option-format classification (12 types) ───────────────────────
    # Options arrive LABEL-FREE here; the (A)/(B) prefix is stripped upstream.
    SHAPES = {
        'single_value':     ['1', '2', '3', '4'],
        'coordinate_pair':  ['(1, 2)', '(3, 4)', '(5, 6)', '(7, 8)'],
        'image_only':       ['', '', '', ''],
        'word_form_number': ['one', 'two', 'three', 'four'],
        'roman_label':      ['i', 'ii', 'iii', 'iv'],
        'value_pair':       ['1-2', '3-4', '5-6', '7-8'],
        'letter_cluster':   ['ABC', 'BCD', 'CDA', 'DAB'],
        'segment_label':    ['A', 'B', 'C', 'D'],
    }
    for want, opts in SHAPES.items():
        check(f'opt_format[{want}]', classify_option_format(opts) == want)
    check('opt_format_empty_list_is_image_only',
          classify_option_format([]) == 'image_only')

    # all_observed MUST be SORTED, and the fixture must span MORE THAN ONE format or
    # the assertion is vacuous — a sorted 1-element list proves nothing, which is the
    # "finding no fixture can distinguish from silence" shape MUTATION_BUDGETS.json
    # exists to name. Three distinct formats across three years.
    # WHY THIS IS PINNED: unsorted set order made section_rules.md irreproducible
    # across processes. Two runs of identical code over an identical corpus produced
    # the same 433,260 bytes with a different sha256, differing at 34 lines — one per
    # subtopic (GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE D4). Set iteration order
    # over str depends on PYTHONHASHSEED, which python randomises PER PROCESS, so a
    # single process cannot see the defect and back-to-back runs look deterministic.
    # SIX formats, supplied in REVERSE-SORTED order. GAP-2026-08-17-B4-MUTATION-FLAKE:
    # this fixture used THREE elements, and `list(set(...))` over three strings comes
    # out already sorted roughly one run in six — so the sorted->list mutant SURVIVED
    # 8 of 40 fresh processes and the budget was unstatable (28 too low one run in
    # five; 29 would be an increase, which the budget file forbids). At six elements
    # accidental sorted order is ~1/720, and with PYTHONHASHSEED pinned by
    # audit_mutation it is not a coin flip at all. Belt and braces, deliberately:
    # the pin makes the SCORE reproducible, the width makes the CHECK honest under any
    # seed — so this fixture still fails if someone later removes the pin.
    # The irony this replaces is on the record: the comment right below documents
    # exactly this hazard for section_rules.md and then relied on a three-element set.
    _six = ['word_form_number', 'single_value', 'roman_label',
            'letter_cluster', 'image_only', 'coordinate_pair']
    qs = [{'options': SHAPES[f], 'year': 2020 + i} for i, f in enumerate(_six)]
    fmt = subtopic_option_format(qs)
    check('all_observed_spans_six_formats', len(fmt['all_observed']) == 6)
    check('all_observed_is_sorted', fmt['all_observed'] == sorted(fmt['all_observed']))
    check('all_observed_exact', fmt['all_observed'] == sorted(_six))
    check('all_observed_not_in_insertion_order', fmt['all_observed'] != _six)
    check('recent_format_is_the_latest_year', fmt['recent_format'] == 'coordinate_pair')
    check('changed_recently_true_when_latest_differs',
          fmt['changed_recently'] is True)
    empty = subtopic_option_format([])
    check('subtopic_option_format_empty_shape',
          empty['all_observed'] == [] and empty['changed_recently'] is False
          and empty['primary'] == 'single_value')

    # ── E-9 difficulty scoring RETIRED (GAP-2026-08-27-DIFFICULTY-PROFILE) ──
    check('e9_scorer_retired', not hasattr(bc, 'score_difficulty') and 'score_difficulty' not in globals())
    check('calibration_block_retired', 'PYQ_DIFFICULTY_CALIBRATION' not in open(__file__, encoding='utf-8').read().split('def write_section_rules(')[1].split('\ndef ')[0])

    # ── E-10 strip mode / templates ───────────────────────────────────────
    check('determine_strip_mode_quantitative_for_maths',
          determine_strip_mode('Maths', 'Algebra', 'Quadratics') == 'quantitative')
    check('generate_templates_empty_is_safe', generate_templates([], 'numeric') == [])

    # ── E-11 note handling ────────────────────────────────────────────────
    check('note_freq_all_years_is_mandatory', classify_note_frequency(5, 5) == 'mandatory')
    check('note_freq_zero_total_is_safe', classify_note_frequency(0, 0) is not None)
    check('canonical_note_text_empty_is_empty_string', canonical_note_text({}) == '')

    # ── §4 vision aggregation ─────────────────────────────────────────────
    check('get_vision_candidates_empty_is_empty_list',
          get_vision_candidates([], {}, {}) == [])
    check('VISION_PER_SHEET_is_positive_int',
          isinstance(VISION_PER_SHEET, int) and VISION_PER_SHEET > 0)
    check('VISION_WORKDIR_is_absolute', str(VISION_WORKDIR).startswith('/'))

    # ══ §3 / §5 / §6 (Batch 3) ═══════════════════════════════════════════
    # ── SESSION_RE INJECTION — the one session global in the extraction scope ──
    # Omitting it must RAISE, not fall back. A default would silently mislabel every
    # paper's shift with a guessed keyword: a wrong answer, not an error, which is the
    # failure shape this corpus pays most for (v2.16 RIGID-1).
    try:
        extract_shift_from_filename('IIT_JAM_2026_Shift-2.docx')
        check('session_re_omission_raises', False)
    except RuntimeError as exc:
        check('session_re_omission_raises', 'session_re' in str(exc))
        check('session_re_error_says_what_to_pass', 'SESSION_RE' in str(exc))
    _sre = re.compile(re.escape('Shift') + r'[-_\s]?(\d+)', re.IGNORECASE)
    check('shift_parsed_with_injected_regex',
          extract_shift_from_filename('IIT_JAM_2026_Shift-2.docx', session_re=_sre) == 'S2')
    check('shift_defaults_to_S1_when_absent',
          extract_shift_from_filename('IIT_JAM_2026.docx', session_re=_sre) == 'S1')
    # A DIFFERENT keyword must give a different answer — proves the regex is really
    # the injected one and not a hidden module-level default.
    _slot = re.compile(re.escape('Slot') + r'[-_\s]?(\d+)', re.IGNORECASE)
    check('injected_regex_is_actually_used',
          extract_shift_from_filename('EXAM_Slot_3.docx', session_re=_slot) == 'S3'
          and extract_shift_from_filename('EXAM_Slot_3.docx', session_re=_sre) == 'S1')

    # ── §5 synthesis: the id/axis contract every later step joins on ──────────
    check('section_prefix_is_deterministic',
          section_prefix('Real Analysis') == section_prefix('Real Analysis'))
    _sid = make_subtopic_id('Real Analysis', 'Sequences', 'Convergence')
    check('make_subtopic_id_is_deterministic',
          _sid == make_subtopic_id('Real Analysis', 'Sequences', 'Convergence'))
    check('make_subtopic_id_is_nonempty_str', isinstance(_sid, str) and _sid)
    check('distinct_subtopics_get_distinct_ids',
          _sid != make_subtopic_id('Real Analysis', 'Sequences', 'Divergence'))

    # ── §6 QV: the contract D1 broke — checks are 2-tuples, counters are not ──
    # print_qv must render a counter WITHOUT unpacking it. v2.41.0 added an int under
    # a '_'-prefixed key and every Step 5 run raised TypeError here for eleven days
    # (GAP-2026-08-16-STEP5-SYNTHESIS-UNRUNNABLE D1).
    import io as _io, contextlib as _cl
    _buf = _io.StringIO()
    with _cl.redirect_stdout(_buf):
        _ok = print_qv({'QV-1': ('PASS', 'all covered'), '_counter_demo': 0})
    check('print_qv_tolerates_a_counter_key', _ok is True)
    check('print_qv_renders_the_counter', 'counter_demo' in _buf.getvalue())
    with _cl.redirect_stdout(_io.StringIO()):
        check('print_qv_reports_FAIL',
              print_qv({'QV-1': ('FAIL', 'broken')}) is False)
    # A malformed CHECK must raise at its source, naming the key.
    try:
        with _cl.redirect_stdout(_io.StringIO()):
            print_qv({'QV-2': 7})
        check('print_qv_rejects_a_malformed_check', False)
    except TypeError as exc:
        check('print_qv_rejects_a_malformed_check', 'QV-2' in str(exc))

    import os as _os

    # ══ B4 MUTATION FIXTURES ═════════════════════════════════════════════
    # Each of these was written to KILL a specific surviving mutant found by
    # `audit_mutation.py --engine analyse_engine.py`. The first run scored 25%
    # (12 killed / 48) — 55 assertions over 4,847 lines looked thorough and detected
    # a quarter of meaningful decision changes. A fixture that cannot distinguish the
    # mutant from the original proves nothing, which is the principle
    # MUTATION_BUDGETS.json exists to state.
    #
    # THE RECURRING MISTAKE THESE CORRECT: feeding inputs ALREADY IN ORDER. With
    # sorted input, `sorted(x)` and `list(x)` are indistinguishable and every
    # sorted->list mutant survives. Every ordering fixture below supplies data OUT OF
    # ORDER on purpose.

    # ── E-8 recent_format depends on sorted(by_year) picking the LATEST year ──
    # Years supplied 2025, 2023, 2024 — insertion order != sorted order, so
    # sorted->list picks the last INSERTED year instead of the latest.
    _ooo = [{'options': SHAPES['coordinate_pair'], 'year': 2025, 'stem': 'a'},
            {'options': SHAPES['single_value'],    'year': 2023, 'stem': 'b'},
            {'options': SHAPES['image_only'],      'year': 2024, 'stem': 'c'}]
    check('recent_format_uses_latest_year_not_last_seen',
          subtopic_option_format(_ooo)['recent_format'] == 'coordinate_pair')

    # ── _extract_qualifiers returns a SORTED tuple (id stability) ─────────────
    _q = _extract_qualifiers('symbolic numeric alphanumeric series')
    check('extract_qualifiers_is_sorted', list(_q) == sorted(_q))
    check('extract_qualifiers_is_deterministic',
          _extract_qualifiers('symbolic numeric alphanumeric series') == _q)

    # ── _is_verbal: FIGURAL is never verbal, whatever the section ─────────────
    check('is_verbal_true_for_language_text', _is_verbal('English Language', 'TEXT') is True)
    check('is_verbal_false_for_non_verbal_section',
          _is_verbal('Real Analysis', 'FIGURAL') is False)

    # ── detect_is_msq: the positive branch must actually fire ────────────────
    check('detect_is_msq_positive',
          detect_is_msq('Which of the following are correct?', ['a', 'b', 'c', 'd']) is True)

    # ── load_mechanic_overrides: malformed JSON must HARD STOP ────────────────
    # UNCONDITIONAL, AND IN A DIRECTORY THAT ALWAYS EXISTS.
    # Two earlier versions of this fixture were environment-dependent and each produced
    # a mutation gate that disagreed with itself:
    #   v1 wrote to a FIXED path under /mnt/project — concurrent mutants raced on one
    #      file: 28 survivors on one machine, 29 on another.
    #   v2 made the filename process-unique but kept the directory. /mnt/project is
    #      writable in the dev container and NOT on the GitHub runner, so the whole
    #      assertion was SKIPPED there: baseline 65 locally, 64 in CI, and the
    #      load_mechanic_overrides raise->pass mutant died here and survived there.
    #      The budget of 28 was measured where the assertion runs and could not gate an
    #      environment where it does not (GAP-2026-08-17-B4-ENV-SKEW).
    # tempfile.mkdtemp() is writable everywhere, and search_dirs points the function at
    # it. NO `if` GUARD: a conditional assertion is one that can vanish silently, which
    # is the whole defect above.
    import tempfile as _tf, shutil as _sh
    _tmpd = _tf.mkdtemp(prefix='ae_selftest_')
    try:
        with open(_os.path.join(_tmpd, 'ZZMUT_mechanic_overrides.json'),
                  'w', encoding='utf-8') as _fh:
            _fh.write('{ this is not json')
        globals()['_OVERRIDES'] = None          # module-level cache; must not leak
        try:
            load_mechanic_overrides('ZZMUT', search_dirs=(_tmpd,))
            check('bad_overrides_json_hard_stops', False)
        except SystemExit:
            check('bad_overrides_json_hard_stops', True)
        except Exception:
            check('bad_overrides_json_hard_stops', False)
    finally:
        globals()['_OVERRIDES'] = None
        _sh.rmtree(_tmpd, ignore_errors=True)

    # ── _compute_structural_changes: year ordering ───────────────────────────
    _sc = [{'section': 'S', 'topic': 'T', 'subtopic': 'U', 'observed_count': 3,
            'years_seen': [2025, 2021, 2023]},
           {'section': 'S', 'topic': 'T', 'subtopic': 'V', 'observed_count': 2,
            'years_seen': [2022, 2026]}]
    _res = _compute_structural_changes(_sc)
    check('structural_changes_returns_a_list', isinstance(_res, list))

    # ── generate_templates: year ordering inside the pattern set ─────────────
    _tq = [{'stem': 'Find x when a=2', 'year': 2026, 'num': 1},
           {'stem': 'Find x when a=7', 'year': 2021, 'num': 2},
           {'stem': 'Find x when a=5', 'year': 2024, 'num': 3}]
    _tpl = generate_templates(_tq, 'quantitative')
    check('generate_templates_returns_patterns', isinstance(_tpl, list) and len(_tpl) >= 1)
    # UNCONDITIONAL. The previous form was `if _tpl and ... 'years' in _tpl[0]:`, which
    # silently disappears the day the pattern shape changes — the same vanishing-check
    # defect as the overrides fixture above, one line apart.
    check('template_pattern_has_years',
          bool(_tpl) and isinstance(_tpl[0], dict) and 'years' in _tpl[0])
    _tyears = list(_tpl[0]['years']) if (_tpl and isinstance(_tpl[0], dict)
                                         and 'years' in _tpl[0]) else None
    check('template_years_are_sorted', _tyears is not None and _tyears == sorted(_tyears))

    # ══ B5 GUARD FIXTURES — the six raise->pass survivors ═════════════════
    # WHY THESE FIRST. A guard is a HARD STOP: "this taxonomy is corrupt, refuse to
    # continue". Delete one and nothing crashes — the run accepts bad data and emits a
    # confident, wrong artefact. That is the most expensive failure shape in this
    # corpus, and until now the suite could not tell whether any of these six alarms
    # were still connected. Each fixture below removes exactly that doubt.
    import copy as _cp

    _GE = [{'subtopic_id': 'RA.SEQ.CONV', 'section': 'Real Analysis', 'topic': 'Seq',
            'subtopic': 'Convergence', 'PYQ_STEM_PATTERNS': [], 'format': 'TEXT',
            'observed_count': 3},
           {'subtopic_id': 'RA.SEQ.DIV', 'section': 'Real Analysis', 'topic': 'Seq',
            'subtopic': 'Divergence', 'PYQ_STEM_PATTERNS': [], 'format': 'TEXT',
            'observed_count': 2}]

    def _with_overrides(**ov):
        """Set the module override cache directly — no filesystem, so this runs
        identically in every environment (the GAP-2026-08-17-B4-ENV-SKEW lesson)."""
        base = {'template_sets': None, 'template_sets_by_section': {},
                'subtopic_overrides': {}, 'subtopic_merges': []}
        base.update(ov)
        globals()['_OVERRIDES'] = base

    def _expect_hard_stop(name, fn, *args, **ov):
        _with_overrides(**ov)
        try:
            fn(*args)
            check(name, False)                       # no raise = the alarm is gone
        except SystemExit:
            check(name, True)
        except Exception:
            check(name, False)                       # wrong exception is not the guard
        finally:
            globals()['_OVERRIDES'] = None

    def _expect_hard_stop_text(name, want, fn, *args, **ov):
        """As _expect_hard_stop, but the MESSAGE must match EXACTLY.

        B7 — DIAGNOSTIC TEXT IS PART OF THE CONTRACT. Five surviving mutants were a
        `sorted(...)` inside a hard-stop f-string, and the tempting verdict is
        "equivalent mutant — only the word order inside an error changes". It is not.
        Every one of those messages enumerates a SET, so under list() the SAME corrupt
        override file produces a differently-worded refusal on every run. A curator
        cannot diff two runs, cannot grep a log for a known message, and cannot tell a
        second distinct fault from the same fault re-ordered. Artefact nondeterminism
        (D4) applies to what the operator READS, not only to what the writers EMIT.
        So the text is asserted rather than assumed — and with three-or-more members
        each time, because a two-element set is a coin flip.
        """
        _with_overrides(**ov)
        try:
            fn(*args)
            check(name, False)                       # no raise = the alarm is gone
        except SystemExit as _ex:
            check(name, str(_ex) == want)
        except Exception:
            check(name, False)
        finally:
            globals()['_OVERRIDES'] = None

    # OV-4 — a merge group with fewer than two members is meaningless
    _expect_hard_stop('OV4_merge_group_under_two_hard_stops',
                      apply_subtopic_merges, _cp.deepcopy(_GE), 'ZZEXAM',
                      subtopic_merges=[['RA.SEQ.CONV']])
    # OV-3 — merging an id that does not exist would silently drop a real subtopic
    _expect_hard_stop('OV3_merge_unknown_id_hard_stops',
                      apply_subtopic_merges, _cp.deepcopy(_GE), 'ZZEXAM',
                      subtopic_merges=[['RA.SEQ.CONV', 'NOPE.ID']])
    # OV-4b — overlapping groups: two merges competing to own one id
    _expect_hard_stop('OV4b_overlapping_merge_groups_hard_stop',
                      apply_subtopic_merges, _cp.deepcopy(_GE), 'ZZEXAM',
                      subtopic_merges=[['RA.SEQ.CONV', 'RA.SEQ.DIV'],
                                       ['RA.SEQ.CONV', 'RA.SEQ.DIV']])
    # OV-1 — an override key matching no subtopic_id is a curator typo that would
    # otherwise apply to nothing, silently
    _expect_hard_stop('OV1_unknown_override_key_hard_stops',
                      stamp_mechanic_axes, _cp.deepcopy(_GE), 'ZZEXAM',
                      subtopic_overrides={'NOPE.ID': {'mechanic': 'x'}})
    # OV-6 — template_sets_by_section naming a section that does not exist
    _expect_hard_stop('OV6_unknown_section_hard_stops',
                      stamp_mechanic_axes, _cp.deepcopy(_GE), 'ZZEXAM',
                      template_sets_by_section={'No Such Section': ['verbal']})

    # ── The two BARE re-raises: absent vs unreadable Analysis doc ─────────────
    # v2.31: an unreadable or unapproved Analysis doc is NEVER a WARN. It is not a
    # source that happened to contribute nothing — it is the WRONG TAXONOMY, and every
    # id minted from it is wrong. `raise->pass` here converts that into a silent
    # downgrade, which is precisely the fault the v2.31 comment says had to survive two
    # independent downgrades to be seen at all.
    import corpus_io as _cio
    _orig_lt = _cio.load_taxonomy

    def _raise_doc_error(msg):
        def _f(**_k):
            raise _cio.AnalysisDocError(msg)
        return _f
    try:
        _cio.load_taxonomy = _raise_doc_error('doc is present but unreadable')
        try:
            _extract_taxonomy_tuples_from_analysis_doc('/nonexistent/x.docx')
            check('unreadable_analysis_doc_reraises', False)
        except _cio.AnalysisDocError:
            check('unreadable_analysis_doc_reraises', True)
        except Exception:
            check('unreadable_analysis_doc_reraises', False)
        # ...and the ONE benign case must still be benign, or the guard is just noise
        _cio.load_taxonomy = _raise_doc_error('no Analysis doc found')
        try:
            check('absent_analysis_doc_is_benign',
                  _extract_taxonomy_tuples_from_analysis_doc('/nonexistent/x.docx') == [])
        except Exception:
            check('absent_analysis_doc_is_benign', False)
    finally:
        _cio.load_taxonomy = _orig_lt

    # ── OV-3b: a merge group spanning two sections ───────────────────────────
    # Merging across sections would move a subtopic into a section it was never
    # observed in, and every Step 6/7 join on section+id would then be wrong.
    _GE2 = _cp.deepcopy(_GE)
    _GE2[1]['section'] = 'Linear Algebra'
    _expect_hard_stop('OV3b_merge_across_sections_hard_stops',
                      apply_subtopic_merges, _GE2, 'ZZEXAM',
                      subtopic_merges=[['RA.SEQ.CONV', 'RA.SEQ.DIV']])

    # ── B7: the five diagnostic messages that enumerate a set ────────────────
    # OV-3 across FOUR sections, supplied out of alphabetical order, so the message
    # is a 1-in-24 permutation detector rather than a coin flip.
    _GE4 = [dict(_e, subtopic_id=f'S{_i}.T.X', section=_s)
            for _i, _s in enumerate(['Real Analysis', 'Real Algebra',
                                     'Real Arithmetic', 'Real Applications'])
            for _e in (_GE[0],)]
    _expect_hard_stop_text(
        'OV3b_span_message_lists_sections_alphabetically',
        "FAIL: subtopic_merges group ['S0.T.X', 'S1.T.X', 'S2.T.X', 'S3.T.X'] spans "
        "sections ['Real Algebra', 'Real Analysis', 'Real Applications', "
        "'Real Arithmetic']",
        apply_subtopic_merges, _cp.deepcopy(_GE4), 'ZZEXAM',
        subtopic_merges=[['S0.T.X', 'S1.T.X', 'S2.T.X', 'S3.T.X']])

    # OV-1 prints one line PER unknown key before the hard stop, and that loop is
    # `for u in sorted(unknown)`. Under list() a curator fixing four typos reads them
    # in a different order every run and cannot tell the list apart from a new fault.
    _with_overrides(subtopic_overrides={_k: {'mechanic': 'x'} for _k in
                                        ('ZED.ID', 'ALPHA.ID', 'MID.ID', 'BETA.ID')})
    _ov1buf = _io7b = None
    try:
        import io as _io1b, contextlib as _cl1b
        _ov1buf = _io1b.StringIO()
        try:
            with _cl1b.redirect_stdout(_ov1buf):
                stamp_mechanic_axes(_cp.deepcopy(_GE), 'ZZEXAM')
            check('OV1_lists_unknown_keys_alphabetically', False)   # no raise
        except SystemExit:
            _ov1keys = [_l.split("override key ")[1].split(" matches")[0]
                        for _l in _ov1buf.getvalue().split('\n') if 'override key ' in _l]
            check('OV1_lists_unknown_keys_alphabetically',
                  _ov1keys == ["'ALPHA.ID'", "'BETA.ID'", "'MID.ID'", "'ZED.ID'"])
    finally:
        globals()['_OVERRIDES'] = None

    # OV-5 / OV-5b: an overrides file naming template sets that do not exist. Both
    # messages list the offending values; both were surviving mutants.
    _tmp5 = _tf.mkdtemp(prefix='ae_ovsets_')
    for _nm5, _body5, _want5 in (
        ('OV5_template_sets_message_is_sorted',
         '{"exam_code": "ZZSET", "template_sets": ["zeta", "alpha", "mu", "beta"]}',
         "template_sets has unknown values ['alpha', 'beta', 'mu', 'zeta']"),
        ('OV5b_template_sets_by_section_message_is_sorted',
         '{"exam_code": "ZZSET", "template_sets_by_section": '
         '{"Sec A": ["zeta", "alpha", "mu", "beta"]}}',
         "template_sets_by_section['Sec A'] has unknown values "
         "['alpha', 'beta', 'mu', 'zeta'] (check case/whitespace)"),
    ):
        with open(_os.path.join(_tmp5, 'ZZSET_mechanic_overrides.json'),
                  'w', encoding='utf-8') as _fh5:
            _fh5.write(_body5)
        globals()['_OVERRIDES'] = None
        try:
            load_mechanic_overrides('ZZSET', search_dirs=(_tmp5,))
            check(_nm5, False)                       # no raise = the alarm is gone
        except SystemExit as _ex5:
            check(_nm5, str(_ex5).endswith(_want5))
        except Exception:
            check(_nm5, False)
        finally:
            globals()['_OVERRIDES'] = None

    # print_qv: an unknown status names the legal vocabulary. Four icons, and the
    # operator is expected to compare that list against what they wrote.
    try:
        import io as _iopq, contextlib as _clpq
        with _clpq.redirect_stdout(_iopq.StringIO()):
            print_qv({'QV-X': ('MAYBE', 'detail')})
        check('print_qv_status_vocabulary_is_sorted', False)        # no raise
    except ValueError as _expq:
        check('print_qv_status_vocabulary_is_sorted',
              str(_expq).endswith("expected one of "
                                  "['FAIL', 'PASS', 'WARN']."))
    except Exception:
        check('print_qv_status_vocabulary_is_sorted', False)

    # ── EC-M17: an overrides file declaring the WRONG exam_code ──────────────
    # Applying another exam's overrides is the quietest possible corruption: valid
    # JSON, plausible keys, entirely the wrong exam. Uses search_dirs so it runs in
    # any environment.
    _tmp2 = _tf.mkdtemp(prefix='ae_ovcode_')
    try:
        with open(_os.path.join(_tmp2, 'ZZCODE_mechanic_overrides.json'),
                  'w', encoding='utf-8') as _fh:
            _fh.write('{"exam_code": "SOME_OTHER_EXAM", "subtopic_overrides": {}}')
        globals()['_OVERRIDES'] = None
        try:
            load_mechanic_overrides('ZZCODE', search_dirs=(_tmp2,))
            check('ECM17_wrong_exam_code_hard_stops', False)
        except SystemExit:
            check('ECM17_wrong_exam_code_hard_stops', True)
        except Exception:
            check('ECM17_wrong_exam_code_hard_stops', False)
    finally:
        globals()['_OVERRIDES'] = None
        _sh.rmtree(_tmp2, ignore_errors=True)

    # ── v2.31: taxonomy_sync_entries must RE-RAISE on an unreadable Analysis doc ─
    # This is the second of the two bare re-raises, and the one with the longest
    # history: the old code caught this fault HERE and inside the extractor, so it had
    # to survive two independent downgrades to be seen at all. raise->pass restores
    # exactly that double-downgrade.
    import glob as _glob_mod
    _orig_glob = _glob_mod.glob
    _orig_extract = globals()['_extract_taxonomy_tuples_from_analysis_doc']
    try:
        _glob_mod.glob = lambda _p: (['/nonexistent/ZZSYNC_analysis.docx']
                                     if '/mnt/project/' in str(_p) else [])

        def _boom(_path):
            raise _cio.AnalysisDocError('present but unreadable')
        globals()['_extract_taxonomy_tuples_from_analysis_doc'] = _boom
        try:
            taxonomy_sync_entries([], 'ZZSYNC')
            check('sync_unreadable_analysis_doc_reraises', False)
        except _cio.AnalysisDocError:
            check('sync_unreadable_analysis_doc_reraises', True)
        except Exception:
            check('sync_unreadable_analysis_doc_reraises', False)
    finally:
        _glob_mod.glob = _orig_glob
        globals()['_extract_taxonomy_tuples_from_analysis_doc'] = _orig_extract

    # ══ B6 ORDERING FIXTURES — subtopic_id stability ══════════════════════
    # THE MOST CONSEQUENTIAL SURVIVOR IN THE SET. build_section_prefix_map is called
    # with sorted({sections}) from THREE places — mint_subtopic_ids, stamp_mechanic_axes
    # and run_qv. When two section names collide on a prefix the map appends a counter
    # (ra, ra2, ra3...), so WHICH section gets the bare prefix is decided by iteration
    # order. Replace sorted( with list( and the assignment follows PYTHONHASHSEED
    # instead of the alphabet: `Real Algebra` is `ra` on one run and `ra2` on the next.
    #
    # subtopic_id is the key EVERY later step joins on — Step 6 blueprints against it,
    # Step 7 generates against it, the manifest is indexed by it. Non-deterministic ids
    # do not fail loudly; they silently fail to match, which is the D4 defect class one
    # level up from ordering inside a file.
    #
    # FOUR colliding sections, supplied out of alphabetical order. The mutant would have
    # to reproduce the alphabet by luck: 1 chance in 24.
    _SECS = ['Real Analysis', 'Real Algebra', 'Real Arithmetic', 'Real Applications']
    _PMAP = build_section_prefix_map(sorted(set(_SECS)))
    check('prefix_collision_resolves_alphabetically',
          _PMAP == {'Real Algebra': 'ra', 'Real Analysis': 'ra2',
                    'Real Applications': 'ra3', 'Real Arithmetic': 'ra4'})
    _IDE = [{'section': _s, 'topic': 'T', 'subtopic': f'S{_i}', 'observed_count': _i + 1,
             'years_seen': [2024], 'PYQ_STEM_PATTERNS': [], 'format': 'TEXT'}
            for _i, _s in enumerate(_SECS)]
    _minted = _cp.deepcopy(_IDE)
    mint_subtopic_ids(_minted)
    _bysec = {e['section']: e['subtopic_id'].split('.')[0] for e in _minted}
    check('subtopic_id_prefix_is_alphabetical_not_hash_order',
          _bysec == {'Real Algebra': 'ra', 'Real Analysis': 'ra2',
                     'Real Applications': 'ra3', 'Real Arithmetic': 'ra4'})
    # ...and minting the SAME entries in a different input order must give the SAME ids,
    # because a subtopic_id that depends on how the papers happened to be read is a
    # cross-step join that breaks for reasons no one can reproduce.
    _minted2 = _cp.deepcopy(_IDE[::-1])
    mint_subtopic_ids(_minted2)
    check('subtopic_ids_are_input_order_independent',
          {e['section']: e['subtopic_id'] for e in _minted2}
          == {e['section']: e['subtopic_id'] for e in _minted})

    # ── generate_templates: the year list inside a pattern ───────────────────
    # Four years supplied out of order; sorted output is 1-in-24 by luck.
    _tq = [{'stem': 'Find x when a=2', 'year': 2026, 'num': 1},
           {'stem': 'Find y when b=3', 'year': 2021, 'num': 2},
           {'stem': 'Find z when c=4', 'year': 2024, 'num': 3},
           {'stem': 'Find x when a=9', 'year': 2023, 'num': 4}]
    _tp = generate_templates(_tq, 'quantitative')
    check('template_years_span_four', bool(_tp) and len(_tp[0].get('years', [])) == 4)
    check('template_years_exact_sorted',
          bool(_tp) and list(_tp[0]['years']) == [2021, 2023, 2024, 2026])

    # ── stamp_mechanic_axes: collision_domain comes from the same prefix map ─
    # form_key uniqueness is asserted PER collision_domain, so a domain that changes
    # between runs silently moves the goalposts of that invariant.
    _st = _cp.deepcopy(_IDE)
    mint_subtopic_ids(_st)
    globals()['_OVERRIDES'] = None
    try:
        import io as _io2, contextlib as _cl2
        with _cl2.redirect_stdout(_io2.StringIO()):
            stamp_mechanic_axes(_st, 'ZZEXAM')
    finally:
        globals()['_OVERRIDES'] = None
    check('collision_domain_follows_alphabetical_prefix',
          {e['section']: e.get('collision_domain') for e in _st}
          == {'Real Algebra': 'ra', 'Real Analysis': 'ra2',
              'Real Applications': 'ra3', 'Real Arithmetic': 'ra4'})

    # ── subtopic_option_format: the NO-YEARS branch, and DEFECT D7 ────────────
    # B6 added these two checks believing they reached the no-years return. They did
    # not: the old `q.get('year', '?')` gave every year-less question the key '?', so
    # `years` was ['?'] — truthy — and the branch stayed unreachable. The mutation gate
    # said so plainly by leaving that line's mutant ALIVE while the fixture passed, and
    # a fixture that passes without reaching the code it names is worse than no fixture.
    # That trail led to D7: the same sentinel mixed str and int keys in a dict whose keys
    # are sorted on the next line, so ONE year-less question crashed synthesis outright.
    # Four formats here: 1-in-24 that a hash-ordered list reproduces the alphabet.
    _noyear = [{'options': SHAPES['single_value']},
               {'options': SHAPES['coordinate_pair']},
               {'options': SHAPES['image_only']},
               {'options': SHAPES['roman_label']}]
    _nf = subtopic_option_format(_noyear)
    check('no_year_branch_all_observed_sorted',
          _nf['all_observed'] == ['coordinate_pair', 'image_only',
                                  'roman_label', 'single_value'])
    check('no_year_branch_not_changed_recently', _nf['changed_recently'] is False)
    check('no_year_branch_recent_format_falls_back_to_primary',
          _nf['recent_format'] == _nf['primary'])

    # ── generate_templates: PRODUCTION-SHAPE year=None (v2.54.1 regression) ──
    # The D7-era fixtures modelled a year-less question as a dict WITHOUT the
    # 'year' key; extraction stamps the key WITH None, and a .get default made
    # the two shapes behave differently — which is how the sibling crash in
    # generate_templates survived the D7 release. These pin the production shape.
    _gt_mixed = [
        {'stem': 'Find x when a=2', 'year': 2026, 'num': 1},
        {'stem': 'Find x when a=7', 'year': None, 'num': 2},  # key PRESENT, None
        {'stem': 'Find x when a=5', 'year': 2024, 'num': 3},
    ]
    try:
        _gt1 = generate_templates(_gt_mixed, 'quantitative')
        check('yearless_mixed_templates_no_crash', True)
        check('yearless_mixed_years_exclude_none',
              all(None not in p['years'] for p in _gt1))
    except TypeError:
        check('yearless_mixed_templates_no_crash', False)
    _gt_all_none = [{'stem': 'Find x when a=2', 'year': None, 'num': 1},
                    {'stem': 'Find y when b=3', 'year': None, 'num': 2}]
    _gt2 = generate_templates(_gt_all_none, 'quantitative')
    check('yearless_only_no_observed_recent',
          all(p['confidence'] != 'observed_recent' and p['years'] == []
              for p in _gt2))
    check('yearless_only_not_deprecated',
          all(not p['deprecated'] for p in _gt2))
    _gt_dated = [{'stem': 'Find x when a=2', 'year': 2026, 'num': 1},
                 {'stem': 'Find x when a=7', 'year': 2021, 'num': 2}]
    check('dated_corpus_years_intact',
          all(p['years'] for p in generate_templates(_gt_dated, 'quantitative')))
    # D7 REGRESSION GUARD — a MIXED corpus. Under the old sentinel this raised
    # TypeError from sorted({2021, '?'}); it must now read the recency vote from the
    # year-bearing questions alone and ignore the year-less one.
    _mixed = [{'options': SHAPES['single_value'], 'year': 2019},
              {'options': SHAPES['single_value'], 'year': 2019},
              {'options': SHAPES['coordinate_pair'], 'year': 2024},
              {'options': SHAPES['image_only']}]
    _mf = subtopic_option_format(_mixed)
    check('d7_mixed_year_corpus_does_not_crash', _mf['primary'] == 'single_value')
    check('d7_recency_vote_ignores_year_less_questions',
          _mf['recent_format'] == 'coordinate_pair' and _mf['changed_recently'] is True)
    check('d7_year_less_question_still_counted_in_all_observed',
          _mf['all_observed'] == ['coordinate_pair', 'image_only', 'single_value'])

    # ══ B7 — THE WRITER CORPUS ═══════════════════════════════════════════════
    # Nine of the twenty surviving mutants lived in five functions this suite had
    # NEVER CALLED: synthesise_subtopic, compute_section_axis_distribution,
    # write_section_rules, write_subtopic_manifest and
    # rebuild_subtopic_manifest_from_section_rules. They were untested for a
    # mechanical reason, not a judgement one — each writes to
    # '/mnt/user-data/outputs/', which does not exist on a CI runner, so no fixture
    # could call them. That is why B7 begins with `out_dir` (see OUTPUT_DIR above)
    # rather than with assertions: the environment was the blocker.
    #
    # This fixture is a different SHAPE from every fixture before it. B5 and B6
    # asserted on directly-callable pure functions; the remaining ordering defects
    # live behind the writers, where the observable is EMITTED TEXT. So the corpus is
    # driven end to end — questions -> synthesise_subtopic -> write_section_rules ->
    # write_subtopic_manifest -> rebuild — and the assertions read the artefacts back.
    #
    # THE CORPUS IS ADVERSARIAL BY CONSTRUCTION, not merely representative:
    #   - FOUR sections colliding on the prefix `ra`, supplied out of alphabetical
    #     order — 1 chance in 24 that a hash-ordered map reproduces the alphabet.
    #   - FOUR papers whose FIGURAL counts are all DIFFERENT (1/2/3/4), so the
    #     per-paper series is a permutation detector rather than a bag of equal
    #     numbers — again 1 in 24.
    #   - FOUR years whose set-iteration order is NOT their sorted order
    #     (list({2021,2023,2024,2026}) == [2024,2026,2021,2023]).
    # A fixture whose values are all equal cannot see an ordering defect at all; that
    # is precisely how these nine survived a suite of eighty-six checks.
    import copy as _cp7, io as _io7, contextlib as _cl7, json as _js7, tempfile as _tf7
    _W7SECS = ['Real Analysis', 'Real Algebra', 'Real Arithmetic', 'Real Applications']
    _W7PAPERS = [(2026, 'S1'), (2021, 'S2'), (2024, 'S1'), (2023, 'S3')]
    _W7FIG = {2021: 1, 2023: 2, 2024: 3, 2026: 4}

    def _w7q(sec, topic, sub, year, shift, num, img='none'):
        return {'section': sec, 'topic': topic, 'subtopic': sub, 'year': year,
                'shift': shift, 'num': num, 'q_num': num, 'question_type': 'MCQ',
                'stem': f'Find x in {sub} case {num}', 'answer': '1',
                'options': ['1', '2', '3', '4'], 'option_label': '1/2/3/4',
                'format': 'TEXT', 'difficulty': {'level': 'Medium'}, 'image_role': img}

    _w7prog, _w7n = {}, 0
    for _s7 in _W7SECS:
        for _j7 in range(2):
            _qs7 = []
            for (_y7, _sh7) in _W7PAPERS:
                _w7n += 1
                _qs7.append(_w7q(_s7, f'T{_j7}', f'{_s7} Sub {_j7}', _y7, _sh7, _w7n))
                if _j7 == 0:                       # only Sub 0 carries figures
                    for _k7 in range(_W7FIG[_y7]):
                        _w7n += 1
                        _qs7.append(_w7q(_s7, f'T{_j7}', f'{_s7} Sub {_j7}',
                                         _y7, _sh7, _w7n, img='essential'))
            _w7prog[(_s7, f'T{_j7}', f'{_s7} Sub {_j7}')] = _qs7

    _w7buf = _io7.StringIO()
    with _cl7.redirect_stdout(_w7buf):
        _w7ent = [synthesise_subtopic(_k7[0], _k7[1], _k7[2], _v7, _w7prog)
                  for _k7, _v7 in _w7prog.items()]

    # synthesise_subtopic L1793 — the year list carried into every stem pattern.
    # These years are the observation window Step 6 reads back.
    check('synth_pattern_years_sorted_not_set_order',
          all(list(_p7.get('years', [])) == [2021, 2023, 2024, 2026]
              for _e7 in _w7ent for _p7 in _e7['PYQ_STEM_PATTERNS']
              if _p7.get('years')))

    # compute_section_axis_distribution L1652 / L1730 / L1754. The per-paper SERIES is
    # what Step 6 turns into a quota; a permuted series assigns a paper's figural count
    # to a different paper, so the MEAN survives unchanged and every individual target
    # is wrong — the silent-wrong-answer shape, not a crash.
    with _cl7.redirect_stdout(_w7buf):
        _w7ax = compute_section_axis_distribution(
            [_e7 for _e7 in _w7ent if _e7['section'] == _W7SECS[0]], _w7prog)
    check('axis_recent_years_descending',
          _w7ax['recent_years'] == [2026, 2024, 2023, 2021])
    check('axis_figural_series_follows_sorted_papers',
          _w7ax['figural_per_paper_observed'] == [1, 2, 3, 4])
    check('axis_per_class_series_matches_figural_series',
          _w7ax['per_paper_observed_by_class'].get('FIGURAL') == [1, 2, 3, 4])
    check('axis_n_papers_recent_is_four', _w7ax['n_papers_recent'] == 4)

    _w7dir = _tf7.mkdtemp()
    _w7meta = {'papers_analysed': 4, 'questions_analysed': _w7n,
               'years_covered': [2021, 2023, 2024, 2026],
               'generation_date': '2026-08-17', 'options_count': 4}
    _w7cfg = {'exam_code': 'ZZW7', 'total_questions': 999,
              'sections': [{'name': 'Paper', 'q_range': [1, 999], 'q_count': 999}]}
    with _cl7.redirect_stdout(_w7buf):
        _w7rules = write_section_rules(_cp7.deepcopy(_w7ent), 'ZZW7', exam_meta=_w7meta,
                                       progress=_w7prog, out_dir=_w7dir)
    _w7txt = io_open_utf8(_w7rules).read()

    # write_section_rules L2991 / L3008 — both sorts are BY DESCENDING observed_count,
    # so the emitted sub_types_observed list and the per-entry blocks come out in
    # frequency order, which is the order Step 6 reads as significance.
    _w7subs = [_l7.strip()[2:] for _l7 in _w7txt.split('\n')
               if _l7.startswith('  - ') and 'Sub ' in _l7]
    check('section_rules_lists_subtypes_by_descending_count',
          _w7subs == [f'{_s7} Sub {_j7} (n={14 if _j7 == 0 else 4})'
                      for _s7 in _W7SECS for _j7 in (0, 1)])
    check('section_rules_emits_all_four_sections',
          sum(1 for _l7 in _w7txt.split('\n') if _l7.startswith('=== SECTION: ')) == 4)
    check('section_rules_carries_axis_distribution',
          _w7txt.count('  recent_years: [2026, 2024, 2023, 2021]') == 4)

    # write_subtopic_manifest L3829 / L3846. The entries are deliberately UNSTAMPED so
    # the id-minting fallback runs — that precondition is asserted, because a future
    # change that mints earlier would leave this fixture green while testing nothing.
    check('w7_entries_are_unstamped_so_id_fallback_runs',
          all('subtopic_id' not in _e7 for _e7 in _w7ent))
    with _cl7.redirect_stdout(_w7buf):
        _w7mpath = write_subtopic_manifest(_cp7.deepcopy(_w7ent), 'ZZW7',
                                           exam_meta=_w7meta, progress=_w7prog,
                                           exam_config=_w7cfg, out_dir=_w7dir)
    _w7man = _js7.loads(io_open_utf8(_w7mpath).read())
    check('manifest_ids_use_alphabetical_prefix_map',
          sorted(_w7man['subtopics']) == [
              'ra.t0.real_algebra_sub_0', 'ra.t1.real_algebra_sub_1',
              'ra2.t0.real_analysis_sub_0', 'ra2.t1.real_analysis_sub_1',
              'ra3.t0.real_applications_sub_0', 'ra3.t1.real_applications_sub_1',
              'ra4.t0.real_arithmetic_sub_0', 'ra4.t1.real_arithmetic_sub_1'])
    check('manifest_pattern_eras_papers_in_sorted_order',
          list(_w7man['pattern_eras']['papers']) ==
          ['2021|S2', '2023|S3', '2024|S1', '2026|S1'])

    # rebuild_subtopic_manifest_from_section_rules L4088 — the dedent test that ends the
    # axis_distribution block. Neutralise it and the parser stops on the first sub-line,
    # so the rebuilt manifest carries an EMPTY axis_distribution: a Step 6 run against a
    # rebuilt manifest would silently lose every target rather than fail.
    with _cl7.redirect_stdout(_w7buf):
        _w7rb = rebuild_subtopic_manifest_from_section_rules(_w7rules, 'ZZW7R',
                                                             out_dir=_w7dir)
    _w7rbm = _js7.loads(io_open_utf8(_w7rb).read())
    check('rebuild_recovers_axis_distribution_for_every_section',
          sorted(_w7rbm['axis_distribution']) == sorted(_W7SECS))
    check('rebuild_axis_block_keeps_all_eight_keys',
          all(sorted(_v7) == ['axis1_per_paper', 'axis2_audit_mode', 'axis2_per_paper',
                              'axis3_per_paper', 'mocks_per_window', 'n_papers_recent',
                              'negative_rate', 'recent_years']
              for _v7 in _w7rbm['axis_distribution'].values()))
    check('rebuild_recovers_every_subtopic_id',
          sorted(_w7rbm['subtopics']) == sorted(_w7man['subtopics']))

    # ── generate_templates L528: `years` decides the RECENCY WINDOW ───────────
    # last2 = set(years[-2:]). Under list() the "two most recent years" become two
    # arbitrary years, which flips `deprecated` and `confidence` on every pattern —
    # Step 6 reads `deprecated` as "do not generate from this pattern any more", so a
    # permuted window retires the live patterns and revives the dead ones. No error is
    # raised anywhere; the mock is simply built from the wrong decade.
    _gt = generate_templates(
        [{'stem': 'Find the limit of the sequence a_n as n tends to 2',
          'year': 2021, 'num': 1, 'options': ['1', '2', '3', '4']},
         {'stem': 'Find the limit of the sequence a_n as n tends to 3',
          'year': 2023, 'num': 2, 'options': ['1', '2', '3', '4']},
         {'stem': 'Evaluate the integral of f over the interval 4',
          'year': 2024, 'num': 3, 'options': ['1', '2', '3', '4']},
         {'stem': 'Evaluate the integral of f over the interval 5',
          'year': 2026, 'num': 4, 'options': ['1', '2', '3', '4']}], 'reasoning')
    _gtby = {tuple(_p['years']): _p for _p in _gt}
    check('recency_window_is_the_two_LATEST_years',
          sorted(_gtby) == [(2021, 2023), (2024, 2026)]
          and _gtby[(2024, 2026)]['confidence'] == 'observed_recent'
          and _gtby[(2024, 2026)]['deprecated'] is False
          and _gtby[(2021, 2023)]['deprecated'] is True)

    # generate_templates L562 — the rounding-deficit top-up. Three equal clusters give
    # 33+33+33 = 99, so the branch runs and hands the spare point to the largest
    # remainder. It had never been reached: every earlier fixture summed to 100.
    _gtd = generate_templates(
        [{'stem': 'Alpha alpha alpha one two three', 'year': 2024, 'num': 1,
          'options': ['1', '2', '3', '4']},
         {'stem': 'Beta beta beta four five six seven', 'year': 2024, 'num': 2,
          'options': ['1', '2', '3', '4']},
         {'stem': 'Gamma gamma gamma eight nine ten eleven', 'year': 2024, 'num': 3,
          'options': ['1', '2', '3', '4']}], 'reasoning')
    check('rounding_deficit_tops_up_to_exactly_100',
          [_p['frequency'] for _p in _gtd] == [34, 33, 33])

    # synthesise_subtopic L1793 — the FIGURAL fallback, reached only when no question
    # yields a text skeleton. Its `years` is the sole year record such a subtopic has.
    _fbe = synthesise_subtopic('S', 'T', 'U',
                               [{'year': _y7f, 'num': _i7f, 'stem': '',
                                 'options': ['1', '2', '3', '4'],
                                 'image_role': 'essential',
                                 'difficulty': {'level': 'Medium'}}
                                for _i7f, _y7f in enumerate((2026, 2021, 2024, 2023))], {})
    check('figural_fallback_pattern_years_are_sorted',
          [_p['years'] for _p in _fbe['PYQ_STEM_PATTERNS']] == [[2021, 2023, 2024, 2026]])

    # ── run_qv: QV-13, the mechanic-identity gate ────────────────────────────
    # L4615 — QV-13 re-derives every form_key and compares it against the stamped one.
    # That derivation needs the SAME prefix map the stamp used; hash-order it and the
    # gate reports every subtopic as nondeterministic. The gate would then be crying
    # wolf on correct data, which is how a FAIL gate gets switched off.
    _qvent = _cp7.deepcopy(_w7ent)
    with _cl7.redirect_stdout(_w7buf):
        mint_subtopic_ids(_qvent)
        globals()['_OVERRIDES'] = None
        try:
            stamp_mechanic_axes(_qvent, 'ZZQV')
        finally:
            globals()['_OVERRIDES'] = None
        _qvres = run_qv(_qvent, {_s7: [{'topic': f'T{_j7}', 'subtopic': f'{_s7} Sub {_j7}'}
                                       for _j7 in range(2)] for _s7 in _W7SECS},
                        {'_meta': {}})
    check('qv13_passes_on_correctly_stamped_entries', _qvres['QV-13'][0] == 'PASS')

    # ...and the SAME entries with one collision_domain nudged off its derived value
    # must FAIL. This is the assertion that makes `_po` load-bearing: without it the
    # gate computes a prefix map and ignores it, and the mutant that hash-orders that
    # map survives — which is precisely how this hole was found.
    _driftent = _cp7.deepcopy(_qvent)
    _driftent[0]['collision_domain'] = 'zz_drifted'
    globals()['_OVERRIDES'] = None
    with _cl7.redirect_stdout(_w7buf):
        _driftres = run_qv(_driftent, {}, {'_meta': {}})
    check('qv13_detects_collision_domain_drift',
          _driftres['QV-13'][0] == 'FAIL'
          and _driftent[0]['subtopic_id'] in _driftres['QV-13'][1])
    # A curator who sets collision_domain deliberately is not drift — the escape hatch
    # must work, or the gate becomes unusable on any exam with an override.
    globals()['_OVERRIDES'] = {
        'template_sets': None, 'template_sets_by_section': {}, 'subtopic_merges': [],
        'subtopic_overrides': {_driftent[0]['subtopic_id']:
                               {'collision_domain': 'zz_drifted'}}}
    try:
        with _cl7.redirect_stdout(_w7buf):
            _okres = run_qv(_cp7.deepcopy(_driftent), {}, {'_meta': {}})
    finally:
        globals()['_OVERRIDES'] = None
    check('qv13_honours_explicit_collision_domain_override',
          _okres['QV-13'][0] == 'PASS')

    # L4632 — and when form_keys DO collide, the report names the colliding ids. Four
    # of them, out of alphabetical order. stamp_mechanic_axes hard-stops on a collision,
    # so these entries are hand-built: run_qv is the second line of defence and has to
    # be tested as one.
    _colent = [{'subtopic_id': _i7c, 'section': 'Real Analysis', 'topic': 'T',
                'subtopic': f'Sub {_k7c}', 'sub_type_label': f'Sub {_k7c}',
                'format': 'TEXT', 'observed_count': 2, 'concept_group': 'c',
                'PYQ_STEM_PATTERNS': [{'id': 'P1', 'template': f'template number {_k7c}',
                                       'years': [2024], 'frequency': 100, 'raw_count': 3,
                                       'confidence': 'observed', 'deprecated': False}],
                'form_key': 'dup_key', 'collision_domain': 'ra',
                'question_mechanic': 'm', 'axis2_capability': {},
                'observed_axis2': {}, 'presentation_family': 'x'}
               for _k7c, _i7c in enumerate(['ra.zeta', 'ra.alpha', 'ra.mid', 'ra.beta'])]
    globals()['_OVERRIDES'] = None
    with _cl7.redirect_stdout(_w7buf):
        _colres = run_qv(_colent, {}, {'_meta': {}})
    check('qv13_collision_report_lists_ids_alphabetically',
          _colres['QV-13'][0] == 'FAIL' and
          "collisions=[\"ra:dup_key=['ra.alpha', 'ra.beta', 'ra.mid', 'ra.zeta']\"]"
          in _colres['QV-13'][1])

    # ── export surface: the stubs in the spec import exactly these ────────
    check('all_exports_exist', all(n in globals() for n in __all__))

    # ── META-ASSERTION: THE SUITE MUST HAVE RUN EVERY CHECK IT CONTAINS ──────
    # GAP-2026-08-17-B4-ENV-SKEW. A conditional assertion that silently skips is
    # indistinguishable, in the printed result, from one that passed: the suite said
    # "0 failed" in both environments while running 65 checks here and 64 in CI. The
    # mutation gate then measured two different budgets and CI went red on a release
    # whose code was correct.
    # A count is the cheapest possible oracle for "did anything vanish?". If a future
    # edit adds or removes an assertion, update this number DELIBERATELY — that edit is
    # then visible in the diff, which is the whole point.
    # ═══ v2.56 STYLE-FIDELITY LAYER (GAP-2026-08-29-STYLE-FIDELITY §9.2) ═══════
    # 49 checks. Every mechanic has a positive fixture whose expected value is the
    # exact class (a mis-fire of ANY earlier rule fails the equality, so each
    # positive fixture is simultaneously the negative control for every rule above
    # it). Three unknown fixtures prove unknown is REPORTED, never defaulted.

    def _mq(stem, opts=None, **kw):
        q = {'stem': stem, 'options': opts or [], 'medium': 'en'}
        q.update(kw)
        return q

    _mfix = [
        ('data_sufficiency', _mq('What is the value of x?',
            ['Statement I alone is sufficient to answer',
             'Statement II alone is sufficient to answer',
             'Both statements together are needed', 'Neither is sufficient'])),
        ('assertion_reason', _mq('Assertion (A): the sky appears blue. '
            'Reason (R): shorter wavelengths scatter more.',
            ['Both A and R are true', 'A is true, R is false',
             'A is false', 'Both false'], axis2='ASSERTION_REASON')),
        ('match', _mq('Match List-I with List-II and select the correct answer.',
            ['A-1, B-2', 'A-2, B-1', 'A-3, B-4', 'A-4, B-3'], axis2='MATCH')),
        ('syllogism', _mq('Statements: All pens are books. Some books are cars. '
            'Conclusions: I. Some cars are pens. II. No car is a pen.',
            ['Only conclusion I follows', 'Only conclusion II follows',
             'Either I or II follows', 'Neither follows'])),
        ('decode', _mq('In a certain code language, PLANT is written as QMBOU. '
            'How will CHAIR be written in that code?',
            ['DIBJS', 'SJBID', 'DIBJT', 'DJBIS'])),
        ('constraint_arrangement', _mq('Six friends Anil, Binod, Chetan, Dinesh, '
            'Ekta and Farhan are sitting in a row facing north. Anil sits to the '
            'left of Binod. Chetan sits between Dinesh and Ekta. Who sits at the '
            'right end?', ['Anil', 'Binod', 'Ekta', 'Farhan'])),
        ('procedure_trace', _mq('Step 1: multiply the input by 2. Step 2: add 3 '
            'to the result. Input: 5. What is the output?',
            ['10', '13', '16', '11'])),
        ('text_reorder', _mq('Arrange the parts P. went to Q. the market R. she '
            'S. early in the morning to form a meaningful sentence.',
            ['RPQS', 'PQRS', 'RSPQ', 'QPRS'])),
        ('sentence_edit', _mq('Improve the underlined part of the sentence: She '
            'go to school daily.', ['goes to', 'gone to', 'going to', 'No error'])),
        ('word_meaning', _mq('Choose the synonym of the word ABUNDANT.',
            ['plentiful', 'scarce', 'meagre', 'hollow'])),
        ('passage_comprehension', _mq('According to the passage, the author\u2019s '
            'tone can best be described as',
            ['critical', 'laudatory', 'neutral', 'ironic'], linked_group_id='LG1',
            axis2='LINKED')),
        ('spatial_figure', _mq('Which of the answer figures is the exact mirror '
            'image of the problem figure?', ['', '', '', ''],
            image_role='options_only', axis1='FIGURAL')),
        ('series_completion', _mq('Find the next term in the series 3, 7, 15, 31, ?',
            ['63', '62', '65', '57'])),
        ('pattern_analogy', _mq('Dog is related to Kennel in the same way as '
            'Horse is related to:', ['Stable', 'Field', 'Cart', 'Saddle'])),
        ('relational_reasoning', _mq('Pointing to a photograph, Ram said, '
            '\u201cShe is the daughter of my father\u2019s only son.\u201d How is '
            'the lady in the photograph related to Ram?',
            ['Daughter', 'Sister', 'Niece', 'Mother'])),
        ('interpret_data', _mq('According to the data in the table, what is the '
            'average production over the given period?',
            ['120', '135', '150', '145'], axis1='DI')),
        ('apply_rule_to_case', _mq('Amit agreed to sell his car to Bharat for a '
            'stated sum. Bharat paid the advance but Amit refused to deliver the '
            'car. What is the legal position?',
            ['The agreement is void', 'Bharat may claim specific performance',
             'Amit may keep the advance', 'The contract never existed'])),
        ('evaluate_statements', _mq('Consider the following statements: '
            'I. The process conserves energy. II. The process is reversible. '
            'Which of the above is/are correct?',
            ['I only', 'II only', 'Both I and II', 'Neither I nor II'])),
        ('rank_order', _mq('Arrange the following in increasing order of strength.',
            ['P < Q < R', 'Q < P < R', 'R < Q < P', 'P < R < Q'])),
        ('identify', _mq('Identify the major product of the reaction shown below.',
            ['', '', '', ''], image_role='options_only', option_shape='structure_image')),
        ('predict', _mq('What happens when the metal piece is dropped into cold '
            'water?', ['It floats quietly', 'A vigorous reaction occurs',
             'Nothing is observed', 'It dissolves slowly'])),
        ('multi_step_derivation', _mq('A gas expands from 2.0 L to 6.0 L at 300 K '
            'against a constant external pressure of 1.0 atm, and the heat '
            'absorbed is 500 J. Hence calculate the change in internal energy.',
            [])),
        ('single_formula', _mq('The half-life of a first-order reaction is 20 s. '
            'What is the rate constant?', [])),
        ('recall', _mq('Which of the following has the greatest electronegativity?',
            ['F', 'O', 'N', 'Cl'])),
    ]
    for _name, _q in _mfix:
        if 'axis2' not in _q:
            _q['axis2'] = classify_axis2(_q)
        if 'option_shape' not in _q:
            _q['option_shape'] = detect_option_shape(_q)
        check('mechanic_' + _name, detect_mechanic(_q) == _name)

    _u1 = _mq('Consider the object shown alongside carefully.', ['A1', 'B2', 'C3', 'D4'])
    _u2 = _mq('The value 42 was recorded near the site during the survey.',
              ['Alpha site', 'Beta site', 'Gamma site', 'Delta site'])
    _u3 = _mq('Choose wisely from the given items.',
              ['Item pack 4', 'Item pack 7', 'Item pack 9', 'Item pack 2'])
    for _i, _q in enumerate((_u1, _u2, _u3), 1):
        _q['axis2'] = classify_axis2(_q)
        _q['option_shape'] = detect_option_shape(_q)
        check(f'mechanic_unknown_{_i}',
              detect_mechanic(_q) == bc.MECHANIC_UNKNOWN)

    # P-1 — no cue token is a subject word (fixture-enforced, both corpora classes)
    _subject_probe = ('acid base enzyme oxide tort ledger voltage plateau dynasty '
                      'integral photosynthesis contract momentum entropy verb '
                      'treaty mitosis polymer isotope theorem statute').split()
    _all_cues = []
    for _spec in bc.MECHANIC_CUES.values():
        for _v in _spec.values():
            if isinstance(_v, (list, tuple)):
                _all_cues += [str(x).lower() for x in _v]
    # a cue set that is EMPTY would pass this vacuously
    check('P1_probe_sets_non_empty', len(_all_cues) >= 40 and len(_subject_probe) >= 15)
    check('P1_cues_not_subject_words',
          not any(w in c.split() or w == c for c in _all_cues for w in _subject_probe))

    # E-10 — the number tokeniser: digit runs can never concatenate
    _tok_text = 'Mix 25 mL of a 0.2 M solution with 1,000 mL of water at 298 K.'
    _toks = [m.group(0) for m in _NUMBER_TOKEN_RE.finditer(_tok_text)]
    _vals = [float(re.match(r'-?[\d,]+(?:\.\d+)?', t).group(0).replace(',', ''))
             for t in _toks]
    check('E10_number_tokeniser',
          25.0 in _vals and 0.2 in _vals and 1000.0 in _vals and 298.0 in _vals
          and 250.2 not in _vals and len(_vals) == 4)

    # E-13 — OMML text in DOCUMENT ORDER via corpus_io.text_of
    class _FEl:
        def __init__(self, tag, text=None, kids=()):
            self.tag, self.text, self._kids = tag, text, list(kids)
        def iter(self):
            yield self
            for k in self._kids:
                for e in k.iter():
                    yield e
    _p = _FEl('p', kids=[_FEl(corpus_io.W_T_TAG, 'rate = '),
                         _FEl(corpus_io.M_T_TAG, 'k[A]'),
                         _FEl(corpus_io.W_T_TAG, ' at temperature T')])
    check('E13_omml_in_order',
          corpus_io.text_of(_p) == 'rate = k[A] at temperature T')

    # §6.1.1 — caps convention on three mini-corpora
    _cw = [_mq('In a certain code, PLANT is written as QMBOU.', ['X', 'Y']),
           _mq('In that code, CHAIR is written as DIBJS.', ['X', 'Y'])]
    _em = [_mq('Which of the following is NOT correct?', ['a', 'b']),
           _mq('Which statement is TRUE for the sample?', ['a', 'b'])]
    _ac = [_mq('The IUPAC name of the compound is required.', ['a', 'b']),
           _mq('The NMR spectrum shows two signals.', ['a', 'b'])]
    check('sig_caps_convention',
          derive_content_signature(_cw)['caps_convention'] == 'codeword'
          and derive_content_signature(_em)['caps_convention'] == 'emphasis'
          and derive_content_signature(_ac)['caps_convention'] == 'acronym')

    # §6.1.1 — emphasis tokens survive the legacy reasoning mask
    _sig_cw = derive_content_signature(_cw)
    _m = strip_variables_v2('Which coded WORD is NOT the answer?', _sig_cw, 'reasoning')
    check('emphasis_never_masked',
          _m == 'Which coded _WORD_ is NOT the answer?')

    # §6.1.1 — content masking: numbers masked with unit kept, notation retained
    _sig_ct = {'proper_noun_rate': 0.06, 'caps_convention': 'emphasis',
               'notation_density': 0.4}
    _m2 = strip_variables_v2('Benzene reacts with 25 mL of HNO3 at 298 K.',
                             _sig_ct, None)
    check('strip_v2_content', '_NUM_ mL' in _m2 and '_NUM_ K' in _m2
          and 'Benzene' in _m2 and 'HNO3' in _m2)

    # Q9 — legacy_mode derived FROM the signature, never a name
    def _s(**kw):
        base = {'numeric_density': 0, 'notation_density': 0, 'caps_convention': 'none',
                'proper_noun_rate': 0, 'blank_rate': 0, 'statement_rate': 0,
                'polarity_rate': 0, 'stimulus_rate': 0, 'label_scheme': 'mixed',
                'medium_en_share': 1.0}
        base.update(kw)
        return base
    _apt = 'aptitude'
    _qq = [_mq('A shopkeeper sold the article for 240 after a discount and made '
               'a profit of 20 in the sale.', ['200', '220', '210', '190'])]
    _qw = [_mq('Choose the synonym of the word ABUNDANT.',
               ['plentiful', 'scarce', 'meagre', 'hollow'])]      # word_meaning
    _qc = [_mq('In a certain code language, PLANT is written as QMBOU. How will '
               'CHAIR be written?', ['DIBJS', 'SJBID', 'DIBJT', 'DJBIS'])]  # decode
    _qs2 = [_mq('Statements: All pens are books. Some books are cars. '
                'Conclusions: I. Some cars are pens. II. No car is a pen.',
                ['Only conclusion I follows', 'Only conclusion II follows',
                 'Either I or II follows', 'Neither follows'])]   # syllogism
    for _fl in (_qw, _qc, _qs2, _qq):
        for _fq in _fl:
            _fq['axis2'] = classify_axis2(_fq)
            _fq['option_shape'] = detect_option_shape(_fq)
    check('legacy_mode_derivation',
          derive_legacy_mode(_s(), questions=_qq, exam_class=_apt) == 'quantitative'
          and derive_legacy_mode(_s(), questions=_qw, exam_class=_apt) == 'english'
          and derive_legacy_mode(_s(), questions=_qc, exam_class=_apt) == 'reasoning'
          and derive_legacy_mode(_s(), questions=_qs2, exam_class=_apt) == 'logical'
          and derive_legacy_mode(_s(proper_noun_rate=0.15), exam_class=_apt) == 'factual'
          and derive_legacy_mode(_s(notation_density=0.5)) is None
          and derive_legacy_mode(_s(), exam_class='content') is None)
    # exam class: NAT/MSQ or aggregate notation => content; cache on progress
    _pc = {'_meta': {'nat_allowed': True}}
    check('exam_style_class',
          _exam_style_class(_pc) == 'content'
          and '_style_exam_class' not in _pc          # progress NEVER mutated
          and _exam_style_class({'_meta': {}}, _s()) == 'aptitude')

    # §6.1.4 — the three new Axis-2 classes
    _qi = _mq('Identify the major product of the reaction shown below.',
              ['', '', '', ''], image_role='options_only',
              option_shape='structure_image')
    _qp = _mq('Which graph shows the variation of rate with concentration?',
              ['', '', '', ''], image_role='options_only', option_shape='plot_image')
    _qr = _mq('The increasing order of acidic strength is:',
              ['P < Q < R', 'Q < P < R', 'R < P < Q', 'R < Q < P'])
    check('axis2_new_classes',
          classify_axis2(_qi) == 'IDENTIFY' and classify_axis2(_qp) == 'SELECT_PLOT'
          and classify_axis2(_qr) == 'RANK')

    # §6.1.4 — widened STATEMENT fires on a labelled Statement block (E-7 family)
    _qs = _mq('Statement (A): the enthalpy change is negative. Statement (B): '
              'the entropy of the system decreases.',
              ['Both statements are correct and B explains A',
               'Both are correct but B does not explain A',
               'A is correct, B is incorrect', 'A is incorrect, B is correct'])
    check('axis2_statement_widened', classify_axis2(_qs) == 'STATEMENT')

    # §6.1.4 — v2.55 classifications preserved (ladder above the insertions)
    _qm = _mq('Match List-I with List-II.', ['A-1, B-2', 'A-2, B-1', 'A-3', 'A-4'])
    _qsq = _mq('Arrange the following steps in the correct sequence.',
               ['1,2,3,4', '2,1,4,3', '4,3,2,1', '1,3,2,4'])
    _qd = _mq('What is the capital of France?', ['Paris', 'Lyon', 'Nice', 'Lille'])
    _qf = _mq('The reagent used here is ____.', ['HCl', 'NaOH', 'KMnO4', 'H2O'],
              blank_pos='middle')
    check('axis2_v255_preserved',
          classify_axis2(_qm) == 'MATCH' and classify_axis2(_qsq) == 'SEQUENCE'
          and classify_axis2(_qd) == 'DIRECT' and classify_axis2(_qf) == 'FILL_BLANK')

    # §6.1.2 — option shapes (representative set across the 15-value vocabulary)
    check('option_shapes',
          detect_option_shape(_mq('x', ['4', '6', '8', '12'])) == 'value'
          and detect_option_shape(_mq('x', ['4 mL', '6 mL', '8 mL', '12 mL'])) == 'value_with_unit'
          and detect_option_shape(_mq('x', ['E = mc^2', 'E = mc', 'E = m/c', 'E = c^2'])) == 'expression'
          and detect_option_shape(_mq('x', ['I only', 'II only', 'Both I and II',
                                            'Neither I nor II'])) == 'combination_label'
          and detect_option_shape(_mq('x', ['A-1, B-2', 'A-2, B-1', 'A-3, B-4',
                                            'A-4, B-3'])) == 'pair_map'
          and detect_option_shape(_mq('x', ['plentiful', 'scarce', 'meagre',
                                            'hollow'])) == 'word'
          and detect_option_shape(_mq('x', [])) == 'none'
          and detect_option_shape(_mq('x', ['', '', '', '']),
                                  {'object_types': {'dominant': ['molecular structures']}}
                                  ) == 'structure_image')

    # §6.1.2 — pattern keys: two phrasings of one abstraction land in ONE key
    _pk_qs = []
    for _i in range(3):
        _pk_qs.append(_mq(f'In a certain code language, WORD{_i} is written as '
                          f'XPSE{_i}. How is GATE written?',
                          ['HBUF', 'FZSD', 'HBUE', 'GZSD'], year=2024,
                          mechanic='decode', axis2='DIRECT',
                          skeleton='In a certain code language, _WORD_ is written '
                                   'as _WORD_. How is _WORD_ written?'))
    for _i in range(3):
        _pk_qs.append(_mq(f'If TREE{_i} = USFF{_i} in a code, then LEAF equals',
                          ['MFBG', 'KDZE', 'MFBH', 'KEBG'], year=2025,
                          mechanic='decode', axis2='DIRECT',
                          skeleton='If _WORD_ = _WORD_ in a code, then _WORD_ equals'))
    _pks = build_pattern_keys(_pk_qs)
    check('pattern_keys_group',
          len(_pks) == 1 and _pks[0]['raw_count'] == 6
          and _pks[0]['confidence'].startswith('observed')
          and len(_pks[0]['exemplars']) == 2)

    # EC-9 — low-entropy detection
    _le = [_mq(f'In a certain code language, ALPHA{_i} is written as BETA{_i}. '
               f'How is GAMMA{_i} written in that code language today?')
           for _i in range(5)]
    _he = [_mq('The rate constant doubles when temperature rises by ten kelvin.'),
           _mq('Aromatic rings undergo electrophilic substitution readily.'),
           _mq('The crystal field splitting depends on the ligand strength.'),
           _mq('Entropy increases in every spontaneous isolated process.')]
    check('low_entropy', detect_low_entropy(_le) is True
          and detect_low_entropy(_he) is False)

    # §6.3 — normaliser, shingles, duplicate_of (EC-38) and short-stem shingles
    _dupA = _mq('Q.5 The rate of the reaction doubles when the temperature is '
                'raised from 298 K to 308 K. Calculate the activation energy.',
                ['50 kJ', '53 kJ', '58 kJ', '60 kJ'], paper_id='P1', num=5)
    _dupB = dict(_dupA, paper_id='P2', num=17,
                 stem='Q.17 ' + _dupA['stem'][4:])
    _recs, _nu = build_pyq_index_questions([_dupA, _dupB])
    check('pyq_index_duplicate_of',
          _nu == 1 and _recs[0]['duplicate_of'] is None
          and _recs[1]['duplicate_of'] == 'P1:5'
          and _recs[0]['stem_md5'] == _recs[1]['stem_md5'])
    _short = _mq('Define lattice enthalpy.', [], paper_id='P1', num=1)
    _srec, _ = build_pyq_index_questions([_short])
    check('pyq_index_short_stem_k4', _srec[0]['stem_shingles_4'] is not None)

    # §6.1.9 — medium split
    check('medium_of', medium_of('The quick brown fox jumps') == 'en'
          and medium_of('\u092f\u0939 \u092a\u094d\u0930\u0936\u094d\u0928 '
                        '\u0939\u093f\u0902\u0926\u0940 \u092e\u0947\u0902 '
                        '\u0939\u0948') == 'other'
          and medium_of('25 + 17 = ?') == 'en')

    # §6.5.2 — style distance: identity 0, disjoint mixes > 0, bounded
    _cA = {'mechanic_mix': {'recall': 0.6, 'identify': 0.4},
           'stem_len': {'p50': 22}, 'polarity_rate': 0.1}
    _cB = {'mechanic_mix': {'decode': 1.0}, 'stem_len': {'p50': 44},
           'polarity_rate': 0.6}
    check('style_distance',
          bc.style_distance(_cA, _cA) == 0.0
          and 0.0 < bc.style_distance(_cA, _cB) <= 1.0)

    # §6.1.3 — numeric distractor mining
    _mn = _mq('Calculate the work done in the process described above.',
              ['20 J', '-20 J', '2000 J', '35 J'], key='1',
              option_shape='value_with_unit')
    _mix = mine_distractor_mechanisms([_mn])
    check('mine_numeric',
          isinstance(_mix, dict) and _mix.get('sign_error', 0) > 0
          and _mix.get('order_of_magnitude', 0) > 0
          and _mix.get('near_miss', 0) > 0)
    check('mine_unavailable',
          mine_distractor_mechanisms([_mq('x', ['1', '2', '3', '4'])]) == 'unavailable')

    # §6.2 — item rules measured, suspension at >= 0.10
    _ir_qs = [_mq(f'Question number {_i} asks about the topic.',
                  ['Alpha', 'Beta', 'Gamma', 'None of the above'], key='1')
              for _i in range(25)]
    _ir = measure_item_rules(_ir_qs)
    check('item_rules_measured',
          _ir['I-2']['measured'] and _ir['I-2']['violation_rate'] == 1.0
          and _ir['I-2']['suspended'] is True
          and _ir['I-3']['measured'] and _ir['I-3']['suspended'] is True
          and _ir['I-1']['measured'] is True)

    # §6.2 / §6.3 — writers: activation gating + shared corpus_hash round-trip
    import tempfile as _tf
    with _tf.TemporaryDirectory() as _td:
        _wq = []
        for _i in range(30):
            _wq.append(_mq(f'The measured value in trial {_i} was 25 mL exactly.',
                           ['20 mL', '25 mL', '30 mL', '35 mL'],
                           paper_id='P1' if _i < 15 else 'P2', num=_i + 1,
                           year=2024, mechanic='recall', axis2='DIRECT',
                           option_shape='value_with_unit'))
        _ent = [{'section': 'S', 'topic': 'T', 'subtopic': 'U', 'subtopic_id': 'S-U'}]
        _qmap = {('S', 'T', 'U'): _wq}
        _pp, _prof = write_style_profile('ZZSTYLE', _ent, _qmap, ['P1', 'P2'],
                                         ['h1', 'h2'], out_dir=_td)
        _ip, _idx = write_pyq_index('ZZSTYLE', _wq, ['h2', 'h1'], out_dir=_td)
        check('style_profile_activation',
              _prof['activation']['status'] == 'DORMANT'
              and 'papers=2' in (_prof['activation']['dormant_reason'] or ''))
        # window metadata travels with the artefact (EC-4/EC-10): a consumer
        # scoring a real sitting must know which sittings fed the cells
        # §6.2 SUBTOPIC SCHEMA — the writing side (Create S3-12c) reads exactly
        # these three from a subtopic cell; a cell without them starves the brief
        # silently. Pinned so the schema cannot regress to "looks valid, feeds
        # nothing".
        _sc = _prof['subtopics'][_ent[0]['subtopic_id']]
        check('style_subtopic_schema_complete',
              set(('number_ranges', 'exclude_values', 'distractor_mix',
                   'low_entropy', 'mechanic_mix', 'form_mix', 'polarity_rate',
                   'n')) <= set(_sc))
        # keys ABSENT for these fixtures -> 'unavailable', never a guess (EC-12)
        # SAME MEASUREMENT, SAME INPUTS across artefacts: the profile mines the
        # subtopic's distractor mix with the SAME figural descriptor
        # section_rules used, derived from the entry itself so a caller cannot
        # forget it. Without this the two artefacts publish different mixes for
        # an image-option subtopic and the writing side reads whichever it holds.
        _fig_ent = [dict(_ent[0], PYQ_IMAGE_ANALYSIS={
            'object_types': {'dominant': ['graph', 'curve']}})]
        _fig_qs = [_mq('Which of the following best represents the behaviour?',
                       ['', '', '', ''], paper_id='F1', num=_i + 1, year=2024)
                   for _i in range(14)]
        _, _fprof = write_style_profile('ZZFIG', _fig_ent,
                                        {('S', 'T', 'U'): _fig_qs}, ['F1'],
                                        ['hf'], out_dir=_td)
        _fcell = _fprof['subtopics']['S-U']
        check('style_profile_uses_entry_figural_descriptor',
              _fcell['option_shape_mix'].get('plot_image', 0) > 0
              and 'figure' not in _fcell['option_shape_mix'])
        # and an explicit override still wins
        _, _fprof2 = write_style_profile(
            'ZZFIG2', _fig_ent, {('S', 'T', 'U'): _fig_qs}, ['F1'], ['hf2'],
            out_dir=_td,
            figural_by_subtopic={'S-U': {'object_types': {'dominant': ['structure']}}})
        check('style_profile_figural_override_honoured',
              _fprof2['subtopics']['S-U']['option_shape_mix'].get('structure_image', 0) > 0)

        check('style_subtopic_distractor_unavailable_without_keys',
              _sc['distractor_mix'] == 'unavailable')
        # APTITUDE-class subtopic: the legacy extractor governs, exactly as
        # v2.55 did — 'reasoning' mode has no ranges, so None is CORRECT and
        # must not be "fixed" into a v2 measurement (that would be an EC-26
        # regression on every aptitude exam).
        check('style_subtopic_ranges_legacy_path_none',
              _sc['number_ranges'] is None and _sc['exclude_values'] == [])
        # CONTENT-class subtopic: notation present => v2 measurement, so the
        # brief actually receives ranges and exclude_values.
        _cq = [_mq('At 300 K the enthalpy change was 25 kJ mol^-1 in trial %d.' % _i,
                   ['20 kJ', '25 kJ', '30 kJ', '35 kJ'],
                   paper_id='C1' if _i < 15 else 'C2', num=_i + 1, year=2024,
                   mechanic='single_formula', axis2='DIRECT',
                   option_shape='value_with_unit') for _i in range(30)]
        for _q in _cq:
            _q['omml_present'] = True          # notation => content class
        _, _cprof = write_style_profile('ZZCONTENT', _ent,
                                        {('S', 'T', 'U'): _cq}, ['C1', 'C2'],
                                        ['hc1', 'hc2'], out_dir=_td)
        _csc = _cprof['subtopics']['S-U']
        check('style_subtopic_ranges_content_path_measured',
              isinstance(_csc['number_ranges'], dict) and _csc['number_ranges']
              and all(set(v) >= {'min', 'max', 'p10', 'p50', 'p90', 'n'}
                      for v in _csc['number_ranges'].values()))
        check('style_subtopic_exclude_values_content_path',
              isinstance(_csc['exclude_values'], list)
              and 300.0 in _csc['exclude_values'])   # in >=20% of stems
        # §6.2 SECTION SCHEMA — S3-12c reads status/cell to decide EC-3
        # fallback; without them every slot silently used the paper cell.
        _sec = _prof['sections']['S']
        # §6.2/§6.3 SCHEMA CONFORMANCE — every field the GAP promises must be
        # present in the artefact, whether or not today's code reads it. A
        # promised-but-absent field is a contract a future consumer will trust.
        _CELL_REQ = ('n','status','dormant_reason','mechanic_mix','form_mix',
                     'option_shape_mix','polarity_rate','nat_rate','msq_rate',
                     'stem_len','option_len','option_count','ask_forms',
                     'instruction_phrases','openers','lexicon','notation',
                     'distractor_mix','stimulus_stats')
        check('style_cell_schema_complete', set(_CELL_REQ) <= set(_prof['paper']))
        # distractor_mix is on EVERY cell (paper + section), because EC-3 makes
        # the paper cell a real fallback source for briefs
        check('style_cell_distractor_mix_everywhere',
              'distractor_mix' in _prof['paper']
              and 'distractor_mix' in _prof['sections']['S']['cell']
              and _prof['paper']['distractor_mix'] == 'unavailable')
        _IDX_REQ = ('schema','exam_code','generated_by','generated_at',
                    'corpus_hash','n_questions','n_unique_stems','normaliser',
                    'shingle_k')
        check('pyq_index_meta_schema_complete', set(_IDX_REQ) <= set(_idx['_meta']))
        # EC-25 needs a comparable timestamp on BOTH artefacts
        check('artefacts_carry_generated_at',
              _prof['_meta']['generated_at'].endswith('Z')
              and _idx['_meta']['generated_at'].endswith('Z'))
        # legacy count keys kept alongside the schema names (additive, P-9)
        check('pyq_index_legacy_count_keys_kept',
              _idx['_meta']['questions'] == _idx['_meta']['n_questions']
              and _idx['_meta']['unique_stems'] == _idx['_meta']['n_unique_stems'])
        check('style_section_schema_complete',
              set(('status', 'n', 'cell', 'by_answer_type', 'by_medium')) <= set(_sec)
              and isinstance(_sec['cell'], dict)
              and set(('mechanic_mix', 'form_mix', 'polarity_rate')) <= set(_sec['cell'])
              and _sec['n'] == 30)
        # 30 english questions < min_questions_section (default) => DORMANT with
        # a stated reason, which is what makes EC-3 fall back HONESTLY
        check('style_section_status_reflects_activation',
              _sec['status'] in ('ACTIVE', 'DORMANT')
              and (_sec['status'] == 'ACTIVE' or _sec['dormant_reason']))
        # ORDER-INDEPENDENCE (metamorphic, 2026-08-31). The SAME corpus read in
        # a different question order must produce the SAME profile. Counter's
        # most_common ties on insertion order, so before _top_n the published
        # ask_forms / openers / instruction_phrases / lexicon lists changed with
        # extraction order — and those lists are consumed by S3-12c briefs and
        # G-STYLE component 4.
        import random as _rnd_mt
        _mt_qs = []
        for _i in range(40):
            # two n-grams deliberately TIED so the cut is decided by tie-break
            _txt = ('Determine the order of reaction here.' if _i % 2 else
                    'Determine the species of reaction here.')
            _mt_qs.append(_mq(_txt, ['a', 'b', 'c', 'd'],
                              paper_id='P%d' % (_i % 2), num=_i + 1, year=2024,
                              mechanic='recall', axis2='DIRECT',
                              option_shape='word'))
        _mt_a = compute_style_cell(_mt_qs)
        _mt_b = compute_style_cell(_rnd_mt.Random(9).sample(_mt_qs, len(_mt_qs)))
        check('style_cell_order_independent',
              _mt_a['lexicon'] == _mt_b['lexicon']
              and _mt_a['ask_forms'] == _mt_b['ask_forms']
              and _mt_a['openers'] == _mt_b['openers']
              and _mt_a['instruction_phrases'] == _mt_b['instruction_phrases'])
        # and the tie-break is BY KEY, so it is stable and inspectable
        from collections import Counter as _C_mt
        check('top_n_ties_broken_by_key',
              _top_n(_C_mt({'b': 2, 'a': 2, 'c': 5}), 3)
              == [('c', 5), ('a', 2), ('b', 2)])
        check('style_profile_window_meta',
              _prof['_meta']['papers_in_window'] == ['P1', 'P2']
              and _prof['_meta']['window_years'] == [2024]
              and _prof['_meta']['papers_excluded_from_style'] == [])
        # thresholds carry the dispersion FLAG even on the default path, because
        # G-STYLE gates components 4-5 on it
        check('style_thresholds_flag_present',
              _prof['thresholds']['computed_from_dispersion'] is False
              and _prof['thresholds']['style_distance_fail'] == 0.40)
        # DISPERSION FLOOR (§6.2 / §8.6 proof 3): a TIGHT corpus — every paper
        # stylistically identical, so leave-one-out distances are ~0 — must NOT
        # compute a fail band narrower than the framework default. Before the
        # floor this produced fail=0.354 on the real JAM corpus and three real
        # sittings scored HIGH against their own exam.
        #
        # ADVERSARIAL BY CONSTRUCTION: paper ids arrive in NON-alphabetical
        # encounter order and the years are a set whose iteration order is not
        # its sorted order, so an emission that forgets to sort is DETECTED
        # rather than accidentally correct (the whole class of defect the
        # mutation gate exists to find).
        _pids = ['T3', 'T1', 'T0', 'T2', 'T6', 'T4', 'T5']
        _yrs4 = [2013, 2005, 2024, 2016]
        _tq = []
        for _i in range(420):
            _tq.append(_mq(f'The measured value in trial {_i} was 25 mL exactly.',
                           ['20 mL', '25 mL', '30 mL', '35 mL'],
                           paper_id=_pids[_i % 7], num=(_i % 60) + 1,
                           year=_yrs4[_i % 4], mechanic='recall',
                           axis2='DIRECT', option_shape='value_with_unit'))
        _, _tprof = write_style_profile('ZZTIGHT', _ent,
                                        {('S', 'T', 'U'): _tq}, list(_pids),
                                        ['h%d' % _k for _k in range(7)],
                                        out_dir=_td)
        check('style_threshold_dispersion_floor',
              _tprof['activation']['thresholds_source'].startswith('dispersion')
              and _tprof['thresholds']['style_distance_fail'] >= 0.40
              and _tprof['thresholds']['computed_from_dispersion'] is True
              and abs(_tprof['thresholds']['style_distance_warn']
                      - _tprof['thresholds']['style_distance_fail'] * 0.625) < 1e-6)
        # every artefact list is EMITTED SORTED although its input order is not
        check('style_profile_lists_sorted',
              _tprof['thresholds']['basis_papers'] == sorted(_pids)
              and _tprof['_meta']['papers_in_window'] == sorted(_pids)
              and _tprof['_meta']['window_years'] == sorted(_yrs4)
              and list(_pids) != sorted(_pids)          # the input really is unsorted
              and list({_y for _y in _yrs4}) != sorted(_yrs4))
        # the p95 pick reads a SORTED distance list: with a spread of
        # leave-one-out distances the selected quantile must be the 95th of the
        # ORDER STATISTICS, not of the encounter order.
        _dsp = []
        for _i in range(480):
            _pid = _pids[_i % 7]
            if _pid == 'T3':          # one stylistically alien sitting
                _dsp.append(_mq('Statements: I. %d holds. II. it fails.' % _i,
                                ['Only I', 'Only II', 'Both', 'Neither'],
                                paper_id=_pid, num=(_i % 60) + 1, year=2016,
                                mechanic='evaluate_statements',
                                axis2='STATEMENT', option_shape='statement'))
            else:
                _dsp.append(_mq('The measured value in trial %d was 25 mL.' % _i,
                                ['20 mL', '25 mL', '30 mL', '35 mL'],
                                paper_id=_pid, num=(_i % 60) + 1, year=2016,
                                mechanic='recall', axis2='DIRECT',
                                option_shape='value_with_unit'))
        _, _dprof = write_style_profile('ZZSPREAD', _ent,
                                        {('S', 'T', 'U'): _dsp}, list(_pids),
                                        ['g%d' % _k for _k in range(7)],
                                        out_dir=_td)
        # ONE alien sitting (T3) sits far from the rest, so the 95th ORDER
        # STATISTIC is that outlier (~0.62) while the same INDEX of the
        # unsorted encounter list is an ordinary sitting (~0.18). Asserting only
        # '>= 0.40' would pass either way — the floor would supply 0.40 — so the
        # assertion pins the outlier value itself.
        check('style_threshold_p95_from_order_statistics',
              _dprof['thresholds']['computed_from_dispersion'] is True
              and _dprof['thresholds']['style_distance_fail'] > 0.55
              and _dprof['thresholds']['basis_papers'] == sorted(_pids)
              and abs(_dprof['thresholds']['style_distance_warn']
                      - _dprof['thresholds']['style_distance_fail'] * 0.625) < 1e-4)
              # (1e-4, not 1e-6: both values are STORED rounded to 4 dp, so a
              #  tighter tolerance tests the rounding, not the relationship.)
        # the dispersion BASIS is recorded, and a partial sitting is not in it
        _tq2 = list(_tq) + [_mq('Tiny fragment sitting question %d.' % _j,
                                ['a', 'b', 'c', 'd'], paper_id='TINY',
                                num=_j + 1, year=2024, mechanic='recall',
                                axis2='DIRECT', option_shape='word')
                            for _j in range(9)]
        _, _bprof = write_style_profile('ZZBASIS', _ent,
                                        {('S', 'T', 'U'): _tq2},
                                        list(_pids) + ['TINY'],
                                        ['h%d' % _k for _k in range(8)],
                                        out_dir=_td)
        check('style_threshold_basis_excludes_partial_sitting',
              'TINY' not in _bprof['thresholds']['basis_papers']
              and len(_bprof['thresholds']['basis_papers']) == 7
              and _bprof['thresholds']['basis_min_questions'] == bc.STYLE_DISPERSION_MIN_PAPER_Q)
        check('artefact_corpus_hash_shared',
              _prof['_meta']['corpus_hash'] == _idx['_meta']['corpus_hash']
              and os.path.exists(_pp) and os.path.exists(_ip))
        # universal constants: emitted SORTED although first-seen order (300
        # before 8.314 in every stem) is reverse-sorted
        _uc_qs = [_mq(f'At 300 K with R = 8.314 J, trial {_j} was run.',
                      ['a', 'b'], paper_id='P1', num=_j + 1, year=2024,
                      mechanic='recall', axis2='DIRECT', option_shape='value')
                  for _j in range(20)]
        _, _uprof = write_style_profile('ZZCONST', _ent,
                                        {('S', 'T', 'U'): _uc_qs}, ['P1'],
                                        ['h9'], out_dir=_td)
        check('universal_constants_sorted',
              _uprof['universal_constants'] == [8.314, 300.0])


    # ═══ JAM-measured rule refinements (2026-08-31) ════════════════════════════
    # recall: notation alone never blocks — a stem that only NAMES species
    check('recall_notation_not_blocking',
          detect_mechanic(_mq('The shape of SF\u2084 is:',
                              ['see-saw', 'square planar', 'tetrahedral',
                               'trigonal bipyramidal'])) == 'recall')
    # recall: NAT counting ask with nothing to compute
    check('recall_nat_counting',
          detect_mechanic(_mq('The number of possible isomers for the molecular '
                              'formula C\u2086H\u2081\u2084 is __.', [])) == 'recall')
    # recall NEGATIVE: a compute-ask is never recall even with zero givens
    check('recall_blocks_compute_ask',
          detect_mechanic(_mq('Calculate the lattice enthalpy of the compound '
                              'shown.', [])) != 'recall')
    # identify: copular ask over expression options
    check('identify_copular_expression',
          detect_mechanic(_mq('The complex conjugate of the wavefunction '
                              'N[exp(ikx) + exp(-ikx)] is',
                              ['N*[exp(-ikx) + exp(ikx)]',
                               'N[exp(-ikx) - exp(ikx)]',
                               'N*[exp(ikx) - exp(-ikx)]',
                               '2N cos(kx)'],
                              option_shape='expression')) == 'identify')
    # E-10 comma grouping: a comma-separated list never fuses; thousands still work
    _ct = [m.group(0) for m in _NUMBER_TOKEN_RE.finditer(
        'The series 1000, 0100, 0010, 0001 repeats; add 1,000 mL now.')]
    _cv = [float(re.match(r'-?[\d,]+', t).group(0).replace(',', '')) for t in _ct]
    check('E10_comma_list_never_fuses',
          1000.0 in _cv and 100.0 in _cv and 10.0 in _cv and 1.0 in _cv
          and _cv.count(1000.0) == 2       # the list's 1000 and the 1,000 mL
          and not any(v > 100000 for v in _cv))

    # extract_number_ranges_v2's exclude list is ASCENDING by contract — the
    # profile writer relies on it instead of re-sorting (see write_style_profile).
    _asc_qs = [_mq('Given 300 K and 18 g, trial %d used 25 kJ and 7 mL.' % _i,
                   ['a', 'b']) for _i in range(25)]
    _asc_r, _asc_e = extract_number_ranges_v2(_asc_qs, {'numeric_density': 1.0,
                                                        '_universal_constants': []})
    check('ranges_exclude_ascending',
          _asc_e == sorted(_asc_e) and len(_asc_e) >= 2)

    # ORDERING: the Axis-2 ladder reads option_shape, and only the figural
    # descriptor refines a bare 'figure' into 'plot_image'/'structure_image'.
    # Tagging axes BEFORE stamping the shape made a choose-the-graph question
    # classify DIRECT instead of SELECT_PLOT.
    _fq = {'stem': 'Which of the following best represents the observed behaviour?',
           'options': ['', '', '', ''], 'medium': 'en', 'num': 8}
    _fd = {'object_types': {'dominant': ['graph', 'curve']}}
    check('axis2_needs_figural_shape_first',
          detect_option_shape(_fq) == 'figure'
          and detect_option_shape(_fq, _fd) == 'plot_image')
    _fq_wrong = dict(_fq); tag_axes(_fq_wrong)
    _fq_right = dict(_fq)
    _fq_right['option_shape'] = detect_option_shape(_fq_right, _fd)
    tag_axes(_fq_right)
    check('axis2_select_plot_requires_stamped_shape',
          _fq_wrong.get('axis2') == 'DIRECT'          # the defect, pinned
          and _fq_right.get('axis2') == 'SELECT_PLOT')

    # ranges: a flattened matrix literal (13+ bare digits) is never a magnitude
    _mx_qs = [_mq('The matrices P = 1000010000100001 and volume 25 mL are given '
                  'for trial %d today.' % _j, ['a', 'b']) for _j in range(3)]
    _mrg, _mex = extract_number_ranges_v2(_mx_qs, {'numeric_density': 0.5,
                                                   '_universal_constants': []})
    check('ranges_skip_matrix_literals',
          _mrg is not None and 'mL' in _mrg
          and all(v <= 1e12 for u in _mrg for v in (_mrg[u]['min'], _mrg[u]['max']))
          and all(abs(v) <= 1e12 for v in _mex))

    # E-3 — approach coverage: every mechanic + unknown has a display string
    check('mechanic_to_approach_total',
          set(MECHANIC_TO_APPROACH) == set(bc.MECHANICS) | {bc.MECHANIC_UNKNOWN})


    # ═══ v2.56 MUTATION-KILL PACK (B4 pipeline model — every style-layer decision
    #     point holds a fixture that can tell a broken sort/filter from a working
    #     one; encounter orders are crafted REVERSE of sorted so an unsorted
    #     emission is deterministically visible) ═══════════════════════════════

    # _pctiles: percentiles are positional reads of the SORTED sample
    check('pctiles_sorts_input',
          _pctiles([30, 10, 20])['p50'] == 20
          and _pctiles([3.0, 1.0, 2.0], as_int=False)['p50'] == 2.0)

    # ranges: p50 via _pctiles + exclude_values sorted from encounter order
    _rg_qs = [_mq('The burette delivered 30 mL in the first trial.', ['a', 'b']),
              _mq('The burette delivered 10 mL in the second trial.', ['a', 'b']),
              _mq('The burette delivered 20 mL in the final trial.', ['a', 'b'])]
    _rg, _ex = extract_number_ranges_v2(_rg_qs, {'numeric_density': 0.5,
                                                 '_universal_constants': []})
    check('ranges_v2_p50_and_exclude_sorted',
          _rg['mL']['p50'] == 20.0 and _ex == [10.0, 20.0, 30.0])

    # pattern keys: recency window is the two NEWEST years however questions
    # arrive; emitted years list is sorted
    _rk = []
    for _yr, _mech in ((2025, 'decode'), (2019, 'recall'), (2024, 'decode')):
        for _j in range(3):
            _rk.append(_mq(f'Stem {_yr} {_j} for the recency fixture.',
                           ['a', 'b'], year=_yr, mechanic=_mech, axis2='DIRECT',
                           option_shape='word', skeleton=f'skel {_mech}'))
    _rko = {p['pattern_key'].split('|')[0]: p for p in build_pattern_keys(_rk)}
    check('pattern_keys_recency_sorted',
          _rko['recall']['confidence'] == 'observed'
          and _rko['recall']['deprecated'] is True
          and _rko['decode']['confidence'] == 'observed_recent'
          and _rko['decode']['years'] == [2024, 2025])

    # style cell: every mix/list field is emitted in sorted order even though
    # first-seen order is crafted reverse-sorted
    _sc_qs = [_mq('First stem measured 5 mL at once.',
                  ['w1', 'w2', 'w3', 'w4', 'w5'], mechanic='recall',
                  axis2='STATEMENT', option_shape='word'),
              _mq('Second stem shows Δ then → with 3 K given in a table: '
                  'Col | Col | Col',
                  ['e one', 'e two', 'e three', 'e four'], mechanic='decode',
                  axis2='DIRECT', option_shape='entity',
                  linked_group_id=None)]
    _cell = compute_style_cell(_sc_qs, cell_min=1)
    _sorted_keys = lambda d: list(d.keys()) == sorted(d.keys())
    check('style_cell_mixes_sorted',
          _sorted_keys(_cell['mechanic_mix']) and _sorted_keys(_cell['form_mix'])
          and _sorted_keys(_cell['option_shape_mix'])
          and _sorted_keys(_cell['option_count'])
          and _cell['mechanic_mix'] and list(_cell['mechanic_mix'])[0] == 'decode')

    # style cell: units and conventions sorted from reverse-sorted encounter
    _un_qs = [_mq('Add 5 mL of the titrant slowly to the flask now.', ['a', 'b']),
              _mq('Keep it at 300 K with Δ noted and → marked after ± checks.',
                  ['a', 'b'])]
    _uc = compute_style_cell(_un_qs, cell_min=1)
    check('style_cell_units_convs_sorted',
          _uc['notation']['units'] == ['K', 'mL']
          and _uc['notation']['conventions'] == sorted(_uc['notation']['conventions'])
          and len(_uc['notation']['conventions']) == 3)

    # stimulus stats: kind/position mixes sorted (first-seen: table before passage,
    # inline before before)
    _st_qs = [_mq('Values: Col | Col | Col in the table below now.', ['a', 'b']),
              _mq('Read the passage and answer this linked item.', ['a', 'b'],
                  linked_group_id='G1')]
    _st = compute_style_cell(_st_qs, cell_min=1)['stimulus_stats']
    check('stimulus_mixes_sorted',
          list(_st['stimulus_kind_mix']) == sorted(_st['stimulus_kind_mix'])
          and len(_st['stimulus_kind_mix']) == 2
          and list(_st['stimulus_position']) == sorted(_st['stimulus_position'])
          and len(_st['stimulus_position']) == 2)

    # context pool v2: dominant/common/rare each sorted from reverse encounter
    _cp_qs = []
    for _j in range(5):
        _cp_qs.append(_mq('The towns Zurich and Basel appear with Aarau in every '
                          'survey of Rivera during the audit.', ['a', 'b']))
    _cp = extract_context_pool_v2(_cp_qs, {'proper_noun_rate': 0.10})
    check('context_pool_v2_sorted',
          _cp is not None and _cp['dominant'] == sorted(_cp['dominant'])
          and len(_cp['dominant']) >= 3)

    # context pool v2: common and rare bands sorted too (40-question corpus so the
    # <5% band is reachable; encounter order reverse-sorted inside every band)
    _cb_qs = []
    _cb_qs.append(_mq('The panel met in Zurich for the opening session.',
                      ['a', 'b']))          # rare, first-seen Z (1/40)
    _cb_qs.append(_mq('Observers came from Basel for the closing session.',
                      ['a', 'b']))          # rare, then B (1/40)
    for _j in range(3):
        _cb_qs.append(_mq('The audit ran in Vienna again this spring.',
                          ['a', 'b']))       # common, first-seen V (3/40)
    for _j in range(3):
        _cb_qs.append(_mq('The review ran in Berlin again this autumn.',
                          ['a', 'b']))       # common, then B (3/40)
    for _j in range(32):
        _cb_qs.append(_mq(f'Trial {_j} was recorded without a named observer.',
                          ['a', 'b']))
    _cb = extract_context_pool_v2(_cb_qs, {'proper_noun_rate': 0.05})
    check('context_pool_v2_bands_sorted',
          _cb is not None
          and _cb['rare'] == sorted(_cb['rare']) and len(_cb['rare']) >= 2
          and _cb['common'] == sorted(_cb['common']) and len(_cb['common']) >= 2)

    # pyq index: numeric values are CANONICALLY ordered so reordered stems match
    _ixA = _mq('Mix 5 mL of acid with 3 g of salt for the test.', ['a'],
               paper_id='P1', num=1)
    _ixB = _mq('Mix 3 g of salt with 5 mL of acid for the test.', ['a'],
               paper_id='P2', num=2)
    _ixr, _ = build_pyq_index_questions([_ixA, _ixB])
    check('pyq_index_values_canonical',
          _ixr[0]['values'] == _ixr[1]['values'] and len(_ixr[0]['values']) == 2)

    # item rule I-8: unsorted numeric options are the violation it measures
    _i8 = [_mq(f'Pick the reading for trial {_j} of the run.',
               ['30', '10', '20', '40'], key='1') for _j in range(25)]
    check('item_rule_I8_detects_unsorted',
          measure_item_rules(_i8)['I-8']['violation_rate'] == 1.0
          and measure_item_rules(_i8)['I-8']['suspended'] is True)

    # distractor mining: transposed expression == reversed_relationship
    _tx = _mq('Which expression gives the required quantity here?',
              ['x^2+y', 'y+x^2', 'x^2-y', 'x^2/y'], key='1',
              option_shape='expression')
    check('mine_transposition',
          mine_distractor_mechanisms([_tx]).get('reversed_relationship', 0) > 0)

    # synthesise: additive mechanic_mix sorted; internal '_' keys filtered out of
    # the content_signature entry field while present in the raw signature
    _sy_qs = []
    for _j in range(3):
        _sy_qs.append(_mq(f'Which of the following salts was studied in run {_j}?',
                          ['NaCl', 'KBr', 'CsI', 'LiF'], year=2024, num=_j + 1))
    for _j in range(3):
        _sy_qs.append(_mq(f'In a certain code, WORD{_j} is written as XPSE{_j}. '
                          'How is GATE written in that code?',
                          ['HBUF', 'FZSD', 'HBUE', 'GZSD'], year=2025, num=_j + 4))
    _sy_sig = derive_content_signature(_sy_qs)
    _sy_ent = synthesise_subtopic('SecZ', 'TopZ', 'SubZ', list(_sy_qs), {}, {})
    check('synthesise_style_fields',
          '_style_en_count' in _sy_sig
          and not any(k.startswith('_') for k in _sy_ent['content_signature'])
          and list(_sy_ent['mechanic_mix']) == sorted(_sy_ent['mechanic_mix'])
          and len(_sy_ent['mechanic_mix']) >= 2)

    EXPECTED_CHECKS = 215  # v2.56: +50 style +13 mutation-kill +6 JAM +1 exam-class +3 threshold/window (GAP-2026-08-29-STYLE-FIDELITY §9.2)
    total = passed + len(fails)
    if total != EXPECTED_CHECKS:
        fails.append(
            f'suite_ran_every_check (ran {total}, expected {EXPECTED_CHECKS} — an '
            f'assertion was SKIPPED or added without updating EXPECTED_CHECKS; a '
            f'skipped check is not a passing one)')
    else:
        passed += 1

    print(f"analyse_engine self-test: {passed} passed, {len(fails)} failed"
          + ("  — " + "; ".join(fails) if fails else ""))
    return not fails




# ═══════════════════════════════════════════════════════════════════════════════
# STYLE-FIDELITY LAYER (reading side) — GAP-2026-08-29-STYLE-FIDELITY Rev 2, §6.1.
# Measure, never look up (P-1): nothing below consults a section/topic/subject/exam
# name. Every classification is computed from the question's own tokens, options,
# stimulus and answer type; where it cannot be made the value is 'unknown', counted
# and reported (never a default that reads as a finding).
# ═══════════════════════════════════════════════════════════════════════════════

_SENT_SPLIT_RE   = re.compile(r'(?<=[.!?])\s+')
_LATIN_RE        = re.compile(r'[A-Za-z]')
_LETTER_RE       = re.compile(r'[^\W\d_]', re.UNICODE)
_ALLCAPS_RE      = re.compile(r'\b[A-Z]{2,}\b')
_CAP_TOKEN_RE    = re.compile(r'\b[A-Z][a-z]+\b')
_NOTATION_RE     = re.compile(r'[₀-₉⁰-⁹⁺⁻ⁿ]|[\u0370-\u03FF]|→|⇌|Δ|±|≤|≥|≠|√|∞|'
                              r'×\s*10|x\s*10\^|\^\s*-?\d|_\{|\\frac|°')
_NUMBER_TOKEN_RE = re.compile(r'(?<![A-Za-z0-9_])' + bc.NUMBER_TOKEN_PATTERN)
_POLARITY_RE     = re.compile(r'\b(' + '|'.join(bc.POLARITY_MARKERS) + r')\b',
                              re.IGNORECASE)
_CODEWORD_CUE_RE = re.compile('|'.join(re.escape(c) for c in bc.CODEWORD_CUES),
                              re.IGNORECASE)

# Retained-notation set (§6.1.1): element symbols + unit symbols + named-law tokens
# stay verbatim under _NAME_ masking. This is NOTATION, not a subject-cue table:
# it never classifies anything — it only protects tokens from erasure.
_ELEMENT_SYMBOLS = frozenset((
    'H He Li Be B C N O F Ne Na Mg Al Si P S Cl Ar K Ca Sc Ti V Cr Mn Fe Co Ni '
    'Cu Zn Ga Ge As Se Br Kr Rb Sr Y Zr Nb Mo Tc Ru Rh Pd Ag Cd In Sn Sb Te I '
    'Xe Cs Ba La Ce Pr Nd Pm Sm Eu Gd Tb Dy Ho Er Tm Yb Lu Hf Ta W Re Os Ir Pt '
    'Au Hg Tl Pb Bi Po At Rn Fr Ra Ac Th Pa U Np Pu Am Cm Bk Cf Es Fm Md No Lr '
    'Rf Db Sg Bh Hs Mt Ds Rg Cn Nh Fl Mc Lv Ts Og').split())
_UNIT_WORDS = frozenset(('K J N Pa V A W Hz T C F S Wb H mol cd kg km cm mm nm pm '
                         'mL L atm bar torr eV kJ MJ kPa MPa GHz MHz kHz ppm').split())
_RETAINED_NOTATION = _ELEMENT_SYMBOLS | _UNIT_WORDS


def medium_of(text):
    """§6.1.9 — per-question, two-way: 'en' if >= 80% of letter characters are
    Latin; otherwise 'other'. No script table; no language names. A rendering with
    no letters at all (pure numeric/symbolic) counts as 'en' — generation is
    English and symbols are universal."""
    letters = _LETTER_RE.findall(text or '')
    if not letters:
        return 'en'
    latin = sum(1 for ch in letters if _LATIN_RE.match(ch))
    return 'en' if latin / len(letters) >= bc.EN_LATIN_SHARE else 'other'


def _sentence_initial_positions(stem):
    """Character offsets that begin a sentence (position 0 + after .!? whitespace)."""
    pos = {0}
    for m in re.finditer(r'[.!?]\s+', stem):
        pos.add(m.end())
    return pos


def _is_full_sentence_option(o):
    return len((o or '').split()) >= 5 and bool(re.search(r'[a-z]\s+[a-z]', o or ''))


def _caps_convention_of_question(stem, options):
    """Per-question caps classification (§6.1.1). Returns (n_emphasis, n_codeword,
    n_acronym) counts over ALL-CAPS tokens of the stem."""
    toks = _ALLCAPS_RE.findall(stem or '')
    if not toks:
        return (0, 0, 0)
    opt_tokens = Counter(t for o in (options or []) for t in _ALLCAPS_RE.findall(o))
    has_map_cue = bool(_CODEWORD_CUE_RE.search(stem or ''))
    n_e = n_c = n_a = 0
    for t in toks:
        if t in bc.EMPHASIS_LEXICON:
            n_e += 1
        elif has_map_cue or opt_tokens.get(t, 0) >= 2:
            n_c += 1
        else:
            n_a += 1
    return (n_e, n_c, n_a)


def derive_content_signature(questions):
    """§6.1.1 — the content signature, computed from the questions, never from
    names. Computed on the 'en' renderings only (§6.1.9); q['medium'] is stamped by
    the caller (stamp_medium)."""
    en_qs = [q for q in questions if q.get('medium', 'en') == 'en']
    n = len(en_qs)
    sig = {'numeric_density': 0.0, 'notation_density': 0.0,
           'caps_convention': 'none', 'proper_noun_rate': 0.0, 'blank_rate': 0.0,
           'statement_rate': 0.0, 'polarity_rate': 0.0, 'stimulus_rate': 0.0,
           'label_scheme': 'mixed',
           'medium_en_share': (len(en_qs) / len(questions)) if questions else 0.0}
    if not n:
        return sig
    num = notn = blank = stmt = pol = stim = 0
    opt_bearing = 0
    cap_nonini = tok_total = 0
    n_e = n_c = n_a = 0
    label_votes = Counter()
    for q in en_qs:
        stem = _strip_segment_labels(q.get('stem') or '')
        opts = q.get('options') or []
        if _NUMBER_TOKEN_RE.search(stem):
            num += 1
        if _NOTATION_RE.search(stem) or q.get('omml_present'):
            notn += 1
        if opts:
            opt_bearing += 1
            if q.get('blank_pos', 'none') not in ('none', None) or re.search(r'_{3,}', stem):
                blank += 1
            full = sum(1 for o in opts if _is_full_sentence_option(o))
            if full >= 2 or _opts_are_combination_labels(opts):
                stmt += 1
        if _POLARITY_RE.search(stem):
            pol += 1
        if (q.get('linked_group_id') or q.get('image_role', 'none') != 'none'
                or _looks_like_table_stimulus(stem)):
            stim += 1
        sini = _sentence_initial_positions(stem)
        for m in _CAP_TOKEN_RE.finditer(stem):
            tok_total += 1
            if m.start() not in sini:
                cap_nonini += 1
        tok_total += len(re.findall(r'\b[a-z]+\b', stem))
        e, c, a = _caps_convention_of_question(stem, opts)
        n_e += e; n_c += c; n_a += a
        # v2.15 stamps this as 'option_label' (BUG-D07); reading the stamped key,
        # not a near-miss name, is what audit_seam's ORPHAN-READ check enforces.
        ls = q.get('option_label')
        if ls:
            label_votes[ls] += 1
    sig['_style_en_count']  = n   # internal (leading '_'): consumed by the style
                                  # writers; MUST be filtered out of the additive
                                  # content_signature entry field (fixture-pinned).
    sig['numeric_density']  = round(num / n, 4)
    sig['notation_density'] = round(notn / n, 4)
    sig['proper_noun_rate'] = round(cap_nonini / tok_total, 4) if tok_total else 0.0
    sig['blank_rate']       = round(blank / opt_bearing, 4) if opt_bearing else 0.0
    sig['statement_rate']   = round(stmt / opt_bearing, 4) if opt_bearing else 0.0
    sig['polarity_rate']    = round(pol / n, 4)
    sig['stimulus_rate']    = round(stim / n, 4)
    total_caps = n_e + n_c + n_a
    if total_caps == 0:
        sig['caps_convention'] = 'none'
    elif n_c / total_caps >= 0.50:
        sig['caps_convention'] = 'codeword'
    elif n_e >= 1:
        sig['caps_convention'] = 'emphasis'
    else:
        sig['caps_convention'] = 'acronym'
    if label_votes:
        top, ct = label_votes.most_common(1)[0]
        canon = {'numeric': '1234', 'upper': 'ABCD', 'lower': 'abcd'}
        sig['label_scheme'] = canon.get(top, top) if ct / sum(label_votes.values()) >= 0.8 \
                              else 'mixed'
    else:
        sig['label_scheme'] = 'mixed'
    return sig


_EXAM_CLASS_CACHE = {}


def _exam_style_class(progress, sig=None):
    """Corpus-level class, computed ONCE and cached on progress: 'content'
    (science/engineering/etc — v2 masking, mode None) vs 'aptitude' (the class
    the legacy layer was built for — legacy path, EC-26). Structural signals
    only: aggregate notation density, or NAT/MSQ answer types (aptitude exams in
    this estate are MCQ-only). A single-subject MCQ content exam without
    notation (law, commerce) classes 'aptitude' — that is exactly v2.55's
    behaviour for it, so it is a no-regression known limitation, documented in
    DEPLOY_NOTES rather than guessed at."""
    if isinstance(progress, dict):
        # Cache OUTSIDE progress: callers legitimately iterate progress while
        # synthesising, and inserting a cache key mid-iteration raises
        # RuntimeError. Keyed by id() with a content fingerprint so a recycled
        # id can never serve a stale class.
        meta = progress.get('_meta', {}) or {}
        fp = (meta.get('total_questions'), len(meta.get('papers_processed', ())))
        hit = _EXAM_CLASS_CACHE.get(id(progress))
        if hit and hit[0] == fp:
            return hit[1]
        if meta.get('nat_allowed') or meta.get('multi_select_allowed'):
            cls = 'content'
        else:
            all_qs = [q for k, v in list(progress.items())
                      if isinstance(k, tuple) for q in v]
            agg = derive_content_signature(all_qs) if all_qs else (sig or {})
            cls = ('content' if (agg.get('notation_density') or 0) >= 0.10
                   else 'aptitude')
        _EXAM_CLASS_CACHE[id(progress)] = (fp, cls)
        return cls
    return ('content' if (sig or {}).get('notation_density', 0) >= 0.10
            else 'aptitude')


def _mode_features(questions):
    """Structural per-subtopic rates the mode ladder reads (P-1: posing-structure
    only — given clauses, currency/operator/series tokens, option shapes)."""
    en = [q for q in (questions or []) if q.get('medium', 'en') == 'en']
    n = max(len(en), 1)
    f = {'given': 0.0, 'sentopt': 0.0, 'wordopt': 0.0}
    if not en:
        return f
    f['given'] = sum(1 for q in en
                     if _count_given_clauses(q.get('stem') or '') >= 1) / n
    f['sentopt'] = sum(1 for q in en if q.get('options') and
                       sum(1 for o in q['options']
                           if _is_full_sentence_option(o)) >= 2) / n
    f['wordopt'] = sum(1 for q in en if q.get('options') and
                       sum(1 for o in q['options']
                           if re.match(r'^[A-Za-z\-]+$', o.strip())) >=
                       max(2, len(q['options']) - 1)) / n
    return f


def derive_legacy_mode(sig, questions=None, exam_class=None,
                       section='', topic='', subtopic=''):
    """§6.1.1 / Q9 / EC-26 — the five legacy mode names, DERIVED from measurement.

    Two-level. The EXAM class decides the path: content => None (v2 masking).
    Within an aptitude-class exam every subtopic lands one of the five legacy
    modes — never None — because the legacy layer must reproduce v2.55
    byte-for-byte there (EC-26). The subtopic's mode is the MAJORITY VOTE of its
    questions' measured mechanics (bc.MECHANIC_MODE_VOTE — the framework's own
    posing-structure vocabulary, P-1). Contested mechanics and the residual are
    resolved by signature tie-breaks fitted on the SSC CGL corpus and judged by
    FIELD parity (§9.3 proof 2), not label aesthetics.
    """
    if exam_class is None:
        exam_class = _exam_style_class(None, sig)
    if exam_class == 'content':
        # CONTENT PATH IS TABLE-FREE (ruling Q9): the keyword tables' harm was
        # silently mislabelling every content exam to 'reasoning'; this early
        # return is the structural guarantee they can never touch one again.
        # mock_sync_audit MS-16 pins this ordering.
        return None
    # ── APTITUDE COMPATIBILITY SHIM (EC-26 / Q13 must-pass) ────────────────────
    # On an aptitude-class corpus the v2.55 table IS the definition of v2.55
    # behaviour, and §9.3 proof 2 demands byte-identity on every pre-existing
    # field there. The measured mechanic-vote ladder below reaches 78% label
    # parity on the SSC corpus (2026-08-31) — real, but not zero-diff — so the
    # table answers first ON THIS PATH ONLY, and the measured ladder is the
    # fallback for any subtopic name the table cannot place. The table's reach
    # is thereby shrunk from 'global classifier of every exam' to 'aptitude-only
    # legacy shim': the root-cause failure mode is structurally impossible, and
    # aptitude continuity is exact.
    _tbl = None
    if section or topic or subtopic:
        try:
            _tbl = determine_strip_mode(section or '', topic or '', subtopic or '')
        except Exception:
            _tbl = None
    if _tbl in ('quantitative', 'english', 'logical', 'reasoning', 'factual'):
        return _tbl
    f = _mode_features(questions)
    pn = sig.get('proper_noun_rate', 0)
    votes = Counter()
    for q in (questions or []):
        if q.get('medium', 'en') != 'en':
            continue
        mech = q.get('mechanic') or detect_mechanic(q)
        v = bc.MECHANIC_MODE_VOTE.get(mech)
        if v is None and mech == 'evaluate_statements':
            # GS statement-evaluation is factual; LOGICAL statement questions
            # (syllogism / statement-argument) carry labelled blocks and land on
            # the syllogism mechanic, which votes logical directly.
            v = 'factual'
        elif v is None and mech == 'relational_reasoning':
            stem_l = (q.get('stem') or '').lower()
            if any(c in stem_l for c in ('calendar', 'day of the week',
                                         'angle between', 'hands of',
                                         'direction', 'north', 'south')):
                v = 'reasoning'
            else:
                v = 'logical' if pn >= 0.05 else 'reasoning'
        elif v is None:                       # unknown mechanic
            continue
        # a cloze question is an english vote whatever its mechanic — the blank
        # is the legacy english layer's own signal
        if (q.get('blank_pos', 'none') not in ('none', None)
                or re.search(r'_{3,}', q.get('stem') or '')):
            v = 'english'
        votes[v] += 1
    if votes:
        top, top_n = votes.most_common(1)[0]
        # quantitative NAT-style word problems can out-vote via recall on short
        # definitional items; a strong given-clause rate overrides a weak vote
        # given-clause override applies only over WEAK tops (recall-driven
        # factual/english): it must never outrank a measured reasoning or
        # logical structure (seating puzzles state positions numerically).
        if top in ('factual', 'english') and f['given'] >= 0.35 and pn < 0.10:
            return 'quantitative'
        return top
    # residual — no mechanic voted: signature ladder
    if sig.get('blank_rate', 0) >= 0.15:
        return 'english'
    if f['given'] >= 0.35 and pn < 0.10:
        return 'quantitative'
    if pn >= 0.10:
        return 'factual'
    if sig.get('numeric_density', 0) >= 0.30:
        return 'reasoning'
    return 'english'


def _protect_emphasis(stem, legacy_mode):
    """§6.1.1 — emphasis tokens are NEVER masked, on any path. Implemented by
    pre-substituting each emphasis-lexicon ALL-CAPS token with a placeholder that
    no legacy mask pattern can match, running the legacy mask, then restoring.
    Byte-identical to the legacy output whenever the stem carries no emphasis
    token (the measured SSC case)."""
    subs = []
    def _stash(m):
        tok = m.group(0)
        if tok in bc.EMPHASIS_LEXICON:
            ph = '\x00e%d\x00' % len(subs)
            subs.append((ph, tok))
            return ph
        return tok
    guarded = _ALLCAPS_RE.sub(_stash, stem)
    masked = strip_variables(guarded, legacy_mode)
    for ph, tok in subs:
        masked = masked.replace(ph, tok)
    return masked


def strip_variables_v2(stem, sig, legacy_mode=None):
    """§6.1.1 — masking driven ONLY by the signature. For an aptitude-class
    signature (legacy_mode derived, not looked up) the legacy per-mode token
    vocabulary is preserved verbatim (currency/percent/time/series tokens as
    today) — this is what EC-26 byte-identity measures. For a content-class
    signature the v2 rules apply: numbers → _NUM_ with unit text retained;
    ALL-CAPS masked only under the codeword convention and NEVER when in the
    emphasis lexicon; capitalised words → _NAME_ only when non-sentence-initial
    AND proper_noun_rate >= 0.05 AND not retained notation; cloze on
    option-bearing questions only (handled by caller flag)."""
    if legacy_mode:
        return _protect_emphasis(stem, legacy_mode)
    t = stem
    # Cloze first (before number masking can touch underscores).
    t = re.sub(r'_{3,}', '_BLANK_', t)
    # Numbers → _NUM_, unit text retained: mask only the numeric part of the token.
    def _num_sub(m):
        tok = m.group(0)
        um = re.match(r'(-?\d[\d,]*(?:\.\d+)?(?:\s*[×x]\s*10\^?-?\d+|e-?\d+)?)(.*)$', tok)
        return '_NUM_' + (um.group(2) if um else '')
    t = _NUMBER_TOKEN_RE.sub(_num_sub, t)
    # ALL-CAPS: only under codeword convention, never emphasis tokens.
    if sig.get('caps_convention') == 'codeword':
        t = re.sub(r'\b[A-Z]{3,}\b',
                   lambda m: m.group(0) if m.group(0) in bc.EMPHASIS_LEXICON else '_WORD_',
                   t)
    # Capitalised → _NAME_ under the three-part condition.
    if sig.get('proper_noun_rate', 0.0) >= 0.05:
        sini = _sentence_initial_positions(t)
        out, last = [], 0
        for m in _CAP_TOKEN_RE.finditer(t):
            if m.start() in sini or m.group(0) in _RETAINED_NOTATION:
                continue
            out.append(t[last:m.start()]); out.append('_NAME_'); last = m.end()
        out.append(t[last:])
        t = ''.join(out)
    t = re.sub(r'^Q\.?\d+\.?\s*', '', t)
    t = re.sub(r'\[\d{1,2}-\w+-\d{4}[^\]]*\]', '', t)
    return t.strip()


# ── Option shape (§6.1.2 OPTION_SHAPES; single-question classifier) ─────────────

_SUFFICIENCY_RE = re.compile(r'statement\s+(i{1,3}|[12ab])\s+alone\s+is\s+sufficient', re.I)
_PERM_RE        = re.compile(r'^[A-Z](\s*[,;\-–>]\s*[A-Z]){2,}$')
_CODE_STR_RE    = re.compile(r'^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9@#$%&*]{2,}$')
_EXPR_RE        = re.compile(r'[=+^√]|→|⇌|[₀-₉⁰-⁹]|\\frac|\b[a-z]\s*/\s*[a-z]\b')
_VALUE_RE       = re.compile(r'^-?\d[\d,]*(?:\.\d+)?(?:\s*[×x]\s*10\^?-?\d+|e-?\d+)?$')
_VALUE_UNIT_RE  = re.compile(r'^-?\d[\d,]*(?:\.\d+)?(?:\s*[×x]\s*10\^?-?\d+|e-?\d+)?'
                             r'\s*[A-Za-zµ°%][A-Za-z⁻¹²³/·\s]*$')
_OPT_IMG_ROLES_STYLE = ('options_only', 'stem_and_options')


def detect_option_shape(q, figural_descriptor=None):
    """§6.1.2 — one of OPTION_SHAPES for this question. Shape carries meaning that
    no subject noun is allowed to (P-1)."""
    opts = [o.strip() for o in (q.get('options') or [])]
    if not opts:
        return 'none'
    if all(o == '' for o in opts):
        # image options; refine from the figural descriptor when present (EC-6)
        d = (figural_descriptor or {})
        kinds = ' '.join(str(v) for v in d.get('object_types', {}).get('dominant', []))
        if re.search(r'structur|molecul|skelet', kinds, re.I):
            return 'structure_image'
        if re.search(r'graph|plot|curve|spectr|axis', kinds, re.I):
            return 'plot_image'
        return 'figure'
    low = [o.lower() for o in opts if o]
    if sum(1 for o in low if _SUFFICIENCY_RE.search(o)) >= 2:
        return 'sufficiency_set'
    if all(_PERM_RE.match(o.upper().replace(' ', '')) or
           re.match(r'^[A-Z]{3,6}$', o.upper().replace(',', '').replace(' ', ''))
           for o in opts if o) and len(set(''.join(sorted(o.upper().replace(' ', '').replace(',', '')))
                                           for o in opts if o)) == 1:
        return 'permutation'
    if _opts_are_match_pairs(opts):
        return 'pair_map'
    if _opts_are_combination_labels(opts):
        return 'combination_label'
    if sum(1 for o in opts if _is_full_sentence_option(o)) >= 2:
        return 'statement'
    n_val  = sum(1 for o in opts if _VALUE_RE.match(o.replace(' ', '')))
    n_valu = sum(1 for o in opts if _VALUE_UNIT_RE.match(o))
    if n_val >= max(2, len(opts) - 1):
        return 'value'
    if n_valu >= max(2, len(opts) - 1):
        return 'value_with_unit'
    if sum(1 for o in opts if _EXPR_RE.search(o)) >= max(2, len(opts) - 1):
        return 'expression'
    if sum(1 for o in opts if o and _CODE_STR_RE.match(o.replace(' ', ''))) >= max(2, len(opts) - 1):
        return 'code_string'
    if all(len(o.split()) == 1 for o in opts if o):
        return 'word'
    return 'entity'


# ── Mechanic detection (§6.1.2) — ordered rule list; first match wins; the order
#    IS the rule (EC-57). Cues come ONLY from bc.MECHANIC_CUES. ─────────────────

_GIVEN_CLAUSE_RE = re.compile(r'(-?\d[\d,]*(?:\.\d+)?\s*[A-Za-zµ°%][A-Za-z⁻¹²³/·]*)|'
                              r'\b[a-zA-Z]\s*=\s*-?\d', )
_SEQ_RUN_RE      = re.compile(r'(?:\b[A-Za-z0-9]{1,6}\s*,\s*){2,}[A-Za-z0-9?_]{1,6}')
_FRAGMENT_RE     = re.compile(r'\(?\b([PQRS]|[A-D]|[1-4])\s*[\.\):]\s+\S')


_NON_UNIT_WORDS = frozenset(('was were is are be been the a an of to in on at for and '
                             'or with from by as that this it its if then than each per '
                             'has have had who men days years items marks questions '
                             'people times').split())


def _unit_of_token(tok):
    """The unit part of a number token, or '' when the trailing word is ordinary
    English rather than a unit (E-10 guard: '42 was' is a bare number)."""
    um = re.match(r'-?\d[\d,]*(?:\.\d+)?(?:\s*[×x]\s*10\^?-?\d+|e-?\d+)?\s*(.*)$', tok)
    unit = (um.group(1) if um else '').strip()
    if not unit or unit.lower() in _NON_UNIT_WORDS:
        return ''
    return unit


_SEG_LABEL_RE = re.compile(r'\(\s*\d\s*\)')


def _strip_segment_labels(stem):
    """'(1) / (2)' segment markers in error-spotting stems are LABELS, not
    values — they inflated numeric_density to 1.0 on grammar subtopics and sent
    them down the quantitative path (SSC measurement, 2026-08-31)."""
    return _SEG_LABEL_RE.sub(' ', stem or '')


def _count_given_clauses(stem):
    n = 0
    stem = _strip_segment_labels(stem)
    for cl in re.split(r'[;,.]|\band\b', stem):
        hit = False
        for m in _NUMBER_TOKEN_RE.finditer(cl):
            if _unit_of_token(m.group(0)):
                hit = True
                break
        if hit or re.search(r'\b[A-Za-z]\w*\s*=\s*-?\d', cl):
            n += 1
    return n


def _named_entities_ordered(stem):
    """Named entities in STEM ORDER, deduped — the deterministic source the pool
    counters insert from. A set here would randomise Counter insertion order and
    make the ordered emissions unkillable by any fixture."""
    sini = _sentence_initial_positions(stem)
    return list(dict.fromkeys(m.group(0) for m in _CAP_TOKEN_RE.finditer(stem)
                              if m.start() not in sini))


def _named_entities(stem):
    return set(_named_entities_ordered(stem))


def detect_mechanic(q, section_stem_p50=None, deduction_steps=None,
                    figural_descriptor=None):
    """§6.1.2 — the 24-mechanic ordered classifier. Computed on the 'en'
    rendering only; a question that matches nothing is 'unknown' — REPORTED,
    counted, never defaulted (P-1)."""
    C = bc.MECHANIC_CUES
    stem = q.get('stem') or ''
    s = stem.lower()
    opts = [o.strip() for o in (q.get('options') or [])]
    low = [o.lower() for o in opts if o]
    shape = q.get('option_shape') or detect_option_shape(q, figural_descriptor)
    axis2 = q.get('axis2') or ''
    words = s.split()

    def any_cue(cues, text):  return any(c in text for c in cues)
    def any_re(res, text):    return any(re.search(r, text) for r in res)

    # 1 data_sufficiency — the canonical sufficiency option set.
    if shape == 'sufficiency_set' or sum(1 for o in low if
            any_re(C['data_sufficiency']['option_res'], o)) >= 2:
        return 'data_sufficiency'
    # 2 assertion_reason — Axis-2.
    if axis2 == 'ASSERTION_REASON':
        return 'assertion_reason'
    # 3 match — Axis-2.
    if axis2 == 'MATCH':
        return 'match'
    # 4 syllogism — labelled Statement(s)+Conclusion(s) blocks or follows-options.
    if (all(re.search(r, s) for r in C['syllogism']['stem_res'])
            or sum(1 for o in low if any_re(C['syllogism']['option_res'], o)) >= 1):
        return 'syllogism'
    # 5 decode — mapping cue with code-string options, or codeword caps on stem.
    if ((any_cue(C['decode']['stem_cues'], s) and shape == 'code_string')
            or (_caps_convention_of_question(stem, opts)[1] > 0
                and any_cue(C['decode']['stem_cues'], s))):
        return 'decode'
    # 6 constraint_arrangement — >=3 named entities AND >=2 constraint clauses.
    cc = C['constraint_arrangement']
    n_constraints = sum(1 for c in cc['constraint_cues'] if re.search(r'\b%s\b' % c, s))
    if len(_named_entities(stem)) >= cc['min_entities'] and n_constraints >= cc['min_constraints']:
        return 'constraint_arrangement'
    # 7 procedure_trace — step/input/output stimulus with an output-ask.
    pc = C['procedure_trace']
    if any_re(pc['stem_res'], s) and any_cue(pc['ask_cues'], s):
        return 'procedure_trace'
    # 8 text_reorder — >=3 labelled fragments with permutation options.
    if shape == 'permutation' and len(set(m.group(1) for m in _FRAGMENT_RE.finditer(stem))) \
            >= C['text_reorder']['min_fragments']:
        return 'text_reorder'
    # 9 sentence_edit — segment labels + "no error" option, or an edit cue.
    se = C['sentence_edit']
    if (any(any_cue(se['option_cues'], o) for o in low)
            or any_cue(se['stem_cues'], s)):
        return 'sentence_edit'
    # 10 word_meaning — meaning cue with word/short-phrase options. The 12-word
    #    stem guard applies only when the options are not themselves short: an
    #    idiom-meaning stem routinely exceeds it while its options stay short
    #    (SSC measurement 2026-08-31 — the guard sent idiom/phrase subtopics to
    #    recall and their mode to factual).
    wm = C['word_meaning']
    _short_opts = bool(opts) and all(len(o.split()) <= 6 for o in opts if o.strip())
    if (any_cue(wm['stem_cues'], s)
            and (shape in ('word', 'entity') and len(words) <= wm['max_stem_words']
                 or _short_opts)):
        return 'word_meaning'
    # 11 passage_comprehension — PASSAGE stimulus, non-cloze, passage-ask.
    if (q.get('axis1') == 'PASSAGE' or q.get('linked_group_id')) and axis2 != 'FILL_BLANK' \
            and (any_cue(C['passage_comprehension']['ask_cues'], s) or not words):
        return 'passage_comprehension'
    # 12 spatial_figure — figure options + spatial cue, or figural with figure
    #    options and no numeric ask.
    sf = C['spatial_figure']
    if shape in ('figure',) and (any_cue(sf['stem_cues'], s)
            or (q.get('axis1') == 'FIGURAL' and not _NUMBER_TOKEN_RE.search(stem))):
        return 'spatial_figure'
    # 13 series_completion — a comma-run of >=3 short tokens ending in ?/_ , or a
    #    series cue WITH a comma-run (a bare '?' can never fire this rule).
    sc = C['series_completion']
    _run = _SEQ_RUN_RE.search(stem)
    if _run and (('?' in _run.group(0) or '_' in _run.group(0)
                  or stem[_run.end():_run.end() + 3].strip().startswith(('?', '_')))
                 or any_cue(sc['stem_cues'], s)):
        return 'series_completion'
    # 14 pattern_analogy — :: or relatedness/odd-one cues.
    pa = C['pattern_analogy']
    if any_re(pa['stem_res'], stem) or any_cue(pa['stem_cues'], s):
        return 'pattern_analogy'
    # 15 relational_reasoning — kinship/direction/clock-calendar composites.
    rr = C['relational_reasoning']
    _kin_hits = sum(1 for k in rr['kin_cues'] if re.search(r'\b%s\b' % k, s))
    if ((_kin_hits >= 1 and any_cue(rr['kin_ask'], s)) or _kin_hits >= 2
            or (any(re.search(r'\b%s\b' % d, s) for d in rr['dir_cues']) and any_cue(rr['dir_ask'], s))
            or any_cue(rr['clock_cues'], s)):
        return 'relational_reasoning'
    # 16 interpret_data — DI/TABLE/PLOT stimulus with a read-off ask.
    idc = C['interpret_data']
    if (q.get('axis1') == 'DI' or _looks_like_table_stimulus(stem)) and \
            (any_re(idc['ask_res'], s) or any_cue(idc['ask_cues'], s)):
        return 'interpret_data'
    # 17 apply_rule_to_case — scenario (>=2 parties, >=2 past-tense actions) with a
    #    consequence ask. Consequence verbs, never subject nouns (Pass-4 P-1 fix).
    ar = C['apply_rule_to_case']
    past_actions = len(re.findall(r'\b\w+ed\b', s))
    if (len(_named_entities(stem)) >= ar['min_parties'] and past_actions >= ar['min_actions']
            and any_cue(ar['ask_cues'], s)
            and shape in ('statement', 'entity', 'value', 'combination_label')):
        return 'apply_rule_to_case'
    # 18 evaluate_statements — sentence options, roman/label combos, or a
    #    statement(s) stem cue — any one suffices.
    if shape in ('statement', 'combination_label') or \
            any_re(C['evaluate_statements']['stem_res'], s):
        return 'evaluate_statements'
    # 19 rank_order — ordering cue with >=3 entities in stem or options.
    ro = C['rank_order']
    ents = max(len(_named_entities(stem)), len(opts))
    if any_cue(ro['stem_cues'], s) and ents >= ro['min_entities']:
        return 'rank_order'
    # 20 identify — identify cue AND a shape that carries the meaning; also the
    #    copular form ('The complex conjugate of psi(x) is:') when the options
    #    are expressions/structures and the stem gives nothing to compute (JAM
    #    measurement 2026-08-31 — the ask-verb form alone missed every copular
    #    identification).
    idf = C['identify']
    if shape in idf['option_shapes']:
        if any_re(idf['stem_res'], s):
            return 'identify'
        if (re.search(r'\b(is|are)\s*:?\s*$', s.strip())
                and _count_given_clauses(stem) == 0):
            return 'identify'
    # 21 predict — outcome cues.
    if any_cue(C['predict']['stem_cues'], s):
        return 'predict'
    # 22 multi_step_derivation — >=2 given clauses, OR chaining, OR measured
    #    deduction_steps >= 3 when a PYQ-Explain record exists.
    ms = C['multi_step_derivation']
    if (_count_given_clauses(stem) >= ms['min_given_clauses']
            or any(re.search(r'\b%s\b' % c, s) for c in ms['chain_cues'])
            or (deduction_steps or 0) >= ms['min_deduction_steps']):
        return 'multi_step_derivation'
    # 23 single_formula — one given clause with NAT or numeric options.
    if _count_given_clauses(stem) >= 1 and shape in ('none', 'value', 'value_with_unit'):
        return 'single_formula'
    # 24 recall — NOTHING TO COMPUTE: no free-standing number, no given clause,
    #    no compute-ask; plain short options. Notation alone never blocks recall —
    #    a stem that merely NAMES species (SF4, B2H6) has nothing to derive, and
    #    the old notation veto sent 20.3% of the JAM corpus to 'unknown'
    #    (measured 2026-08-31). Free-standing numbers still block: they are the
    #    signal that a value is there to be worked with.
    if (not _NUMBER_TOKEN_RE.search(stem)
            and _count_given_clauses(stem) == 0
            and not any(c in s for c in C['recall']['exclude_asks'])
            and shape in ('word', 'entity', 'value', 'structure_image', 'figure',
                          'expression', 'none')
            and (section_stem_p50 is None or len(words) <= section_stem_p50)):
        return 'recall'
    return bc.MECHANIC_UNKNOWN


# ── Distractor-mechanism mining (§6.1.3) — vocabulary is bc.DISTRACTOR_MECHANISMS ──

_KNOWN_UNIT_FACTORS = (1000.0, 100.0, 60.0, 3600.0, 1e3, 1e6, 1e9, 2.54, 4.184,
                       101.325, 760.0, 1.602e-19, 6.022e23)


def _parse_magnitude(text):
    """SI-prefix and ×10ⁿ normalised magnitude; None on parse failure (EC-14)."""
    t = (text or '').strip().replace(',', '')
    m = re.match(r'^(-?\d+(?:\.\d+)?)(?:\s*[×x]\s*10\^?(-?\d+)|e(-?\d+))?\s*'
                 r'([A-Za-zµ°%][A-Za-z⁻¹²³/·]*)?$', t)
    if not m:
        return None
    val = float(m.group(1))
    exp = m.group(2) or m.group(3)
    if exp:
        val *= 10.0 ** int(exp)
    unit = m.group(4) or ''
    _SI = {'p': 1e-12, 'n': 1e-9, 'µ': 1e-6, 'u': 1e-6, 'm': 1e-3,
           'k': 1e3, 'M': 1e6, 'G': 1e9}
    if len(unit) >= 2 and unit[0] in _SI and unit[1].isalpha():
        val *= _SI[unit[0]]
    return val


def _mine_numeric(opt, key):
    ov, kv = _parse_magnitude(opt), _parse_magnitude(key)
    if ov is None or kv is None or kv == 0:
        return bc.MECHANIC_UNKNOWN
    if ov == kv:
        return bc.MECHANIC_UNKNOWN
    r = ov / kv
    if abs(ov - kv) < 0.01 * abs(kv):
        return 'rounding_trap'
    if r < 0 and abs(abs(r) - 1.0) < 1e-9:
        return 'sign_error'
    la = math.log10(abs(r)) if r != 0 else 0.0
    if abs(la - round(la)) < 1e-6 and round(la) != 0:
        return 'order_of_magnitude'
    for f in _KNOWN_UNIT_FACTORS:
        if abs(abs(r) - f) <= 0.01 * f or abs(abs(1.0 / r) - f) <= 0.01 * f:
            return 'unit_error'
    if 0.5 <= abs(r) <= 2.0:
        return 'near_miss'
    return bc.MECHANIC_UNKNOWN


def mine_distractor_mechanisms(questions, figural_descriptor=None):
    """§6.1.3 — per option-bearing PYQ with a known key. When keys are absent the
    subtopic reports 'unavailable', never guessed (EC-12). Cancelled questions with
    unknown keys are excluded (EC-11)."""
    mix = Counter()
    n = 0
    for q in questions:
        opts = [o.strip() for o in (q.get('options') or [])]
        key = q.get('key') or q.get('correct_option') or q.get('answer')
        if not opts or key in (None, ''):
            continue
        shape = q.get('option_shape') or detect_option_shape(q, figural_descriptor)
        key_texts = []
        if isinstance(key, (list, tuple, set)):
            idxs = [int(k) - 1 for k in key if str(k).isdigit()]
            key_texts = [opts[i] for i in idxs if 0 <= i < len(opts)]
        elif str(key).isdigit() and 1 <= int(key) <= len(opts):
            key_texts = [opts[int(key) - 1]]
        elif isinstance(key, str) and len(key.strip()) == 1 and key.strip().upper() in 'ABCDE':
            i = 'ABCDE'.index(key.strip().upper())
            if i < len(opts):
                key_texts = [opts[i]]
        if not key_texts:
            continue
        kset = {k.lower() for k in key_texts}
        for o in opts:
            if o.lower() in kset or not o:
                continue
            n += 1
            if shape in ('value', 'value_with_unit'):
                mix[_mine_numeric(o, key_texts[0])] += 1
            elif shape in ('expression', 'structure_image'):
                d = (figural_descriptor or {})
                kinds = ' '.join(str(v) for v in d.get('transformation_types', []))
                if re.search(r'stereo|configuration|chiral', kinds, re.I):
                    mix['stereochemistry_error'] += 1
                elif re.search(r'regio|position|site', kinds, re.I):
                    mix['regiochemistry_error'] += 1
                elif shape == 'expression':
                    a, b = o.replace(' ', ''), key_texts[0].replace(' ', '')
                    if a == b[::-1] or (len(a) == len(b) and sorted(a) == sorted(b)):
                        mix['reversed_relationship'] += 1
                    else:
                        mix['formula_error'] += 1
                else:
                    mix['structural_variant'] += 1
            elif shape == 'statement':
                ow = set(o.lower().split()); kw = set(key_texts[0].lower().split())
                ov = len(ow & kw) / max(len(ow | kw), 1)
                if _POLARITY_RE.search(o) != _POLARITY_RE.search(key_texts[0]):
                    mix['polarity_flip'] += 1
                elif ov >= 0.5:
                    mix['partial_truth'] += 1
                else:
                    mix['overgeneralised_rule'] += 1
            elif shape in ('entity', 'word'):
                mix['name_swap' if shape == 'entity' else 'near_synonym'] += 1
            elif shape == 'permutation':
                a = o.replace(' ', '').replace(',', '')
                b = key_texts[0].replace(' ', '').replace(',', '')
                swaps = sum(1 for x, y in zip(a, b) if x != y)
                mix['adjacent_swap' if swaps == 2 else 'anchor_misplaced'] += 1
            elif shape == 'code_string':
                a = o.replace(' ', ''); b = key_texts[0].replace(' ', '')
                if len(a) == len(b) and sum(1 for x, y in zip(a, b) if x != y) <= 2:
                    mix['off_by_one_shift'] += 1
                else:
                    mix['partial_mapping'] += 1
            elif shape in ('sufficiency_set', 'combination_label', 'pair_map'):
                mix['constraint_dropped' if len(o) < len(key_texts[0])
                    else 'constraint_inverted'] += 1
            elif shape in ('figure', 'plot_image'):
                d = (figural_descriptor or {})
                kinds = ' '.join(str(v) for v in d.get('transformation_types', []))
                if re.search(r'mirror', kinds, re.I):
                    mix['mirror_variant'] += 1
                elif re.search(r'rotat', kinds, re.I):
                    mix['rotation_variant'] += 1
                elif d:
                    mix['element_missing'] += 1
                else:
                    return 'unavailable'          # EC-6: no descriptors at all
            else:
                mix[bc.MECHANIC_UNKNOWN] += 1
    if n == 0:
        return 'unavailable'
    return {k: round(v / n, 4) for k, v in mix.items()}


# ── Rebuilt extractors (§6.1.5) — ungated, single-regex tokeniser (E-10) ─────────

def extract_number_ranges_v2(questions, sig):
    """Runs whenever numeric_density > 0. Ranges PER UNIT {min,max,p10,p50,p90,n}
    plus exclude_values (exact given values), EXCLUDING universal constants —
    a value in >= 20% of the exam-corpus numeric stems (computed by the caller and
    passed via sig['_universal_constants']) is a constant, exempt (EC-49)."""
    if sig.get('numeric_density', 0) <= 0:
        return None, []
    by_unit, givens = {}, Counter()
    for q in questions:
        if q.get('medium', 'en') != 'en':
            continue
        for m in _NUMBER_TOKEN_RE.finditer(q.get('stem') or ''):
            tok = m.group(0)
            um = re.match(r'(-?\d[\d,]*(?:\.\d+)?)((?:\s*[×x]\s*10\^?-?\d+|e-?\d+)?)'
                          r'\s*(.*)$', tok)
            if not um:
                continue
            try:
                v = float(um.group(1).replace(',', ''))
            except ValueError:
                continue
            # A bare integer run of >12 digits with no decimal point, comma
            # grouping or exponent is a FLATTENED LITERAL, not a magnitude —
            # measured on JAM: OMML matrices linearise as digit-row
            # concatenations ('1000010000100001' is the 4x4 identity matrix).
            # Feeding it to ranges poisons min/max for the unit.
            if ('.' not in um.group(1) and ',' not in um.group(1)
                    and not um.group(2) and len(um.group(1).lstrip('-')) > 12):
                continue
            if um.group(2):
                em = re.search(r'10\^?(-?\d+)|e(-?\d+)', um.group(2))
                if em:
                    v *= 10.0 ** int(em.group(1) or em.group(2))
            unit_raw = (um.group(3) or '').strip()
            if unit_raw and unit_raw.lower() in _NON_UNIT_WORDS:
                unit_raw = ''
            unit = unit_raw or '_dimensionless_'
            by_unit.setdefault(unit, []).append(v)
            givens[round(v, 10)] += 1
    if not by_unit:
        return None, []
    def _p6(vs):
        p = _pctiles(vs, as_int=False)
        return {k: round(v, 6) for k, v in p.items()}
    ranges = {u: dict({'min': min(vs), 'max': max(vs), 'n': len(vs)}, **_p6(vs))
              for u, vs in by_unit.items()}
    constants = set(sig.get('_universal_constants') or ())
    exclude = sorted(v for v in givens if v not in constants)
    return ranges, exclude


def extract_context_pool_v2(questions, sig):
    """Runs whenever proper_noun_rate > 0: the observed named-entity pool, by
    frequency band. Shape mirrors the legacy pool so existing readers keep
    working; the entities themselves are measured, never listed."""
    if sig.get('proper_noun_rate', 0) <= 0:
        return None
    counts = Counter()
    for q in questions:
        if q.get('medium', 'en') != 'en':
            continue
        for e in _named_entities_ordered(q.get('stem') or ''):
            if e not in _RETAINED_NOTATION:
                counts[e] += 1
    if not counts:
        return None
    tot = max(len([q for q in questions if q.get('medium', 'en') == 'en']), 1)
    return {'dominant': sorted(c for c, n in counts.items() if n / tot > 0.20),
            'common':   sorted(c for c, n in counts.items() if 0.05 <= n / tot <= 0.20),
            'rare':     sorted(c for c, n in counts.items() if 0 < n / tot < 0.05),
            'avoid':    []}


# ── §6.1.7 — lexicon, surface and stimulus statistics (per section × answer_type;
#    English rendering only) ─────────────────────────────────────────────────────

_STOPLIST = frozenset(('the a an of to in for on is are was were be been which what '
                       'who whom whose when where why how do does did done with by at '
                       'from as that this these those it its if then than and or not '
                       'no nor so such can could will would shall should may might '
                       'must have has had having following correct incorrect true '
                       'false given above below only all none both each every most '
                       'least other another same different among between statement '
                       'statements option options question questions answer choose '
                       'select respectively is are').split())


def _pctiles(vals, as_int=True):
    if not vals:
        return {'p10': 0, 'p50': 0, 'p90': 0}
    vs = sorted(vals)   # LOAD-BEARING: percentiles are positional reads of the
    def pv(p):          # SORTED sample; the mutation fixture feeds unsorted input.
        k = (len(vs) - 1) * p; f, c = int(k), min(int(k) + 1, len(vs) - 1)
        return vs[f] if f == c else vs[f] + (vs[c] - vs[f]) * (k - f)
    if as_int:
        return {'p10': int(round(pv(.10))), 'p50': int(round(pv(.50))),
                'p90': int(round(pv(.90)))}
    return {'p10': pv(.10), 'p50': pv(.50), 'p90': pv(.90)}


def _last_clause(stem):
    parts = re.split(r'[.;:?]', stem.strip())
    tail = next((p.strip() for p in reversed(parts) if p.strip()), '')
    tail = re.sub(r'-?\d[\d,]*(?:\.\d+)?', '_NUM_', tail).lower()
    return ' '.join(tail.split()[-10:])


_INSTRUCTION_RE = re.compile(r'((?:round(?:ed)? (?:off )?to|correct to|nearest|'
                             r'closest to|choose the|select the|in terms of|'
                             r'up to \w+ decimal)[^.;\n]{0,60})', re.IGNORECASE)


def _content_ngrams(text, nmax=3):
    toks = [t for t in re.findall(r"[a-z][a-z\-']+", text.lower())
            if t not in _STOPLIST and len(t) > 2]
    out = []
    for n in range(1, nmax + 1):
        out += [' '.join(toks[i:i + n]) for i in range(len(toks) - n + 1)]
    return out


def _top_n(counter, n):
    """Counter.most_common(n) with TIES BROKEN BY KEY, not by encounter order.

    Counter.most_common is stable in insertion order, so two runs over the same
    corpus in a different question order publish DIFFERENT top-N lists whenever
    counts tie at the cut — measured on the JAM corpus, where 'order' and
    'species' both scored tf 0.009615 and whichever was read first survived.
    These lists are CONSUMED (S3-12c draws ask_form and instruction from them;
    G-STYLE component 4 scores against the lexicon), so an order-dependent cut
    is an order-dependent brief. (Metamorphic test, 2026-08-31.)
    """
    return sorted(counter.items(), key=lambda kv: (-kv[1], str(kv[0])))[:n]


def compute_style_cell(questions, cell_min=None, figural_descriptor=None):
    """§6.1.7 + §6.2 StyleCell for one (section × answer_type) or paper cell.
    Callers stamp status/activation; DORMANT cells are still WRITTEN (P-3).
    Stimulus text of a linked group is excluded from members' lexicon (EC-8)."""
    cell_min = cell_min if cell_min is not None else bc.STYLE_ACTIVATION['min_questions_cell']
    en = [q for q in questions if q.get('medium', 'en') == 'en']
    n = len(en)
    cell = {'n': n, 'status': 'ACTIVE' if n >= cell_min else 'DORMANT',
            'dormant_reason': None if n >= cell_min else 'thin_cell'}
    if not en:
        return cell
    mech = Counter(q.get('mechanic', bc.MECHANIC_UNKNOWN) for q in en)
    form = Counter(q.get('axis2', 'DIRECT') for q in en)
    # COMPUTE when not stamped, with the figural descriptor — do not silently
    # report 'none'. The production path stamps option_shape in
    # synthesise_subtopic before this runs, but a cell that is WRONG rather than
    # ABSENT for an unstamped caller is the kind of defect that only shows up as
    # a strange brief three steps later. (Found 2026-08-31 while pinning the
    # descriptor contract.)
    shp  = Counter(q.get('option_shape') or detect_option_shape(q, figural_descriptor)
                   for q in en)
    cell['mechanic_mix']     = {k: round(v / n, 4) for k, v in sorted(mech.items())}
    cell['form_mix']         = {k: round(v / n, 4) for k, v in sorted(form.items())}
    cell['option_shape_mix'] = {k: round(v / n, 4) for k, v in sorted(shp.items())}
    cell['polarity_rate'] = round(sum(1 for q in en if _POLARITY_RE.search(q.get('stem') or '')) / n, 4)
    cell['nat_rate'] = round(sum(1 for q in en if not q.get('options')) / n, 4)
    cell['msq_rate'] = round(sum(1 for q in en if q.get('is_msq')) / n, 4)
    cell['note_block_rate'] = round(sum(1 for q in en if q.get('has_note')) / n, 4)
    cell['image_option_rate'] = round(sum(1 for q in en if q.get('image_role', 'none')
                                          in _OPT_IMG_ROLES_STYLE) / n, 4)
    cell['stem_len']   = _pctiles([len((q.get('stem') or '').split()) for q in en])
    cell['option_len'] = _pctiles([len(o.split()) for q in en
                                   for o in (q.get('options') or []) if o.strip()])
    oc = Counter(str(len(q.get('options') or [])) for q in en if q.get('options'))
    tot_oc = sum(oc.values()) or 1
    cell['option_count'] = {k: round(v / tot_oc, 4) for k, v in sorted(oc.items())}
    ask = Counter(_last_clause(q.get('stem') or '') for q in en)
    cell['ask_forms'] = [{'text': t, 'pct': round(c / n, 4)}
                         for t, c in _top_n(ask, 10) if t]
    instr = Counter(m.group(1).strip() for q in en
                    for m in _INSTRUCTION_RE.finditer(q.get('stem') or ''))
    cell['instruction_phrases'] = [{'text': t, 'pct': round(c / n, 4)}
                                   for t, c in _top_n(instr, 10)]
    op = Counter(' '.join((q.get('stem') or '').lower().split()[:3]) for q in en)
    cell['openers'] = [{'text': t, 'pct': round(c / n, 4)}
                       for t, c in _top_n(op, 20) if t]
    tf = Counter()
    seen_groups = set()
    for q in en:
        text = q.get('stem') or ''
        gid = q.get('linked_group_id')
        if gid and gid in seen_groups:
            pass                                        # EC-8: shared stimulus once
        seen_groups.add(gid)
        for g in _content_ngrams(text):
            tf[g] += 1
    tot_tf = sum(tf.values()) or 1
    cell['lexicon'] = [{'ngram': g, 'tf': round(c / tot_tf, 6)}
                       for g, c in _top_n(tf, 200)]
    units = sorted(dict.fromkeys(
        u for q in en for m in _NUMBER_TOKEN_RE.finditer(q.get('stem') or '')
        for u in [_unit_of_token(m.group(0))] if u))
    convs = sorted(dict.fromkeys(sym for q in en for sym in
                                 re.findall(r'→|⇌|Δ|±|≤|≥|×\s*10', q.get('stem') or '')))
    cell['notation'] = {'units': units[:40], 'conventions': convs}
    # stimulus_stats (ruling Q11) — CLASS-4 groups and single-question stimuli alike.
    groups = {}
    n_stim = 0
    kinds, slens, positions = Counter(), [], Counter()
    for q in en:
        stem = q.get('stem') or ''
        gid = q.get('linked_group_id')
        has_tbl = _looks_like_table_stimulus(stem)
        has_img = q.get('image_role', 'none') not in ('none', None)
        has_pass = bool(gid)
        if not (has_tbl or has_img or has_pass):
            continue
        n_stim += 1
        kind = ('passage' if has_pass and not has_tbl and not has_img else
                'table' if has_tbl else 'figure' if has_img else 'case')
        kinds[kind] += 1
        slens.append(len(stem.split()))
        positions['before' if has_pass else 'inline'] += 1
        if gid:
            groups.setdefault(gid, 0)
            groups[gid] += 1
    tot_k = sum(kinds.values()) or 1
    qps = list(groups.values()) if groups else [1] * n_stim if n_stim else []
    cell['stimulus_stats'] = {
        'stimulus_rate': round(n_stim / n, 4),
        'stimulus_kind_mix': {k: round(v / tot_k, 4) for k, v in sorted(kinds.items())},
        'stimulus_len': _pctiles(slens),
        'questions_per_stimulus': {'p50': (_pctiles(qps)['p50'] if qps else 0),
                                   'max': (max(qps) if qps else 0)},
        'stimulus_position': {k: round(v / (sum(positions.values()) or 1), 4)
                              for k, v in sorted(positions.items())}}
    # §6.2 StyleCell promises distractor_mix on EVERY cell, not only on subtopic
    # cells: S3-12c falls back to the PAPER cell whenever a section cell is
    # DORMANT (EC-3), and a fallback cell without it would silently force the
    # EC-12 default prior even on an exam whose keys are fully known.
    # 'unavailable' is the honest value when no question carries a key (EC-12) —
    # never a guess.
    # SAME MEASUREMENT, SAME INPUTS. section_rules mines with the figural
    # descriptor (EC-6 refines image-option mechanisms); a cell that mined
    # WITHOUT it would publish a different distractor_mix for the same subtopic
    # in the two artefacts, and the writing side reads whichever it happens to
    # hold. (Found by reading the diff, 2026-08-31.)
    _dm = mine_distractor_mechanisms(en, figural_descriptor)
    cell['distractor_mix'] = _dm if _dm else 'unavailable'
    return cell


# ── §6.1.2 pattern-key aggregation (additive; the legacy per-cluster
#    PYQ_STEM_PATTERNS list keeps its old granularity and fields for byte-identity
#    on aptitude corpora, and each entry gains pattern_key/mechanic/polarity/
#    option_shape; the KEY-level view — frequency, recency, confidence per KEY —
#    is the new `pattern_keys` block Step 7's S7-STYLE reads) ────────────────────

def build_pattern_keys(questions):
    keys = {}
    # dict.fromkeys preserves ENCOUNTER order, so this sorted() is killable by a
    # fixture whose questions arrive newest-year-first (a set here would iterate
    # small ints in value order and hide a broken sort).
    years_all = sorted(dict.fromkeys(q.get('year') for q in questions
                                     if q.get('year') is not None))
    last2 = set(years_all[-2:]) if len(years_all) >= 2 else set(years_all)
    for q in questions:
        if q.get('medium', 'en') != 'en' or not q.get('stem'):
            continue
        k = (q.get('mechanic', bc.MECHANIC_UNKNOWN), q.get('axis2', 'DIRECT'),
             'negative' if _POLARITY_RE.search(q['stem']) else 'positive',
             ('nat' if not q.get('options') else 'msq' if q.get('is_msq') else 'mcq'),
             q.get('option_shape', 'none'))
        e = keys.setdefault(k, {'raw_count': 0, 'w_count': 0, 'years': [],
                                'exemplars': []})
        w = 2 if q.get('year') in last2 else 1
        e['raw_count'] += 1
        e['w_count'] += w
        if q.get('year') is not None and q['year'] not in e['years']:
            e['years'].append(q['year'])   # encounter order; sorted at emit
        skel = q.get('skeleton') or ''
        for ex in e['exemplars']:
            if SequenceMatcher(None, ex['skel'], skel).ratio() >= 0.90:
                ex['raw_count'] += 1
                break
        else:
            if skel and len(e['exemplars']) < 8:
                e['exemplars'].append({'skel': skel, 'raw_count': 1})
    total_w = sum(e['w_count'] for e in keys.values()) or 1
    out = []
    for k, e in sorted(keys.items(), key=lambda kv: -kv[1]['w_count']):
        conf = 'observed' if e['raw_count'] >= 3 else 'inferred'
        ys = sorted(e['years'])
        if ys and all(y in last2 for y in ys) and len(ys) <= 2:
            conf = 'observed_recent'
        out.append({'pattern_key': '|'.join(k),
                    'mechanic': k[0], 'form': k[1], 'polarity': k[2],
                    'answer_type': k[3], 'option_shape': k[4],
                    'frequency': round(e['w_count'] / total_w * 100),
                    'raw_count': e['raw_count'], 'confidence': conf,
                    'deprecated': bool(ys and not any(y in last2 for y in ys)),
                    'years': ys,
                    'exemplars': sorted(e['exemplars'],
                                        key=lambda x: -x['raw_count'])[:5]})
    return out


def detect_low_entropy(questions):
    """EC-9 — low_entropy: True when >= 80% of the subtopic's UNIQUE stems
    (duplicate_of excluded, >= 3 unique required) are pairwise >= 0.50 Jaccard on
    8-token shingles."""
    stems = []
    seen = set()
    for q in questions:
        s = normalise_v1(q.get('stem') or '')
        if not s or s in seen:
            continue
        seen.add(s)
        # token-set Jaccard with digits masked: near-identical stems that differ
        # only in their values are exactly the low-entropy shape EC-9 names.
        stems.append(set(re.sub(r'\d+', '#', s).split()))
    if len(stems) < 3:
        return False
    close = 0
    pairs = 0
    for i in range(len(stems)):
        best = 0.0
        for j in range(len(stems)):
            if i == j or not stems[i] or not stems[j]:
                continue
            jac = len(stems[i] & stems[j]) / max(len(stems[i] | stems[j]), 1)
            best = max(best, jac)
        pairs += 1
        if best >= bc.LOW_ENTROPY_JACCARD:
            close += 1
    return (close / pairs) >= bc.LOW_ENTROPY_SHARE if pairs else False


# ── §6.3 — PYQ index: normaliser v1, shingles, builder ──────────────────────────

def normalise_v1(text):
    """§6.3 normaliser v1: NFKC → lower → strip Q-number/date/shift tags →
    collapse whitespace → keep digits, letters, unit strings, math tokens; strip
    punctuation."""
    t = unicodedata.normalize('NFKC', text or '').lower()
    t = re.sub(r'^q\.?\s*\d+\.?\s*', '', t)
    t = re.sub(r'\[\d{1,2}-\w+-\d{4}[^\]]*\]', '', t)
    t = re.sub(r'\b(shift|session)\s*[-:i1-3]+\b', '', t)
    t = re.sub(r'[^\w\s×^./%°µ⁻¹²³·-]', ' ', t)
    return ' '.join(t.split())


def _shingles(norm_text, k):
    toks = norm_text.split()
    if len(toks) < k:
        return []
    return [hashlib.md5(' '.join(toks[i:i + k]).encode()).hexdigest()[:12]
            for i in range(len(toks) - k + 1)]


def build_pyq_index_questions(all_questions, subtopic_of=None):
    """§6.3 — one record per PYQ, no text. Exact duplicates across sittings keep
    their own pyq_id with duplicate_of → first occurrence (EC-38)."""
    recs, first_by_md5 = [], {}
    for q in all_questions:
        stem_raw = q.get('stem') or ''
        norm = normalise_v1(stem_raw)
        md5 = hashlib.md5(norm.encode()).hexdigest()
        pid = f"{q.get('paper_id', '?')}:{q.get('num', '?')}"
        toks = norm.split()
        vals = sorted(m.group(0) for m in _NUMBER_TOKEN_RE.finditer(stem_raw))
        pol = 'negative' if _POLARITY_RE.search(stem_raw) else 'positive'
        rec = {'pyq_id': pid, 'year': q.get('year'),
               'subtopic_id': (subtopic_of or {}).get(id(q)) or q.get('subtopic_id') or '',
               'stem_md5': md5,
               'stem_shingles_8': _shingles(norm, bc.SHINGLE_K),
               'stem_shingles_4': (_shingles(norm, bc.SHINGLE_K_SHORT)
                                   if len(toks) < bc.SHORT_STEM_TOKENS else None),
               'stimulus_shingles_8': (_shingles(normalise_v1(q.get('stimulus_text') or ''),
                                                 bc.SHINGLE_K)
                                       if q.get('linked_group_id') else None),
               'semantic_tuple': [(subtopic_of or {}).get(id(q)) or q.get('subtopic_id') or '',
                                  '|'.join([q.get('mechanic', bc.MECHANIC_UNKNOWN)
                                            if q.get('medium', 'en') == 'en'
                                            else bc.MECHANIC_UNKNOWN,
                                            q.get('axis2', 'DIRECT'), pol]),
                                  vals],
               'values': vals,
               'option_md5s': [hashlib.md5(normalise_v1(o).encode()).hexdigest()
                               for o in (q.get('options') or []) if o.strip()],
               'image_phash': q.get('image_dhash'),
               'duplicate_of': first_by_md5.get(md5)}
        if md5 not in first_by_md5:
            first_by_md5[md5] = pid
        recs.append(rec)
    return recs, len(first_by_md5)


# ── MECHANIC → approach display string (E-3 fix; content path only — the legacy
#    aptitude path keeps infer_approach verbatim for EC-26 byte-identity) ────────

MECHANIC_TO_APPROACH = {
    'data_sufficiency'     : 'Test each statement for sufficiency independently, then together',
    'assertion_reason'     : 'Judge each clause, then the explanatory link',
    'match'                : 'Map each List-I item to its List-II partner',
    'syllogism'            : 'Evaluate statement-conclusion pairs using syllogism rules',
    'decode'               : 'Decode substitution pattern',
    'constraint_arrangement': 'Build the arrangement from the constraints, then read off',
    'procedure_trace'      : 'Trace the procedure step by step to the asked value',
    'text_reorder'         : 'Order the fragments into a coherent whole',
    'sentence_edit'        : 'Apply the governing usage rule to the marked segment',
    'word_meaning'         : 'Select the closest meaning',
    'passage_comprehension': 'Locate and read the governing lines of the passage',
    'spatial_figure'       : 'Transform the figure mentally and compare',
    'series_completion'    : 'Find the generating rule, apply it to the next term',
    'pattern_analogy'      : 'Find operation A->B, apply same to C',
    'relational_reasoning' : 'Chain the stated relations to the asked one',
    'interpret_data'       : 'Read the required values off the data, then compute',
    'apply_rule_to_case'   : 'Apply the governing rule to the stated facts',
    'evaluate_statements'  : 'Judge each statement independently, then combine',
    'rank_order'           : 'Order the entities by the governing property',
    'identify'             : 'Identify the entity that satisfies every stated condition',
    'predict'              : 'Predict the outcome from the governing principle',
    'multi_step_derivation': 'Derive the result over the required steps',
    'single_formula'       : 'Apply the governing relation once',
    'recall'               : 'Recall factual information',
    bc.MECHANIC_UNKNOWN    : 'Apply appropriate strategy',
}


# ── §6.2 + §6.3 — artefact writers. Same command, same run, same corpus_hash. ────

def compute_corpus_hash(paper_file_hashes):
    """sha256 over the SORTED list of PYQ file hashes — order-independent, so the
    same corpus always stamps the same hash on all three artefacts."""
    h = hashlib.sha256()
    for fh in sorted(paper_file_hashes):
        h.update(fh.encode())
    return h.hexdigest()


def measure_item_rules(all_questions):
    """§6.2 item_rules I-1..I-8 — the measured rate at which the exam's own PYQ
    violate each rule; a rule with violation_rate >= 0.10 is SUSPENDED for the
    exam (the house style legitimately breaks it). Rules needing answer keys are
    written measured:false when keys are absent — never guessed."""
    en = [q for q in all_questions if q.get('medium', 'en') == 'en' and q.get('options')]
    n = len(en)
    def rate(k):  return round(k / n, 4) if n else None
    have_keys = [q for q in en if (q.get('key') or q.get('correct_option')
                                   or q.get('answer')) not in (None, '')]
    def key_texts(q):
        opts = [o.strip() for o in q.get('options') or []]
        key = q.get('key') or q.get('correct_option') or q.get('answer')
        if isinstance(key, (list, tuple, set)):
            idxs = [int(k) - 1 for k in key if str(k).isdigit()]
            return [opts[i] for i in idxs if 0 <= i < len(opts)]
        ks = str(key).strip()
        if ks.isdigit() and 1 <= int(ks) <= len(opts):
            return [opts[int(ks) - 1]]
        if len(ks) == 1 and ks.upper() in 'ABCDE' and 'ABCDE'.index(ks.upper()) < len(opts):
            return [opts['ABCDE'.index(ks.upper())]]
        return []
    rules = {}
    # I-1 longest-option-is-key
    if have_keys:
        v = 0
        for q in have_keys:
            kt = key_texts(q)
            opts = [o.strip() for o in q['options'] if o.strip()]
            if kt and opts and max(opts, key=len) in kt and len(max(opts, key=len)) > \
                    1.5 * (sum(len(o) for o in opts) - len(max(opts, key=len))) / max(len(opts) - 1, 1):
                v += 1
        rules['I-1'] = {'rule': 'key is conspicuously the longest option',
                        'violation_rate': round(v / len(have_keys), 4), 'measured': True}
    else:
        rules['I-1'] = {'rule': 'key is conspicuously the longest option',
                        'violation_rate': None, 'measured': False}
    # I-2 all-of-the-above / none-of-the-above present
    v = sum(1 for q in en if any(re.search(r'\b(all|none) of (the above|these)\b',
                                           o, re.I) for o in q['options']))
    rules['I-2'] = {'rule': 'all/none-of-the-above options', 'violation_rate': rate(v),
                    'measured': True}
    # I-3 key position imbalance (needs keys)
    if len(have_keys) >= 20:
        pos = Counter()
        for q in have_keys:
            key = str(q.get('key') or q.get('correct_option') or q.get('answer')).strip()
            if key.isdigit():
                pos[int(key)] += 1
            elif len(key) == 1 and key.upper() in 'ABCDE':
                pos['ABCDE'.index(key.upper()) + 1] += 1
        tot = sum(pos.values())
        top = max(pos.values()) / tot if tot else 0
        rules['I-3'] = {'rule': 'answer-position imbalance (>40% one slot)',
                        'violation_rate': round(top, 4) if top > 0.40 else 0.0,
                        'measured': True}
    else:
        rules['I-3'] = {'rule': 'answer-position imbalance (>40% one slot)',
                        'violation_rate': None, 'measured': False}
    # I-4 grammatical cueing (a/an mismatch with options)
    v = 0
    for q in en:
        m = re.search(r'\b(a|an)\s*_{3,}', (q.get('stem') or ''), re.I)
        if m:
            art = m.group(1).lower()
            opts = [o.strip() for o in q['options'] if o.strip()]
            bad = [o for o in opts if (o[:1].lower() in 'aeiou') != (art == 'an')]
            if bad and len(bad) < len(opts):
                v += 1
    rules['I-4'] = {'rule': 'article/grammar cues eliminate options',
                    'violation_rate': rate(v), 'measured': True}
    # I-5 negative stem without emphasis
    v = sum(1 for q in en
            if _POLARITY_RE.search(q.get('stem') or '')
            and not any(t in (q.get('stem') or '') for t in bc.EMPHASIS_LEXICON))
    rules['I-5'] = {'rule': 'negative stem without emphasised marker',
                    'violation_rate': rate(v), 'measured': True}
    # I-6 overlapping numeric options
    v = 0
    for q in en:
        vals = []
        for o in q['options']:
            pv = _parse_magnitude(o)
            if pv is not None:
                vals.append(pv)
        if len(vals) >= 3 and len(set(vals)) < len(vals):
            v += 1
    rules['I-6'] = {'rule': 'duplicate/overlapping numeric options',
                    'violation_rate': rate(v), 'measured': True}
    # I-7 stem asks two things
    v = sum(1 for q in en if (q.get('stem') or '').count('?') >= 2)
    rules['I-7'] = {'rule': 'double-barrelled stem', 'violation_rate': rate(v),
                    'measured': True}
    # I-8 option order not canonical (numeric options unsorted)
    v = 0
    for q in en:
        vals = [_parse_magnitude(o) for o in q['options']]
        vals = [x for x in vals if x is not None]
        if len(vals) == len([o for o in q['options'] if o.strip()]) and len(vals) >= 3:
            if vals != sorted(vals) and vals != sorted(vals, reverse=True):
                v += 1
    rules['I-8'] = {'rule': 'numeric options not in monotone order',
                    'violation_rate': rate(v), 'measured': True}
    for r in rules.values():
        r['suspended'] = bool(r['measured'] and (r['violation_rate'] or 0) >= 0.10)
    return rules


def _utc_stamp():
    """UTC ISO-8601, second precision — the EC-25 recency comparator. A bare
    datetime.now() would compare local clocks across sessions."""
    import datetime as _dt
    return _dt.datetime.now(_dt.timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')


def write_style_profile(exam_code, entries, questions_by_subtopic, paper_ids,
                        paper_file_hashes, out_dir=None, exam_meta=None,
                        figural_by_subtopic=None):
    """§6.2 — <EXAM>_style_profile.json. DORMANT is a written status with a
    reason, never a missing file (P-3). Thresholds are dispersion-computed when
    questions_en >= 300 across >= 4 papers, else defaults with
    thresholds_source: 'default'."""
    all_qs = [q for qs in questions_by_subtopic.values() for q in qs]
    en_qs = [q for q in all_qs if q.get('medium', 'en') == 'en']
    n_papers = len(set(paper_ids))
    act = bc.STYLE_ACTIVATION
    active = (n_papers >= act['min_papers'] and len(all_qs) >= act['min_questions_paper'])
    reason = None
    if not active:
        reason = ('papers=%d < %d' % (n_papers, act['min_papers'])
                  if n_papers < act['min_papers']
                  else 'questions=%d < %d' % (len(all_qs), act['min_questions_paper']))
    # universal constants for exclude_values (EC-49) — computed once per corpus
    given_counts = Counter()
    numeric_stems = 0
    for q in en_qs:
        vals = list(dict.fromkeys(          # stem order, deduped — deterministic
            round(float(m.group(1).replace(',', '')), 10)
            for m in re.finditer(r'(-?\d[\d,]*(?:\.\d+)?)', q.get('stem') or '')))
        if vals:
            numeric_stems += 1
        for v in vals:
            given_counts[v] += 1
    constants = [v for v, c in given_counts.items()   # Counter = encounter order
                 if numeric_stems and c / numeric_stems >= bc.UNIVERSAL_CONSTANT_SHARE]
    # section × answer_type cells over English renderings; by_medium.other DORMANT
    sections = {}
    for e in entries:
        sec = e['section']
        for q in questions_by_subtopic.get((e['section'], e['topic'], e['subtopic']), []):
            sections.setdefault(sec, []).append(q)
    section_cells = {}
    for sec, qs in sections.items():   # dict is entry-ordered; JSON sorts keys
        en = [q for q in qs if q.get('medium', 'en') == 'en']
        cell_all = {}
        for at in ('mcq', 'msq', 'nat'):
            sub = [q for q in en if
                   ('nat' if not q.get('options') else 'msq' if q.get('is_msq') else 'mcq') == at]
            if sub:
                c = compute_style_cell(sub)
                if len(en) < act['min_questions_section']:
                    c['status'], c['dormant_reason'] = 'DORMANT', \
                        'section_en=%d < %d' % (len(en), act['min_questions_section'])
                cell_all[at] = c
        other = [q for q in qs if q.get('medium', 'en') != 'en']
        # §6.2 SECTION SCHEMA: status / n / cell are REQUIRED alongside
        # by_answer_type. Create S3-12c reads sect['status'] to decide whether a
        # section cell is usable and sect['cell'] as the fallback within the
        # section (EC-3). Emitting only by_answer_type made both reads falsy, so
        # EVERY slot silently drew from the PAPER cell and section-level style
        # never took effect — a defect invisible from the artefact alone,
        # because the profile still validated. Found by seam trace 2026-08-31.
        _sec_cell = compute_style_cell(en, cell_min=act['min_questions_section'])
        section_cells[sec] = {'status': _sec_cell.get('status', 'DORMANT'),
                              'dormant_reason': _sec_cell.get('dormant_reason'),
                              'n': len(en),
                              'cell': _sec_cell,
                              'by_answer_type': cell_all,
                              'by_medium': {'other': {'n': len(other), 'status': 'DORMANT',
                                                      'dormant_reason': 'medium_other_v1'}}}
    subtopic_cells = {}
    for e in entries:
        key = (e['section'], e['topic'], e['subtopic'])
        qs = questions_by_subtopic.get(key, [])
        if not qs:
            continue
        # THE FRAMEWORK OWNS THIS DERIVATION, not the caller. Each entry already
        # carries the subtopic's figural descriptor as PYQ_IMAGE_ANALYSIS (it is
        # what section_rules mined with), so reading it here makes the two
        # artefacts consistent BY CONSTRUCTION. A caller-supplied map is honoured
        # as an override but is never required — a call site that forgets it is
        # the exact way the two artefacts drift apart.
        _fd_cell = ((figural_by_subtopic or {}).get(
                        e.get('subtopic_id') or e['subtopic'])
                    or e.get('PYQ_IMAGE_ANALYSIS'))
        c = compute_style_cell(qs, figural_descriptor=_fd_cell)
        c['low_entropy'] = detect_low_entropy(qs)
        # §6.2 SUBTOPIC SCHEMA — the three fields the WRITING side reads from a
        # subtopic cell and cannot obtain anywhere else (Create S3-12c):
        #   number_ranges  -> the brief's per-unit value ranges
        #   exclude_values -> G-PYQ-DIST value reuse; without it the pre-write
        #                     redraw (EC-49) and REJECT reason (d) can NEVER fire
        #   distractor_mix -> the brief's distractor mechanisms; absent, EC-12's
        #                     default prior is taken even where real keys exist
        # They were computed for section_rules and never placed on the cell, so
        # the profile satisfied its own schema check while silently starving the
        # writing side. Found by tracing the seam field-by-field, 2026-08-31.
        _c_sig = derive_content_signature(qs)
        _c_mode = derive_legacy_mode(_c_sig, questions=qs,
                                     exam_class=_exam_style_class(None, _c_sig),
                                     section=e['section'], topic=e['topic'],
                                     subtopic=e['subtopic'])
        if _c_mode is not None:
            _rng, _exc = extract_number_ranges(qs, _c_mode), []
        else:
            _rng, _exc = extract_number_ranges_v2(qs, _c_sig)
        c['number_ranges'] = _rng
        # extract_number_ranges_v2 GUARANTEES an ascending exclude list — that
        # is its contract, pinned by the ranges_exclude_ascending fixture. A
        # second sort here would be an emission no fixture can distinguish from
        # its absence (a permanent mutation survivor), so the guarantee is
        # tested at the producer instead of re-asserted at the consumer.
        c['exclude_values'] = list(_exc) if _exc else []
        _dmix = mine_distractor_mechanisms(qs, _fd_cell)
        c['distractor_mix'] = _dmix if _dmix else 'unavailable'
        subtopic_cells[e.get('subtopic_id') or e['subtopic']] = c
    paper_cell = compute_style_cell(all_qs, cell_min=act['min_questions_paper'])
    # thresholds — leave-one-paper-out dispersion of D when the basis suffices
    thresholds = dict(bc.STYLE_THRESHOLD_DEFAULTS)
    thresholds['computed_from_dispersion'] = False
    thresholds_source = 'default'
    if len(en_qs) >= bc.STYLE_DISPERSION_MIN_EN and n_papers >= 4:
        dists = []
        basis = []
        for pid in dict.fromkeys(paper_ids):   # mean/sd are order-invariant
            held = [q for q in en_qs if q.get('paper_id') == pid]
            rest = [q for q in en_qs if q.get('paper_id') != pid]
            if len(held) >= bc.STYLE_DISPERSION_MIN_PAPER_Q and len(rest) >= 50:
                dists.append(bc.style_distance(compute_style_cell(held),
                                               compute_style_cell(rest)))
                basis.append(pid)
        if len(dists) >= 4:
            # §6.2: fail = max(DEFAULT_FAIL, p95 of real-paper distance);
            # warn = fail x 0.625. The FLOOR is the point: a tightly clustered
            # corpus would otherwise compute a fail band NARROWER than the
            # framework default and start reporting HIGH on ordinary papers.
            # (Measured 2026-08-31 on IIT_JAM_CHEMISTRY: mu+3sd gave 0.354,
            # below the 0.40 default, and §8.6 proof 3 caught it.)
            ordered = sorted(dists)
            p95 = ordered[min(len(ordered) - 1,
                              int(round(0.95 * (len(ordered) - 1))))]
            fail = max(bc.STYLE_THRESHOLD_DEFAULTS['style_distance_fail'], p95)
            # The BASIS is part of the artefact: a partial sitting (EC-36 —
            # e.g. a 10-question fragment) is too small to calibrate a
            # threshold, and therefore too small to be JUDGED by one. Recording
            # which sittings calibrated the band is what lets a consumer — and
            # §8.6 proof 3 — apply the same rule instead of guessing at it.
            thresholds = {'style_distance_warn': round(fail * 0.625, 4),
                          'style_distance_fail': round(fail, 4),
                          'computed_from_dispersion': True,
                          'basis_papers': sorted(basis),
                          'basis_min_questions': bc.STYLE_DISPERSION_MIN_PAPER_Q}
            thresholds_source = 'dispersion(n=%d)' % len(dists)
    # EC-4 / EC-10 — the STYLE WINDOW is part of the artefact's meaning: a
    # consumer that scores a real sitting must know which sittings fed the
    # cells. Years may be absent entirely (coverage_mode no_year_info, EC-4),
    # in which case the window is every sitting and the list is empty.
    _yrs = sorted({q.get('year') for q in all_qs if q.get('year')})
    profile = {'_meta': {'schema': bc.STYLE_PROFILE_SCHEMA, 'exam_code': exam_code,
                         'corpus_hash': compute_corpus_hash(paper_file_hashes),
                         'papers': n_papers, 'questions': len(all_qs),
                         'questions_en': len(en_qs),
                         'generated_at': _utc_stamp(),   # EC-25 recency rule
                         'window_years': _yrs,
                         'papers_in_window': sorted(dict.fromkeys(paper_ids)),
                         'papers_excluded_from_style': [],
                         'generated_by': 'analyse_engine v2.56'},
               'activation': {'status': 'ACTIVE' if active else 'DORMANT',
                              'dormant_reason': reason,
                              'thresholds_source': thresholds_source},
               'thresholds': thresholds,
               'paper': paper_cell,
               'sections': section_cells,
               'subtopics': subtopic_cells,
               'item_rules': measure_item_rules(all_qs),
               'universal_constants': sorted(constants),   # artefact list; fixture-pinned
               'self_repeat_rate': None,   # stamped below when index built same run
               'papers_excluded_from_style': []}
    path = _artefact_path(out_dir, f'{exam_code}_style_profile.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(profile, f, ensure_ascii=False, indent=1, sort_keys=True)
    return path, profile


def write_pyq_index(exam_code, all_questions, paper_file_hashes, subtopic_of=None,
                    out_dir=None):
    """§6.3 — <EXAM>_pyq_index.json. No question text; shingle hashes, hashes and
    tuples only."""
    recs, n_unique = build_pyq_index_questions(all_questions, subtopic_of)
    # self-repeat rate: unique stems that recur across different papers
    by_md5 = {}
    for r in recs:
        by_md5.setdefault(r['stem_md5'], set()).add(r['pyq_id'].split(':')[0])
    n_repeat = sum(1 for papers in by_md5.values() if len(papers) >= 2)
    self_repeat = round(n_repeat / max(len(by_md5), 1), 4)
    # §6.3 names these n_questions / n_unique_stems / generated_at. The legacy
    # 'questions'/'unique_stems' keys are KEPT ALONGSIDE (additive, P-9) so any
    # reader written against the shipped artefact keeps working; the schema
    # names are what EC-1 (empty index => n_questions: 0) and EC-25 (the newer
    # generated_at wins when two profiles are otherwise equal) are stated in.
    # Without generated_at, EC-25 is not implementable at all.
    index = {'_meta': {'schema': bc.PYQ_INDEX_SCHEMA, 'exam_code': exam_code,
                       'corpus_hash': compute_corpus_hash(paper_file_hashes),
                       'n_questions': len(recs), 'n_unique_stems': n_unique,
                       'questions': len(recs), 'unique_stems': n_unique,
                       'generated_at': _utc_stamp(),
                       'normaliser': 'v1', 'shingle_k': bc.SHINGLE_K,
                       'shingle_k_short': bc.SHINGLE_K_SHORT,
                       'self_repeat_rate': self_repeat,
                       'generated_by': 'analyse_engine v2.56'},
             'questions': recs}
    path = _artefact_path(out_dir, f'{exam_code}_pyq_index.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=1, sort_keys=True)
    return path, index


def stamp_medium(questions):
    """§6.1.9 — per-question medium stamp; the style layer computes on 'en' only."""
    for q in questions:
        if 'medium' not in q:
            q['medium'] = medium_of((q.get('stem') or '') + ' ' +
                                    ' '.join(q.get('options') or []))
    return questions


# ── pattern annotation (§6.1.8 additive per-pattern fields) ─────────────────────

def _annotate_patterns_with_style(patterns, questions):
    """v2.56 — stamp each legacy 0.90-cluster pattern with its measured style
    abstraction: majority mechanic / polarity / option_shape of the member
    questions and the composed pattern_key. Membership is decided by the SAME
    0.90 skeleton-similarity rule that formed the cluster, so the annotation can
    never re-cluster anything (byte-identity of the pre-existing fields, EC-26)."""
    if not patterns:
        return patterns
    for p in patterns:
        members = [q for q in questions
                   if q.get('skeleton')
                   and SequenceMatcher(None, p.get('template', ''),
                                       q['skeleton']).ratio() >= 0.90]
        if not members:
            members = [q for q in questions if q.get('stem')]
        en = [q for q in members if q.get('medium', 'en') == 'en']
        mech = Counter(q.get('mechanic', bc.MECHANIC_UNKNOWN) for q in en)
        pol = Counter('negative' if _POLARITY_RE.search(q.get('stem') or '')
                      else 'positive' for q in members)
        shp = Counter(q.get('option_shape', 'none') for q in members)
        at = Counter(('nat' if not q.get('options') else
                      'msq' if q.get('is_msq') else 'mcq') for q in members)
        form = Counter(q.get('axis2', 'DIRECT') for q in members)
        p['mechanic'] = mech.most_common(1)[0][0] if mech else bc.MECHANIC_UNKNOWN
        p['polarity'] = pol.most_common(1)[0][0] if pol else 'positive'
        p['option_shape'] = shp.most_common(1)[0][0] if shp else 'none'
        p['pattern_key'] = '|'.join((
            p['mechanic'],
            form.most_common(1)[0][0] if form else 'DIRECT',
            p['polarity'],
            at.most_common(1)[0][0] if at else 'mcq',
            p['option_shape']))
    return patterns


if __name__ == "__main__":
    import sys
    if "--self-test" in sys.argv:
        sys.exit(0 if self_test() else 1)
    print("analyse_engine.py — Step 5 shared extraction engine. Run with --self-test.")
