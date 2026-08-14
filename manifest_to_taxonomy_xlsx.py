#!/usr/bin/env python3
"""
manifest_to_taxonomy_xlsx.py — turn a Step-5 subtopic_manifest.json into a human-readable
Excel you can browse to pick Step-6 scope values (no JSON reading required).

Usage:
    python3 manifest_to_taxonomy_xlsx.py  <ExamCode>_subtopic_manifest.json  [output.xlsx]

Output sheet "Taxonomy": Subject | Topic | Sub Topic Name | Sub Topic Id | Upload Order
(one row per sub-topic, in MANIFEST ORDER — v2: GAP-2026-08-14-TAXONOMY-ORDER).
Sheet "How to use" maps each column to the Step-6 trigger and states the portal
upload sequence.

ROW ORDER IS THE TEACHING ORDER, NEVER RE-SORTED. v1 sorted alphabetically
"for readability", which silently discarded the manifest's curated order —
the order notes_core.assign_numbering freezes into the permanent S/T/ST unit
numbers and the order the portal upload must follow. This converter and the
spec-inline writer (Framework_MockTestAnalyse S5-3 write_taxonomy_xlsx) are
changed identically in the same release — two writers, one behaviour.
Columns A-D keep their positions and meanings; Upload Order is column E.
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER = ['Subject', 'Topic', 'Sub Topic Name', 'Sub Topic Id', 'Upload Order']


def build_rows(manifest):
    """Flatten manifest['subtopics'] in MANIFEST ORDER (insertion order == the
    taxonomy/teaching order that assign_numbering freezes into unit numbers).
    NEVER re-sort: an alphabetical resort shows the operator a different order
    from the one every delivered filename already carries."""
    subs = manifest.get('subtopics', {})
    rows = []
    for i, (sid, v) in enumerate(subs.items(), start=1):
        rows.append([v.get('section', ''), v.get('topic', ''),
                     v.get('display_name', ''), sid, i])
    return rows


def write_xlsx(manifest, out_path):
    exam = manifest.get('exam_code', 'Exam')
    rows = build_rows(manifest)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Taxonomy'

    arial = Font(name='Arial', size=11)
    head_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    head_fill = PatternFill('solid', fgColor='1F4E79')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header
    for c, name in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border

    # data — zebra striping so a long list stays readable; light band per alternate SUBJECT block
    band = PatternFill('solid', fgColor='F2F6FB')
    prev_subject = None
    shade = False
    for i, row in enumerate(rows, start=2):
        if row[0] != prev_subject:
            shade = not shade
            prev_subject = row[0]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = arial
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            if shade:
                cell.fill = band

    # widths (roomy for long names/ids)
    for c, width in enumerate([26, 30, 46, 30, 14], start=1):
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = 'A2'                     # keep the header visible while scrolling
    ws.auto_filter.ref = f'A1:E{len(rows) + 1}'  # click-to-filter by Subject / Topic

    # ── "How to use" sheet — plain instructions for a non-technical reader ──
    hw = wb.create_sheet('How to use')
    guide = [
        (f'{exam} — how to use this list in Step 6', True),
        ('', False),
        ('This sheet lists every Subject, Topic and Sub-Topic in your exam.', False),
        ('Use the "Taxonomy" tab to pick what to test, then copy the value into the', False),
        ('Step-6 trigger. Match the text EXACTLY (spelling and capital letters matter).', False),
        ('', False),
        ('PORTAL UPLOAD ORDER — the rows are already in the correct teaching sequence', True),
        ('(basics first). Upload sub-topics to the portal from TOP TO BOTTOM, exactly as', False),
        ('listed; the "Upload Order" column (E) numbers that sequence. Do NOT re-sort', False),
        ('this sheet before uploading — the row order matches the numbering printed in', False),
        ('every delivered file.', False),
        ('', False),
        ('SUBJECT test  → copy the Subject cell (column A):', True),
        ('    ScopedBlueprint --level subject  --scope "<Subject>"            --count N --qs_per_paper Q', False),
        ('', False),
        ('TOPIC test    → join Subject and Topic with "::" (columns A and B):', True),
        ('    ScopedBlueprint --level topic    --scope "<Subject>::<Topic>"   --count N --qs_per_paper Q', False),
        ('', False),
        ('SUB-TOPIC test → join Subject + Topic + Sub Topic Name with "::" (columns A, B, C):', True),
        ('    ScopedBlueprint --level subtopic --scope "<Subject>::<Topic>::<Sub Topic Name>" --count N --qs_per_paper Q', False),
        ('    The Topic in the middle keeps same-named sub-topics apart — e.g. "Kinematics"', False),
        ('    under Mechanics vs under Rotational Motion resolve to two DIFFERENT sub-topics.', False),
        ('    Use the Sub Topic Id (column D) ONLY if the same name appears twice under the', False),
        ('    same Topic:   --scope <Sub Topic Id>', False),
        ('', False),
        ('Tip: use the little filter arrows on the Taxonomy header to narrow by Subject or Topic.', False),
    ]
    for i, (text, bold) in enumerate(guide, start=1):
        cell = hw.cell(row=i, column=1, value=text)
        cell.font = Font(name='Arial', size=11, bold=bold)
    hw.column_dimensions['A'].width = 100

    wb.save(out_path)
    return len(rows), out_path


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    manifest = json.load(open(sys.argv[1], encoding='utf-8'))
    exam = manifest.get('exam_code', 'Exam')
    out = sys.argv[2] if len(sys.argv) > 2 else f'{exam}_taxonomy.xlsx'
    n, path = write_xlsx(manifest, out)
    print(f'Wrote {n} sub-topics to {path}')


if __name__ == '__main__':
    main()
