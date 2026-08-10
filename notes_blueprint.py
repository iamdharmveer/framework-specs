"""
notes_blueprint.py v2.0 — Engine for Notes Step NB (Framework_NotesBlueprint).

v2.0 — 2026-08-10 — TAXONOMY CONSUMER (Framework_NotesBlueprint v3.0.0). The
    Step-5 [ExamCode]_subtopic_manifest.json is the single source of truth for
    unit identity; NB never mints an id (Mock-pipeline Cross-Step Subtopic
    Contract RULES 1/2/2a applied to Notes verbatim). New: build_blueprint_v2()
    consumes the loaded manifest — every unit row references a manifest sid
    (HARD STOP ValueError on any sid absent from the manifest, including the
    evidence-added path: an evidence row's questions were sorted against this
    same taxonomy, so a resolution failure is a real upstream defect, never a
    mint-it-here case); names/section/topic are the manifest's exact bytes;
    S/T/ST numbering comes from notes_core.assign_numbering (manifest row
    order, prior assignment preserved verbatim — persisted numbering); the
    F-1 slug is notes_core.sid_slug(sid); the blueprint carries taxonomy_ref
    (notes_core.taxonomy_ref_for) beside bank_ref. write_blueprint_and_registry
    passes taxonomy_ref into the v2 registry (sid-keyed). build_blueprint (the
    v1 positional builder) is retained UNCHANGED for legacy reads/migration
    only — the NB spec no longer calls it. All v1.4 self-tests retained
    verbatim; v2.0 adds its own.

v1.4 — 2026-08-10 — DEPLOYMENT-REVIEW FIX 1 (bank_ref). build_blueprint now
    accepts and EMITS a bank_ref, and bank_ref_for(bank_path) produces it
    ({path, sha256, questions, generated}). The blueprint that NC/NA read is now
    provably tied to the exact bank NB built, so NC's stale-bank stop
    (notes_core.verify_bank_ref) has evidence to fire. Nothing else in v1.3
    changed.

v1.3 — 2026-08-10 — INGEST BASE (Framework_NotesBlueprint v2.0.0). NB now
    performs the eager full-corpus ingest and owns the PYQ bank. Added:
    assemble_bank() (validates + wraps notes_core.bank_*), write_bank(), and
    counts_from_bank() (delegates to notes_core.derive_taxonomy_counts so the
    subtopic pyq_count / recent3_count fed into build_blueprint come from the
    ingested corpus, not a separate Analysis doc — owner decision 5i). The
    Drive enumeration/download/image work itself is done by corpus_io with
    Claude-supplied Drive-MCP resolvers (spec §3A); this engine assembles and
    validates the result. build_blueprint / read_exam_pattern /
    extract_allowed_types / resolve_sources are unchanged; all v1.2 self-tests
    are retained verbatim.

v1.2 — 2026-08-08 — DEPLOYMENT-REVIEW FIXES. The EVIDENCE_ADDED path no
    longer crashes: a folded-in analysis-only subtopic requires explicit
    placement (s_no/t_no/st_no) supplied by the blueprint author, validated
    with a clear error naming the unit instead of a KeyError; its slug
    defaults from its name; and it joins its topic's seq_in_topic counter so
    outline numbering stays gapless. Excluded rows still need no placement.
    self_test() added per CLAUDE.md engine rule with a fixture that fails on
    the rectified defect.

v1.1 — 2026-08-08 — REFINEMENT SUPPORT. extract_allowed_types() reads the
    Range tab Type column through notes_core.normalize_types (HARD STOP on an
    empty set); build_blueprint() accepts and emits allowed_question_types,
    computes per-unit seq_in_topic from list order, and passes through
    optional prose_ban_exemptions; the registry is initialised with the
    allowed type set.

v1.0 — 2026-08-08 — INITIAL RELEASE. Exam Pattern xlsx reader (Overview /
    Sections / Range / optional Sources tabs), Option-B evidence-expansion
    filter, blueprint writer, exclusion report. Syllabus/PYQ-Analysis parsing
    is Claude-driven per spec §2; this engine validates and assembles.
"""
import json, os
from datetime import datetime, timezone
import notes_core


def _now():
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def read_exam_pattern(xlsx_path):
    """Returns dict with overview / sections / ranges / sources. HARD STOP if
    the Overview Level field is absent (spec §2 S-2)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    out = {"overview": {}, "sections": [], "ranges": [], "sources": []}
    for row in wb["Overview"].iter_rows(values_only=True):
        if row and row[0] is not None:
            out["overview"][str(row[0]).strip()] = row[1]
    if not any(k.lower().startswith("level") for k in out["overview"]):
        raise SystemExit("HARD STOP: Exam Pattern Overview has no Level field")
    for name, key in (("Sections", "sections"), ("Range", "ranges")):
        if name in wb.sheetnames:
            rows = list(wb[name].iter_rows(values_only=True))
            hdr = [str(h).strip() for h in rows[0]]
            out[key] = [dict(zip(hdr, r)) for r in rows[1:] if any(r)]
    if "Sources" in wb.sheetnames:
        rows = list(wb["Sources"].iter_rows(values_only=True))
        hdr = [str(h).strip().lower() for h in rows[0]]
        out["sources"] = [dict(zip(hdr, r)) for r in rows[1:] if any(r)]
    return out


def extract_allowed_types(pattern):
    """Ordered unique canonical types from the Range tab. HARD STOP if empty
    (Framework_NotesBlueprint §6 O-1)."""
    col = [r.get("Type") for r in pattern.get("ranges", [])]
    types = notes_core.normalize_types(col)
    if not types:
        raise SystemExit("HARD STOP: no recognisable question types in the "
                         "Exam Pattern Range tab")
    return types


def resolve_sources(chat_link, pattern_sources, project_files):
    """Spec §3 priority: chat link > Sources tab > project Files."""
    if chat_link:
        return {"mode": "chat", "entries": [{"label": "chat", "url": chat_link}]}
    if pattern_sources:
        return {"mode": "xlsx-sources-tab", "entries": pattern_sources}
    docx = [f for f in project_files if f.lower().endswith(".docx")]
    return {"mode": "project-files", "entries": [{"label": f} for f in docx]}


def build_blueprint(exam_code, level, syllabus_hash, sources, unit_rows,
                    analysis_only_rows, allowed_question_types=None,
                    bank_ref=None):
    """unit_rows: in-syllabus rows {s_no,t_no,st_no,name,slug,pyq_count,
    is_bridge,bridge_reason?,provenance}. analysis_only_rows: PYQ-analysis
    subtopics absent from syllabus {name,pyq_count,recent3_count}; a row
    that folds in (Option B satisfied) MUST also carry its placement
    {s_no,t_no,st_no} and may carry slug and prose_ban_exemptions.

    bank_ref (fix 1): {path, sha256, questions, generated} tying this blueprint
    to the exact bank it was built from; emitted so NC/NA can detect a stale
    pairing (notes_core.verify_bank_ref)."""
    units, excluded = [], []
    _topic_seq = {}
    for r in unit_rows:
        key = (r["s_no"], r["t_no"])
        _topic_seq[key] = _topic_seq.get(key, 0) + 1
        r = dict(r, _seq=_topic_seq[key])
        role = notes_core.assign_role(True, r["pyq_count"], r.get("is_bridge", False), 0)
        units.append({
            "unit_code": notes_core.unit_code(exam_code, r["s_no"], r["t_no"], r["st_no"]),
            "name": r["name"], "slug": r["slug"], "role": role,
            "tier": notes_core.assign_tier(role, r["pyq_count"]),
            "pyq_count": r["pyq_count"],
            "provenance": r.get("provenance", "syllabus"),
            "bridge_reason": r.get("bridge_reason"),
            "seq_in_topic": r["_seq"],
            "prose_ban_exemptions": r.get("prose_ban_exemptions", []),
        })
    for r in analysis_only_rows:
        role = notes_core.assign_role(False, r["pyq_count"], False, r["recent3_count"])
        if role is None:
            excluded.append({"name": r["name"], "pyq_count": r["pyq_count"],
                             "recent3_count": r["recent3_count"],
                             "reason": "out-of-syllabus, <2 PYQs in latest 3 years"})
        else:
            missing = [k for k in ("s_no", "t_no", "st_no") if r.get(k) is None]
            if missing:
                raise ValueError(
                    "EVIDENCE_ADDED unit %r folds into the blueprint and "
                    "needs placement: missing %s" % (r["name"], ", ".join(missing)))
            key = (r["s_no"], r["t_no"])
            _topic_seq[key] = _topic_seq.get(key, 0) + 1
            units.append({
                "unit_code": notes_core.unit_code(exam_code, r["s_no"], r["t_no"], r["st_no"]),
                "name": r["name"], "slug": r.get("slug") or r["name"], "role": role,
                "tier": notes_core.assign_tier(role, r["pyq_count"]),
                "pyq_count": r["pyq_count"], "provenance": "evidence-added",
                "seq_in_topic": _topic_seq[key],
                "prose_ban_exemptions": r.get("prose_ban_exemptions", []),
            })
    return {
        "schema": notes_core.BLUEPRINT_SCHEMA,
        "exam_code": exam_code, "level": level,
        "allowed_question_types": allowed_question_types or [],
        "bank_ref": bank_ref,
        "syllabus_sha256": syllabus_hash, "generated": _now(),
        "sources": sources, "units": units, "excluded": excluded,
    }


def build_blueprint_v2(exam_code, level, syllabus_hash, sources, manifest,
                       in_syllabus_rows, evidence_rows,
                       allowed_question_types=None, bank_ref=None,
                       taxonomy_ref=None, prior_numbering=None):
    """v2.0 taxonomy-consumer builder (Framework_NotesBlueprint v3.0.0 §1A).

    manifest: the LOADED Step-5 subtopic manifest
    (notes_core.load_subtopic_manifest — exam_code already verified there).
    in_syllabus_rows: [{sid, pyq_count, is_bridge?, bridge_reason?,
    prose_ban_exemptions?}] — the manifest sids the syllabus scope-matched IN.
    evidence_rows: [{sid, pyq_count, recent3_count, prose_ban_exemptions?}] —
    out-of-syllabus sids present in the bank (Option-B test decides fold-in vs
    excluded; an excluded row is reported with its manifest names).

    HARD STOP (ValueError) on any row whose sid is not in the manifest — NB
    never mints, never falls back, never display-name-matches here (resolution
    to a sid happened during scope matching; this builder consumes sids only).
    Names/section/topic are copied from the manifest VERBATIM. Numbering is
    notes_core.assign_numbering(manifest, prior_numbering) — prior assignments
    preserved so filenames/title numbers never churn. seq_in_topic == st_no
    (numbering is within (section, topic) by construction)."""
    subs = manifest["subtopics"]
    for r in list(in_syllabus_rows) + list(evidence_rows):
        if r["sid"] not in subs:
            raise ValueError(
                "HARD STOP: sid %r is not in the subtopic manifest. Notes "
                "never mints an id — re-run Step 5 (PYQExtract) so the "
                "manifest includes it, then re-run NB." % r["sid"])
    numbering = notes_core.assign_numbering(manifest, prior_numbering)
    units, excluded = [], []

    def _unit(sid, role, pyq_count, provenance, r):
        v, num = subs[sid], numbering[sid]
        return {
            "sid": sid,
            "unit_code": notes_core.unit_code(
                exam_code, num["s_no"], num["t_no"], num["st_no"]),
            "name": v["display_name"], "section": v["section"],
            "topic": v["topic"], "slug": notes_core.sid_slug(sid),
            "role": role, "tier": notes_core.assign_tier(role, pyq_count),
            "pyq_count": pyq_count, "provenance": provenance,
            "bridge_reason": r.get("bridge_reason"),
            "seq_in_topic": num["st_no"],
            "prose_ban_exemptions": r.get("prose_ban_exemptions", []),
        }

    for r in in_syllabus_rows:
        role = notes_core.assign_role(True, r["pyq_count"],
                                      r.get("is_bridge", False), 0)
        units.append(_unit(r["sid"], role, r["pyq_count"], "syllabus", r))
    for r in evidence_rows:
        role = notes_core.assign_role(False, r["pyq_count"], False,
                                      r["recent3_count"])
        v = subs[r["sid"]]
        if role is None:
            excluded.append({"sid": r["sid"], "name": v["display_name"],
                             "section": v["section"], "topic": v["topic"],
                             "pyq_count": r["pyq_count"],
                             "recent3_count": r["recent3_count"],
                             "reason": "out-of-syllabus, <2 PYQs in latest 3 years"})
        else:
            units.append(_unit(r["sid"], role, r["pyq_count"],
                               "evidence-added", r))
    return {
        "schema": notes_core.BLUEPRINT_SCHEMA,
        "exam_code": exam_code, "level": level,
        "allowed_question_types": allowed_question_types or [],
        "bank_ref": bank_ref, "taxonomy_ref": taxonomy_ref,
        "syllabus_sha256": syllabus_hash, "generated": _now(),
        "sources": sources, "units": units, "excluded": excluded,
    }


def bank_ref_for(bank_path):
    """Fix 1: the {path, sha256, questions, generated} reference embedded in the
    blueprint. sha256 is over the bank bytes on disk, so any later change to
    notes_pyq_bank.json is detectable by notes_core.verify_bank_ref."""
    bank = notes_core.bank_load(bank_path)
    return {"path": bank_path, "sha256": notes_core.file_sha256(bank_path),
            "questions": len(bank.get("questions", [])), "generated": _now()}


def assemble_bank(exam_code, papers, questions):
    """Build + validate the PYQ bank from ingested material (spec §3A/§3B).

    papers:    [{paper_key, exam_date, exam_year, name, n_questions,
                 image_report?}]  — one per sorted paper successfully ingested.
    questions: [rec, ...]         — see notes_core.bank_add_question; each rec
                 carries verbatim correct_answer + explanation and the
                 stem_figures / solution_figures split (owner decisions 2,3,4).

    Returns the validated bank dict. Raises ValueError on a malformed record or
    duplicate bank_id — a corrupt bank must never reach NC/NA."""
    bank = notes_core.bank_new(exam_code)
    for p in papers:
        notes_core.bank_add_paper(bank, p["paper_key"], p["exam_date"],
                                  p["exam_year"], p.get("name", ""),
                                  p.get("n_questions", 0), p.get("image_report"))
    for rec in questions:
        notes_core.bank_add_question(bank, rec)
    notes_core.bank_validate(bank)
    return bank


def write_bank(bank, out_dir="."):
    """Persist notes_pyq_bank.json (checkpoint + final). Append-only ingest
    writes this after every batch of 3 papers so an interrupted run resumes."""
    return notes_core.bank_save(bank, os.path.join(out_dir, "notes_pyq_bank.json"))


def counts_from_bank(bank, latest_years=3):
    """Subtopic-wise pyq_count + recent3_count derived from the bank — the
    single source of truth for blueprint weighting (owner decision 5i,
    replaces the PYQ Analysis doc). Keyed by notes_core.subtopic_key."""
    return notes_core.derive_taxonomy_counts(bank, latest_years)


def write_blueprint_and_registry(bp, out_dir="."):
    bp_path = os.path.join(out_dir, "notes_blueprint.json")
    with open(bp_path, "w", encoding="utf-8") as f:
        json.dump(bp, f, indent=2, ensure_ascii=False); f.write("\n")
    reg = notes_core.registry_init(bp["exam_code"], bp["syllabus_sha256"],
                                   bp["level"], bp["units"],
                                   taxonomy_ref=bp.get("taxonomy_ref"))
    reg["allowed_question_types"] = bp.get("allowed_question_types") or None
    reg_path = os.path.join(out_dir, "notes_registry.json")
    notes_core.registry_save(reg, reg_path)
    return bp_path, reg_path


# ---------------------------------------------------------------- self-test
def self_test():
    passed, fails = 0, []

    def check(name, cond):
        nonlocal passed
        if cond:
            passed += 1
        else:
            fails.append(name)

    rows = [dict(s_no=1, t_no=2, st_no=8, name="EK", slug="EK", pyq_count=37),
            dict(s_no=1, t_no=2, st_no=9, name="B", slug="B", pyq_count=4)]

    # Defect fixture: EVIDENCE_ADDED path crashed with KeyError (the shipped bug)
    ev = [dict(name="Oscillations", pyq_count=18, recent3_count=3,
               s_no=6, t_no=1, st_no=4)]
    try:
        bp = build_blueprint("EX", "Grad", "h", {"mode": "chat", "entries": []},
                             rows, ev, allowed_question_types=["MCQ"])
        u = [x for x in bp["units"] if x["provenance"] == "evidence-added"]
        check("evidence path builds", len(u) == 1
              and u[0]["unit_code"] == "EX_S6_T1_ST04"
              and u[0]["slug"] == "Oscillations"
              and u[0]["seq_in_topic"] == 1
              and u[0]["role"] == "EVIDENCE_ADDED")
    except KeyError:
        check("evidence path builds", False)

    try:
        build_blueprint("EX", "G", "h", {"mode": "chat", "entries": []}, rows,
                        [dict(name="Orphan", pyq_count=5, recent3_count=2)])
        check("missing placement raises ValueError", False)
    except ValueError as e:
        check("missing placement raises ValueError", "Orphan" in str(e))
    except KeyError:
        check("missing placement raises ValueError", False)

    bp = build_blueprint("EX", "G", "h", {"mode": "chat", "entries": []}, rows,
                         [dict(name="Weak", pyq_count=1, recent3_count=1)])
    check("excluded row needs no placement",
          bp["excluded"] and bp["excluded"][0]["name"] == "Weak")

    bp = build_blueprint("EX", "G", "h", {"mode": "chat", "entries": []}, rows,
                         [dict(name="Joined", pyq_count=6, recent3_count=2,
                               s_no=1, t_no=2, st_no=10)])
    seqs = {u["name"]: u["seq_in_topic"] for u in bp["units"]}
    check("evidence unit continues topic sequence",
          seqs == {"EK": 1, "B": 2, "Joined": 3})

    check("types extracted",
          extract_allowed_types({"ranges": [{"Type": "MCQ"}, {"Type": "NAT"}]})
          == ["MCQ", "NAT"])
    for bad in ([{"Kind": "MCQ"}], []):
        try:
            extract_allowed_types({"ranges": bad})
            check("empty type set hard-stops", False)
        except SystemExit:
            check("empty type set hard-stops", True)

    # ---- v1.3 ingest base -------------------------------------------------
    papers = [dict(paper_key="k26", exam_date="2026-02-15", exam_year=2026,
                   name="EX_15-Feb-2026.docx", n_questions=2),
              dict(paper_key="k13", exam_date="2013-02-10", exam_year=2013,
                   name="EX_10-Feb-2013.docx", n_questions=1)]
    qs = [dict(bank_id="EX-26-1", paper_key="k26", exam_date="2026-02-15",
               exam_year=2026, q_no=1, type="MCQ", subject="Physics",
               topic="Optics", subtopic="Polarization", stem="a",
               correct_answer="2", explanation="AXIOM...",
               stem_figures=["m1.png"], solution_figures=["m2.png"]),
          dict(bank_id="EX-26-2", paper_key="k26", exam_date="2026-02-15",
               exam_year=2026, q_no=2, type="NAT", subject="Physics",
               topic="Optics", subtopic="Polarization", stem="b",
               correct_answer="0.41"),
          dict(bank_id="EX-13-5", paper_key="k13", exam_date="2013-02-10",
               exam_year=2013, q_no=5, type="MSQ", subject="Physics",
               topic="Thermo", subtopic="Carnot", stem="c",
               correct_answer="1,3")]
    bank = assemble_bank("IITJAM_PH", papers, qs)
    check("bank assembles + validates", len(bank["questions"]) == 3
          and len(bank["papers"]) == 2)
    check("verbatim answer + explanation stored",
          bank["questions"][0]["correct_answer"] == "2"
          and bank["questions"][0]["explanation"] == "AXIOM...")
    check("figure split preserved",
          bank["questions"][0]["stem_figures"] == ["m1.png"]
          and bank["questions"][0]["solution_figures"] == ["m2.png"]
          and bank["questions"][0]["figure"] is True)
    counts = counts_from_bank(bank, latest_years=1)
    pol = counts[notes_core.subtopic_key("Physics", "Optics", "Polarization")]
    check("counts_from_bank derives subtopic weighting",
          pol["pyq_count"] == 2 and pol["recent3_count"] == 2)
    try:
        assemble_bank("EX", papers, qs + [dict(bank_id="EX-26-1",
            paper_key="k26", exam_date="2026-02-15", exam_year=2026, q_no=9,
            type="MCQ", subject="s", topic="t", subtopic="u", stem="d")])
        check("assemble rejects duplicate bank_id", False)
    except ValueError:
        check("assemble rejects duplicate bank_id", True)
    import tempfile
    bp_out = tempfile.mkdtemp()
    p = write_bank(bank, bp_out)
    check("bank round-trips", notes_core.bank_load(p)["exam_code"] == "IITJAM_PH")

    # fix 1: bank_ref produced + carried into the blueprint, detects staleness
    ref = bank_ref_for(p)
    check("bank_ref_for yields path+sha+count",
          ref["path"] == p and len(ref["sha256"]) == 64 and ref["questions"] == 3)
    bp2 = build_blueprint("IITJAM_PH", "Graduate", "h",
                          {"mode": "chat", "entries": []},
                          [dict(s_no=1, t_no=1, st_no=1, name="Pol", slug="Pol",
                                pyq_count=2)], [], bank_ref=ref)
    check("blueprint carries bank_ref", bp2["bank_ref"]["sha256"] == ref["sha256"])
    ok_, _ = notes_core.verify_bank_ref(p, bp2["bank_ref"])
    check("fresh blueprint verifies against its bank", ok_ is True)
    with open(p, "a", encoding="utf-8") as _f:
        _f.write(" ")
    check("mutated bank is caught as stale",
          notes_core.verify_bank_ref(p, bp2["bank_ref"])[0] is False)

    # ---- v2.0 taxonomy consumer ------------------------------------------
    man = {"exam_code": "EXBT", "subtopics": {
        "gb.cell.membranes": {"display_name": "Membrane Structure and Function",
                              "section": "General Biology", "topic": "Cell"},
        "gb.cell.signalling": {"display_name": "Cell Signalling",
                               "section": "General Biology", "topic": "Cell"},
        "gb.gen.linkage": {"display_name": "Linkage and Mapping",
                           "section": "General Biology", "topic": "Genetics"},
        "ch.bond.vsepr": {"display_name": "VSEPR Theory",
                          "section": "Chemistry (10+2+3 level)",
                          "topic": "Bonding"}}}
    tref = {"path": "m.json", "sha256": "0" * 64, "subtopics": 4,
            "generated": "g"}
    bpv2 = build_blueprint_v2(
        "EXBT", "Graduate", "h", {"mode": "chat", "entries": []}, man,
        in_syllabus_rows=[dict(sid="gb.cell.membranes", pyq_count=20),
                          dict(sid="gb.cell.signalling", pyq_count=1),
                          dict(sid="gb.gen.linkage", pyq_count=0,
                               is_bridge=True, bridge_reason="prereq")],
        evidence_rows=[dict(sid="ch.bond.vsepr", pyq_count=6, recent3_count=3)],
        allowed_question_types=["MCQ", "NAT"], bank_ref=None,
        taxonomy_ref=tref)
    by_sid = {u["sid"]: u for u in bpv2["units"]}
    check("v2: schema + taxonomy_ref carried",
          bpv2["schema"] == "notes-blueprint/2.0"
          and bpv2["taxonomy_ref"]["sha256"] == "0" * 64)
    check("v2: names/section/topic are manifest bytes; slug from sid",
          by_sid["gb.cell.membranes"]["name"] == "Membrane Structure and Function"
          and by_sid["ch.bond.vsepr"]["section"] == "Chemistry (10+2+3 level)"
          and by_sid["gb.cell.membranes"]["slug"] == "membranes")
    check("v2: numbering from manifest order; seq_in_topic == st_no",
          by_sid["gb.cell.membranes"]["unit_code"] == "EXBT_S1_T1_ST01"
          and by_sid["gb.cell.signalling"]["unit_code"] == "EXBT_S1_T1_ST02"
          and by_sid["gb.gen.linkage"]["unit_code"] == "EXBT_S1_T2_ST01"
          and by_sid["ch.bond.vsepr"]["unit_code"] == "EXBT_S2_T1_ST01"
          and by_sid["gb.cell.signalling"]["seq_in_topic"] == 2)
    check("v2: roles/tiers per rules incl. bridge + evidence",
          by_sid["gb.cell.membranes"]["role"] == "PYQ_WEIGHTED"
          and by_sid["gb.cell.membranes"]["tier"] == "TIER-1"
          and by_sid["gb.cell.signalling"]["role"] == "COVERAGE"
          and by_sid["gb.gen.linkage"]["role"] == "BRIDGE"
          and by_sid["ch.bond.vsepr"]["role"] == "EVIDENCE_ADDED"
          and by_sid["ch.bond.vsepr"]["provenance"] == "evidence-added")
    bpv2x = build_blueprint_v2(
        "EXBT", "G", "h", {"mode": "chat", "entries": []}, man,
        in_syllabus_rows=[],
        evidence_rows=[dict(sid="ch.bond.vsepr", pyq_count=1, recent3_count=1)])
    check("v2: failed Option-B excluded WITH manifest names",
          bpv2x["excluded"][0]["sid"] == "ch.bond.vsepr"
          and bpv2x["excluded"][0]["name"] == "VSEPR Theory")
    try:
        build_blueprint_v2("EXBT", "G", "h", {}, man,
                           [dict(sid="gb.cell.NOT_IN_MANIFEST", pyq_count=3)],
                           [])
        check("v2: unknown sid hard-stops (no self-mint)", False)
    except ValueError as e:
        check("v2: unknown sid hard-stops (no self-mint)",
              "re-run Step 5" in str(e))
    try:
        build_blueprint_v2("EXBT", "G", "h", {}, man, [],
                           [dict(sid="zz.not.there", pyq_count=9,
                                 recent3_count=3)])
        check("v2: unknown EVIDENCE sid hard-stops too (no bypass)", False)
    except ValueError:
        check("v2: unknown EVIDENCE sid hard-stops too (no bypass)", True)
    # persisted numbering: re-blueprint with an inserted subtopic
    man2 = {"exam_code": "EXBT", "subtopics": {}}
    man2["subtopics"]["gb.cell.transport"] = {
        "display_name": "Membrane Transport", "section": "General Biology",
        "topic": "Cell"}
    man2["subtopics"].update(man["subtopics"])
    prior = {u["sid"]: {"s_no": int(u["unit_code"].split("_S")[1].split("_")[0]),
                        "t_no": int(u["unit_code"].split("_T")[1].split("_")[0]),
                        "st_no": int(u["unit_code"].split("_ST")[1])}
             for u in bpv2["units"]}
    bpv3 = build_blueprint_v2(
        "EXBT", "G", "h", {"mode": "chat", "entries": []}, man2,
        in_syllabus_rows=[dict(sid="gb.cell.membranes", pyq_count=20),
                          dict(sid="gb.cell.transport", pyq_count=4)],
        evidence_rows=[], prior_numbering=prior)
    by3 = {u["sid"]: u for u in bpv3["units"]}
    check("v2: prior numbering preserved; inserted sid appends",
          by3["gb.cell.membranes"]["unit_code"] == "EXBT_S1_T1_ST01"
          and by3["gb.cell.transport"]["unit_code"] == "EXBT_S1_T1_ST03")
    # v2 registry write: sid-keyed + taxonomy_ref stored
    reg_dir = tempfile.mkdtemp()
    _, regp = write_blueprint_and_registry(bpv2, reg_dir)
    reg2 = notes_core.registry_load(regp)
    check("v2: registry keyed by sid and carries taxonomy_ref",
          "gb.cell.membranes" in reg2["units"]
          and reg2["units"]["gb.cell.membranes"]["unit_code"] == "EXBT_S1_T1_ST01"
          and reg2["taxonomy_ref"]["sha256"] == "0" * 64)
    check("v2: operator resolution works against the registry units",
          notes_core.resolve_unit(reg2["units"],
                                  "membrane structure & function")["sid"]
          == "gb.cell.membranes")

    print(f"notes_blueprint self-test: {passed} passed, {len(fails)} failed"
          + (" — " + "; ".join(fails) if fails else ""))
    return not fails


if __name__ == "__main__":
    import sys
    if "--self-test" in sys.argv:
        sys.exit(0 if self_test() else 1)
    print("notes_blueprint.py — Notes Step NB engine. Run with --self-test.")
